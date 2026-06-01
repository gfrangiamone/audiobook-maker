#!/usr/bin/env python3
"""
Audiobook Maker  -  Web app to convert EPUB/PDF into MP3 audiobooks.

Requirements:
    pip install flask edge-tts ebooklib beautifulsoup4 lxml Pillow pymupdf

Usage:
    python audiobook_app.py
    Then open http://localhost:5601
"""

import asyncio
import concurrent.futures
import logging
import re
import json
import os
import shutil
import sys
import threading
import time
import uuid
import hmac
import secrets
import html as html_mod
from collections import defaultdict, OrderedDict
from datetime import datetime, timezone
from copy import copy
from pathlib import Path

from flask import (
    Flask, render_template_string, request, jsonify,
    send_file, Response, stream_with_context, redirect
)
from werkzeug.middleware.proxy_fix import ProxyFix

#  -  -  Import epub_to_tts (must be in the same folder)  -  - 
SCRIPT_DIR = Path(__file__).parent.resolve()
sys.path.insert(0, str(SCRIPT_DIR))

# Startup timestamp for Last-Modified / Cache-Control on pre-rendered HTML pages
_STARTUP_TIME = datetime.now(timezone.utc)

try:
    from epub_to_tts import parse_epub
except ImportError:
    print("ERROR: epub_to_tts.py not found in the same folder.", file=sys.stderr)
    print(f"  Script folder: {SCRIPT_DIR}", file=sys.stderr)
    sys.exit(1)

try:
    from pdf_to_tts import parse_pdf
except ImportError:
    parse_pdf = None
    print("WARNING: pdf_to_tts.py not found  -  PDF support disabled.", file=sys.stderr)

#  -  -  Google Cloud TTS (Chirp3-HD)  -  opzionale  -  -
try:
    import google_tts
except ImportError:
    google_tts = None
    print("WARNING: google_tts.py not found  -  Google Cloud TTS disabled.", file=sys.stderr)

#  -  -  Gemini TTS (Flash 2.5 / 3.1)  -  opzionale  -  -
try:
    import gemini_tts
except ImportError:
    gemini_tts = None
    print("[startup] gemini_tts module not available (google-genai not installed)")

from audio_utils import (
    _zip_safe_read, _extract_cover_from_epub, _generate_fallback_cover,
    _extract_cover_for_preview, _include_cover_in_dir, _generate_podcast_rss,
    _generate_silence_mp3, _concatenate_mp3, _get_audio_duration_ms,
    _convert_mp3_to_m4b, _prepare_m4b_cover_path, _safe_filename,
    _check_audio_dependencies, pcm_to_mp3,
)
from tts_split import (
    CHUNK_MAX_CHARS, split_text_into_chunks, _is_multilingual_voice,
    _TTS_MIN_SENT_CHARS, _TTS_MAX_SENT_CHARS, _split_sentences_for_tts,
    _edge_tts_call, generate_chunk_mp3, generate_chunk_mp3_google,
    _strip_parenthetical, _ensure_heading_pause, _plan_chunks,
    _pick_chunk_max_chars, _pick_chunk_max_bytes,
)

import email_service
import payment
import generation_engine
import community_store
import community_translator
import community_moderator

# Carica traduzioni pagine di download da file JSON esterno
_DL_PAGES_I18N = {}
try:
    with open(SCRIPT_DIR / "i18n" / "download_pages.json", encoding="utf-8") as _f:
        _DL_PAGES_I18N = json.load(_f)
except Exception as _e:
    print(f"WARNING: Could not load i18n/download_pages.json: {_e}", file=sys.stderr)

#  -  -  LLM per ottimizzazione testo TTS  -  opzionale  -  -
# (Configurati e gestiti in generation_engine.py; lette qui solo per startup log)
LLM_API_KEY = os.environ.get("ABM_LLM_API_KEY", "")
LLM_MODEL = os.environ.get("ABM_LLM_MODEL", "deepseek-chat")
LLM_THINKING = os.environ.get("ABM_LLM_THINKING", "false").lower() == "true"
LLM_REASONING_EFFORT = os.environ.get("ABM_LLM_REASONING_EFFORT", "none").lower()

def _llm_available():
    """True se l'ottimizzazione LLM è disponibile."""
    return generation_engine._llm_available()


def _estimate_chapter_seconds(ch, language):
    """Stima durata audio di un capitolo in secondi.

    Usa la stessa funzione di stima del pannello "Voci PREMIUM"
    (gemini_tts.estimate_audio_seconds con rate empirico) per allineare
    il dato mostrato nel pannello selezione capitoli con quello del
    pannello Premium. Fallback a 150 WPM se gemini_tts non disponibile.
    """
    try:
        if gemini_tts is not None:
            secs = gemini_tts.estimate_audio_seconds(
                getattr(ch, "text", "") or "",
                language=language,
                model_key="flash25",
                rate_pct=0,
            )
            if secs and secs > 0:
                return float(secs)
    except Exception:
        pass
    # Fallback: 150 WPM (storica)
    return (getattr(ch, "word_count", 0) or 0) * 60.0 / 150.0


# Payment / voucher state and operations live in payment.py.
# Re-exported names below keep the rest of audiobook_app.py working unchanged.
# Mutable state dicts (payment._payments, payment._vouchers, payment._paid_opt_done) are accessed via
# `payment.<name>` because _load_*() rebinds the module globals at startup.
from payment import (
    PAYPAL_CLIENT_ID, PAYPAL_SECRET, PAYPAL_MODE, PAYPAL_API_BASE,
    LLM_RATE_EUR_PER_MCHAR, LLM_FREE_THRESHOLD_EUR,
    VOUCHER_EXPIRY_DAYS, VOUCHER_BONUS_PERCENT, PAYMENT_RETENTION_DAYS,
    VOUCHER_RL_PER_MIN, VOUCHER_RL_PER_HOUR,
    VOUCHER_EMAIL_FAIL_LIMIT, VOUCHER_EMAIL_LOCKOUT_SEC,
    _paypal_available, _estimate_llm_cost_eur,
    _paypal_get_access_token, _paypal_create_order, _paypal_capture_order,
    _voucher_rl_check, _voucher_rl_record_result,
    _save_payments, _load_payments,
    _save_vouchers, _load_vouchers,
    _generate_voucher_code, _create_voucher,
    _voucher_remaining, _voucher_consume, _voucher_refund,
    _save_paid_opt_done, _load_paid_opt_done,
    _mark_paid_opt_done, _cleanup_paid_opt_done,
)


#  -  -  Import version and template builder  -  -
from version import __version__, get_formatted_date
from templates.index_page import build_html_template
from guide_content import build_guide_html
import seo_reviews
import seo_content

#  -  -  Import favicon data (embedded, served via Flask routes for SEO)  -  - 
from favicon_data import (
    get_favicon_ico, get_favicon_png_192,
    get_apple_touch_icon, get_favicon_svg,
)
from og_image_data import get_og_image



# ----------------------------------------------------------------------
# LOGGING CONFIG
# ----------------------------------------------------------------------

class HeartbeatFilter(logging.Filter):
    def filter(self, record):
        msg = record.getMessage()
        return "/api/heartbeat" not in msg and "/api/job_status/" not in msg

logging.getLogger("werkzeug").addFilter(HeartbeatFilter())

# ----------------------------------------------------------------------
# APP CONFIG
# ----------------------------------------------------------------------

app = Flask(__name__)
# Reverse proxy (nginx) sits in front: trust one hop of X-Forwarded-* so request.remote_addr
# reflects the real client IP instead of 127.0.0.1 (needed for logs, rate limiting, fail2ban).
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)
app.config["MAX_CONTENT_LENGTH"] = int(os.environ.get("ABM_MAX_UPLOAD_MB", "50")) * 1024 * 1024
# Static assets are cache-busted via ?v=__APP_VERSION__ so a 1-year max-age is safe.
app.config["SEND_FILE_MAX_AGE_DEFAULT"] = 31536000  # 1 year

# Endpoint esenti da CSRF check (whitelist esplicita).
# Da aggiungere SOLO endpoint che ricevono webhook server-to-server (es. PayPal
# webhook firmato): attualmente nessuno.
_CSRF_EXEMPT_PATHS: set[str] = set()


@app.before_request
def _csrf_protect():
    """CSRF protection: verifica Origin/Referer su metodi mutating.

    - GET/HEAD/OPTIONS: nessun check (operazioni read-only).
    - POST/PUT/PATCH/DELETE: se ``Origin`` presente, deve matchare ``host_url``;
      altrimenti se ``Referer`` presente, stesso check. Se entrambi assenti
      (client non-browser come curl/script), passa.
    - I cookie ``SameSite=Strict`` (admin) e ``SameSite=Lax`` (abm_cid) gia'
      offrono difesa parziale; questo check chiude il gap residuo per browser
      vecchi e per ``SameSite=Lax`` su navigazioni top-level.
    """
    if request.method not in ("POST", "PUT", "PATCH", "DELETE"):
        return None
    if request.path in _CSRF_EXEMPT_PATHS:
        return None
    origin = request.headers.get("Origin", "")
    referer = request.headers.get("Referer", "")
    expected = request.host_url.rstrip("/")
    if origin:
        if not (origin == expected or origin.startswith(expected + "/")):
            return jsonify({"error": "CSRF: origin mismatch"}), 403
    elif referer:
        if not referer.startswith(expected + "/") and referer != expected:
            return jsonify({"error": "CSRF: referer mismatch"}), 403
    return None


@app.after_request
def add_security_headers(response):
    """Aggiunge header di sicurezza alle risposte HTTP."""
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    # HSTS: forza HTTPS sui browser per 1 anno (incl. subdomain).
    # Nginx in produzione probabilmente lo aggiunge gia'; lo settiamo qui per
    # defense-in-depth ed evitare regressioni se la config nginx cambia.
    # Solo su HTTPS per non bloccare dev locale HTTP.
    if request.is_secure:
        response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    # Content Security Policy (base)
    # Permettiamo script inline per la nostra app (SPA-like) ma blocchiamo fonti esterne non autorizzate.
    # Nota: per una configurazione più rigida, bisognerebbe usare i nonce.
    response.headers['Content-Security-Policy'] = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' https://www.paypal.com https://www.googletagmanager.com; "
        "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; "
        "font-src 'self' https://fonts.gstatic.com; "
        "img-src 'self' data: https://api.producthunt.com; "
        "media-src 'self' blob:; "
        "connect-src 'self' https://api-m.sandbox.paypal.com https://api-m.paypal.com https://*.google-analytics.com; "
        "frame-src https://www.paypal.com;"
    )
    # Cache-Control:
    #  - HTML: 1h cache + 1d stale-while-revalidate (lets CDN serve stale while
    #    refreshing in the background; reduces TTFB on Google crawls)
    #  - sitemap.xml / robots.txt / llms.txt: 1h cache (short — they may change
    #    when guides/translations are added)
    ct = response.content_type or ''
    path = (request.path or '') if request else ''
    if 'Cache-Control' not in response.headers:
        # Admin/API non devono mai essere cachati (cambi di stato real-time).
        if path.startswith('/admin') or path.startswith('/api/'):
            response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
            response.headers['Pragma'] = 'no-cache'
        elif 'text/html' in ct:
            # 5-min cache + 1h stale-while-revalidate: lets new approved
            # reviews surface in the embedded JSON-LD/visible block within a
            # few minutes while keeping repeat-visit perf high.
            response.headers['Cache-Control'] = (
                'public, max-age=300, stale-while-revalidate=3600'
            )
            response.headers['Last-Modified'] = _STARTUP_TIME.strftime('%a, %d %b %Y %H:%M:%S GMT')
        elif path in ('/sitemap.xml', '/robots.txt', '/llms.txt'):
            response.headers['Cache-Control'] = 'public, max-age=3600'
    return response

# Directory di lavoro persistente (sopravvive ai restart del servizio)
# Configurabile via ABM_DATA_DIR, default: /var/lib/audiobook-maker/data
_DATA_DIR = os.environ.get("ABM_DATA_DIR", "/var/lib/audiobook-maker/data")
UPLOAD_DIR = Path(_DATA_DIR)
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

# Inizializza Google Cloud TTS (tracking utilizzo nella data dir)
if google_tts is not None:
    google_tts.init(_DATA_DIR)
    # Forza l'invalidazione della cache voci locale per includere Google all'avvio
    _voices_cache = None

# Inizializza Gemini TTS (Flash 2.5/3.1)
if gemini_tts is not None:
    try:
        gemini_tts.init(_DATA_DIR)
        if gemini_tts.is_available():
            print("[startup] Gemini TTS enabled")
        else:
            print("[startup] Gemini TTS initialized but disabled (ABM_GEMINI_API_KEY not set)")
    except Exception as e:
        print(f"[startup] Gemini TTS init failed: {e}")
        gemini_tts = None

# Inizializza JSON store community (news, feedback)
community_store.init(_DATA_DIR)

jobs = {}
_jobs_lock = threading.Lock()  # Protects all reads/writes of `jobs` dict


def _has_active_google_tts_jobs():
    """True se c'è almeno un job Google TTS in corso (caratteri prenotati ma
    non ancora visibili al Cloud Monitoring). Usato per decidere se è sicuro
    riconciliare al ribasso il contatore locale.
    """
    if google_tts is None:
        return False
    with _jobs_lock:
        try:
            for j in jobs.values():
                if j.get("status") in ("queued", "running", "generating"):
                    if j.get("google_tts_reserved", 0) > 0:
                        return True
                    voice = j.get("voice", "")
                    if voice and google_tts.is_google_voice(voice):
                        return True
        except Exception:
            return True  # safe default
        return False


if google_tts is not None and hasattr(google_tts, "set_active_jobs_callback"):
    google_tts.set_active_jobs_callback(_has_active_google_tts_jobs)

from email_service import (
    _smtp_available, _send_email, _admin_notify_generation,
    _try_send_admin_digest, _send_payment_receipt_email, _send_voucher_email,
    SMTP_HOST, SMTP_PORT, SMTP_USER, SMTP_PASS, SMTP_FROM, ADMIN_EMAIL,
    BASE_URL, ADMIN_DIGEST_INTERVAL_SEC
)

EMAIL_FILE_RETENTION_SEC = int(os.environ.get("ABM_JOB_RETENTION_SEC", "64800"))  # 18h default
# Override per job con voce PREMIUM (Gemini): retention piu' lunga perche'
# i pagamenti Premium meritano una finestra di download/email piu' ampia.
GEMINI_FILE_RETENTION_SEC = int(os.environ.get("ABM_GEMINI_JOB_RETENTION_SEC", "172800"))  # 48h default
# Finestra "calda" locale: dopo questo tempo dal completamento, i file output
# vengono evacuati dal disco locale e serviti via redirect dal cold storage S3.
# La retention TOTALE (disponibilità al download) resta governata da
# EMAIL_FILE_RETENTION_SEC / GEMINI_FILE_RETENTION_SEC (cold delete su S3).
HOT_WINDOW_SEC = int(os.environ.get("ABM_HOT_WINDOW_SEC", "7200"))            # 2h voci standard
HOT_WINDOW_GEMINI_SEC = int(os.environ.get("ABM_HOT_WINDOW_GEMINI_SEC", "14400"))  # 4h voci PREMIUM
# Hard cap caratteri per audiolibro completo (taglia output audio):
# - standard (edge-tts/Google): ABM_MAX_TEXT_CHARS
# - PREMIUM (gemini:): ABM_MAX_GEMINI_TEXT_CHARS, tipicamente piu' basso perche'
#   le voci Gemini hanno cost-per-char piu' alto e RPM/RPD piu' restrittive.
MAX_TEXT_CHARS = int(os.environ.get("ABM_MAX_TEXT_CHARS", "1500000"))
MAX_GEMINI_TEXT_CHARS = int(os.environ.get("ABM_MAX_GEMINI_TEXT_CHARS", "800000"))

# Tolleranza di crescita del testo dovuta all'ottimizzazione AI. Un libro che
# era ENTRO il cap prima dell'ottimizzazione (precondizione garantita dal cap
# enforced in /api/optimize e /api/optimize_estimate sul testo originale) puo'
# superare il cap fino a questa frazione DOPO l'espansione LLM ed essere
# comunque generato, invece di essere rifiutato a valle. Default 5%.
try:
    LLM_OPT_GROWTH_TOLERANCE = float(
        os.environ.get("ABM_LLM_OPT_GROWTH_TOLERANCE", "0.05").replace(",", ".")
    )
except (TypeError, ValueError):
    LLM_OPT_GROWTH_TOLERANCE = 0.05
LLM_OPT_GROWTH_TOLERANCE = max(0.0, LLM_OPT_GROWTH_TOLERANCE)


def _is_gemini_voice(voice):
    """True se la voce e' una voce PREMIUM Gemini (formato gemini:<model>:<voice>)."""
    return bool(voice) and isinstance(voice, str) and voice.startswith("gemini:")


def _max_text_chars_for_voice(voice):
    """Cap caratteri appropriato per la voce: Gemini -> MAX_GEMINI_TEXT_CHARS, altrimenti MAX_TEXT_CHARS."""
    return MAX_GEMINI_TEXT_CHARS if _is_gemini_voice(voice) else MAX_TEXT_CHARS


def _effective_max_text_chars(voice, job=None):
    """Cap caratteri EFFETTIVO per la generazione/pagamento.

    Identico a _max_text_chars_for_voice per i job non ottimizzati. Per i job
    gia' ottimizzati con AI (`job["ai_optimized"]`) concede la tolleranza
    LLM_OPT_GROWTH_TOLERANCE (default 5%) sul cap base: un libro che era entro i
    limiti prima dell'ottimizzazione e che l'espansione LLM ha portato di poco
    oltre viene comunque elaborato. NON usare questo helper per il cap PRE-
    ottimizzazione (/api/optimize, /api/optimize_estimate): li' va applicato il
    cap base sul testo originale, che e' la precondizione di questa tolleranza."""
    base = _max_text_chars_for_voice(voice)
    if isinstance(job, dict) and job.get("ai_optimized"):
        return int(base * (1.0 + LLM_OPT_GROWTH_TOLERANCE))
    return base


def _retention_for_job(job):
    """Retention sec applicabile al job: GEMINI_FILE_RETENTION_SEC se voce Gemini, altrimenti EMAIL_FILE_RETENTION_SEC.
    Fallback su `opt_voice` per il flusso optimize-only/batch dove `voice` non e' ancora settato."""
    if not isinstance(job, dict):
        return EMAIL_FILE_RETENTION_SEC
    v = job.get("voice", "") or job.get("opt_voice", "")
    return GEMINI_FILE_RETENTION_SEC if _is_gemini_voice(v) else EMAIL_FILE_RETENTION_SEC


def _retention_for_token_info(info):
    """Retention sec applicabile a un download token: usa is_gemini se salvato sul token."""
    if isinstance(info, dict) and info.get("is_gemini"):
        return GEMINI_FILE_RETENTION_SEC
    return EMAIL_FILE_RETENTION_SEC


# Protezione no-download per voci PREMIUM (costose): se il job/token Gemini
# non ha mai registrato un download, raddoppiamo la retention prima di
# cancellare gli output. Salvaguardia per utenti che ricevono l'email tardi
# o non aprono subito il link.
GEMINI_NO_DOWNLOAD_RETENTION_MULTIPLIER = 2


def _effective_retention_for_job(job):
    """Retention con protezione no-download per job PREMIUM/Gemini.
    Se il job e' Gemini e non risulta alcun download (job["downloaded_at"] vuoto),
    raddoppia la retention base. Per voci standard: identica a _retention_for_job."""
    base = _retention_for_job(job)
    if not isinstance(job, dict):
        return base
    v = job.get("voice", "") or job.get("opt_voice", "")
    if _is_gemini_voice(v) and not job.get("downloaded_at"):
        return base * GEMINI_NO_DOWNLOAD_RETENTION_MULTIPLIER
    return base


def _effective_retention_for_token_info(info):
    """Retention con protezione no-download per token PREMIUM/Gemini.
    Se il token e' is_gemini e nessun /dl/<token>/* ha mai servito il file
    (downloaded_at vuoto), raddoppia la retention base."""
    base = _retention_for_token_info(info)
    if isinstance(info, dict) and info.get("is_gemini") and not info.get("downloaded_at"):
        return base * GEMINI_NO_DOWNLOAD_RETENTION_MULTIPLIER
    return base


def _mark_token_downloaded(token_info):
    """Registra che il file del token e' stato servito (download reale, non
    probe/HEAD/Range). Aggiorna token_info in-place e persiste su disco per
    sopravvivere ai restart, disattivando la protezione no-download
    (_effective_retention_for_token_info). Idempotente: skip su probe/HEAD/
    Range e se gia' marcato."""
    try:
        if _is_resume_or_probe_request():
            return
    except Exception:
        pass
    if not isinstance(token_info, dict):
        return
    if token_info.get("downloaded_at"):
        return
    token_info["downloaded_at"] = time.time()
    try:
        _save_tokens()
    except Exception as e:
        print(f"[tokens] _mark_token_downloaded persist failed: {e}")


def _has_active_download_tokens(job_id, now=None):
    """True se esiste almeno un download token non scaduto per il job_id.

    Protegge job in stato analysed/cancelled dalla rimozione prematura
    della directory quando esistono ancora link email validi.
    I token vengono confrontati con _effective_retention_for_token_info
    (+300s di margine come nel resto del cleanup)."""
    if now is None:
        now = time.time()
    try:
        for _t, info in list(_download_tokens.items()):
            if info.get("job_id") != job_id:
                continue
            if (now - info.get("created_at", 0)) <= _effective_retention_for_token_info(info) + 300:
                return True
    except Exception:
        pass
    return False


#  -  -  Admin activity digest (email log)  -  -
# Set ABM_ADMIN_EMAIL to enable. Leave empty to disable.
#   export ABM_ADMIN_EMAIL=gfrangiamone@gmail.com
# Rate limited: max 1 digest email per hour, batches all pending events.
# Token admin per UI web /admin/vouchers. Se vuoto, l'endpoint è disabilitato.
ADMIN_TOKEN = os.environ.get("ABM_ADMIN_TOKEN", "").strip()

#  -  -  Client tracking & rate limiting  -  - 
# Max concurrent generating jobs per client device (cookie-based).
# Set via ABM_MAX_CONCURRENT_PER_CLIENT env var; default 2.
MAX_CONCURRENT_PER_CLIENT = int(os.environ.get("ABM_MAX_CONCURRENT_PER_CLIENT", "2"))

# Max concurrent LLM optimization jobs per client device.
# Set via ABM_MAX_CONCURRENT_LLM_PER_CLIENT env var; default 1.
MAX_CONCURRENT_LLM_PER_CLIENT = int(os.environ.get("ABM_MAX_CONCURRENT_LLM_PER_CLIENT", "1"))

# Cookie name and max-age for client identification
_CLIENT_COOKIE_NAME = "abm_cid"
_CLIENT_COOKIE_MAX_AGE = 365 * 24 * 60 * 60  # 1 year


def _get_client_id():
    """Return the client_id from cookie, or generate a new one (will be set later)."""
    return request.cookies.get(_CLIENT_COOKIE_NAME, "")


def _new_job_id():
    """Generate a high-entropy job identifier (128 bit, URL-safe).
    Never starts with ``_`` — the ``_`` prefix is the protected system
    namespace (``_payments.json``, ``_vouchers.json``, …) and the cleanup
    loop skips all dirs starting with it.  ``_`` occurs in 1/64
    url-safe-base64 IDs; this loop keeps rolling until we get one
    without it."""
    import secrets as _secrets
    while True:
        jid = _secrets.token_urlsafe(16)
        if not jid.startswith("_"):
            return jid


def _check_job_owner(job_id):
    """Validate that the calling client owns the requested job.

    Returns (job, error_response, status_code). On success error_response is None
    and the caller may use `job`. On failure caller must `return error_response, status_code`.

    Ownership rule: the cookie-stored client_id must match jobs[job_id]['client_id'].
    Jobs predating this enforcement (no client_id stored) are allowed through to preserve
    backward compatibility, but new jobs always store it at creation.

    Admin bypass: una richiesta autenticata come admin (header X-Admin-Token o cookie
    abm_admin_session valido) passa il controllo. Necessario per la pagina /admin/log-activity che
    fa polling di /api/job_status/<job_id> per mostrare la % delle conversioni in corso.
    """
    if job_id not in jobs:
        return None, jsonify({"error": "Job not found"}), 404
    job = jobs[job_id]
    owner = job.get("client_id", "")
    if owner:
        caller = _get_client_id()
        if not caller or caller != owner:
            # Admin bypass: l'admin puo' osservare lo stato di qualunque job.
            if _admin_auth_ok(_admin_auth_from_request()):
                return job, None, 0
            return None, jsonify({"error": "Forbidden"}), 403
    else:
        # Legacy job senza client_id (creato prima dell'enforcement). Lascia
        # passare per compat ma logga warning: l'admin puo' monitorare e
        # decidere di rimuovere il bypass dopo che la coorte legacy e' esaurita.
        print(f"[SECURITY-WARN] Legacy job senza client_id: {job_id} "
              f"(status={job.get('status', '?')}) - bypass ownership check")
    return job, None, 0


def _get_client_ip():
    """Return client IP address, respecting reverse proxy headers."""
    # X-Forwarded-For: client, proxy1, proxy2  →  take the first
    forwarded = request.headers.get("X-Forwarded-For", "")
    if forwarded:
        return forwarded.split(",")[0].strip()
    return request.remote_addr or ""


def _get_browser_lang():
    """Return primary browser language from Accept-Language header (e.g. 'it', 'en', 'fr')."""
    accept = request.headers.get("Accept-Language", "")
    if not accept:
        return ""
    # Parse first language tag: "it-IT,it;q=0.9,en;q=0.8"  →  "it"
    first = accept.split(",")[0].split(";")[0].strip()
    # Return just the primary subtag (e.g. "it-IT"  →  "it")
    return first.split("-")[0].lower() if first else ""


def _active_generating_for_client(client_id):
    """Count how many jobs are currently generating for the given client_id. Thread-safe."""
    with _jobs_lock:
        return _active_generating_for_client_unlocked(client_id)


def _active_generating_for_client_unlocked(client_id):
    """Internal: caller MUST hold _jobs_lock."""
    if not client_id:
        return 0
    return sum(
        1 for j in jobs.values()
        if j.get("client_id") == client_id and j.get("status") == "generating"
    )


def _refund_payment_on_orphan(job_id, job, reason):
    """Refund Gemini payment if /api/generate rejected after the token was consumed.

    Mirrors generation_engine._refund_gemini_payment: voucher → _voucher_refund,
    paypal → emit refund voucher. Best-effort; non-fatal on errors.
    Also clears job['payment'] so a retry doesn't see stale state.
    """
    payment_meta = job.get("payment") or {}
    tok = payment_meta.get("token")
    amt = float(payment_meta.get("total_eur", 0) or 0)
    method = payment_meta.get("method", "")
    if not tok or amt <= 0:
        return
    try:
        if method == "voucher":
            payment._voucher_refund(tok, amt, job_id=job_id, reason=reason)
        elif method == "paypal":
            pay = payment._payments.get(tok, {})
            email = pay.get("email", "") or ""
            if email:
                payment._create_voucher(
                    email, amt, origin_order_id=tok, origin_job_id=job_id,
                    kind="refund", note=f"refund {reason} job {job_id}",
                )
            else:
                print(
                    f"[{job_id}] WARNING: orphan refund voucher not emitted — "
                    f"PayPal order {tok} has no buyer email "
                    f"(amount {amt:.2f} EUR, reason {reason})"
                )
            # Free up the PayPal order to be re-spent (or leave used=True and let
            # the refund voucher carry the value forward; we choose refund voucher
            # to keep idempotency simple)
    except Exception as e:
        print(f"[{job_id}] orphan refund failed ({reason}, non-fatal): {e}")
    finally:
        job.pop("payment", None)


def _active_optimizing_for_client(client_id):
    """Count how many jobs are currently optimizing for the given client_id. Thread-safe."""
    with _jobs_lock:
        return _active_optimizing_for_client_unlocked(client_id)


def _active_optimizing_for_client_unlocked(client_id):
    """Internal: caller MUST hold _jobs_lock."""
    if not client_id:
        return 0
    return sum(
        1 for j in jobs.values()
        if j.get("client_id") == client_id and j.get("status") == "optimizing"
    )


FAVICON_B64 = "data:image/svg+xml;base64,PHN2ZyB4bWxucz0iaHR0cDovL3d3dy53My5vcmcvMjAwMC9zdmciIHZpZXdCb3g9IjAgMCA2NCA2NCI+CiAgPGRlZnM+CiAgICA8bGluZWFyR3JhZGllbnQgaWQ9ImJnIiB4MT0iMCUiIHkxPSIwJSIgeDI9IjEwMCUiIHkyPSIxMDAlIj4KICAgICAgPHN0b3Agb2Zmc2V0PSIwJSIgc3R5bGU9InN0b3AtY29sb3I6I2MyOWE2YyIvPgogICAgICA8c3RvcCBvZmZzZXQ9IjEwMCUiIHN0eWxlPSJzdG9wLWNvbG9yOiNhMDc4NTAiLz4KICAgIDwvbGluZWFyR3JhZGllbnQ+CiAgPC9kZWZzPgogIDxyZWN0IHdpZHRoPSI2NCIgaGVpZ2h0PSI2NCIgcng9IjE0IiBmaWxsPSJ1cmwoI2JnKSIvPgogIDxwYXRoIGQ9Ik0xNiA0NFYyMGMwLTIgMS41LTMuNSAzLjUtMy41QzIzIDE2LjUgMjggMTcgMzIgMTljNC0yIDktMi41IDEyLjUtMi41IDIgMCAzLjUgMS41IDMuNSAzLjV2MjQiIGZpbGw9Im5vbmUiIHN0cm9rZT0id2hpdGUiIHN0cm9rZS13aWR0aD0iMi41IiBzdHJva2UtbGluZWNhcD0icm91bmQiIHN0cm9rZS1saW5lam9pbj0icm91bmQiLz4KICA8cGF0aCBkPSJNMzIgMTl2MjUiIHN0cm9rZT0id2hpdGUiIHN0cm9rZS13aWR0aD0iMiIgc3Ryb2tlLWxpbmVjYXA9InJvdW5kIi8+CiAgPHBhdGggZD0iTTE3IDM2YzAtOSA2LjctMTUgMTUtMTVzMTUgNiAxNSAxNSIgZmlsbD0ibm9uZSIgc3Ryb2tlPSJ3aGl0ZSIgc3Ryb2tlLXdpZHRoPSIyLjgiIHN0cm9rZS1saW5lY2FwPSJyb3VuZCIvPgogIDxyZWN0IHg9IjEzIiB5PSIzNCIgd2lkdGg9IjciIGhlaWdodD0iMTAiIHJ4PSIzIiBmaWxsPSJ3aGl0ZSIvPgogIDxyZWN0IHg9IjQ0IiB5PSIzNCIgd2lkdGg9IjciIGhlaWdodD0iMTAiIHJ4PSIzIiBmaWxsPSJ3aGl0ZSIvPgogIDxwYXRoIGQ9Ik0yMiAzNy41YzEuMi0xIDEuMi0zIDAtNCIgZmlsbD0ibm9uZSIgc3Ryb2tlPSIjYzI5YTZjIiBzdHJva2Utd2lkdGg9IjEuMyIgc3Ryb2tlLWxpbmVjYXA9InJvdW5kIi8+CiAgPHBhdGggZD0iTTQyIDM3LjVjLTEuMi0xLTEuMi0zIDAtNCIgZmlsbD0ibm9uZSIgc3Ryb2tlPSIjYzI5YTZjIiBzdHJva2Utd2lkdGg9IjEuMyIgc3Ryb2tlLWxpbmVjYXA9InJvdW5kIi8+Cjwvc3ZnPg=="

_download_tokens = {}  # token -> {job_id, created_at, download_type, base_url, ...}
_TOKENS_FILE = UPLOAD_DIR / "_download_tokens.json"
_tokens_lock = threading.Lock()

_download_tracking = {}  # file_path -> {"count": int, "last_download": float}
_DL_THROTTLE_SEC = 30
_DL_MAX_DOWNLOADS = 5


def _check_download_throttle(file_path):
    """Check and update download throttle for a file path.
    Returns (status, info):
      ('ok', remaining)        -> allowed; remaining = additional downloads still permitted after this one
      ('last', 0)              -> allowed but this is the last permitted download (file will be removed on next attempt)
      ('cooldown', seconds_left) -> too soon, wait
      ('deleted', None)        -> file deleted after max downloads
    """
    if not file_path or not os.path.exists(file_path):
        return ("ok", None)
    now = time.time()
    rec = _download_tracking.get(file_path)

    if rec:
        elapsed = now - rec["last_download"]
        if elapsed < _DL_THROTTLE_SEC:
            return ("cooldown", int(_DL_THROTTLE_SEC - elapsed))

    current = rec["count"] if rec else 0
    if current >= _DL_MAX_DOWNLOADS:
        try:
            os.remove(file_path)
            print(f"[throttle] Deleted {file_path} after {_DL_MAX_DOWNLOADS} downloads")
        except Exception as e:
            print(f"[throttle] Error deleting {file_path}: {e}")
        _download_tracking.pop(file_path, None)
        return ("deleted", None)

    new_count = current + 1
    if rec:
        rec["count"] = new_count
        rec["last_download"] = now
    else:
        _download_tracking[file_path] = {"count": new_count, "last_download": now}

    if new_count >= _DL_MAX_DOWNLOADS:
        return ("last", 0)
    return ("ok", _DL_MAX_DOWNLOADS - new_count)


def _apply_no_cache(response):
    """Disabilita la cache HTTP sulla risposta (per contenuti rigenerati on-demand con URL stabile).
    SEND_FILE_MAX_AGE_DEFAULT è 1 anno: senza questa override il browser servirebbe la
    prima versione scaricata anche dopo che il server ha rigenerato il file con contenuto
    aggiornato (es. .abm cumulativo dopo successive ottimizzazioni)."""
    try:
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
    except Exception:
        pass
    return response


def _iter_output_dirs(job_dir):
    """Yield all output directories for a job, newest first.

    Convention: each /api/generate call creates `output_{gen_epoch}/` inside the
    job_dir. Earlier epochs are preserved so active email tokens keep working.
    Legacy names (`output/`, `output_archive_*`) are included for backwards
    compatibility with files created before this layout.
    """
    if not job_dir.exists():
        return
    epoch_dirs = []
    legacy = []
    for d in job_dir.iterdir():
        if not d.is_dir():
            continue
        n = d.name
        if n == "output" or n.startswith("output_archive_"):
            legacy.append(d)
        elif n.startswith("output_"):
            try:
                epoch_dirs.append((int(n.split("_", 1)[1]), d))
            except (ValueError, IndexError):
                legacy.append(d)
    epoch_dirs.sort(key=lambda t: t[0], reverse=True)
    for _, d in epoch_dirs:
        yield d
    for d in legacy:
        yield d


def _find_files_in_outputs(job_dir, pattern):
    """Glob `pattern` across all output dirs (recursive). Returns list of Paths."""
    results = []
    for d in _iter_output_dirs(job_dir):
        results.extend(d.rglob(pattern))
    return results


def _send_file_throttled(file_path, as_attachment=True, download_name=None, mimetype=None, no_cache=False, bypass_throttle=False, **kwargs):
    # HEAD e Range request (anteprima/resume del browser o client email) non devono
    # consumare il quota di 5 download: serviamo il file senza toccare il counter.
    is_probe = False
    try:
        is_probe = request.method == "HEAD" or bool(request.headers.get("Range"))
    except Exception:
        pass
    # bypass_throttle: per le rotte UI autenticate (cookie owner via
    # _check_job_owner). Il throttle è disegnato per i link email pubblici via
    # token, non per il proprietario del job. Senza bypass, un download via
    # email link 30s prima blocca quello via UI sullo stesso file.
    if is_probe or bypass_throttle:
        response = send_file(file_path, as_attachment=as_attachment, download_name=download_name, mimetype=mimetype, **kwargs)
        if no_cache:
            _apply_no_cache(response)
        return response

    status, info = _check_download_throttle(file_path)
    if status == "cooldown":
        lang = _get_browser_lang() or "en"
        return _render_dl_cooldown_page(lang, info), 429
    if status == "deleted":
        lang = _get_browser_lang() or "en"
        return _render_dl_deleted_page(lang), 410
    response = send_file(file_path, as_attachment=as_attachment, download_name=download_name, mimetype=mimetype, **kwargs)
    try:
        if status == "last":
            response.headers["X-Download-Last"] = "1"
            response.headers["X-Download-Remaining"] = "0"
        elif status == "ok" and info is not None:
            response.headers["X-Download-Remaining"] = str(info)
        response.headers["Access-Control-Expose-Headers"] = "X-Download-Last, X-Download-Remaining, Content-Disposition"
    except Exception:
        pass
    if no_cache:
        _apply_no_cache(response)
    return response


def _read_tokens_file():
    """Read raw token dict from disk. Returns {} on missing/invalid file."""
    if not _TOKENS_FILE.exists():
        return {}
    try:
        with open(_TOKENS_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except Exception as e:
        print(f"[tokens] Failed to read tokens file: {e}")
        return {}


def _merge_tokens_from_disk():
    """Pick up tokens persisted by other workers (Gunicorn multi-worker safety).

    Each worker keeps its own in-memory `_download_tokens`; the only shared
    state is `_TOKENS_FILE`. Without periodic merge, worker B's cleanup loop
    cannot see tokens created by worker A and would delete their job dirs as
    orphan. In-memory entries always win on conflict (worker may have data
    not yet flushed to disk).
    """
    disk = _read_tokens_file()
    if not disk:
        return
    now = time.time()
    with _tokens_lock:
        for tok, info in disk.items():
            try:
                created = float(info.get("created_at", 0) or 0)
            except (TypeError, ValueError):
                continue
            # Per token PREMIUM (is_gemini) la retention e' GEMINI_FILE_RETENTION_SEC,
            # raddoppiata se non risulta alcun download (_effective_*).
            if (now - created) > _effective_retention_for_token_info(info) + 300:
                continue
            if tok not in _download_tokens:
                _download_tokens[tok] = info


def _save_tokens():
    """Persist download tokens to disk (survives restart).

    Re-reads the file first and merges to avoid clobbering tokens written by
    other workers since our last sync.
    """
    _merge_tokens_from_disk()
    try:
        with _tokens_lock:
            # Save only serializable data
            data = {}
            for tok, info in _download_tokens.items():
                data[tok] = {
                    "job_id": info["job_id"],
                    "created_at": info["created_at"],
                    "download_type": info.get("download_type", "audio"),
                    "base_url": info.get("base_url", ""),
                    # Snapshot of job data needed for download after restart
                    "book_title": info.get("book_title", ""),
                    "output_zip": info.get("output_zip", ""),
                    "output_name": info.get("output_name", ""),
                    "output_file": info.get("output_file", ""),
                    "epub_path": info.get("epub_path", ""),
                    "podcast_safe_name": info.get("podcast_safe_name", ""),
                    "podcast_ready": info.get("podcast_ready", False),
                    "podcast_mp3s": info.get("podcast_mp3s", []),
                    "podcast_info_title": info.get("podcast_info_title", ""),
                    "podcast_info_author": info.get("podcast_info_author", ""),
                    "podcast_info_language": info.get("podcast_info_language", ""),
                    "original_filename": info.get("original_filename", ""),
                    "lang": info.get("lang", "en"),
                    "optimized_abm_path": info.get("optimized_abm_path", ""),
                    "optimized_abm_name": info.get("optimized_abm_name", ""),
                    # Marker per scegliere retention: True se job ha generato con voce PREMIUM.
                    "is_gemini": bool(info.get("is_gemini", False)),
                    # Timestamp primo download reale del file via /dl/<token>/*.
                    # 0/None = mai scaricato (attiva protezione 2x per voci PREMIUM).
                    "downloaded_at": info.get("downloaded_at") or 0,
                    # Fields required by /dl/<token> rendering after worker restart
                    # or cross-worker token merge (Gunicorn multi-process).
                    "output_format": info.get("output_format", ""),
                    "output_m4b": info.get("output_m4b", ""),
                    "ai_optimized": info.get("ai_optimized", False),
                }
            # Atomic write: tmp + fsync + rename per evitare corruzione su crash
            _tmp_tokens = _TOKENS_FILE.with_suffix(_TOKENS_FILE.suffix + ".tmp")
            with open(_tmp_tokens, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
                f.flush()
                try:
                    os.fsync(f.fileno())
                except OSError:
                    pass
            os.replace(str(_tmp_tokens), str(_TOKENS_FILE))
    except Exception as e:
        print(f"[tokens] Failed to save tokens: {e}")


def _load_tokens():
    """Reload download tokens from disk on startup."""
    global _download_tokens
    if not _TOKENS_FILE.exists():
        return
    try:
        with open(_TOKENS_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        now = time.time()
        loaded = 0
        expired = 0
        for tok, info in data.items():
            # Skip expired tokens (retention dipende da is_gemini sul token,
            # raddoppiata se downloaded_at non e' settato).
            if (now - info.get("created_at", 0)) > _effective_retention_for_token_info(info) + 300:
                expired += 1
                continue
            # Verify that job files still exist
            job_dir = UPLOAD_DIR / info.get("job_id", "")
            if not job_dir.exists():
                expired += 1
                continue
            _download_tokens[tok] = info
            loaded += 1
        if loaded or expired:
            print(f"[tokens] Loaded {loaded} tokens from disk ({expired} expired/invalid)")
    except Exception as e:
        print(f"[tokens] Failed to load tokens: {e}")


# ----------------------------------------------------------------------
# CLIENT EMAILS — persistenza email di notifica per client_id (fallback UI)
# ----------------------------------------------------------------------

_CLIENT_EMAILS_FILE = UPLOAD_DIR / "_client_emails.json"
_client_emails = {}
_client_emails_lock = threading.Lock()


def _load_client_emails():
    """Carica la mappatura client_id → email, per fallback cross-job."""
    global _client_emails
    if not _CLIENT_EMAILS_FILE.exists():
        return
    try:
        with open(_CLIENT_EMAILS_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, dict):
            _client_emails = {k: v for k, v in data.items() if k and v}
        print(f"[client_emails] Loaded {len(_client_emails)} entries")
    except Exception as e:
        print(f"[client_emails] Failed to load: {e}")


def _save_client_emails():
    """Persiste la mappatura client_id → email in scrittura atomica."""
    try:
        with _client_emails_lock:
            tmp_file = _CLIENT_EMAILS_FILE.with_suffix(
                _CLIENT_EMAILS_FILE.suffix + ".tmp"
            )
            with open(tmp_file, "w", encoding="utf-8") as f:
                json.dump(_client_emails, f, ensure_ascii=False, indent=2)
                f.flush()
                try:
                    os.fsync(f.fileno())
                except OSError:
                    pass
            os.replace(str(tmp_file), str(_CLIENT_EMAILS_FILE))
    except Exception as e:
        print(f"[client_emails] Failed to save: {e}")


def _lookup_client_email(client_id):
    """Cerca l'email associata a un client_id (thread-safe)."""
    if not client_id:
        return ""
    with _client_emails_lock:
        return _client_emails.get(client_id, "")


# ----------------------------------------------------------------------
# PAYMENTS & VOUCHERS (for LLM optimization) — state lives in payment.py
# ----------------------------------------------------------------------

# Sospensione avvio nuovi processi (attivabile da admin via /admin/log-activity)
_suspend_new_jobs = False
_suspend_lock = threading.Lock()


#  -  -  Admin activity digest  -  - 

# (Functions imported from email_service)

LOCALE_NAMES = {
    "af": "Afrikaans", "am": "Amarico", "ar": "Arabo", "az": "Azero",
    "bg": "Bulgaro", "bn": "Bengalese", "bs": "Bosniaco", "ca": "Catalano",
    "cs": "Ceco", "cy": "Gallese", "da": "Danese", "de": "Tedesco",
    "el": "Greco", "en": "Inglese", "es": "Spagnolo", "et": "Estone",
    "fa": "Persiano", "fi": "Finlandese", "fil": "Filippino", "fr": "Francese",
    "ga": "Irlandese", "gu": "Gujarati", "he": "Ebraico", "hi": "Hindi",
    "hr": "Croato", "hu": "Ungherese", "id": "Indonesiano", "is": "Islandese",
    "it": "Italiano", "ja": "Giapponese", "jv": "Giavanese", "ka": "Georgiano",
    "kk": "Kazako", "km": "Khmer", "kn": "Kannada", "ko": "Coreano",
    "lo": "Lao", "lt": "Lituano", "lv": "Lettone", "mk": "Macedone",
    "ml": "Malayalam", "mr": "Marathi", "ms": "Malese", "mt": "Maltese",
    "my": "Birmano", "nb": "Norvegese", "ne": "Nepalese", "nl": "Olandese",
    "pl": "Polacco", "ps": "Pashto", "pt": "Portoghese", "ro": "Rumeno",
    "ru": "Russo", "si": "Singalese", "sk": "Slovacco", "sl": "Sloveno",
    "so": "Somalo", "sq": "Albanese", "sr": "Serbo", "sv": "Svedese",
    "sw": "Swahili", "ta": "Tamil", "te": "Telugu", "th": "Thailandese",
    "tr": "Turco", "uk": "Ucraino", "ur": "Urdu", "uz": "Uzbeco",
    "vi": "Vietnamita", "zh": "Cinese", "zu": "Zulu",
}

_voices_cache = None
_voices_lock = threading.Lock()

async def _fetch_voices():
    """Fetches and categorizes Edge TTS, Google TTS, and Gemini TTS voices."""
    try:
        import edge_tts
        vman = await edge_tts.VoicesManager.create()
        edge_list = vman.voices
    except Exception as e:
        print(f"Error fetching Edge voices: {e}")
        edge_list = []

    languages = {} # lang_code -> { "name": "...", "voices": [] }
    
    # 1. Edge TTS
    for v in edge_list:
        lc_full = v["Locale"]
        lc = lc_full.split("-")[0].lower()
        region = lc_full.split("-")[-1].upper()
        
        if lc not in languages:
            languages[lc] = {
                "name": LOCALE_NAMES.get(lc, lc.upper()),
                "voices": []
            }
        
        # Pulizia nome: "Microsoft Isabella Online (Natural) - Italian (Italy)" -> "Isabella"
        raw_name = v["FriendlyName"]
        clean_name = raw_name.replace("Microsoft ", "").replace(" Online (Natural)", "")
        # Rimuove l'eventuale suffisso della lingua dopo il trattino
        if " - " in clean_name:
            clean_name = clean_name.split(" - ")[0].strip()
        
        gender_icon = "👨" if v["Gender"] == "Male" else "👩"
        languages[lc]["voices"].append({
            "id": v["ShortName"],
            "name": f"{clean_name} ({region})",
            "gender": v["Gender"],
            "gender_icon": gender_icon,
            "locale": lc_full,
            "engine": "edge"
        })

    # 2. Google TTS (Optional)
    if google_tts is not None:
        try:
            # get_voices restituisce { "it": [ {...}, ... ], "en": [...] }
            g_dict = google_tts.get_voices()
            for lc_short, v_list in g_dict.items():
                if lc_short not in languages:
                    languages[lc_short] = {
                        "name": LOCALE_NAMES.get(lc_short, lc_short.upper()),
                        "voices": []
                    }
                languages[lc_short]["voices"].extend(v_list)
        except Exception as e:
            print(f"Error merging Google voices: {e}")

    # 3. Gemini TTS (Optional) — solo se effettivamente abilitato.
    # `gemini_tts is not None` significa solo che il modulo è importato;
    # senza ABM_GEMINI_API_KEY le voci non vanno comunque mostrate.
    # NB: il branch GEMINI espone le voci nel tab "PREMIUM" (la rimozione
    # fatta su main era temporanea per la release pre-feature).
    if gemini_tts is not None and gemini_tts.is_available():
        try:
            gem_dict = gemini_tts.get_voices()
            for lc_short, v_list in gem_dict.items():
                if lc_short not in languages:
                    languages[lc_short] = {
                        "name": LOCALE_NAMES.get(lc_short, lc_short.upper()),
                        "voices": []
                    }
                # Gemini voices are multilingual; gender è impostato in
                # gemini_tts.get_voices() da GEMINI_VOICE_GENDER (doc Google).
                # Lo shim resta come fallback per voci eventuali senza metadata.
                for v in v_list:
                    v.setdefault("gender", "Neutral")
                    v.setdefault("gender_icon", "★")
                languages[lc_short]["voices"].extend(v_list)
        except Exception as e:
            print(f"Error merging Gemini voices: {e}")

    # Sorting
    for lang in languages.values():
        lang["voices"].sort(key=lambda x: (x["gender"], x["name"]))

    # Priority sorting for languages
    priority = {"it": 0, "en": 1, "fr": 2, "de": 3, "es": 4, "pt": 5}
    sorted_langs = dict(sorted(languages.items(), key=lambda x: (priority.get(x[0], 99), x[1]["name"])))
    
    return sorted_langs

def get_voices():
    """Thread-safe access to the voices cache."""
    global _voices_cache
    with _voices_lock:
        if _voices_cache is not None:
            return _voices_cache
            
    import asyncio
    try:
        # Create new loop for this thread (or use existing if in main)
        loop = asyncio.new_event_loop()
        voices = loop.run_until_complete(_fetch_voices())
        loop.close()
        
        with _voices_lock:
            _voices_cache = voices
        return voices
    except Exception as e:
        print(f"Error in get_voices: {e}")
        return {}

def _invalidate_voices_cache():
    """Invalida la cache voci (ricarica al prossimo get_voices())."""
    global _voices_cache
    with _voices_lock:
        _voices_cache = None
    if google_tts is not None:
        google_tts.invalidate_voices_cache()

# ----------------------------------------------------------------------
# HELPER CLASSES & PARSERS (Moved to generation_engine.py)
# ----------------------------------------------------------------------

def parse_txt(file_path):
    return generation_engine.parse_txt(file_path)

def parse_abm(path):
    return generation_engine.parse_abm(path)

# ----------------------------------------------------------------------
# GENERATION & OPTIMIZATION THREADS (Moved to generation_engine.py)
# ----------------------------------------------------------------------

def run_optimization(job_id, selected_chapters=None):
    return generation_engine.run_optimization(job_id, selected_chapters)

def run_generation(job_id, info, voice, rate, single_file, output_format='m4b', podcast_base_url='', gemini_style_instruction=None):
    try:
        return generation_engine.run_generation(job_id, info, voice, rate, single_file,
                                                 output_format=output_format, podcast_base_url=podcast_base_url,
                                                 gemini_style_instruction=gemini_style_instruction)
    except BaseException as e:
        # SystemExit/KeyboardInterrupt devono propagare — non sopprimerli.
        if isinstance(e, (SystemExit, KeyboardInterrupt)):
            raise
        print(f"[{job_id}] CRITICAL: run_generation wrapper crashed: {e}")
        import traceback
        traceback.print_exc()
        job = jobs.get(job_id)
        if job:
            job["status"] = "error"
            job["progress_message"] = f"Internal error: {e}"
            job["last_poll"] = time.time()

#  -  -  Activity log  -  -
_log_lock = threading.Lock()
_logged_month: str = ""
_logged_sids_ops: set[tuple[str, str]] = set()

def _init_log_dedup():
    """Popola il set di dedup dal file di log del mese corrente."""
    global _logged_month, _logged_sids_ops
    from datetime import datetime
    _logged_month = datetime.now().strftime('%Y-%m')
    log_path = SCRIPT_DIR / f"activity_{_logged_month}.log"
    if not log_path.exists():
        return
    with _log_lock:
        _logged_sids_ops.clear()
        with open(log_path, "r", encoding="utf-8") as f:
            for line in f:
                parts = line.strip().split(" # ")
                if len(parts) >= 4:
                    _logged_sids_ops.add((parts[0], parts[3]))

def _log_activity(session_id, filename, operation, client_id='', client_ip='', voice='', browser_lang=''):
    global _logged_month
    from datetime import datetime
    now = datetime.now()
    current_month = now.strftime('%Y-%m')
    log_path = SCRIPT_DIR / f"activity_{current_month}.log"
    ts = now.strftime('%Y-%m-%d %H:%M:%S')
    key = (session_id, operation)
    with _log_lock:
        if current_month != _logged_month:
            _logged_month = current_month
            _logged_sids_ops.clear()
        if key in _logged_sids_ops:
            return
        line = f'{session_id} # {ts} # "{filename}" # {operation} # {client_id} # {client_ip} # {voice} # {browser_lang}\n'
        try:
            with open(log_path, "a", encoding="utf-8") as f:
                f.write(line)
            _logged_sids_ops.add(key)
        except OSError:
            pass


def _is_resume_or_probe_request():
    """True se la richiesta corrente è HEAD o resume con Range header.

    Why: i browser e i client email aprono il link con HEAD/Range per
    prefetch/anteprima/resume. Senza filtro ogni link aperto genera N voci
    nel log anche se l'utente ha cliccato una sola volta.
    How to apply: chiamare prima di _log_activity nelle route di download
    per contare solo i download reali (GET completi).
    """
    try:
        return request.method == "HEAD" or bool(request.headers.get("Range"))
    except Exception:
        return False

CHAPTER_SILENCE_SEC = 3  # secondi di silenzio all'inizio di ogni capitolo


# ----------------------------------------------------------------------
# COMMUNITY STATS — derivate dai log activity_YYYY-MM.log esistenti
# ----------------------------------------------------------------------
# Conta operation=='COMPLETE' (audiolibri generati con successo) e aggrega
# per lingua TTS (voice.split('-')[0]). Cache in-memory: 60s today, 5min mese.

_stats_lock = threading.Lock()
_stats_today_cache = {"value": None, "expires": 0.0}
_stats_month_cache = {"value": None, "expires": 0.0}


def _parse_activity_lines(yyyymm: str):
    """Itera (ts_str, operation, voice) dalle righe del log mensile.
    Formato: '<sid> # <ts> # "<file>" # <op> # <cid> # <ip> # <voice> # <lang>'.
    Resiliente a righe malformate."""
    log_path = SCRIPT_DIR / f"activity_{yyyymm}.log"
    if not log_path.exists():
        return
    try:
        with open(log_path, "r", encoding="utf-8") as f:
            for line in f:
                parts = line.rstrip("\n").split(" # ")
                if len(parts) < 7:
                    continue
                yield parts[1], parts[3], parts[6]
    except OSError:
        return


def _stats_today_count() -> int:
    """Conta COMPLETE e OPT_COMPLETE odierni. Cache 60s."""
    now = time.time()
    with _stats_lock:
        if _stats_today_cache["value"] is not None and now < _stats_today_cache["expires"]:
            return _stats_today_cache["value"]
    today = datetime.now()
    yyyymm = today.strftime("%Y-%m")
    today_str = today.strftime("%Y-%m-%d")
    count = 0
    for ts, op, _voice in _parse_activity_lines(yyyymm):
        if op in ("COMPLETE", "OPT_COMPLETE") and ts.startswith(today_str):
            count += 1
    with _stats_lock:
        _stats_today_cache["value"] = count
        _stats_today_cache["expires"] = now + 60.0
    return count


def _stats_month_by_lang() -> dict:
    """Aggrega COMPLETE e OPT_COMPLETE del mese corrente per lingua TTS.
    Restituisce {monthly: int, top: [{lang, count}], other: int}.
    Cache 5min."""
    now = time.time()
    with _stats_lock:
        if _stats_month_cache["value"] is not None and now < _stats_month_cache["expires"]:
            return _stats_month_cache["value"]
    yyyymm = datetime.now().strftime("%Y-%m")
    by_lang: dict[str, int] = defaultdict(int)
    total = 0
    for _ts, op, voice in _parse_activity_lines(yyyymm):
        if op not in ("COMPLETE", "OPT_COMPLETE"):
            continue
        total += 1
        if not voice:
            continue
        lang = voice.split("-")[0].strip().lower()
        if lang:
            by_lang[lang] += 1
    sorted_langs = sorted(by_lang.items(), key=lambda kv: kv[1], reverse=True)
    top = [{"lang": k, "count": v} for k, v in sorted_langs[:4]]
    other = sum(v for _, v in sorted_langs[4:])
    result = {"monthly": total, "top": top, "other": other}
    with _stats_lock:
        _stats_month_cache["value"] = result
        _stats_month_cache["expires"] = now + 300.0
    return result


# ----------------------------------------------------------------------
# ROUTES
# ----------------------------------------------------------------------

#  -  -  -  Rotte per lingua (/it/, /en/, /fr/, /es/, /de/, /zh/, /hi/)  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -
# Ogni URL ha HTML pre-renderizzato con meta tag, title, hreflang e canonical
# corretti per quella lingua  -  indicizzabili da Google come pagine distinte.

def _inject_reviews(template_html: str, lang: str) -> str:
    """Swap __REVIEWS_LD__ placeholder with fresh AggregateRating + Review
    JSON-LD built from the live feedback store.

    Cost ≤ 1 ms per request. Falls back to empty replacement if the store
    is unavailable so the page still renders cleanly."""
    try:
        rev = seo_reviews.build_reviews(lang)
        ld = rev.get("ld_block", "") or ""
    except Exception as e:
        print(f"[seo_reviews] inject failed: {e!s}")
        ld = ""
    return template_html.replace("__REVIEWS_LD__", ld)


@app.route("/")
def index():
    """Root: serve la lingua rilevata dall'Accept-Language, senza redirect.
    Il redirect 302 penalizzerebbe il PageRank; meglio rispondere con canonical.
    Usa HTML_ROOT_TEMPLATES: canonical punta a BASE_URL/ (non /{lang}/).
    Questo garantisce che l'URL x-default negli hreflang sia auto-canonicalizzante.
    """
    lang = _detect_lang_from_request()
    base = HTML_ROOT_TEMPLATES.get(lang, HTML_ROOT_TEMPLATES["en"])
    resp = app.make_response(_inject_reviews(base, lang))
    resp.headers["Content-Type"] = "text/html; charset=utf-8"
    resp.headers["Vary"] = "Accept-Language"
    return resp

def _serve_lang(lang: str):
    return (
        _inject_reviews(HTML_TEMPLATES[lang], lang),
        200,
        {"Content-Type": "text/html; charset=utf-8"},
    )

@app.route("/it/")
def index_it():
    return _serve_lang("it")

@app.route("/en/")
def index_en():
    return _serve_lang("en")

@app.route("/fr/")
def index_fr():
    return _serve_lang("fr")

@app.route("/es/")
def index_es():
    return _serve_lang("es")

@app.route("/de/")
def index_de():
    return _serve_lang("de")

@app.route("/zh/")
def index_zh():
    return _serve_lang("zh")

@app.route("/hi/")
def index_hi():
    return _serve_lang("hi")


# ── FAQ Page (dedicated) ────────────────────────────────────────────

_FAQ_TITLES = {
    "it": "Domande Frequenti — Audiobook Maker",
    "en": "Frequently Asked Questions — Audiobook Maker",
    "fr": "Questions Fréquentes — Audiobook Maker",
    "es": "Preguntas Frecuentes — Audiobook Maker",
    "de": "Häufig Gestellte Fragen — Audiobook Maker",
    "zh": "常见问题 — Audiobook Maker",
    "hi": "अक्सर पूछे जाने वाले प्रश्न — Audiobook Maker",
}


@app.route("/faq/")
def faq_page_root():
    """Redirect root FAQ to the browser-language or English variant."""
    lang = _detect_lang()
    if not lang or lang not in _SUPPORTED_LANGS:
        lang = "en"
    base = BASE_URL or ""
    if base:
        return redirect(f"{base}/faq/{lang}/", code=301)
    return redirect(f"/faq/{lang}/", code=301)


@app.route("/faq/<lang>/")
def faq_page(lang):
    """Dedicated FAQ page per language. Crawler-facing minimal HTML with
    JSON-LD FAQPage schema and full hreflang alternates."""
    if lang not in _SUPPORTED_LANGS:
        return "Language not supported", 404

    html_lang = {"zh": "zh-Hans"}.get(lang, lang)
    c = seo_content._CONTENT.get(lang, seo_content._CONTENT.get("en", {}))
    title = html_mod.escape(_FAQ_TITLES.get(lang, _FAQ_TITLES["en"]))
    desc = html_mod.escape(c.get("direct_answer", ""))
    base = BASE_URL or ""
    canonical = f"{base}/faq/{lang}/"

    # Build hreflang links for FAQ page
    hreflang_lines = []
    lc_to_hl = {
        "it": "it", "en": "en", "fr": "fr", "es": "es",
        "de": "de", "zh": "zh-Hans", "hi": "hi",
    }
    for lc, hl in lc_to_hl.items():
        href = f"{base}/faq/{lc}/" if base else f"/faq/{lc}/"
        hreflang_lines.append(
            f'<link rel="alternate" hreflang="{hl}" href="{href}">'
        )
    x_default_href = f"{base}/faq/en/" if base else "/faq/en/"
    hreflang_lines.append(
        f'<link rel="alternate" hreflang="x-default" href="{x_default_href}">'
    )
    hreflang_block = "\n    ".join(hreflang_lines)

    # Build FAQ HTML
    faqs_html = ""
    faq_ld_items = []
    for q, a in c.get("faqs", []):
        faqs_html += (
            f'  <details open><summary>{html_mod.escape(q)}</summary>\n'
            f'    <p>{html_mod.escape(a)}</p>\n'
            f'  </details>\n\n'
        )
        faq_ld_items.append({
            "@type": "Question",
            "name": q,
            "acceptedAnswer": {"@type": "Answer", "text": a},
        })

    faq_ld_json = json.dumps({
        "@context": "https://schema.org",
        "@type": "FAQPage",
        "mainEntity": faq_ld_items,
    }, ensure_ascii=False)

    iso_modified = datetime.now().strftime("%Y-%m-%d")

    page = f'''<!DOCTYPE html>
<html lang="{html_lang}">
<head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>{title}</title>
    <meta name="description" content="{desc}">
    <meta name="robots" content="index, follow">
    <link rel="canonical" href="{canonical}">
    {hreflang_block}
    <meta property="og:type" content="website">
    <meta property="og:title" content="{title}">
    <meta property="og:description" content="{desc}">
    <meta property="og:url" content="{canonical}">
    <meta property="article:published_time" content="2022-06-01">
    <meta property="article:modified_time" content="{iso_modified}">
    <script type="application/ld+json">{faq_ld_json}</script>
    <style>
        body{{margin:0;padding:0;background:#f5f3ef;font-family:system-ui,sans-serif;color:#2c2a26}}
        main{{max-width:760px;margin:0 auto;padding:24px 20px 48px}}
        h1{{font-size:1.5rem;font-weight:700;margin:0 0 8px}}
        .sub{{color:#8B7B6B;margin:0 0 32px;font-size:.95rem}}
        details{{border:1px solid #d5d0c8;border-radius:8px;padding:16px 20px;margin-bottom:12px;background:#fff}}
        summary{{font-weight:600;cursor:pointer;font-size:1rem;color:#c29a6c}}
        summary:hover{{color:#a07840}}
        details p{{margin:12px 0 0;line-height:1.6;color:#4a4640;font-size:.95rem}}
        .footer{{text-align:center;color:#8B7B6B;font-size:.82rem;margin-top:40px}}
        .footer a{{color:#c29a6c}}
    </style>
</head>
<body>
<main>
    <h1>{title}</h1>
    <p class="sub">{desc}</p>
{faqs_html}</main>
<div class="footer">
    <a href="/">Audiobook Maker</a> &middot; {iso_modified}
</div>
</body>
</html>'''
    return page, 200, {"Content-Type": "text/html; charset=utf-8"}


@app.route("/content/<lang>/")
def seo_content_page(lang):
    """Dedicated SEO content page per language. Minimal HTML wrapper around
    the rich SEO block generated by build_seo_content_html(). Crawler-facing,
    no app CSS/JS required — the block already carries inline styles."""
    if lang not in ("it", "en", "fr", "es", "de", "zh", "hi"):
        return "Language not supported", 404

    html_lang = {"zh": "zh-Hans"}.get(lang, lang)
    c = seo_content._CONTENT.get(lang, seo_content._CONTENT["en"])
    title = html_mod.escape(c.get("heading", "Audiobook Maker"))
    desc = html_mod.escape(c.get("direct_answer", ""))
    base = BASE_URL or ""
    canonical = f'{base}/content/{lang}/' if base else ""

    seo_block = seo_content.build_seo_content_html(lang)

    head_extra = ""
    if canonical:
        head_extra += f'\n    <link rel="canonical" href="{canonical}" />'

    page = f'''<!DOCTYPE html>
<html lang="{html_lang}">
<head>
    <meta charset="utf-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1" />
    <title>{title}</title>
    <meta name="description" content="{desc}" />
    <meta name="robots" content="index, follow" />{head_extra}
</head>
<body style="margin:0;padding:0;background:var(--bg,#f5f3ef);">
{seo_block}
</body>
</html>'''
    return page, 200, {"Content-Type": "text/html; charset=utf-8"}


#  -  -  SEO Guide Pages  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -
_VALID_GUIDES = {"epub-to-audiobook", "m4b-format", "text-to-speech-audiobook", "podcast"}

@app.route("/guide/<guide_id>/")
def guide_page(guide_id):
    if guide_id not in _VALID_GUIDES:
        return "Guide not found", 404
    lang = request.args.get("lang", _detect_lang_from_request()).strip()
    if lang not in ("it", "en", "fr", "es", "de", "zh", "hi"):
        lang = "en"
    html = build_guide_html(guide_id, lang, BASE_URL, __version__)
    return html, 200, {"Content-Type": "text/html; charset=utf-8"}


#  -  -  -  sitemap.xml  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  - 
@app.route("/sitemap.xml")
def sitemap():
    """Sitemap con tutte le varianti linguistiche.
    Richiede ABM_BASE_URL configurato per gli URL assoluti (obbligatorio per Google).
    """
    if not BASE_URL:
        return "<!-- sitemap non disponibile: impostare ABM_BASE_URL -->", 200, {
            "Content-Type": "text/xml; charset=utf-8"
        }

    import os as _os
    from datetime import date, datetime as _dt

    def _file_lastmod(path: str) -> str:
        """Return ISO date of file mtime, or today as a safe fallback."""
        try:
            return _dt.utcfromtimestamp(_os.path.getmtime(path)).strftime("%Y-%m-%d")
        except OSError:
            return date.today().isoformat()

    _here = _os.path.dirname(_os.path.abspath(__file__))
    # The home page reflects content from app + visible SEO + live user
    # reviews; pick the most recent of all three so Google sees a real change
    # signal whenever new feedback is approved.
    candidates = [
        _file_lastmod(_os.path.join(_here, "audiobook_app.py")),
        _file_lastmod(_os.path.join(_here, "seo_content.py")),
    ]
    try:
        latest_review_ts = seo_reviews.build_reviews("en").get("latest_ts", 0)
        if latest_review_ts:
            candidates.append(
                _dt.utcfromtimestamp(latest_review_ts).strftime("%Y-%m-%d")
            )
    except Exception:
        pass
    home_lastmod = max(candidates)
    guide_lastmod = _file_lastmod(_os.path.join(_here, "guide_content.py"))

    lang_hreflang_map = {
        "it": "it", "en": "en", "fr": "fr",
        "es": "es", "de": "de", "zh": "zh-Hans",
        "hi": "hi"
    }

    # Blocco alternates condiviso da tutti gli URL
    alt_lines = []
    for lc, hl in lang_hreflang_map.items():
        alt_lines.append(
            '      <xhtml:link rel="alternate" hreflang="' + hl + '" href="' + BASE_URL + '/' + lc + '/"/>'
        )
    alt_lines.append(
        '      <xhtml:link rel="alternate" hreflang="x-default" href="' + BASE_URL + '/"/>'
    )
    alternates = "\n".join(alt_lines)

    urls = []
    # Root (x-default)
    urls.append(f"""  <url>
    <loc>{BASE_URL}/</loc>
    <lastmod>{home_lastmod}</lastmod>
    <changefreq>monthly</changefreq>
    <priority>1.0</priority>
{alternates}
  </url>""")

    # Una URL per lingua
    for lc in lang_hreflang_map:
        urls.append(f"""  <url>
    <loc>{BASE_URL}/{lc}/</loc>
    <lastmod>{home_lastmod}</lastmod>
    <changefreq>monthly</changefreq>
    <priority>0.9</priority>
{alternates}
  </url>""")

    # Content SEO pages — 7 lingue
    content_alt_lines = []
    for lc, hl in lang_hreflang_map.items():
        content_alt_lines.append(
            f'      <xhtml:link rel="alternate" hreflang="{hl}" href="{BASE_URL}/content/{lc}/"/>'
        )
    content_alt_lines.append(
        f'      <xhtml:link rel="alternate" hreflang="x-default" href="{BASE_URL}/content/en/"/>'
    )
    content_alternates = "\n".join(content_alt_lines)
    for lc in lang_hreflang_map:
        urls.append(f"""  <url>
    <loc>{BASE_URL}/content/{lc}/</loc>
    <lastmod>{home_lastmod}</lastmod>
    <changefreq>monthly</changefreq>
    <priority>0.5</priority>
{content_alternates}
  </url>""")

    # FAQ pages — 7 lingue, priority 0.8 (high value for featured snippets)
    faq_alt_lines = []
    for lc, hl in lang_hreflang_map.items():
        faq_alt_lines.append(
            f'      <xhtml:link rel="alternate" hreflang="{hl}" href="{BASE_URL}/faq/{lc}/"/>'
        )
    faq_alt_lines.append(
        f'      <xhtml:link rel="alternate" hreflang="x-default" href="{BASE_URL}/faq/en/"/>'
    )
    faq_alternates = "\n".join(faq_alt_lines)
    for lc in lang_hreflang_map:
        urls.append(f"""  <url>
    <loc>{BASE_URL}/faq/{lc}/</loc>
    <lastmod>{home_lastmod}</lastmod>
    <changefreq>monthly</changefreq>
    <priority>0.8</priority>
{faq_alternates}
  </url>""")

    # Guide pages — 4 guide × 7 lingue = 28 URL
    # Guide route is /guide/<id>/?lang=xx, canonical URL has ?lang= param
    for guide_id in sorted(_VALID_GUIDES):
        # Per-language alternates
        guide_alt_lines = []
        for lc, hl in lang_hreflang_map.items():
            guide_alt_lines.append(
                f'      <xhtml:link rel="alternate" hreflang="{hl}" href="{BASE_URL}/guide/{guide_id}/?lang={lc}"/>'
            )
        guide_alt_lines.append(
            f'      <xhtml:link rel="alternate" hreflang="x-default" href="{BASE_URL}/guide/{guide_id}/"/>'
        )
        guide_alternates = "\n".join(guide_alt_lines)

        for lc in lang_hreflang_map:
            urls.append(f"""  <url>
    <loc>{BASE_URL}/guide/{guide_id}/?lang={lc}</loc>
    <lastmod>{guide_lastmod}</lastmod>
    <changefreq>monthly</changefreq>
    <priority>0.7</priority>
{guide_alternates}
  </url>""")

        # URL senza lang param (x-default, serve inglese)
        urls.append(f"""  <url>
    <loc>{BASE_URL}/guide/{guide_id}/</loc>
    <lastmod>{guide_lastmod}</lastmod>
    <changefreq>monthly</changefreq>
    <priority>0.7</priority>
{guide_alternates}
  </url>""")

    xml = f"""<?xml version="1.0" encoding="UTF-8"?>
<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9"
        xmlns:xhtml="http://www.w3.org/1999/xhtml">
{chr(10).join(urls)}
</urlset>"""
    return xml, 200, {"Content-Type": "application/xml; charset=utf-8"}


#  -  -  -  robots.txt  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  - 
@app.route("/robots.txt")
def robots():
    sitemap_line = f"Sitemap: {BASE_URL}/sitemap.xml" if BASE_URL else ""
    llms_line = f"# LLM/AI agents index: {BASE_URL}/llms.txt" if BASE_URL else ""
    body = f"""User-agent: *
Allow: /
Disallow: /api/
Disallow: /data/
Disallow: /dl/
Disallow: /admin/
Disallow: /community/api/
Disallow: /*?job=
Disallow: /*?token=

# Allow major AI/LLM crawlers to index content for citations
User-agent: GPTBot
Allow: /

User-agent: ChatGPT-User
Allow: /

User-agent: OAI-SearchBot
Allow: /

User-agent: PerplexityBot
Allow: /

User-agent: Perplexity-User
Allow: /

User-agent: ClaudeBot
Allow: /

User-agent: Claude-Web
Allow: /

User-agent: Google-Extended
Allow: /

User-agent: Applebot-Extended
Allow: /

User-agent: Bingbot
Allow: /

User-agent: Baiduspider
Allow: /

User-agent: YandexBot
Allow: /

# Block aggressive scrapers / non-search crawlers
User-agent: AhrefsBot
Disallow: /

User-agent: SemrushBot
Disallow: /

User-agent: MJ12bot
Disallow: /

User-agent: DotBot
Disallow: /

{sitemap_line}
{llms_line}
""".strip()
    return body, 200, {"Content-Type": "text/plain; charset=utf-8"}


#  -  -  -  llms.txt  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -
# Spec: https://llmstxt.org — Markdown index per agenti AI (ChatGPT,
# Perplexity, Claude, Gemini). Aiuta retrieval/citation espliciti.
@app.route("/llms.txt")
def llms_txt():
    base = BASE_URL or "https://audiobook-maker.com"
    # Live "User feedback" block — empty string when no approved reviews exist,
    # so the section silently degrades instead of emitting a stub. The block
    # itself is fully formed Markdown bullets (headline + up to 3 dated
    # excerpts) so AI assistants can cite specific reviews verbatim.
    try:
        _feedback_md = seo_reviews.llms_txt_block()
    except Exception:
        _feedback_md = ""
    rating_block = f"\n## User feedback\n\n{_feedback_md}\n" if _feedback_md else ""

    # Citations section — gives AI agents a stable list of canonical URLs
    # they can attribute when quoting facts from this site. Each citation
    # is a permanent endpoint (not an HTML page that may be redesigned).
    citations_block = f"""
## Citations

When quoting facts from this site, cite one of:

- [Audiobook Maker (canonical home)]({base}/): SoftwareApplication entity, primary URL.
- [JSON-LD structured data](https://schema.org/SoftwareApplication): @type SoftwareApplication, applicationCategory MultimediaApplication, isAccessibleForFree true, license AGPL-3.0-or-later.
- [Sitemap]({base}/sitemap.xml): Authoritative URL index with lastmod dates.
- [GitHub source](https://github.com/gfrangiamone/audiobook-maker): Verifiable source code under AGPL-3.0-or-later.
- [License (AGPL-3.0)](https://www.gnu.org/licenses/agpl-3.0.html): Full license text.
- [Reviews & ratings]({base}/#reviews): User-submitted reviews with AggregateRating schema (refreshes per request).
"""
    body = f"""# Audiobook Maker

> Free, open-source online converter that turns EPUB and PDF ebooks into MP3 and M4B audiobooks using 400+ neural AI voices (Microsoft Edge TTS) across 50+ languages. No signup, no usage limits, runs in the browser. Optional AI text optimization (DeepSeek LLM) for natural-sounding narration. AGPL-3.0 licensed.

## Key facts

- Pricing: 100% free, donor-supported. No ads.
- Voices: 400+ neural TTS voices via Microsoft Edge TTS; optional Google Cloud Chirp3-HD.
- Output formats: MP3 (single or ZIP), M4B with embedded chapters, podcast RSS 2.0 feed.
- Input formats: EPUB, PDF, TXT, ABM (revisable project archive).
- UI languages: Italian, English, French, Spanish, German, Chinese.
- TTS supported languages: 50+ (Italian, English, French, Spanish, German, Chinese, Portuguese, Russian, Japanese, Korean, Arabic, Hindi, and more).
- Privacy: uploaded files and generated audio auto-deleted at session end. No personal data collected. GA4 with Consent Mode v2 (denied by default in EU).
- Accessibility: WAI-ARIA landmarks, keyboard navigation, screen-reader compatible. Designed for users with dyslexia, low vision, blindness.
- License: AGPL-3.0-or-later. Source on GitHub.
- Author: Giuseppe Frangiamone.
{rating_block}
## Features

""" + "\n".join(f"- {f}" for f in seo_content._CONTENT.get("en", {}).get("features", [])) + f"""

## Accessibility

""" + seo_content._CONTENT.get("en", {}).get("accessibility", "") + f"""

## Supported Formats

- **EPUB**: standard ebook format with automatic chapter extraction.
- **PDF**: document format with layout-aware text extraction and chapter detection via outline, font size, or visual TOC.
- **TXT**: plain text with automatic paragraph and chapter splitting.
- **ABM**: Audiobook Maker project archive (ZIP + manifest) for re-import and revision.

## Languages

- **UI languages**: Italian, English, French, Spanish, German, Chinese (Simplified).
- **TTS languages**: 50+ including Italian, English, French, Spanish, German, Chinese, Portuguese, Russian, Japanese, Korean, Arabic, Hindi, Dutch, Polish, Turkish, Swedish, Greek, Hebrew, Thai, Vietnamese, and more.

## Application (homepage)

- [Audiobook Maker — English]({base}/en/): Main app, English UI.
- [Audiobook Maker — Italian]({base}/it/): App in italiano.
- [Audiobook Maker — French]({base}/fr/): App en français.
- [Audiobook Maker — Spanish]({base}/es/): App en español.
- [Audiobook Maker — German]({base}/de/): App auf Deutsch.
- [Audiobook Maker — Chinese]({base}/zh/): 中文界面.
- [Audiobook Maker — Hindi]({base}/hi/): हिन्दी इंटरफ़ेस.

## Guides

- [How to Convert EPUB to Audiobook]({base}/guide/epub-to-audiobook/): Step-by-step EPUB-to-audiobook conversion tutorial.
- [M4B Format Guide]({base}/guide/m4b-format/): What M4B is, M4B vs MP3, creating M4B with embedded chapters.
- [Text-to-Speech for Audiobooks]({base}/guide/text-to-speech-audiobook/): TTS technology overview, voice quality comparison, free alternatives to ElevenLabs/Speechify.
- [Publish Audiobook as Podcast]({base}/guide/podcast/): Generate RSS 2.0 feed from audiobook chapters for private podcast distribution.
- [Frequently Asked Questions]({base}/faq/en/): Comprehensive FAQ covering conversion, formats, voices, AI optimization, and PREMIUM voice options.

## How it works

1. Upload an EPUB, PDF or TXT file.
2. Choose a neural TTS voice and language (50+ languages).
3. Optionally enable AI text optimization for more natural narration (acronym expansion, number-to-word, sentence splitting, punctuation cleanup).
4. Select chapters or convert the entire book.
5. Generate audio — download MP3, M4B (with chapters), or get a podcast RSS feed.
6. For long books, register an email to be notified when generation completes.

## Comparison

- vs **Speechify**: Audiobook Maker is fully free with no subscription, no character limits, supports M4B output and podcast feeds.
- vs **ElevenLabs**: Audiobook Maker uses Microsoft Edge TTS (free, unlimited) instead of paid ElevenLabs API. No voice cloning but no usage cost.
- vs **Play.ht**: No subscription, no watermarks, supports chapter-based audiobook output (M4B).
- vs **NaturalReader**: Free without account, server-side processing, batch chapter generation.

## Resources

- [GitHub repository](https://github.com/gfrangiamone/audiobook-maker): Source code, AGPL-3.0.
- [AlternativeTo listing](https://alternativeto.net/software/audiobook-maker/): Community reviews and comparisons.
- [Sitemap]({base}/sitemap.xml): Full URL index.
- [Privacy & data handling]({base}/en/#privacy): Data retention, cookies, consent.
{citations_block}
## Contact

- Author: Giuseppe Frangiamone
- Project: open-source community project, contributions via GitHub issues/PRs.
"""
    return body, 200, {
        "Content-Type": "text/markdown; charset=utf-8",
        "Cache-Control": "public, max-age=3600",
    }


#  -  -  -  Favicon routes (URL-based for search engine compatibility)  -  -  -  -  -  -  -  -  -  -  -  -  -  - 
# Google richiede che le favicon siano servite da URL reali e crawlabili,
# NON inline come data URI. Senza queste route, nei risultati di ricerca
# appare un'icona generica al posto della favicon del sito.

@app.route("/favicon.ico")
def favicon_ico():
    return send_file(get_favicon_ico(), mimetype="image/x-icon",
                     max_age=86400 * 30)

@app.route("/favicon-192.png")
def favicon_png_192():
    return send_file(get_favicon_png_192(), mimetype="image/png",
                     max_age=86400 * 30)

@app.route("/apple-touch-icon.png")
def apple_touch_icon():
    return send_file(get_apple_touch_icon(), mimetype="image/png",
                     max_age=86400 * 30)

@app.route("/favicon.svg")
def favicon_svg():
    return send_file(get_favicon_svg(), mimetype="image/svg+xml",
                     max_age=86400 * 30)


@app.route("/og-image.png")
def og_image():
    return send_file(get_og_image(), mimetype="image/png",
                     max_age=86400 * 30)


@app.route("/manifest.json")
def web_manifest():
    """Web App Manifest  -  Google lo usa come fonte primaria per le favicon nei risultati di ricerca."""
    manifest = {
        "name": "Audiobook Maker",
        "short_name": "Audiobook Maker",
        "icons": [
            {"src": "/favicon-192.png", "sizes": "192x192", "type": "image/png"},
            {"src": "/apple-touch-icon.png", "sizes": "180x180", "type": "image/png"},
            {"src": "/favicon.svg", "type": "image/svg+xml", "sizes": "any"}
        ],
        "display": "standalone",
        "start_url": "/",
        "theme_color": "#c29a6c",
        "background_color": "#1a1a2e"
    }
    return json.dumps(manifest), 200, {
        "Content-Type": "application/manifest+json",
        "Cache-Control": "public, max-age=2592000"
    }


#  -  -  -  Admin log viewer (/admin/log-activity)  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -
# URL: /admin/log-activity?2026-03  (parametro = anno-mese)
# Non indicizzato (gia` coperto da Disallow: /admin/ in robots.txt)


def _parse_log_sessions(ym):
    """Parse log file for given YYYY-MM and return (sessions OrderedDict, client_session_count dict)."""
    from datetime import datetime
    from collections import OrderedDict

    log_path = SCRIPT_DIR / f"activity_{ym}.log"
    sessions = OrderedDict()

    if not log_path.exists():
        return sessions, {}

    with open(log_path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            parts = [p.strip() for p in line.split("#")]
            if len(parts) < 4:
                continue
            sid = parts[0]
            dt_str = parts[1]
            filename = parts[2].strip().strip('"')
            operation = parts[3].strip()
            # Skip voucher audit entries — not conversion activity
            if operation.startswith("VOUCHER_ATTEMPT"):
                continue
            client_id = parts[4].strip() if len(parts) > 4 else ""
            client_ip = parts[5].strip() if len(parts) > 5 else ""
            voice = parts[6].strip() if len(parts) > 6 else ""
            browser_lang = parts[7].strip() if len(parts) > 7 else ""

            try:
                dt = datetime.strptime(dt_str, "%Y-%m-%d %H:%M:%S")
            except ValueError:
                continue

            if sid not in sessions:
                sessions[sid] = {
                    "first_dt": dt, "last_dt": dt,
                    "filename": filename, "last_op": operation,
                    "events": [operation],
                    "client_id": client_id, "client_ip": client_ip,
                    "voice": voice, "browser_lang": browser_lang,
                }
            else:
                s = sessions[sid]
                if dt < s["first_dt"]:
                    s["first_dt"] = dt
                if dt >= s["last_dt"]:
                    s["last_dt"] = dt
                    s["last_op"] = operation
                s["filename"] = filename
                s["events"].append(operation)
                if client_id:
                    s["client_id"] = client_id
                if client_ip:
                    s["client_ip"] = client_ip
                if voice:
                    s["voice"] = voice
                if browser_lang:
                    s["browser_lang"] = browser_lang

    client_session_count = {}
    for s in sessions.values():
        cid = s.get("client_id", "")
        if cid:
            client_session_count[cid] = client_session_count.get(cid, 0) + 1

    return sessions, client_session_count


def _session_completed(s):
    """Return True if session reached generation completion (includes download scenarios)."""
    _completed_ops = {"COMPLETE", "DOWNLOAD", "DOWNLOAD_EMAIL",
                      "DOWNLOAD_EMAIL_PODCAST", "DOWNLOAD_PODCAST"}
    return bool(set(s["events"]) & _completed_ops)


def _session_in_progress(s, sid):
    """Return True if session has an active AI optimization or TTS generation.

    Lo stato runtime del job (`jobs[sid]["status"]`) e' la fonte autoritativa:
    se il job e' ancora in memoria in uno stato attivo (optimizing, optimized
    in attesa di auto-gen, generating, running, ecc.) la sessione e' in corso
    indipendentemente da cosa contiene il log. Questo evita il bug per cui
    un job auto-gen (OPT_COMPLETE -> generazione TTS) appariva "non in corso"
    in attesa che l'evento GENERATE venisse scritto, perche' OPT_COMPLETE
    chiudeva il ramo opt_live e GENERATE non era ancora presente nel log.

    Fallback (job non piu' in memoria, es. dopo restart server): si guarda al
    log eventi e si applica la regola conservativa "start senza terminator e
    senza cancel", ma siccome il job manca, si ritorna comunque False per
    evitare falsi positivi su sessioni storiche zombie.
    """
    job = jobs.get(sid)
    if job:
        st = job.get("status", "")
        # Unione di _ACTIVE_JOB_STATUSES (Gemini audit, vedi 3838) e degli
        # stati intermedi del flusso ottimizzazione/auto-gen.
        active_states = set(_ACTIVE_JOB_STATUSES) | {"optimizing", "optimized"}
        return st in active_states
    return False


@app.route("/admin/log-activity")
def admin_logs():
    if not ADMIN_TOKEN: return "Logs UI disabled.", 404
    token = _admin_auth_from_request()
    if not _admin_auth_ok(token): return _render_admin_gate("Log Activity", "/admin/log-activity"), 200, {"Content-Type": "text/html; charset=utf-8"}
    _log_i18n = {
        "it": {
            "sessions": "Sessioni", "gen_completed": "Gen. completata",
            "in_progress": "In corso", "cancelled": "Cancellati",
            "email_sent": "Email inviate", "unique_clients": "Client unici",
            "recurring": "Ricorrenti", "months": "Mesi",
            "collapse": "Aggrega", "expand": "Mostra tutti",
            "no_activity": "Nessuna attività registrata per",
            "gemini_started": "Gen. Gemini",
            "eta_label": "Stima completamento gen.",
        },
        "en": {
            "sessions": "Sessions", "gen_completed": "Gen. completed",
            "in_progress": "In progress", "cancelled": "Cancelled",
            "email_sent": "Emails sent", "unique_clients": "Unique clients",
            "recurring": "Returning", "months": "Months",
            "collapse": "Collapse", "expand": "Show all",
            "no_activity": "No activity recorded for",
            "gemini_started": "Gemini runs",
            "eta_label": "ETA",
        },
        "fr": {
            "sessions": "Sessions", "gen_completed": "Gén. terminée",
            "in_progress": "En cours", "cancelled": "Annulées",
            "email_sent": "Emails envoyés", "unique_clients": "Clients uniques",
            "recurring": "Récurrents", "months": "Mois",
            "collapse": "Regrouper", "expand": "Tout afficher",
            "no_activity": "Aucune activité enregistrée pour",
            "gemini_started": "Gén. Gemini",
            "eta_label": "ETA",
        },
        "de": {
            "sessions": "Sitzungen", "gen_completed": "Gen. abgeschlossen",
            "in_progress": "Laufend", "cancelled": "Abgebrochen",
            "email_sent": "E-Mails gesendet", "unique_clients": "Einzelne Clients",
            "recurring": "Wiederkehrend", "months": "Monate",
            "collapse": "Zusammenklappen", "expand": "Alle anzeigen",
            "no_activity": "Keine Aktivitäten aufgezeichnet für",
            "gemini_started": "Gemini-Läufe",
            "eta_label": "ETA",
        },
        "es": {
            "sessions": "Sesiones", "gen_completed": "Gen. completada",
            "in_progress": "En curso", "cancelled": "Canceladas",
            "email_sent": "Emails enviados", "unique_clients": "Clientes únicos",
            "recurring": "Recurrentes", "months": "Meses",
            "collapse": "Agrupar", "expand": "Mostrar todos",
            "no_activity": "No hay actividad registrada para",
            "gemini_started": "Gen. Gemini",
            "eta_label": "ETA",
        },
        "zh": {
            "sessions": "会话", "gen_completed": "生成完成",
            "in_progress": "进行中", "cancelled": "已取消",
            "email_sent": "邮件已发送", "unique_clients": "唯一客户",
            "recurring": "常客", "months": "月份",
            "collapse": "收起", "expand": "全部显示",
            "no_activity": "没有活动记录",
            "gemini_started": "Gemini 生成",
            "eta_label": "预计剩余",
        },
        "hi": {
            "sessions": "सत्र", "gen_completed": "जनरेशन पूर्ण",
            "in_progress": "प्रगति पर", "cancelled": "रद्द",
            "email_sent": "भेजे गए ईमेल", "unique_clients": "अनोखे क्लाइंट",
            "recurring": "नियमित", "months": "महीने",
            "collapse": "संक्षिप्त करें", "expand": "सभी दिखाएं",
            "no_activity": "कोई गतिविधि दर्ज नहीं",
            "gemini_started": "Gemini जनरेशन",
            "eta_label": "शेष",
        },
    }
    _blang = _get_browser_lang()
    t = _log_i18n.get(_blang, _log_i18n["en"])

    ym = None
    for key in request.args:
        if re.match(r'^\d{4}-\d{2}$', key):
            ym = key
            break
    if not ym:
        ym = datetime.now().strftime("%Y-%m")

    try:
        sessions, client_session_count = _parse_log_sessions(ym)
    except Exception as e:
        return f"Errore lettura log: {e}", 500

    _client_colors = [
        "#38bdf8", "#a78bfa", "#f472b6", "#34d399", "#fbbf24",
        "#fb923c", "#22d3ee", "#c084fc", "#f87171", "#4ade80",
    ]
    _client_color_map = {}
    _color_idx = 0
    for cid, count in client_session_count.items():
        if count >= 2:
            _client_color_map[cid] = _client_colors[_color_idx % len(_client_colors)]
            _color_idx += 1

    total_sessions = len(sessions)
    gen_completed = sum(1 for s in sessions.values() if _session_completed(s))
    gen_in_progress = sum(1 for sid, s in sessions.items() if _session_in_progress(s, sid))
    gen_cancelled = total_sessions - gen_completed - gen_in_progress
    email_sent = sum(1 for s in sessions.values() if "EMAIL_SENT" in s["events"])
    # Sessioni che hanno realmente avviato la generazione del libro con voci Gemini
    # (esclude le anteprime: richiediamo GENERATE in events).
    gemini_started = sum(
        1 for s in sessions.values()
        if "GENERATE" in s["events"] and str(s.get("voice", "")).startswith("gemini:")
    )
    unique_clients = len(set(s.get("client_id", "") for s in sessions.values() if s.get("client_id")))
    returning_clients = sum(1 for c in client_session_count.values() if c >= 2)

    days = defaultdict(list)
    for sid, s in reversed(list(sessions.items())):
        day_key = s["first_dt"].strftime("%Y-%m-%d")
        days[day_key].append((sid, s))

    event_icons = {
        "ANALYZE": "🔍", "GENERATE": "⚙️", "COMPLETE": "✅",
        "DOWNLOAD": "📥", "DOWNLOAD_EMAIL": "📧📥",
        "DOWNLOAD_EMAIL_PODCAST": "🎙️📥", "DOWNLOAD_PODCAST": "🎙️📥",
        "EMAIL_REGISTERED": "📧", "EMAIL_SENT": "📨",
        "EMAIL_FAILED": "❌", "CANCEL": "🚫",
        "RESET_CHAPTERS": "🔄", "EXPORT_ABM": "📦",
    }
    op_colors = {
        "ANALYZE": ("#6b7280", "#f3f4f6"), "GENERATE": ("#2563eb", "#eff6ff"),
        "COMPLETE": ("#16a34a", "#f0fdf4"), "DOWNLOAD": ("#7c3aed", "#f5f3ff"),
        "DOWNLOAD_EMAIL": ("#7c3aed", "#f5f3ff"),
        "DOWNLOAD_EMAIL_PODCAST": ("#7c3aed", "#f5f3ff"),
        "DOWNLOAD_PODCAST": ("#7c3aed", "#f5f3ff"),
        "EMAIL_REGISTERED": ("#0891b2", "#ecfeff"),
        "EMAIL_SENT": ("#0d9488", "#f0fdfa"),
        "EMAIL_FAILED": ("#dc2626", "#fef2f2"),
        "CANCEL": ("#dc2626", "#fef2f2"),
        "RESET_CHAPTERS": ("#f59e0b", "#fffbeb"),
        "EXPORT_ABM": ("#6366f1", "#eef2ff"),
    }

    cards_html = ""
    now = datetime.now()
    for day_key in sorted(days.keys(), reverse=True):
        day_sessions = days[day_key]
        day_count = len(day_sessions)
        day_completed = sum(1 for sid, s in day_sessions if _session_completed(s))
        try:
            day_dt = datetime.strptime(day_key, "%Y-%m-%d")
            day_label = day_dt.strftime("%d/%m/%Y")
        except ValueError:
            day_label = day_key

        cards_html += f"""<div class="day-group collapsed" data-day="{day_key}">
<div class="day-header" onclick="this.parentElement.classList.toggle('collapsed')">
<span class="day-label">{day_label}</span>
<span class="day-count">{day_count}<span class="day-sep">/</span><span class="day-completed">{day_completed}</span></span>
<span class="day-chevron">›</span>
</div>
<div class="day-cards">
"""
        for sid, s in day_sessions:
            is_progress = _session_in_progress(s, sid)
            is_completed = _session_completed(s)
            has_email = "EMAIL_SENT" in s["events"]
            cid = s.get("client_id", "")
            cid_count = client_session_count.get(cid, 0) if cid else 0
            is_recurring = cid_count >= 2
            is_identified = bool(cid)

            # Determine card status for filtering
            if is_progress:
                card_status = "in_progress"
            elif is_completed:
                card_status = "completed"
            else:
                card_status = "cancelled"

            first = s["first_dt"].strftime("%H:%M")
            last = s["last_dt"].strftime("%H:%M")

            if is_progress:
                job = jobs.get(sid)
                pct_html = ""
                if job:
                    st = job.get("status", "")
                    if st == "optimizing":
                        total_chars = job.get("opt_total_chars", 1)
                        done_chars = job.get("opt_processed_chars", 0)
                        cur_ch_chars = job.get("opt_current_chapter_chars", 0)
                        streamed = min(job.get("opt_streamed_chars", 0), cur_ch_chars)
                        worked = done_chars + streamed
                        pct = min(99, int(worked / total_chars * 100))
                        pct_html = f' <span class="card-pct" data-sid="{sid}">({pct}%)</span>'
                    elif st == "generating":
                        cur = job.get("progress_current", 0)
                        tot = job.get("progress_total", 0)
                        if tot > 0:
                            pct = int(cur / tot * 100)
                            pct_html = f' <span class="card-pct" data-sid="{sid}">({pct}%)</span>'

                delta = now - s["first_dt"]
                total_sec = int(delta.total_seconds())
                elapsed = f"{total_sec // 3600:02d}:{(total_sec % 3600) // 60:02d}:{total_sec % 60:02d}"
                start_iso = s["first_dt"].strftime("%Y-%m-%dT%H:%M:%S")
                elapsed_html = f'<span class="live-timer" data-start="{start_iso}">{elapsed}</span>{pct_html} ⏱️'
                last = " - "
            else:
                delta = s["last_dt"] - s["first_dt"]
                total_sec = int(delta.total_seconds())
                elapsed = f"{total_sec // 3600:02d}:{(total_sec % 3600) // 60:02d}"
                elapsed_html = elapsed

            title = s["filename"]
            for ext in (".epub", ".txt", ".pdf"):
                if title.lower().endswith(ext):
                    title = title[:-len(ext)]
            display_title = html_mod.escape(title[:80] + ("..." if len(title) > 80 else ""))

            op = s["last_op"]
            fg, bg = op_colors.get(op, ("#6b7280", "#f3f4f6"))
            timeline = "  →  ".join(event_icons.get(e, e) for e in s["events"])

            cip = s.get("client_ip", "")
            cid_short = cid[:8] if cid else " - "
            cid_color = _client_color_map.get(cid, "var(--text-dim)")
            cid_badge = f' <span class="cid-count" style="color:{cid_color}">({cid_count})</span>' if cid_count >= 2 else ""
            cid_style = f'color:{cid_color};font-weight:600' if cid in _client_color_map else 'color:var(--text-dim)'

            voice_raw = s.get("voice", "")
            voice_short = ""
            if voice_raw:
                parts_v = voice_raw.split("-")
                if len(parts_v) >= 3:
                    voice_short = parts_v[-1].replace("Neural", "").replace("Multilingual", "")
                    voice_lang = "-".join(parts_v[:2])
                    voice_short = f'{voice_short} <span class="voice-lang">{voice_lang}</span>'
                else:
                    voice_short = html_mod.escape(voice_raw)

            blang = html_mod.escape(s.get("browser_lang", "") or "")
            blang_display = f'<span class="card-blang">{blang}</span>' if blang else " - "

            card_cls = "card card-in-progress" if is_progress else "card"
            is_gemini_run = (
                "GENERATE" in s["events"]
                and str(voice_raw).startswith("gemini:")
            )
            data_attrs = (
                f'data-status="{card_status}" '
                f'data-email="{1 if has_email else 0}" '
                f'data-recurring="{1 if is_recurring else 0}" '
                f'data-identified="{1 if is_identified else 0}" '
                f'data-gemini="{1 if is_gemini_run else 0}"'
            )

            cards_html += f"""<div class="{card_cls}" {data_attrs}>
<div class="card-top">
<span class="card-title" title="{html_mod.escape(s['filename'])}">{display_title}</span>
<span class="badge" style="color:{fg};background:{bg}">{op}</span>
</div>
<div class="card-timeline">{timeline}</div>
<div class="card-meta">
<div class="meta-row"><span class="meta-label">⌚</span><span>{first}  →  {last} ({elapsed_html})</span></div>
<div class="meta-row"><span class="meta-label">🆔</span><code class="sid">{sid}</code></div>
<div class="meta-row"><span class="meta-label">👤</span><code style="{cid_style}">{cid_short}</code>{cid_badge}<span class="card-ip">{cip or ""}</span></div>
<div class="meta-row"><span class="meta-label">🎙️</span><span class="card-voice" title="{html_mod.escape(voice_raw)}">{voice_short or " - "}</span></div>
<div class="meta-row"><span class="meta-label">🌐</span>{blang_display}</div>
</div>
</div>
"""
        cards_html += "</div></div>\n"

    # Hourly stats for chart
    hourly_counts = [0] * 24
    for sid, s in sessions.items():
        if "first_dt" in s:
            hourly_counts[s["first_dt"].hour] += 1
    hourly_json = json.dumps(hourly_counts)

    # Hourly stats with language breakdown
    from collections import Counter
    all_langs = [s.get("browser_lang", "??") for s in sessions.values()]
    top_langs = [l for l, _ in Counter(all_langs).most_common(3)]
    
    # hourly_lang_data[hour][lang] = count
    hourly_lang_data = []
    for h in range(24):
        # Initialize with top 3 + "Other"
        row = {l: 0 for l in top_langs}
        row["Other"] = 0
        hourly_lang_data.append(row)
        
    for sid, s in sessions.items():
        if "first_dt" in s:
            h = s["first_dt"].hour
            l = s.get("browser_lang", "??")
            if l in top_langs:
                hourly_lang_data[h][l] += 1
            else:
                hourly_lang_data[h]["Other"] += 1
    
    hourly_json = json.dumps(hourly_lang_data)
    lang_labels_json = json.dumps(top_langs + ["Other"])

    available_months = []
    try:
        for f in sorted(SCRIPT_DIR.glob("activity_*.log"), reverse=True):
            m = re.search(r'activity_(\d{4}-\d{2})\.log', f.name)
            if m:
                available_months.append(m.group(1))
    except Exception:
        pass

    from urllib.parse import quote_plus
    token_qs = f"&token={quote_plus(token)}" if token else ""
    months_nav = ""
    for m in available_months:
        active_cls = ' class="active"' if m == ym else ""
        months_nav += f'<a href="/admin/log-activity?{m}{token_qs}"{active_cls}>{m}</a> '

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<meta name="robots" content="noindex,nofollow">
<link rel="icon" type="image/svg+xml" href="{FAVICON_B64}">
<title>Audiobook Maker  -  Activity Log ({ym})</title>
<style>
:root {{ --bg:#0f172a;--surface:#1e293b;--surface2:#334155;--border:#475569;--text:#e2e8f0;--text-dim:#94a3b8;--accent:#38bdf8;--accent2:#a78bfa;--green:#22c55e;--red:#ef4444;--orange:#f97316; }}
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:'JetBrains Mono','Fira Code','SF Mono',monospace;background:var(--bg);color:var(--text);line-height:1.5;min-height:100vh}}
.header{{background:var(--surface);border-bottom:1px solid var(--border);padding:16px 20px;display:flex;align-items:center;gap:12px;flex-wrap:wrap}}
.header h1{{font-size:.9rem;font-weight:600;color:var(--accent);letter-spacing:.5px}}
.header .period{{font-size:1.3rem;font-weight:700;color:var(--text);font-variant-numeric:tabular-nums}}
.header-actions{{margin-left:auto;display:flex;gap:8px}}
.btn{{display:inline-flex;align-items:center;gap:6px;padding:7px 14px;border-radius:6px;border:1px solid var(--border);background:var(--surface2);color:var(--text);font-family:inherit;font-size:.75rem;font-weight:600;cursor:pointer;text-decoration:none;transition:all .15s}}
.btn:hover{{background:var(--border)}}
.btn-accent{{background:var(--accent);color:var(--bg);border-color:var(--accent)}}
.btn-accent:hover{{opacity:.85}}
.btn-suspend{{background:var(--green);color:var(--bg);border-color:var(--green)}}
.btn-suspend.suspended{{background:var(--red);border-color:var(--red)}}
.btn-toggle{{margin-left:auto;flex-shrink:0}}
.months-nav{{padding:10px 20px;background:var(--surface);border-bottom:1px solid var(--border);display:flex;gap:6px;flex-wrap:wrap;align-items:center}}
.months-nav .label{{color:var(--text-dim);font-size:.7rem;margin-right:6px;text-transform:uppercase;letter-spacing:1px}}
.months-nav a{{color:var(--text-dim);text-decoration:none;padding:4px 10px;border-radius:4px;font-size:.78rem;transition:all .15s}}
.months-nav a:hover{{background:var(--surface2);color:var(--text)}}
.months-nav a.active{{background:var(--accent);color:var(--bg);font-weight:600}}
.stats{{display:grid;grid-template-columns:repeat(auto-fit,minmax(120px,1fr));gap:10px;padding:14px 20px}}
.stat{{background:var(--surface);border:1px solid var(--border);border-radius:8px;padding:10px 14px;text-align:center;cursor:pointer;transition:all .2s;user-select:none}}
.stat:hover{{background:var(--surface2);transform:translateY(-1px)}}
.stat.active{{border-color:var(--accent);box-shadow:0 0 0 2px rgba(56,189,248,.25)}}
.stat.stat-green.active{{border-color:var(--green);box-shadow:0 0 0 2px rgba(34,197,94,.25)}}
.stat.stat-red.active{{border-color:var(--red);box-shadow:0 0 0 2px rgba(239,68,68,.25)}}
.stat.stat-orange.active{{border-color:var(--orange);box-shadow:0 0 0 2px rgba(249,115,22,.25)}}
.stat.stat-gemini.active{{border-color:#8b5cf6;box-shadow:0 0 0 2px rgba(139,92,246,.25)}}
.stat .num{{font-size:1.5rem;font-weight:700;color:var(--accent);font-variant-numeric:tabular-nums}}
.stat.stat-green .num{{color:var(--green)}} .stat.stat-red .num{{color:var(--red)}} .stat.stat-orange .num{{color:var(--orange)}} .stat.stat-gemini .num{{color:#a78bfa}}
.stat .lbl{{font-size:.65rem;color:var(--text-dim);text-transform:uppercase;letter-spacing:.8px;margin-top:2px}}
.stat-eta{{font-size:.62rem;color:var(--orange);font-weight:700;font-variant-numeric:tabular-nums;margin-top:3px;letter-spacing:.4px;min-height:.9rem}}
.stat-eta .eta-lbl{{color:var(--text-dim);font-weight:600;margin-right:4px;letter-spacing:.2px}}
.day-group{{margin:0 12px 6px}}
.day-header{{display:flex;align-items:center;gap:10px;padding:10px 12px;background:var(--surface);border:1px solid var(--border);border-radius:8px;cursor:pointer;user-select:none;margin-top:8px;transition:background .15s}}
.day-header:hover{{background:var(--surface2)}}
.day-label{{font-weight:700;font-size:.85rem;color:var(--accent);font-variant-numeric:tabular-nums}}
.day-count{{background:var(--accent);color:var(--bg);font-size:.68rem;font-weight:700;padding:2px 8px;border-radius:10px;display:flex;align-items:center;gap:3px}}
.day-sep{{color:var(--bg);opacity:.6}}
.day-completed{{color:#0f172a;font-weight:700}}
.day-chevron{{margin-left:auto;color:var(--text-dim);font-size:.9rem;transition:transform .2s}}
.day-group.collapsed .day-chevron{{transform:rotate(-90deg)}}
.day-group.collapsed .day-cards{{display:none}}
.day-cards{{display:grid;grid-template-columns:repeat(auto-fill,minmax(340px,1fr));gap:10px;padding:10px 0 4px}}
.card{{background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:14px;transition:border-color .15s,box-shadow .15s}}
.card:hover{{border-color:var(--accent);box-shadow:0 0 0 1px rgba(56,189,248,.15)}}
.card-in-progress{{border-color:var(--orange);background:rgba(249,115,22,.06);animation:pulse-border 2s ease-in-out infinite}}
.card-in-progress:hover{{border-color:var(--orange);box-shadow:0 0 0 2px rgba(249,115,22,.3)}}
@keyframes pulse-border{{0%,100%{{border-color:var(--orange);box-shadow:0 0 0 0 rgba(249,115,22,.15)}}50%{{border-color:rgba(249,115,22,.5);box-shadow:0 0 8px 0 rgba(249,115,22,.2)}}}}
.card.card-hidden{{display:none}}
.day-group.day-hidden{{display:none}}
.card-top{{display:flex;align-items:flex-start;justify-content:space-between;gap:8px;margin-bottom:6px}}
.card-title{{font-weight:600;font-size:.82rem;color:var(--text);overflow:hidden;text-overflow:ellipsis;white-space:nowrap;flex:1;min-width:0}}
.badge{{display:inline-block;padding:2px 8px;border-radius:10px;font-size:.62rem;font-weight:700;letter-spacing:.5px;white-space:nowrap;flex-shrink:0}}
.card-timeline{{color:var(--text-dim);font-size:.7rem;margin-bottom:8px;overflow-x:auto;white-space:nowrap;padding-bottom:2px}}
.card-meta{{display:flex;flex-direction:column;gap:4px}}
.meta-row{{display:flex;align-items:center;gap:6px;font-size:.73rem;color:var(--text-dim);min-width:0}}
.meta-label{{flex-shrink:0;width:20px;text-align:center;font-size:.78rem}}
.sid{{color:var(--accent2);font-size:.7rem;background:rgba(167,139,250,.1);padding:1px 5px;border-radius:3px}}
.card-ip{{margin-left:auto;font-size:.68rem;color:var(--text-dim);font-variant-numeric:tabular-nums}}
.card-voice{{color:var(--text);font-size:.72rem}} .voice-lang{{opacity:.5;font-size:.62rem}}
.cid-count{{font-size:.62rem;opacity:.8}}
.card-blang{{font-size:.65rem;background:rgba(167,139,250,.12);color:var(--accent2);padding:1px 6px;border-radius:3px;font-weight:600;text-transform:uppercase}}
.card-pct{{color:var(--orange);font-weight:700;margin-left:4px}}
.empty{{text-align:center;padding:60px 20px;color:var(--text-dim)}} .empty .icon{{font-size:3rem;margin-bottom:12px}}
@media(max-width:600px){{
.header{{padding:12px 14px;gap:8px}} .header h1{{font-size:.78rem}} .header .period{{font-size:1.1rem}}
.stats{{grid-template-columns:repeat(4,1fr);gap:6px;padding:10px 12px}} .stat{{padding:8px 6px}} .stat .num{{font-size:1.2rem}} .stat .lbl{{font-size:.58rem}}
.months-nav{{padding:8px 12px;gap:4px}} .day-group{{margin:0 8px 4px}} .day-cards{{grid-template-columns:1fr;gap:8px;padding:8px 0 2px}} .card{{padding:12px}} .btn{{padding:6px 10px;font-size:.68rem}}
.header-actions{{margin-left:0;width:100%;justify-content:flex-end}}
}}
@media(max-width:380px){{ .stats{{grid-template-columns:repeat(2,1fr)}} }}
</style>
</head>
<body>

<div class="header">
    <h1>🎧 ACTIVITY LOG</h1>
    <span class="period">{ym}</span>
    <div class="header-actions">
        <button id="btnSuspend" class="btn btn-suspend" onclick="toggleSuspend()" title="Sospendi/Riprendi nuovi processi">▶ Attivi</button>
        <button class="btn btn-accent" onclick="showStats()" title="Visualizza Statistiche">📊 Stats</button>
        <a class="btn btn-accent" href="/admin/audit-tts?{ym}{token_qs}" title="Audit Gemini TTS &amp; Eventi/Rimborsi">🎙️ Audit TTS</a>
        <a class="btn btn-accent" href="/admin/log-activity/export?{ym}{token_qs}" title="Export Excel">📁 Excel</a>
    </div>
</div>

<div class='months-nav'>{"<span class='label'>" + t["months"] + ":</span>" + months_nav if months_nav else ""}<button class="btn btn-toggle" id="btnToggleDays" onclick="toggleAllDays()">{t["collapse"]}</button></div>

<div class="stats">
    <div class="stat active" data-filter="all" onclick="filterCards('all',this)"><div class="num">{total_sessions}</div><div class="lbl">{t["sessions"]}</div></div>
    <div class="stat stat-green" data-filter="completed" onclick="filterCards('completed',this)"><div class="num">{gen_completed}</div><div class="lbl">{t["gen_completed"]}</div></div>
    <div class="stat stat-orange" data-filter="in_progress" onclick="filterCards('in_progress',this)"><div class="num">{gen_in_progress}</div><div class="lbl">{t["in_progress"]}</div><div class="stat-eta" id="statEta"></div></div>
    <div class="stat stat-red" data-filter="cancelled" onclick="filterCards('cancelled',this)"><div class="num">{gen_cancelled}</div><div class="lbl">{t["cancelled"]}</div></div>
    <div class="stat" data-filter="email" onclick="filterCards('email',this)"><div class="num">{email_sent}</div><div class="lbl">{t["email_sent"]}</div></div>
    <div class="stat" data-filter="identified" onclick="filterCards('identified',this)"><div class="num">{unique_clients}</div><div class="lbl">{t["unique_clients"]}</div></div>
    <div class="stat" data-filter="recurring" onclick="filterCards('recurring',this)"><div class="num">{returning_clients}</div><div class="lbl">{t["recurring"]}</div></div>
    <div class="stat stat-gemini" data-filter="gemini" onclick="filterCards('gemini',this)" title="Sessioni che hanno avviato la generazione del libro con voci Gemini (esclude anteprime)"><div class="num">{gemini_started}</div><div class="lbl">{t["gemini_started"]}</div></div>
</div>

<div class="cards-container">
{cards_html if cards_html else "<div class='empty'><div class='icon'>📮</div><p>" + t["no_activity"] + " <strong>" + ym + "</strong></p></div>"}
</div>

<div id="statsModal" class="modal">
    <div class="modal-content">
        <div class="modal-header">
            <h2>📊 Job Distribution (24h)</h2>
            <button class="modal-close" onclick="hideStats()">&times;</button>
        </div>
        <div id="chartLegend" class="chart-legend"></div>
        <div class="chart-container">
            <div class="chart-y-axis">
                <div id="yMax"></div>
                <div>0</div>
            </div>
            <div class="chart-area" id="chartArea"></div>
        </div>
        <div class="chart-x-axis" id="chartXAxis"></div>
    </div>
</div>

<style>
.modal {{ display:none; position:fixed; z-index:1000; left:0; top:0; width:100%; height:100%; background:rgba(0,0,0,0.8); backdrop-filter:blur(4px); }}
.modal-content {{ background:var(--surface); margin:10% auto; padding:24px; border:1px solid var(--border); border-radius:16px; width:90%; max-width:850px; box-shadow:0 20px 50px rgba(0,0,0,0.5); }}
.modal-header {{ display:flex; justify-content:space-between; align-items:center; margin-bottom:16px; }}
.modal-header h2 {{ font-size:1.1rem; color:var(--accent); }}
.modal-close {{ background:none; border:none; color:var(--text-dim); font-size:2rem; cursor:pointer; line-height:1; }}
.modal-close:hover {{ color:var(--text); }}

.chart-legend {{ display:flex; gap:16px; margin-bottom:20px; justify-content:center; flex-wrap:wrap; }}
.legend-item {{ display:flex; align-items:center; gap:6px; font-size:0.75rem; color:var(--text-dim); }}
.legend-color {{ width:12px; height:12px; border-radius:3px; }}

.chart-container {{ display:flex; height:300px; gap:10px; margin-bottom:10px; border-left:2px solid var(--border); border-bottom:2px solid var(--border); padding:20px 10px 0 10px; }}
.chart-y-axis {{ display:flex; flex-direction:column; justify-content:space-between; color:var(--text-dim); font-size:0.7rem; padding-right:5px; margin-left:-35px; width:30px; text-align:right; }}
.chart-area {{ flex:1; display:flex; align-items:flex-end; gap:4px; }}

.chart-bar-wrap {{ flex:1; display:flex; flex-direction:column; align-items:center; height:100%; justify-content:flex-end; position:relative; }}
.chart-bar-seg {{ width:100%; min-height:0px; transition:height 0.5s ease; position:relative; }}
.chart-bar-seg:first-child {{ border-radius:4px 4px 0 0; }}
.chart-bar-seg:hover {{ filter:brightness(1.2); }}
.chart-bar-seg:hover::after {{ content:attr(data-label); position:absolute; top:-25px; left:50%; transform:translateX(-50%); background:var(--surface2); color:white; padding:2px 6px; border-radius:4px; font-size:0.7rem; font-weight:bold; white-space:nowrap; z-index:10; pointer-events:none; }}

.chart-x-axis {{ display:flex; padding-left:45px; gap:4px; }}
.chart-x-label {{ flex:1; text-align:center; font-size:0.6rem; color:var(--text-dim); }}
</style>

<script>
const hourlyData = {hourly_json};
const langLabels = {lang_labels_json};
const langColors = ['var(--accent)', 'var(--accent2)', 'var(--green)', '#94a3b8'];

function showStats() {{
    const modal = document.getElementById('statsModal');
    const area = document.getElementById('chartArea');
    const xAxis = document.getElementById('chartXAxis');
    const yMax = document.getElementById('yMax');
    const legend = document.getElementById('chartLegend');
    
    area.innerHTML = '';
    xAxis.innerHTML = '';
    legend.innerHTML = '';
    
    // Legend
    langLabels.forEach((l, i) => {{
        const item = document.createElement('div');
        item.className = 'legend-item';
        item.innerHTML = `<div class="legend-color" style="background:${{langColors[i]}}"></div><span>${{l.toUpperCase()}}</span>`;
        legend.appendChild(item);
    }});
    
    const totals = hourlyData.map(h => Object.values(h).reduce((a, b) => a + b, 0));
    const maxVal = Math.max(...totals, 5);
    yMax.textContent = maxVal;
    
    hourlyData.forEach((hData, hour) => {{
        const wrap = document.createElement('div');
        wrap.className = 'chart-bar-wrap';
        
        // Reverse labels for stacking order (top to bottom)
        [...langLabels].reverse().forEach((lang) => {{
            const val = hData[lang] || 0;
            if (val > 0) {{
                const seg = document.createElement('div');
                seg.className = 'chart-bar-seg';
                const langIdx = langLabels.indexOf(lang);
                seg.style.backgroundColor = langColors[langIdx];
                seg.style.height = (val / maxVal * 100) + '%';
                seg.dataset.label = lang.toUpperCase() + ': ' + val;
                wrap.appendChild(seg);
            }}
        }});
        
        if (totals[hour] === 0) {{
            const stub = document.createElement('div');
            stub.className = 'chart-bar-seg';
            stub.style.height = '2px';
            stub.style.backgroundColor = 'var(--border)';
            stub.style.opacity = '0.2';
            wrap.appendChild(stub);
        }}
        
        area.appendChild(wrap);
        
        const label = document.createElement('div');
        label.className = 'chart-x-label';
        label.textContent = hour.toString().padStart(2, '0');
        xAxis.appendChild(label);
    }});
    
    modal.style.display = 'block';
}}


function hideStats() {{
    document.getElementById('statsModal').style.display = 'none';
}}

window.onclick = function(event) {{
    const modal = document.getElementById('statsModal');
    if (event.target == modal) hideStats();
}};

function toggleAllDays() {{
    const groups = document.querySelectorAll('.day-group');
    const btn = document.getElementById('btnToggleDays');
    const allCollapsed = [...groups].every(g => g.classList.contains('collapsed'));
    groups.forEach(g => {{
        if (allCollapsed) g.classList.remove('collapsed');
        else g.classList.add('collapsed');
    }});
    btn.textContent = allCollapsed ? '{t["collapse"]}' : '{t["expand"]}';
}}

function filterCards(filter, el) {{
    document.querySelectorAll('.stat').forEach(s => s.classList.remove('active'));
    el.classList.add('active');
    const cards = document.querySelectorAll('.card');
    cards.forEach(card => {{
        let show = false;
        if (filter === 'all') {{ show = true; }} 
        else if (filter === 'completed' || filter === 'in_progress' || filter === 'cancelled') {{ show = card.dataset.status === filter; }} 
        else if (filter === 'email') {{ show = card.dataset.email === '1'; }} 
        else if (filter === 'identified') {{ show = card.dataset.identified === '1'; }} 
        else if (filter === 'recurring') {{ show = card.dataset.recurring === '1'; }}
        else if (filter === 'gemini') {{ show = card.dataset.gemini === '1'; }}
        card.classList.toggle('card-hidden', !show);
    }});
    document.querySelectorAll('.day-group').forEach(group => {{
        const visibleCards = group.querySelectorAll('.card:not(.card-hidden)');
        group.classList.toggle('day-hidden', visibleCards.length === 0);
        const countBadge = group.querySelector('.day-count');
        if (countBadge) countBadge.firstChild.textContent = visibleCards.length;
    }});
}}

function updateLiveTimers() {{
    const timers = document.querySelectorAll('.live-timer');
    const now = new Date();
    timers.forEach(timer => {{
        const startStr = timer.dataset.start;
        if (!startStr) return;
        const start = new Date(startStr);
        const diff = Math.max(0, Math.floor((now - start) / 1000));
        const h = String(Math.floor(diff / 3600)).padStart(2, '0');
        const m = String(Math.floor((diff % 3600) / 60)).padStart(2, '0');
        const s = String(diff % 60).padStart(2, '0');
        timer.textContent = h + ':' + m + ':' + s;
    }});
    updateAggregateEta();
}}

function fmtEta(sec) {{
    sec = Math.max(0, Math.floor(sec));
    const h = Math.floor(sec / 3600);
    const m = Math.floor((sec % 3600) / 60);
    const s = sec % 60;
    if (h > 0) {{
        return String(h).padStart(2,'0') + ':' + String(m).padStart(2,'0') + ':' + String(s).padStart(2,'0');
    }}
    return String(m).padStart(2,'0') + ':' + String(s).padStart(2,'0');
}}

function updateAggregateEta() {{
    const box = document.getElementById('statEta');
    if (!box) return;
    const cards = document.querySelectorAll('.card-in-progress');
    const now = new Date();
    let maxRemaining = -1;
    let unknownCount = 0;
    let estimableCount = 0;
    cards.forEach(card => {{
        const timer = card.querySelector('.live-timer');
        const pctEl = card.querySelector('.card-pct');
        if (!timer) return;
        if (pctEl && !pctEl.hasAttribute('data-sid')) return;
        const start = new Date(timer.dataset.start);
        const elapsed = Math.max(0, (now - start) / 1000);
        let pct = 0;
        if (pctEl) {{
            const m = pctEl.textContent.match(/(\\d+)/);
            if (m) pct = parseInt(m[1], 10);
        }}
        if (pct <= 0 || pct >= 100 || elapsed <= 0) {{
            unknownCount++;
            return;
        }}
        const remaining = elapsed * (100 - pct) / pct;
        estimableCount++;
        if (remaining > maxRemaining) maxRemaining = remaining;
    }});
    if (cards.length === 0) {{
        box.textContent = '';
        return;
    }}
    if (estimableCount === 0) {{
        box.innerHTML = '<span class="eta-lbl">{t["eta_label"]}</span>--:--';
        return;
    }}
    const prefix = unknownCount > 0 ? '≥ ' : '';
    box.innerHTML = '<span class="eta-lbl">{t["eta_label"]}</span>' + prefix + fmtEta(maxRemaining);
}}

async function updateLiveProgress() {{
    const pcts = Array.from(document.querySelectorAll('.card-pct')).filter(el => el.dataset.sid);
    if (pcts.length === 0) return;
    await Promise.all(pcts.map(async (el) => {{
        const sid = el.dataset.sid;
        try {{
            const r = await fetch('/api/job_status/' + sid);
            if (!r.ok) return;
            const d = await r.json();
            el.textContent = '(' + d.pct + '%)';
            if (d.status === 'done' || d.status === 'error' || d.status === 'cancelled') {{
                el.removeAttribute('data-sid');
                if (d.status === 'done') el.style.color = 'var(--green)';
            }}
        }} catch(e) {{}}
    }}));
}}

if (document.querySelectorAll('.live-timer').length > 0) {{
    setInterval(updateLiveTimers, 1000);
    setInterval(updateLiveProgress, 5000);
    updateLiveProgress();
    updateAggregateEta();
}}

//  -  -  Admin: sospensione nuovi processi  -  -
const ADMIN_TOKEN = new URLSearchParams(window.location.search).get('token') ||
                    sessionStorage.getItem('abm_admin_token') || '';

function updateSuspendButton(suspended) {{
    const btn = document.getElementById('btnSuspend');
    if (!btn) return;
    if (suspended) {{
        btn.textContent = '⏸ Sospesi';
        btn.classList.add('suspended');
    }} else {{
        btn.textContent = '▶ Attivi';
        btn.classList.remove('suspended');
    }}
}}

async function checkSuspendStatus() {{
    if (!ADMIN_TOKEN) return;
    try {{
        const r = await fetch('/api/admin/suspend?token=' + encodeURIComponent(ADMIN_TOKEN));
        if (r.ok) {{
            const d = await r.json();
            updateSuspendButton(d.suspended);
        }}
    }} catch(e) {{}}
}}

async function toggleSuspend() {{
    const btn = document.getElementById('btnSuspend');
    if (!btn || !ADMIN_TOKEN) return;
    const currentlySuspended = btn.classList.contains('suspended');
    try {{
        const r = await fetch('/api/admin/suspend?token=' + encodeURIComponent(ADMIN_TOKEN), {{
            method: 'POST',
            headers: {{'Content-Type': 'application/json'}},
            body: JSON.stringify({{suspend: !currentlySuspended}})
        }});
        if (r.ok) {{
            const d = await r.json();
            updateSuspendButton(d.suspended);
        }}
    }} catch(e) {{}}
}}

checkSuspendStatus();
</script>

</body>
</html>"""
    return html, 200, {"Content-Type": "text/html; charset=utf-8"}


def _csv_safe(val):
    """Sanitizza un valore per esportazioni CSV/XLSX contro CSV-injection.

    Excel/LibreOffice interpretano celle che iniziano con =, +, -, @, TAB, CR
    come formule. Prefissiamo con apostrofo ('), che disabilita l'interpretazione
    e non viene mostrato nella cella.
    """
    if val is None:
        return ""
    s = str(val)
    if s and s[0] in ("=", "+", "-", "@", "\t", "\r"):
        return "'" + s
    return s


@app.route("/admin/log-activity/export")
def admin_logs_export():
    """Export activity log as Excel (.xlsx) file."""
    if not ADMIN_TOKEN: return "Export disabled.", 404
    token = _admin_auth_from_request()
    if not _admin_auth_ok(token): return "Unauthorized", 401
    from datetime import datetime
    import io, csv

    ym = None
    for key in request.args:
        if re.match(r'^\d{4}-\d{2}$', key):
            ym = key
            break
    if not ym:
        ym = datetime.now().strftime("%Y-%m")

    try:
        sessions, client_session_count = _parse_log_sessions(ym)
    except Exception as e:
        return f"Errore lettura log: {e}", 500

    output = io.StringIO()
    writer = csv.writer(output, delimiter="#")
    writer.writerow([
        "Session ID", "Date Start", "Date End", "Duration (min)",
        "Filename", "Last Status", "Events", "Client ID", "Client IP",
        "Voice", "Browser Lang", "Completed", "In Progress", "Recurring Client"
    ])
    for sid, s in sessions.items():
        delta = s["last_dt"] - s["first_dt"]
        duration_min = round(delta.total_seconds() / 60, 1)
        completed = "Yes" if _session_completed(s) else "No"
        in_progress = "Yes" if _session_in_progress(s, sid) else "No"
        cid = s.get("client_id", "")
        recurring = "Yes" if client_session_count.get(cid, 0) >= 2 else "No"
        writer.writerow([
            _csv_safe(sid), s["first_dt"].strftime("%Y-%m-%d %H:%M:%S"),
            s["last_dt"].strftime("%Y-%m-%d %H:%M:%S"), duration_min,
            _csv_safe(s["filename"]), _csv_safe(s["last_op"]),
            _csv_safe("  →  ".join(s["events"])),
            _csv_safe(cid), _csv_safe(s.get("client_ip", "")),
            _csv_safe(s.get("voice", "")),
            _csv_safe(s.get("browser_lang", "")), completed, in_progress, recurring,
        ])

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = f"Log {ym}"

        total_s = len(sessions)
        gen_c = sum(1 for s_ in sessions.values() if _session_completed(s_))
        gen_p = sum(1 for sid_, s_ in sessions.items() if _session_in_progress(s_, sid_))
        gen_x = total_s - gen_c - gen_p
        em_s = sum(1 for s_ in sessions.values() if "EMAIL_SENT" in s_["events"])
        uniq = len(set(s_.get("client_id", "") for s_ in sessions.values() if s_.get("client_id")))
        ret = sum(1 for c in client_session_count.values() if c >= 2)

        ws.merge_cells("A1:B1")
        ws["A1"] = f"Audiobook Maker  -  Activity Log {ym}"
        ws["A1"].font = Font(name="Arial", bold=True, color="38bdf8", size=14)
        summary = [("Sessioni", total_s), ("Gen. completata", gen_c), ("In corso", gen_p),
                   ("Cancellati", gen_x), ("Email inviate", em_s), ("Client unici", uniq),
                   ("Ricorrenti", ret)]
        for i, (lbl, val) in enumerate(summary):
            ws.cell(row=2, column=1 + i * 2, value=lbl).font = Font(name="Arial", color="94a3b8", size=10)
            ws.cell(row=2, column=2 + i * 2, value=val).font = Font(name="Arial", bold=True, color="e2e8f0", size=12)

        headers = ["Session ID", "Data inizio", "Data fine", "Durata (min)", "Contenuto",
                   "Ultimo stato", "Timeline eventi", "Client ID", "IP", "Voce",
                   "Lingua browser", "Completato", "In corso", "Client ricorrente"]
        hdr_fill = PatternFill("solid", fgColor="334155")
        hdr_font = Font(name="Arial", bold=True, color="e2e8f0", size=10)
        for col_idx, h in enumerate(headers, 1):
            c = ws.cell(row=4, column=col_idx, value=h)
            c.font = hdr_font; c.fill = hdr_fill; c.alignment = Alignment(horizontal="center")

        data_font = Font(name="Arial", size=10, color="e2e8f0")
        for row_idx, (sid, s) in enumerate(reversed(list(sessions.items())), 5):
            delta = s["last_dt"] - s["first_dt"]
            row_data = [_csv_safe(sid), s["first_dt"].strftime("%Y-%m-%d %H:%M:%S"),
                        s["last_dt"].strftime("%Y-%m-%d %H:%M:%S"),
                        round(delta.total_seconds() / 60, 1),
                        _csv_safe(s["filename"]), _csv_safe(s["last_op"]),
                        _csv_safe("  →  ".join(s["events"])),
                        _csv_safe(s.get("client_id", "")), _csv_safe(s.get("client_ip", "")),
                        _csv_safe(s.get("voice", "")), _csv_safe(s.get("browser_lang", "")),
                        "✅" if _session_completed(s) else "❌",
                        "✅" if _session_in_progress(s, sid) else "",
                        "✅" if client_session_count.get(s.get("client_id", ""), 0) >= 2 else ""]
            for col_idx, val in enumerate(row_data, 1):
                c = ws.cell(row=row_idx, column=col_idx, value=val)
                c.font = data_font
                if row_idx % 2 == 0:
                    c.fill = PatternFill("solid", fgColor="1e293b")

        col_widths = [12, 20, 20, 12, 45, 18, 50, 14, 16, 25, 10, 12, 10, 14]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.auto_filter.ref = f"A4:N{4 + len(sessions)}"

        xlsx_io = io.BytesIO()
        wb.save(xlsx_io)
        xlsx_io.seek(0)
        return Response(xlsx_io.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="activity_log_{ym}.xlsx"'})
    except ImportError:
        csv_bytes = output.getvalue().encode("utf-8-sig")
        return Response(csv_bytes, mimetype="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="activity_log_{ym}.csv"'})


#  -  -  -  Admin voucher web UI (/admin/vouchers)  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  - 
# Protetta da token ABM_ADMIN_TOKEN. Se il token non è configurato, endpoint 404.
# Il token viene inviato via header X-Admin-Token (dalle API) o nel form HTML.
# Confronto a tempo costante tramite hmac.compare_digest.

def _render_admin_gate(title, target_url):
    """Render a password gate for admin pages."""
    return f"""<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Admin Auth - {title}</title>
<style>
body{{font-family:system-ui,-apple-system,sans-serif;display:flex;justify-content:center;align-items:center;min-height:100vh;margin:0;background:#f3f4f6;color:#1f2937}}
.card{{background:#fff;padding:2rem;border-radius:12px;box-shadow:0 10px 15px -3px rgba(0,0,0,0.1);width:100%;max-width:360px}}
h2{{margin:0 0 1.5rem;font-size:1.5rem;text-align:center}}
.field{{margin-bottom:1rem}}
label{{display:block;font-size:.875rem;font-weight:600;margin-bottom:.375rem}}
input[type=password]{{width:100%;padding:.625rem;border:1px solid #d1d5db;border-radius:.375rem;box-sizing:border-box}}
.btn{{width:100%;padding:.75rem;background:#2563eb;color:#fff;border:none;border-radius:.375rem;font-weight:600;cursor:pointer}}
.btn:hover{{background:#1d4ed8}}
.hint{{font-size:.75rem;color:#6b7280;margin-top:1rem;text-align:center}}
</style>
<script>
// Sec: il token admin non viaggia più nell'URL. POSTa a /admin/login che imposta
// un cookie HttpOnly+Secure+SameSite=Strict; per le fetch API la pagina admin usa
// header X-Admin-Token da sessionStorage (cookie non leggibile da JS).
(function(){{
    const tok = sessionStorage.getItem('abm_admin_token') || localStorage.getItem('abm_admin_token');
    if(tok && !sessionStorage.getItem('abm_admin_gate_retry')){{
        sessionStorage.setItem('abm_admin_gate_retry', '1');
        fetch('/admin/login', {{
            method: 'POST',
            headers: {{'Content-Type': 'application/json'}},
            body: JSON.stringify({{token: tok}}),
            credentials: 'same-origin',
        }}).then(r => {{
            if(r.ok){{ window.location.reload(); }}
        }}).catch(()=>{{}});
    }}
}})();

function doLogin(){{
    const tok = document.getElementById('pw').value;
    const remember = document.getElementById('rem').checked;
    if(!tok) return;
    fetch('/admin/login', {{
        method: 'POST',
        headers: {{'Content-Type': 'application/json'}},
        body: JSON.stringify({{token: tok, remember: remember}}),
        credentials: 'same-origin',
    }}).then(r => {{
        if(!r.ok){{ alert('Token non valido'); return; }}
        sessionStorage.setItem('abm_admin_token', tok);
        if(remember){{
            localStorage.setItem('abm_admin_token', tok);
            localStorage.setItem('abm_admin_expiry', (Date.now() + 30 * 86400000).toString());
        }}
        sessionStorage.removeItem('abm_admin_gate_retry');
        window.location.reload();
    }}).catch(()=>{{ alert('Errore di rete'); }});
}}
</script>
</head><body>
<div class="card">
    <h2>Admin Access</h2>
    <div class="field">
        <label for="pw">Admin Token</label>
        <input type="password" id="pw" placeholder="Enter token..." onkeydown="if(event.key==='Enter')doLogin()">
    </div>
    <div class="field" style="display:flex;align-items:center;gap:.5rem">
        <input type="checkbox" id="rem">
        <label for="rem" style="margin:0;font-weight:400">Rimani connesso (30 giorni)</label>
    </div>
    <button class="btn" onclick="doLogin()">Entra</button>
    <div class="hint">Autenticazione richiesta per accedere a questa risorsa.</div>
</div>
</body></html>"""


#  -  -  Admin API: sospensione nuovi processi  -  -

@app.route("/api/admin/suspend", methods=["GET", "POST"])
def admin_api_suspend():
    """GET: restituisce stato sospensione (pubblico). POST: imposta sospensione (richiede admin token)."""
    global _suspend_new_jobs
    if request.method == "POST":
        token = _admin_auth_from_request()
        if not _admin_auth_ok(token):
            return jsonify({"error": "Unauthorized"}), 401
        data = request.json or {}
        with _suspend_lock:
            _suspend_new_jobs = bool(data.get("suspend", False))
        return jsonify({"suspended": _suspend_new_jobs})
    with _suspend_lock:
        suspended = _suspend_new_jobs
    return jsonify({"suspended": suspended})


def _admin_auth_ok(provided):
    """Costante-time check del token admin."""
    import hmac
    if not ADMIN_TOKEN or not provided:
        return False
    return hmac.compare_digest(str(provided), ADMIN_TOKEN)


_ADMIN_COOKIE_NAME = "abm_admin_session"


def _admin_auth_from_request():
    """Estrae il token admin dall'header X-Admin-Token o dal cookie HttpOnly abm_admin_session.

    Sec: NON accettiamo più il token da query string (`?token=`) né da form GET/POST:
    il valore comparirebbe in access log nginx, history browser, Referer verso domini esterni.
    Il cookie è settato esclusivamente dall'endpoint POST /admin/login con HttpOnly+SameSite=Strict.
    """
    tok = request.headers.get("X-Admin-Token", "")
    if not tok:
        tok = request.cookies.get(_ADMIN_COOKIE_NAME, "")
    return tok


@app.route("/admin/login", methods=["POST"])
def admin_login():
    """Login admin: valida il token e lo deposita in cookie HttpOnly Secure SameSite=Strict.
    Il client JS POSTa qui invece di mettere il token nell'URL."""
    if not ADMIN_TOKEN:
        return jsonify({"error": "Admin disabled"}), 404
    data = request.json or {}
    tok = (data.get("token") or "").strip()
    remember = bool(data.get("remember", True))
    if not _admin_auth_ok(tok):
        return jsonify({"error": "Invalid token"}), 401
    resp = jsonify({"ok": True})
    is_https = (request.scheme == "https") or (request.headers.get("X-Forwarded-Proto", "") == "https")
    # 30 giorni con "remember" (default), 8 ore altrimenti. Il cookie è HttpOnly+Strict
    # quindi non può essere esfiltrato lato client; estendere la durata evita un round-trip
    # extra (gate → /admin/login → reload) a ogni navigazione successiva alla scadenza.
    max_age = 30 * 86400 if remember else 8 * 3600
    resp.set_cookie(
        _ADMIN_COOKIE_NAME, tok,
        max_age=max_age,
        httponly=True,
        secure=is_https,
        samesite="Strict",
        path="/",
    )
    return resp


@app.route("/admin/logout", methods=["POST", "GET"])
def admin_logout():
    """Logout admin: cancella il cookie di sessione."""
    resp = jsonify({"ok": True}) if request.method == "POST" else ("Logged out", 200)
    if isinstance(resp, tuple):
        from flask import make_response
        body, code = resp
        resp = make_response(body, code)
    resp.set_cookie(_ADMIN_COOKIE_NAME, "", max_age=0, httponly=True, samesite="Strict", path="/")
    return resp


@app.route("/admin/vouchers", methods=["GET"])
def admin_vouchers_page():
    if not ADMIN_TOKEN: return ("Admin voucher UI disabled.", 404, {"Content-Type": "text/plain; charset=utf-8"})
    token = _admin_auth_from_request()
    if not _admin_auth_ok(token): return _render_admin_gate("Voucher Admin", "/admin/vouchers"), 200, {"Content-Type": "text/html; charset=utf-8"}
    html = r"""<!DOCTYPE html>
<html lang="it"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<meta name="robots" content="noindex,nofollow">
<title>Admin  -  Voucher</title>
<style>
  :root{--bg:#0f172a;--panel:#1e293b;--ink:#e2e8f0;--muted:#94a3b8;--accent:#8b5cf6;--ok:#10b981;--err:#ef4444;--warn:#f59e0b;}
  *{box-sizing:border-box}
  body{margin:0;font-family:system-ui,-apple-system,sans-serif;background:var(--bg);color:var(--ink);padding:20px;max-width:1200px;margin:0 auto}
  h1{margin:0 0 20px;font-size:1.5rem;display:flex;align-items:center;gap:10px}
  h1 .lock{font-size:1.2rem}
  .panel{background:var(--panel);border-radius:10px;padding:20px;margin-bottom:20px;box-shadow:0 2px 8px rgba(0,0,0,.2)}
  .panel h2{margin:0 0 14px;font-size:1.1rem;color:var(--accent)}
  label{display:block;font-size:.85rem;color:var(--muted);margin:8px 0 4px}
  input,select,textarea{width:100%;padding:9px 12px;background:#0f172a;border:1px solid #334155;color:var(--ink);border-radius:6px;font-size:.95rem;font-family:inherit}
  input:focus,select:focus,textarea:focus{outline:none;border-color:var(--accent)}
  button{padding:10px 20px;background:var(--accent);color:#fff;border:none;border-radius:6px;cursor:pointer;font-weight:600;font-size:.95rem}
  button:hover{filter:brightness(1.1)}
  button.secondary{background:#334155}
  button.danger{background:var(--err)}
  .row{display:grid;grid-template-columns:1fr 1fr;gap:12px}
  .row3{display:grid;grid-template-columns:1fr 1fr 1fr;gap:12px}
  @media(max-width:600px){.row,.row3{grid-template-columns:1fr}}
  .msg{padding:10px 14px;border-radius:6px;margin:12px 0;font-size:.9rem}
  .msg.ok{background:rgba(16,185,129,.15);color:var(--ok);border:1px solid var(--ok)}
  .msg.err{background:rgba(239,68,68,.15);color:var(--err);border:1px solid var(--err)}
  table{width:100%;border-collapse:collapse;font-size:.85rem}
  th{text-align:left;padding:8px;border-bottom:2px solid #334155;color:var(--muted);font-weight:500}
  td{padding:8px;border-bottom:1px solid #1e293b;vertical-align:top}
  tr:hover{background:rgba(139,92,246,.05)}
  code{font-family:ui-monospace,monospace;background:#0f172a;padding:2px 6px;border-radius:3px;font-size:.85em}
  .badge{display:inline-block;padding:2px 8px;border-radius:10px;font-size:.75rem;font-weight:600}
  .badge.promo{background:rgba(139,92,246,.2);color:#c4b5fd}
  .badge.gift{background:rgba(245,158,11,.2);color:#fbbf24}
  .badge.refund{background:rgba(148,163,184,.2);color:var(--muted)}
  .badge.ACTIVE{background:rgba(16,185,129,.2);color:var(--ok)}
  .badge.PARTIAL{background:rgba(245,158,11,.2);color:var(--warn)}
  .badge.USED{background:rgba(148,163,184,.2);color:var(--muted)}
  .badge.EXPIRED{background:rgba(239,68,68,.2);color:var(--err)}
  .btn-sm{padding:4px 10px;font-size:.8rem}
  .toolbar{display:flex;gap:10px;flex-wrap:wrap;margin-bottom:12px;align-items:flex-end}
  .toolbar > div{flex:1;min-width:150px}
  #tokenGate{max-width:420px;margin:80px auto}
  .hint{font-size:.8rem;color:var(--muted);margin-top:4px}
</style>
</head><body>

<div id="tokenGate" class="panel" style="display:none">
  <h1><span class="lock">&#x1F512;</span> Admin Voucher</h1>
  <label>Admin token</label>
  <input id="tokenInput" type="password" autocomplete="off" placeholder="ABM_ADMIN_TOKEN">
  <div class="hint">Il token viene memorizzato solo in questa scheda del browser (sessionStorage).</div>
  <div style="margin-top:14px"><button onclick="saveToken()">Sblocca</button></div>
  <div id="tokenMsg"></div>
</div>

<div id="app" style="display:none">
  <h1><span class="lock">&#x1F512;</span> Admin Voucher <button class="secondary btn-sm" style="margin-left:auto" onclick="logout()">Logout</button></h1>

  <div class="panel">
    <h2>Crea nuovo voucher</h2>
    <div class="row">
      <div><label>Email destinatario *</label><input id="cEmail" type="email" placeholder="user@example.com"></div>
      <div><label>Tipo</label>
        <select id="cKind">
          <option value="promo">promo (promozionale)</option>
          <option value="gift">gift (regalo)</option>
          <option value="refund">refund (rimborso)</option>
        </select>
      </div>
    </div>
    <div class="row3">
      <div><label>Importo (EUR) *</label><input id="cAmount" type="number" step="0.01" min="0.01" placeholder="2.00"></div>
      <div><label>Validità (giorni)</label><input id="cDays" type="number" min="1" value="180"></div>
      <div><label>&nbsp;</label><button onclick="createVoucher()" style="width:100%">Crea voucher</button></div>
    </div>
    <label>Nota (causale interna)</label>
    <textarea id="cNote" rows="2" placeholder="Es: campagna lancio, influencer X, compenso collaboratore..."></textarea>
    <div id="createMsg"></div>
  </div>

  <div class="panel">
    <h2>Voucher esistenti</h2>
    <div class="toolbar">
      <div><label>Filtra email</label><input id="fEmail" placeholder="contiene..." oninput="renderList()"></div>
      <div><label>Filtra tipo</label>
        <select id="fKind" onchange="renderList()">
          <option value="">tutti</option><option value="promo">promo</option><option value="gift">gift</option><option value="refund">refund</option>
        </select>
      </div>
      <div><label>Stato</label>
        <select id="fStatus" onchange="renderList()">
          <option value="">tutti</option><option value="ACTIVE">attivi</option><option value="PARTIAL">parziali</option><option value="USED">usati</option><option value="EXPIRED">scaduti</option>
        </select>
      </div>
      <div style="flex:0"><button class="secondary" onclick="loadList()">&#x21BB; Ricarica</button></div>
    </div>
    <div style="overflow-x:auto">
      <table id="tbl">
        <thead><tr>
          <th>Codice</th><th>Tipo</th><th>Saldo / Iniziale</th><th>Email</th><th>Creato</th><th>Scade</th><th>Stato</th><th>Nota</th><th></th>
        </tr></thead>
        <tbody id="tbody"></tbody>
      </table>
    </div>
    <div id="listMsg"></div>
  </div>
</div>

<script>
var TK=sessionStorage.getItem('abm_admin_tok')||'';
var VOUCHERS=[];

function show(el){document.getElementById(el).style.display=''}
function hide(el){document.getElementById(el).style.display='none'}
function msg(id,txt,cls){var e=document.getElementById(id);e.innerHTML='<div class="msg '+cls+'">'+txt+'</div>';setTimeout(function(){e.innerHTML=''},5000)}

async function api(path,opts){
  opts=opts||{};
  opts.headers=opts.headers||{};
  opts.headers['X-Admin-Token']=TK;
  if(opts.body&&typeof opts.body==='object'){opts.headers['Content-Type']='application/json';opts.body=JSON.stringify(opts.body)}
  var r=await fetch(path,opts);
  var j=await r.json().catch(function(){return{}});
  if(!r.ok)throw new Error(j.error||('HTTP '+r.status));
  return j;
}

function saveToken(){
  var t=document.getElementById('tokenInput').value.trim();
  if(!t){document.getElementById('tokenMsg').innerHTML='<div class="msg err">Token richiesto</div>';return}
  TK=t;sessionStorage.setItem('abm_admin_tok',t);
  tryEnter();
}

function logout(){TK='';sessionStorage.removeItem('abm_admin_tok');location.reload()}

async function tryEnter(){
  try{
    await loadList();
    hide('tokenGate');show('app');
  }catch(e){
    TK='';sessionStorage.removeItem('abm_admin_tok');
    show('tokenGate');hide('app');
    document.getElementById('tokenMsg').innerHTML='<div class="msg err">Token non valido: '+e.message+'</div>';
  }
}

async function loadList(){
  var j=await api('/admin/api/vouchers');
  VOUCHERS=j.vouchers||[];
  renderList();
}

function fmtTs(ts){if(!ts)return'-';var d=new Date(ts*1000);return d.toLocaleString('it-IT',{year:'numeric',month:'2-digit',day:'2-digit',hour:'2-digit',minute:'2-digit'})}

function statusOf(v){
  var rem=(v.remaining_eur!=null)?v.remaining_eur:(v.used?0:v.amount_eur);
  if(rem<0.01)return'USED';
  if((v.expires_at||0)*1000<Date.now())return'EXPIRED';
  if(rem<(v.amount_eur||0)-0.01)return'PARTIAL';
  return'ACTIVE';
}

function escapeHtml(s){return(s||'').replace(/[&<>"']/g,function(c){return{'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c]})}

function renderList(){
  var fe=(document.getElementById('fEmail').value||'').toLowerCase();
  var fk=document.getElementById('fKind').value;
  var fs=document.getElementById('fStatus').value;
  var rows=VOUCHERS.filter(function(v){
    if(fe&&(v.email||'').toLowerCase().indexOf(fe)<0)return false;
    if(fk&&(v.kind||'refund')!==fk)return false;
    if(fs&&statusOf(v)!==fs)return false;
    return true;
  }).sort(function(a,b){return(b.created_at||0)-(a.created_at||0)});
  var tb=document.getElementById('tbody');
  if(!rows.length){tb.innerHTML='<tr><td colspan="9" style="text-align:center;color:var(--muted);padding:20px">Nessun voucher</td></tr>';return}
  tb.innerHTML=rows.map(function(v){
    var st=statusOf(v);var k=v.kind||'refund';
    var rem=(v.remaining_eur!=null)?v.remaining_eur:(v.used?0:(v.amount_eur||0));
    var tot=v.amount_eur||0;
    var balCell='<strong>'+rem.toFixed(2)+'</strong> / '+tot.toFixed(2);
    if(v.uses&&v.uses.length){balCell+=' <span style="color:var(--muted);font-size:.8em">('+v.uses.length+' us'+(v.uses.length===1?'o':'i')+')</span>';}
    var btn=(rem<0.01)?'':('<button class="danger btn-sm" onclick="revokeVoucher(\''+v.code+'\')">Revoca</button>');
    return '<tr>'+
      '<td><code>'+escapeHtml(v.code)+'</code></td>'+
      '<td><span class="badge '+k+'">'+k+'</span></td>'+
      '<td>'+balCell+'</td>'+
      '<td>'+escapeHtml(v.email||'')+'</td>'+
      '<td>'+fmtTs(v.created_at)+'</td>'+
      '<td>'+fmtTs(v.expires_at)+'</td>'+
      '<td><span class="badge '+st+'">'+st+'</span></td>'+
      '<td style="max-width:240px;font-size:.8em;color:var(--muted)">'+escapeHtml(v.note||'')+'</td>'+
      '<td>'+btn+'</td>'+
    '</tr>';
  }).join('');
}

async function createVoucher(){
  var email=document.getElementById('cEmail').value.trim();
  var amount=parseFloat(document.getElementById('cAmount').value);
  var days=parseInt(document.getElementById('cDays').value||'180');
  var kind=document.getElementById('cKind').value;
  var note=document.getElementById('cNote').value.trim();
  if(!email||!amount||amount<=0){msg('createMsg','Email e importo > 0 obbligatori','err');return}
  try{
    var j=await api('/admin/api/vouchers',{method:'POST',body:{email:email,amount_eur:amount,days:days,kind:kind,note:note}});
    msg('createMsg','Creato <code>'+j.code+'</code> per '+escapeHtml(email)+' ('+j.amount_eur.toFixed(2)+' EUR)','ok');
    document.getElementById('cEmail').value='';document.getElementById('cAmount').value='';document.getElementById('cNote').value='';
    await loadList();
  }catch(e){msg('createMsg','Errore: '+e.message,'err')}
}

async function revokeVoucher(code){
  if(!confirm('Revocare il voucher '+code+' ?'))return;
  var reason=prompt('Motivo (opzionale):','')||'';
  try{
    await api('/admin/api/vouchers/'+encodeURIComponent(code)+'/revoke',{method:'POST',body:{reason:reason}});
    msg('listMsg','Voucher '+code+' revocato','ok');
    await loadList();
  }catch(e){msg('listMsg','Errore: '+e.message,'err')}
}

// bootstrap
if(TK){tryEnter()}else{show('tokenGate')}
</script>
</body></html>"""
    return html, 200, {"Content-Type": "text/html; charset=utf-8",
                       "X-Robots-Tag": "noindex, nofollow"}


@app.route("/admin/api/vouchers", methods=["GET", "POST"])
def admin_api_vouchers():
    """GET: elenca voucher. POST: crea nuovo voucher.

    Autenticazione via header X-Admin-Token (confronto costante-time).
    """
    if not ADMIN_TOKEN:
        return jsonify({"error": "Admin UI disabled"}), 404
    if not _admin_auth_ok(_admin_auth_from_request()):
        time.sleep(0.5)  # rallenta brute-force
        return jsonify({"error": "Unauthorized"}), 401

    ip = _get_client_ip()

    if request.method == "GET":
        items = []
        for code, v in payment._vouchers.items():
            items.append({
                "code": code,
                "email": v.get("email", ""),
                "amount_eur": v.get("amount_eur", 0),
                "remaining_eur": _voucher_remaining(v),
                "base_amount_eur": v.get("base_amount_eur", 0),
                "created_at": v.get("created_at"),
                "expires_at": v.get("expires_at"),
                "used": bool(v.get("used")),
                "used_at": v.get("used_at"),
                "uses": v.get("uses", []),
                "kind": v.get("kind", "refund"),
                "note": v.get("note", ""),
                "created_by": v.get("created_by", ""),
                "origin_order_id": v.get("origin_order_id"),
                "revoked": bool(v.get("revoked")),
            })
        return jsonify({"vouchers": items, "count": len(items)})

    # POST  →  create
    data = request.json or {}
    email = (data.get("email") or "").strip().lower()
    try:
        amount = float(data.get("amount_eur") or 0)
    except (TypeError, ValueError):
        amount = 0
    try:
        days = int(data.get("days") or 180)
    except (TypeError, ValueError):
        days = 180
    kind = (data.get("kind") or "promo").strip().lower()
    note = (data.get("note") or "").strip()
    if not email or "@" not in email:
        return jsonify({"error": "email required"}), 400
    if amount <= 0:
        return jsonify({"error": "amount must be > 0"}), 400
    if days <= 0 or days > 3650:
        return jsonify({"error": "days out of range (1..3650)"}), 400
    if kind not in ("promo", "gift", "refund"):
        return jsonify({"error": "invalid kind"}), 400

    # Codice con prefisso a seconda del tipo
    prefix = "PROMO-" if kind == "promo" else ("GIFT-" if kind == "gift" else "")
    # _generate_voucher_code dà core XXXX-XXXX-XXXX; per prefisso lo componiamo manualmente
    import secrets as _sec
    _alpha = "ABCDEFGHJKLMNPQRSTUVWXYZ23456789"
    custom_code = None
    if prefix:
        for _ in range(20):
            core = "-".join("".join(_sec.choice(_alpha) for _ in range(4)) for _ in range(3))
            cand = prefix + core
            if cand not in payment._vouchers:
                custom_code = cand
                break
        if not custom_code:
            return jsonify({"error": "code generation failed"}), 500

    code, bonus_amount = _create_voucher(
        email, amount,
        kind=kind,
        note=note,
        created_by="admin",
        expiry_days=days,
        apply_bonus=False,       # promo/gift: importo nominale, niente +10%
        code=custom_code,
    )
    _log_activity("", "", f"ADMIN_VOUCHER_CREATE:{kind}", "", ip, code[:8] + "...", email)
    print(f"[admin] voucher created via UI: {code} kind={kind} email={email} amount={amount:.2f} days={days} ip={ip}")
    return jsonify({
        "code": code,
        "amount_eur": bonus_amount,
        "expires_at": payment._vouchers[code].get("expires_at"),
        "kind": kind,
    })


@app.route("/admin/api/vouchers/<code>/revoke", methods=["POST"])
def admin_api_voucher_revoke(code):
    """Revoca (marca usato) un voucher."""
    if not ADMIN_TOKEN:
        return jsonify({"error": "Admin UI disabled"}), 404
    if not _admin_auth_ok(_admin_auth_from_request()):
        time.sleep(0.5)
        return jsonify({"error": "Unauthorized"}), 401
    code = (code or "").strip().upper()
    reason = ((request.json or {}).get("reason") or "").strip()[:200]
    with payment._vouchers_lock:
        if code not in payment._vouchers:
            return jsonify({"error": "Not found"}), 404
        v = payment._vouchers[code]
        v["used"] = True
        v["used_at"] = time.time()
        v["remaining_eur"] = 0.0
        v["revoked"] = True
        v["revoke_reason"] = reason or "admin revoke"
        _save_vouchers()
    _log_activity("", "", "ADMIN_VOUCHER_REVOKE", "", _get_client_ip(), code[:8] + "...", reason[:40])
    print(f"[admin] voucher revoked via UI: {code} reason={reason!r}")
    return jsonify({"ok": True, "code": code})


@app.route("/admin/job/<path:job_id>/forensic.zip", methods=["GET"])
def admin_forensic_zip(job_id):
    """Scarica ZIP della work_dir di un job Gemini fallito per analisi post-mortem.
    Richiede admin auth via cookie HttpOnly (set da /admin/login) o header
    X-Admin-Token. Disponibile finché la dir è protetta dal marker forense
    (default 7 giorni dal refund; ABM_GEMINI_FORENSIC_RETENTION_DAYS).
    """
    if not ADMIN_TOKEN:
        return ("Admin disabled.", 404, {"Content-Type": "text/plain; charset=utf-8"})
    token = _admin_auth_from_request()
    if not _admin_auth_ok(token):
        return ("Unauthorized. Effettua login su /admin/audit-tts e ritorna a questo link.",
                401, {"Content-Type": "text/plain; charset=utf-8"})
    if "/" in job_id or "\\" in job_id or ".." in job_id:
        return ("Invalid job_id", 400, {"Content-Type": "text/plain; charset=utf-8"})
    work_dir = UPLOAD_DIR / job_id
    if not work_dir.exists() or not work_dir.is_dir():
        return ("Work dir not found (cleanup completed or never existed).",
                404, {"Content-Type": "text/plain; charset=utf-8"})
    import io
    import zipfile
    buf = io.BytesIO()
    try:
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED, allowZip64=True) as zf:
            for root, _dirs, files in os.walk(str(work_dir)):
                for fn in files:
                    fp = Path(root) / fn
                    try:
                        arc = fp.relative_to(work_dir.parent)
                        zf.write(str(fp), str(arc))
                    except (OSError, ValueError):
                        continue
    except OSError as e:
        return (f"Zip build failed: {e}", 500, {"Content-Type": "text/plain; charset=utf-8"})
    payload = buf.getvalue()
    safe = re.sub(r'[^a-zA-Z0-9_-]', '_', job_id).strip("_") or "job"
    return Response(
        payload,
        mimetype="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="forensic_{safe}.zip"',
            "Content-Length": str(len(payload)),
            "Cache-Control": "no-store",
        },
    )


@app.route("/admin/audit-tts", methods=["GET"])
def admin_logs_page():
    """Admin TTS audit dashboard. Hosts the Gemini Cost Audit and Events/Refunds tabs."""
    if not ADMIN_TOKEN:
        return ("Admin audit TTS UI disabled.", 404, {"Content-Type": "text/plain; charset=utf-8"})
    token = _admin_auth_from_request()
    if not _admin_auth_ok(token):
        return _render_admin_gate("Audit TTS", "/admin/audit-tts"), 200, {"Content-Type": "text/html; charset=utf-8"}
    html = r"""<!DOCTYPE html>
<html lang="it"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<meta name="robots" content="noindex,nofollow">
<title>Admin - Audit TTS</title>
<style>
  :root{--bg:#0f172a;--panel:#1e293b;--ink:#e2e8f0;--muted:#94a3b8;--accent:#8b5cf6;--ok:#10b981;--err:#ef4444;--warn:#f59e0b;}
  *{box-sizing:border-box}
  body{margin:0;font-family:system-ui,-apple-system,sans-serif;background:var(--bg);color:var(--ink);padding:20px;max-width:1400px;margin:0 auto}
  h1{margin:0 0 20px;font-size:1.5rem}
  .panel{background:var(--panel);border-radius:10px;padding:20px;margin-bottom:20px}
  .panel h2{margin:0 0 14px;font-size:1.1rem;color:var(--accent)}
  .tab-bar{display:flex;gap:8px;margin-bottom:16px;border-bottom:2px solid #334155}
  .tab-btn{padding:10px 16px;background:transparent;color:var(--muted);border:none;cursor:pointer;font-size:.95rem;border-bottom:2px solid transparent;margin-bottom:-2px}
  .tab-btn.active{color:var(--accent);border-bottom-color:var(--accent)}
  .tab-panel{display:none}
  .tab-panel.active{display:block}
  .filters{display:grid;grid-template-columns:repeat(5,1fr);gap:10px;margin-bottom:16px}
  @media(max-width:900px){.filters{grid-template-columns:1fr 1fr}}
  label{display:block;font-size:.8rem;color:var(--muted);margin-bottom:4px}
  input,select{width:100%;padding:8px 10px;background:#0f172a;border:1px solid #334155;color:var(--ink);border-radius:6px;font-size:.9rem}
  button{padding:9px 16px;background:var(--accent);color:#fff;border:none;border-radius:6px;cursor:pointer;font-weight:600}
  button.secondary{background:#334155}
  table{width:100%;border-collapse:collapse;font-size:.82rem;margin-top:12px}
  th{text-align:left;padding:8px;border-bottom:2px solid #334155;color:var(--muted);font-weight:500}
  td{padding:6px 8px;border-bottom:1px solid #1e293b}
  .agg-grid{display:grid;grid-template-columns:repeat(5,1fr);gap:10px;margin-top:12px}
  .agg-grid-6{grid-template-columns:repeat(6,1fr)}
  @media(max-width:900px){.agg-grid,.agg-grid-6{grid-template-columns:1fr 1fr}}
  .agg-box{background:#0f172a;border:1px solid #334155;border-radius:6px;padding:10px;text-align:center}
  .agg-label{font-size:.75rem;color:var(--muted);text-transform:uppercase}
  .agg-value{font-size:1.3rem;font-weight:600;color:var(--ink);margin-top:4px}
  .delta-positive{color:var(--ok)}
  .delta-negative{color:var(--err)}
  .empty-msg{text-align:center;color:var(--muted);padding:20px}
  pre{background:#0f172a;padding:12px;border-radius:6px;overflow:auto;font-size:.8rem}
  tr.row-refund{background:rgba(239,68,68,0.10)}
  tr.row-refund td{border-bottom-color:rgba(239,68,68,0.35)}
  tr.row-preflight{background:rgba(37,99,235,0.10)}
  .badge{display:inline-block;padding:2px 8px;border-radius:10px;font-size:.7rem;font-weight:600;text-transform:uppercase;letter-spacing:.5px}
  .badge-ok{background:rgba(16,185,129,.18);color:#10b981}
  .badge-err{background:rgba(239,68,68,.18);color:#ef4444}
  .badge-warn{background:rgba(245,158,11,.18);color:#f59e0b}
  .badge-info{background:rgba(59,130,246,.18);color:#60a5fa}
  .badge-muted{background:rgba(148,163,184,.18);color:var(--muted)}
  .badge-live{background:rgba(139,92,246,.22);color:var(--accent);animation:livePulse 1.8s ease-in-out infinite}
  @keyframes livePulse{0%,100%{opacity:1}50%{opacity:.55}}
  tr.row-live{background:rgba(139,92,246,.10)}
  tr.row-live td{border-bottom-color:rgba(139,92,246,.30)}
  .toggle-row{display:flex;align-items:center;gap:10px;margin-bottom:14px;font-size:.9rem;color:var(--muted)}
  .toggle-row input{width:auto}
</style></head>
<body>
<h1>Admin - Audit TTS <a href="/admin/log-activity" style="float:right;font-size:.8rem;color:var(--accent);text-decoration:none;font-weight:500">&larr; Activity Log</a></h1>

<div class="panel" id="killSwitchPanel" style="display:flex;flex-wrap:wrap;align-items:center;gap:14px">
  <div style="flex:1;min-width:280px">
    <h2 style="margin:0 0 4px">Voci PREMIUM (Gemini TTS)</h2>
    <div id="ksStatus" style="font-size:.9rem;color:var(--muted)">Caricamento stato...</div>
    <div id="ksReason" style="font-size:.8rem;color:var(--muted);margin-top:4px;display:none"></div>
  </div>
  <div style="display:flex;gap:8px;align-items:center">
    <input type="text" id="ksReasonInput" placeholder="Motivo (facoltativo)"
           maxlength="200" style="width:240px;padding:8px 10px;background:#0f172a;border:1px solid #334155;color:var(--ink);border-radius:6px;font-size:.85rem">
    <button type="button" id="ksToggleBtn" disabled>...</button>
  </div>
</div>

<div class="tab-bar">
  <button type="button" class="tab-btn active" data-tab="gemini_audit">Audit Gemini TTS</button>
  <button type="button" class="tab-btn" data-tab="gemini_events">Eventi &amp; Rimborsi</button>
</div>

<div class="tab-panel active" id="tab_gemini_audit">
  <div class="panel">
    <h2>Filtri</h2>
    <div class="filters">
      <div>
        <label for="auditModelFilter">Modello</label>
        <select id="auditModelFilter">
          <option value="all">Tutti</option>
          <option value="flash25">Gemini 2.5 Flash TTS</option>
          <option value="flash31">Gemini 3.1 Flash TTS</option>
        </select>
      </div>
      <div>
        <label for="auditLangFilter">Lingua</label>
        <select id="auditLangFilter">
          <option value="all">Tutte</option>
        </select>
      </div>
      <div>
        <label for="auditOutcomeFilter">Esito</label>
        <select id="auditOutcomeFilter">
          <option value="all">Tutti</option>
          <option value="running">In corso</option>
          <option value="completed">Completato</option>
          <option value="failed_refunded">Fallito generico (rimborsato)</option>
          <option value="failed_quota_refunded">Fallito quota (rimborsato)</option>
          <option value="failed_budget_refunded">Fallito budget (rimborsato)</option>
          <option value="failed_quality_refunded">Fallito qualit&agrave; (rimborsato)</option>
          <option value="preflight_blocked_refunded">Bloccato preventivamente</option>
          <option value="cancelled_refunded">Annullato (rimborsato)</option>
          <option value="cancelled_partial">Annullato (parziale)</option>
        </select>
      </div>
      <div>
        <label for="auditDateFrom">Dal</label>
        <input type="date" id="auditDateFrom">
      </div>
      <div>
        <label for="auditDateTo">Al</label>
        <input type="date" id="auditDateTo">
      </div>
    </div>
    <button type="button" id="auditRefreshBtn">Aggiorna</button>
    <button type="button" id="auditRecalcBtn" class="btn">Calcola parametri suggeriti</button>
    <pre id="auditRecalcOutput" class="recalc-output" style="display:none;white-space:pre-wrap;margin-top:10px;padding:12px;background:#0f172a;color:#e2e8f0;border:1px solid #334155;border-radius:6px;font-family:ui-monospace,SFMono-Regular,Menlo,monospace;font-size:.85rem;line-height:1.5"></pre>
  </div>

  <div class="panel">
    <h2>Aggregati</h2>
    <div class="agg-grid agg-grid-6" id="auditAggregates">
      <div class="agg-box"><div class="agg-label">Job</div><div class="agg-value" id="aggCount">-</div></div>
      <div class="agg-box"><div class="agg-label">Ricavi</div><div class="agg-value" id="aggRevenue">-</div></div>
      <div class="agg-box"><div class="agg-label">Costo Google</div><div class="agg-value" id="aggCost">-</div></div>
      <div class="agg-box"><div class="agg-label" title="Margine lordo = Ricavi − Costo Google (include fee PayPal)">Margine</div><div class="agg-value" id="aggMargin">-</div></div>
      <div class="agg-box"><div class="agg-label" title="Margine netto = Margine − fee PayPal (zero per voucher)">Margine netto</div><div class="agg-value" id="aggNetMargin">-</div></div>
      <div class="agg-box"><div class="agg-label">Margine % medio</div><div class="agg-value" id="aggDelta">-</div></div>
    </div>
  </div>

  <div class="panel">
    <h2>Record (ultimi 200)</h2>
    <table>
      <thead><tr>
        <th>Data</th><th>Job</th><th>Modello</th><th>Lingua</th>
        <th>Char</th><th>Sec audio</th><th>Costo G.</th>
        <th>Prezzo &euro;</th><th title="Margine = Prezzo − Costo Google (lordo, include fee PayPal)">Margine &euro;</th><th title="Margine netto = Margine − fee PayPal. Zero per voucher (PayPal non coinvolto). Per PayPal: revenue × % + fee fissa.">Margine netto &euro;</th><th title="Margine % = Margine netto / Costo Google · markup applicato">Margine %</th><th>Esito</th>
      </tr></thead>
      <tbody id="auditRecordsBody">
        <tr><td colspan="12" class="empty-msg">Premi "Aggiorna" per caricare i record.</td></tr>
      </tbody>
    </table>
  </div>
</div>

<div class="tab-panel" id="tab_gemini_events">
  <div class="panel">
    <h2>Filtri</h2>
    <div class="filters">
      <div>
        <label for="evModelFilter">Modello</label>
        <select id="evModelFilter">
          <option value="all">Tutti</option>
          <option value="flash25">Gemini 2.5 Flash TTS</option>
          <option value="flash31">Gemini 3.1 Flash TTS</option>
        </select>
      </div>
      <div>
        <label for="evLangFilter">Lingua</label>
        <select id="evLangFilter">
          <option value="all">Tutte</option>
        </select>
      </div>
      <div>
        <label for="evDateFrom">Dal</label>
        <input type="date" id="evDateFrom">
      </div>
      <div>
        <label for="evDateTo">Al</label>
        <input type="date" id="evDateTo">
      </div>
      <div>
        <label>&nbsp;</label>
        <button type="button" id="evRefreshBtn">Aggiorna</button>
      </div>
    </div>
    <div class="toggle-row">
      <label><input type="checkbox" id="evOnlyRefunds"> Solo eventi con rimborso</label>
      <span style="margin-left:auto;font-size:.85em">Eventi con rimborso evidenziati in rosso; blocchi preventivi in blu.</span>
    </div>
  </div>

  <div class="panel">
    <h2>Aggregati Eventi</h2>
    <div class="agg-grid">
      <div class="agg-box"><div class="agg-label">Totale eventi</div><div class="agg-value" id="evCount">-</div></div>
      <div class="agg-box"><div class="agg-label">Completati</div><div class="agg-value" id="evCompleted">-</div></div>
      <div class="agg-box"><div class="agg-label">Rimborsi</div><div class="agg-value" id="evRefunds" style="color:#ef4444">-</div></div>
      <div class="agg-box"><div class="agg-label">Bloccati preflight</div><div class="agg-value" id="evPreflight" style="color:#60a5fa">-</div></div>
      <div class="agg-box"><div class="agg-label">Annullati</div><div class="agg-value" id="evCancelled">-</div></div>
    </div>
  </div>

  <div class="panel">
    <h2>Eventi (ultimi 500)</h2>
    <table>
      <thead><tr>
        <th>Data</th><th>Job</th><th>Esito</th><th>Modello</th>
        <th>Lingua</th><th>Char</th><th>Prezzo &euro;</th><th>Costo G. &euro;</th>
      </tr></thead>
      <tbody id="evRecordsBody">
        <tr><td colspan="8" class="empty-msg">Premi "Aggiorna" per caricare gli eventi.</td></tr>
      </tbody>
    </table>
  </div>
</div>

<script>
(function(){
  const ADMIN_TOKEN = """ + '"' + token + '"' + r""";
  const $ = (id) => document.getElementById(id);

  function fmtEur(n){ return (Number(n)||0).toFixed(2) + " €"; }
  function fmtPct(n){
    const v = Number(n)||0;
    const cls = v >= 0 ? "delta-positive" : "delta-negative";
    return `<span class="${cls}">${v.toFixed(2)}%</span>`;
  }

  // Italian display names per i codici lingua usati dal motore Gemini TTS.
  // Coprire l'intero set ufficiale (24) cosi` la dropdown e` sensata anche
  // quando in futuro verranno generati audio in lingue oggi non esposte in UI.
  const LANG_NAMES = {
    "it":"Italiano","en":"Inglese","fr":"Francese","es":"Spagnolo",
    "de":"Tedesco","pt":"Portoghese","ru":"Russo","ja":"Giapponese",
    "ko":"Coreano","zh":"Cinese","hi":"Hindi","ar":"Arabo",
    "id":"Indonesiano","nl":"Olandese","pl":"Polacco","th":"Thai",
    "tr":"Turco","vi":"Vietnamita","ro":"Rumeno","uk":"Ucraino",
    "bn":"Bengalese","mr":"Marathi","ta":"Tamil","te":"Telugu",
  };
  function langLabel(code){
    const c = (code||"").toLowerCase();
    return LANG_NAMES[c] ? `${LANG_NAMES[c]} (${c})` : c;
  }

  async function loadLanguageOptions(){
    let codes = [];
    try {
      const r = await fetch("/admin/api/gemini_cost_audit/languages",
                            {headers: {"X-Admin-Token": ADMIN_TOKEN}});
      if (r.ok) {
        const d = await r.json();
        codes = Array.isArray(d.languages) ? d.languages : [];
      }
    } catch(e) { /* silenzioso: dropdown resta con solo "Tutte" */ }
    codes = codes.slice().sort((a,b) =>
      langLabel(a).localeCompare(langLabel(b), "it"));
    for (const selId of ["auditLangFilter", "evLangFilter"]) {
      const sel = $(selId);
      if (!sel) continue;
      const prev = sel.value;
      // Rimuove option diverse da "all" e ripopola
      Array.from(sel.querySelectorAll('option:not([value="all"])')).forEach(o => o.remove());
      for (const code of codes) {
        const opt = document.createElement("option");
        opt.value = code;
        opt.textContent = langLabel(code);
        sel.appendChild(opt);
      }
      if (prev && Array.from(sel.options).some(o => o.value === prev)) {
        sel.value = prev;
      }
    }
  }

  async function fetchAudit(){
    const params = new URLSearchParams();
    const m = $("auditModelFilter").value;
    const l = $("auditLangFilter").value;
    const o = $("auditOutcomeFilter").value;
    const df = $("auditDateFrom").value;
    const dt = $("auditDateTo").value;
    if (m && m !== "all") params.set("model", m);
    if (l && l !== "all") params.set("language", l);
    if (o && o !== "all") params.set("outcome", o);
    if (df) params.set("date_from", df);
    if (dt) params.set("date_to", dt);
    params.set("limit", "200");
    const r = await fetch("/admin/api/gemini_cost_audit?" + params.toString(),
                         {headers: {"X-Admin-Token": ADMIN_TOKEN}});
    if (!r.ok) { alert("Errore caricamento audit: " + r.status); return; }
    const d = await r.json();
    renderAggregates(d.aggregates || {});
    renderRecords(d.records || []);
  }

  function renderAggregates(agg){
    $("aggCount").textContent = agg.count ?? 0;
    $("aggRevenue").textContent = fmtEur(agg.revenue_eur);
    $("aggCost").textContent = fmtEur(agg.google_cost_eur);
    $("aggMargin").textContent = fmtEur(agg.margin_eur);
    // Margine netto: gia` decurtato delle fee PayPal lato server (voucher=0).
    const netMargin = (agg.net_margin_eur != null) ? Number(agg.net_margin_eur)
                      : ((Number(agg.margin_eur)||0) - (Number(agg.paypal_fees_eur)||0));
    const netEl = $("aggNetMargin");
    if (netEl) {
      netEl.textContent = fmtEur(netMargin);
      netEl.title = "Fee PayPal totali: " + fmtEur(agg.paypal_fees_eur || 0);
    }
    const _margPctAvg = (Number(agg.google_cost_eur)||0) > 0
      ? (Number(agg.net_margin_eur)||0) / Number(agg.google_cost_eur) * 100
      : 0;
    $("aggDelta").innerHTML = fmtPct(_margPctAvg);
  }

  function esc(s){
    const d = document.createElement('div');
    d.textContent = (s == null ? "" : String(s));
    return d.innerHTML;
  }

  function renderRecords(recs){
    const tbody = $("auditRecordsBody");
    if (!recs.length) {
      tbody.innerHTML = '<tr><td colspan="12" class="empty-msg">Nessun record trovato.</td></tr>';
      return;
    }
    // Ordina mettendo i live "running" sempre in cima, poi gli altri per ts desc.
    recs = recs.slice().sort((a,b) => {
      const la = a._live ? 1 : 0, lb = b._live ? 1 : 0;
      if (la !== lb) return lb - la;
      return (b.ts||"").localeCompare(a.ts||"");
    });
    tbody.innerHTML = recs.map(r => {
      const ts = esc((r.ts || "").slice(0, 19).replace("T", " "));
      // Prezzo effettivo: per cancel anticipato usa quanto trattenuto
      // (cancel_retained_eur) invece dell'originale pre-pagato; per record
      // normali coincide con user_price_eur_charged.
      const revenue = (r._eff_revenue_eur != null) ? Number(r._eff_revenue_eur)
                                                  : Number(r.user_price_eur_charged || 0);
      const gCost = Number(r.google_cost_eur_actual || 0);
      // Margine = Prezzo − Costo Google (lordo, include fee PayPal).
      const margEur = revenue - gCost;
      // Margine netto: prende _net_margin_eur dal server (gia` decurta fee
      // PayPal per i record paypal, zero per voucher/free). Fallback locale:
      // se il campo manca (record antichi non passati per _apply_cancel_effective),
      // assume nessuna fee.
      const fee = Number(r._paypal_fee_eur || 0);
      const netMarg = (r._net_margin_eur != null) ? Number(r._net_margin_eur)
                                                  : (margEur - fee);
      const method = (r.payment_method || "").toLowerCase();
      const netTip = method === "paypal"
        ? `PayPal fee: ${fmtEur(fee)} (revenue × % + fissa)`
        : (method === "voucher" ? "Voucher: nessuna fee PayPal"
            : (revenue > 0 ? "Pagamento non PayPal: nessuna fee" : "Free: nessun pagamento"));
      const margPct = gCost > 0 ? (netMarg / gCost * 100) : 0;
      const dCls = margEur >= 0 ? "delta-positive" : "delta-negative";
      const nCls = netMarg >= 0 ? "delta-positive" : "delta-negative";
      const isLive = !!r._live;
      const rowCls = isLive ? "row-live" : "";
      const [bcls, blab] = OUTCOME_BADGE[r.outcome] || ["badge-muted", r.outcome || "?"];
      // Per cancel: mostra prezzo effettivo + tooltip con il pagato originale.
      const cancelTip = (r.cancel_retained_eur != null && r.user_price_eur_charged != null)
        ? ` title="Pagato: ${Number(r.user_price_eur_charged).toFixed(2)} € · Rimborso: ${Number(r.cancel_refund_eur || 0).toFixed(2)} €"`
        : "";
      return `<tr class="${rowCls}">
        <td>${ts}</td>
        <td><code>${esc(r.job_id)}</code></td>
        <td>${esc(r.model_key)}</td>
        <td>${esc(langLabel(r.language))}</td>
        <td>${(Number(r.chars_total) || 0).toLocaleString()}</td>
        <td>${(Number(r.audio_seconds_actual) || 0).toFixed(1)}</td>
        <td>${fmtEur(gCost)}</td>
        <td${cancelTip}>${fmtEur(revenue)}</td>
        <td class="${dCls}">${fmtEur(margEur)}</td>
        <td class="${nCls}" title="${esc(netTip)}">${fmtEur(netMarg)}</td>
        <td class="${dCls}">${margPct.toFixed(2)}%</td>
        <td><span class="badge ${bcls}">${esc(blab)}</span></td>
      </tr>`;
    }).join("");
  }

  $("auditRefreshBtn").addEventListener("click", fetchAudit);

  async function recalcParams(){
    const out = $("auditRecalcOutput");
    out.style.display = "block";
    out.textContent = "Caricamento...";
    try {
      const r = await fetch("/admin/api/gemini_cost_audit/recalc-params",
                            {headers: {"X-Admin-Token": ADMIN_TOKEN}});
      if (!r.ok) {
        out.textContent = "Errore: " + r.status;
        return;
      }
      const d = await r.json();
      const sugg = d.suggestions || [];
      if (!sugg.length) {
        out.textContent = "Nessun suggerimento (campioni insufficienti).";
      } else {
        out.textContent = sugg.join("\n");
      }
    } catch (e) {
      out.textContent = "Errore: " + e;
    }
  }
  $("auditRecalcBtn").addEventListener("click", recalcParams);

  // ---- Tab switching ----
  function showTab(name){
    document.querySelectorAll(".tab-btn").forEach(b => {
      b.classList.toggle("active", b.dataset.tab === name);
    });
    document.querySelectorAll(".tab-panel").forEach(p => {
      p.classList.toggle("active", p.id === "tab_" + name);
    });
    if (name === "gemini_events" && !window._evLoaded) {
      window._evLoaded = true;
      fetchEvents();
    }
    if (location.hash !== "#tab-" + name) {
      location.hash = "#tab-" + name;
    }
  }
  document.querySelectorAll(".tab-btn").forEach(b => {
    b.addEventListener("click", () => showTab(b.dataset.tab));
  });
  if (location.hash === "#tab-gemini" || location.hash === "#tab-gemini_events") {
    showTab("gemini_events");
  }

  // ---- Eventi & Rimborsi tab ----
  const REFUND_OUTCOMES = new Set([
    "failed_refunded", "failed_quota_refunded", "failed_budget_refunded",
    "failed_quality_refunded", "preflight_blocked_refunded",
    "cancelled_refunded", "cancelled_partial",
  ]);
  const OUTCOME_BADGE = {
    "running":                    ["badge-live", "In corso"],
    "completed":                  ["badge-ok",   "Completato"],
    "failed_refunded":            ["badge-err",  "Fallito (rimborso)"],
    "failed_quota_refunded":      ["badge-err",  "Quota esaurita"],
    "failed_budget_refunded":     ["badge-err",  "Budget superato"],
    "failed_quality_refunded":    ["badge-warn", "Qualità ins."],
    "preflight_blocked_refunded": ["badge-info", "Bloccato preflight"],
    "cancelled_refunded":         ["badge-muted","Annullato"],
    "cancelled_partial":          ["badge-muted","Annullato (parz.)"],
    "cancelled":                  ["badge-muted","Annullato"],
  };

  async function fetchEvents(){
    const params = new URLSearchParams();
    const m = $("evModelFilter").value;
    const l = $("evLangFilter").value;
    const df = $("evDateFrom").value;
    const dt = $("evDateTo").value;
    if (m && m !== "all") params.set("model", m);
    if (l && l !== "all") params.set("language", l);
    if (df) params.set("date_from", df);
    if (dt) params.set("date_to", dt);
    params.set("limit", "500");
    const r = await fetch("/admin/api/gemini_cost_audit?" + params.toString(),
                         {headers: {"X-Admin-Token": ADMIN_TOKEN}});
    if (!r.ok) { alert("Errore caricamento eventi: " + r.status); return; }
    const d = await r.json();
    renderEvents(d.records || []);
  }

  function renderEvents(recs){
    const onlyRefunds = $("evOnlyRefunds").checked;
    let filtered = recs;
    if (onlyRefunds) filtered = recs.filter(r => REFUND_OUTCOMES.has(r.outcome));
    // Aggregati
    let completed=0, refunds=0, preflight=0, cancelled=0;
    for (const r of recs) {
      if (r.outcome === "completed") completed++;
      else if (r.outcome === "preflight_blocked_refunded") { preflight++; refunds++; }
      else if (REFUND_OUTCOMES.has(r.outcome)) {
        refunds++;
        if (r.outcome.startsWith("cancelled")) cancelled++;
      }
    }
    $("evCount").textContent = recs.length;
    $("evCompleted").textContent = completed;
    $("evRefunds").textContent = refunds;
    $("evPreflight").textContent = preflight;
    $("evCancelled").textContent = cancelled;

    const tbody = $("evRecordsBody");
    if (!filtered.length) {
      tbody.innerHTML = '<tr><td colspan="8" class="empty-msg">Nessun evento.</td></tr>';
      return;
    }
    // Live in cima, poi resto desc per ts.
    filtered = filtered.slice().sort((a,b) => {
      const la = a._live ? 1 : 0, lb = b._live ? 1 : 0;
      if (la !== lb) return lb - la;
      return (b.ts||"").localeCompare(a.ts||"");
    });
    tbody.innerHTML = filtered.map(r => {
      const ts = esc((r.ts || "").slice(0, 19).replace("T", " "));
      const out = r.outcome || "?";
      const [cls, label] = OUTCOME_BADGE[out] || ["badge-muted", out];
      const rowCls = r._live ? "row-live"
        : (REFUND_OUTCOMES.has(out)
            ? (out === "preflight_blocked_refunded" ? "row-preflight" : "row-refund")
            : "");
      const revenue = (r._eff_revenue_eur != null) ? Number(r._eff_revenue_eur)
                                                  : Number(r.user_price_eur_charged || 0);
      const cancelTip = (r.cancel_retained_eur != null && r.user_price_eur_charged != null)
        ? ` title="Pagato: ${Number(r.user_price_eur_charged).toFixed(2)} € · Rimborso: ${Number(r.cancel_refund_eur || 0).toFixed(2)} €"`
        : "";
      return `<tr class="${rowCls}">
        <td>${ts}</td>
        <td><code>${esc(r.job_id)}</code></td>
        <td><span class="badge ${cls}">${esc(label)}</span></td>
        <td>${esc(r.model_key)}</td>
        <td>${esc(langLabel(r.language))}</td>
        <td>${(Number(r.chars_total) || 0).toLocaleString()}</td>
        <td${cancelTip}>${fmtEur(revenue)}</td>
        <td>${fmtEur(r.google_cost_eur_actual)}</td>
      </tr>`;
    }).join("");
  }

  $("evRefreshBtn").addEventListener("click", fetchEvents);
  $("evOnlyRefunds").addEventListener("change", () => fetchEvents());

  // ---- Kill-switch voci PREMIUM ----
  async function ksRefresh(){
    try {
      const r = await fetch("/admin/api/gemini_kill_switch",
                            {headers: {"X-Admin-Token": ADMIN_TOKEN}});
      if (!r.ok) {
        $("ksStatus").textContent = "Errore caricamento stato (" + r.status + ")";
        return;
      }
      const s = await r.json();
      ksApply(s);
    } catch (e) {
      $("ksStatus").textContent = "Errore: " + e;
    }
  }
  function ksApply(s){
    const btn = $("ksToggleBtn");
    const status = $("ksStatus");
    const reasonRow = $("ksReason");
    btn.disabled = false;
    if (!s.capability_ok && !s.disabled) {
      status.innerHTML = '<span style="color:var(--muted)">Gemini TTS non configurato (nessun backend valido).</span>';
      btn.textContent = "Non disponibile";
      btn.disabled = true;
      reasonRow.style.display = "none";
      return;
    }
    if (s.disabled) {
      status.innerHTML = '<span style="color:var(--err);font-weight:600">DISATTIVO</span> · pannello voci PREMIUM nascosto agli utenti';
      btn.textContent = "Riattiva voci PREMIUM";
      btn.style.background = "var(--ok)";
      if (s.reason) {
        reasonRow.textContent = "Motivo: " + s.reason + (s.updated_at ? " (" + s.updated_at.slice(0,19).replace("T"," ") + ")" : "");
        reasonRow.style.display = "block";
      } else {
        reasonRow.style.display = "none";
      }
    } else {
      status.innerHTML = '<span style="color:var(--ok);font-weight:600">ATTIVO</span> · voci PREMIUM offerte agli utenti';
      btn.textContent = "Disattiva voci PREMIUM";
      btn.style.background = "var(--err)";
      reasonRow.style.display = "none";
    }
    btn.dataset.targetDisabled = s.disabled ? "0" : "1";
  }
  async function ksToggle(){
    const btn = $("ksToggleBtn");
    const target = btn.dataset.targetDisabled === "1";
    const reason = $("ksReasonInput").value.trim();
    if (target && !confirm("Disattivare il pannello Voci PREMIUM per tutti gli utenti?\n\nLe stime e i pagamenti Premium risponderanno 503 finché non viene riattivato.")) {
      return;
    }
    btn.disabled = true;
    btn.textContent = "...";
    try {
      const r = await fetch("/admin/api/gemini_kill_switch", {
        method: "POST",
        headers: {"X-Admin-Token": ADMIN_TOKEN, "Content-Type": "application/json"},
        body: JSON.stringify({disabled: target, reason: reason}),
      });
      if (!r.ok) { alert("Errore: " + r.status); await ksRefresh(); return; }
      const s = await r.json();
      ksApply(s);
      if (!target) $("ksReasonInput").value = "";
    } catch (e) {
      alert("Errore: " + e);
      await ksRefresh();
    }
  }
  $("ksToggleBtn").addEventListener("click", ksToggle);

  // Auto-load on page open: prima popola la dropdown lingue, poi carica i record.
  loadLanguageOptions().finally(fetchAudit);
  ksRefresh();
})();
</script>
</body></html>"""
    return html, 200, {"Content-Type": "text/html; charset=utf-8"}


_ACTIVE_JOB_STATUSES = ("queued", "running", "generating", "paused", "starting")


def _synth_running_gemini_audit_records():
    """Snapshot dei job Gemini attivi sintetizzato in forma audit-shaped.

    Permette a /admin/audit-tts di mostrare una riga immediatamente all'avvio
    di una generazione Premium, aggiornata ad ogni refresh con i dati di
    costo accumulati in `job["gemini_actual"]`. Quando il job termina, la riga
    "running" sparisce e viene rimpiazzata dal record persistito nel JSONL.
    """
    out = []
    if gemini_tts is None:
        return out
    now_iso = datetime.now(timezone.utc).isoformat()
    with _jobs_lock:
        snapshot = list(jobs.items())
    for job_id, job in snapshot:
        try:
            if not isinstance(job, dict):
                continue
            status = job.get("status", "")
            if status not in _ACTIVE_JOB_STATUSES:
                continue
            voice = job.get("voice") or job.get("opt_voice") or ""
            if not _is_gemini_voice(voice):
                continue
            ga = job.get("gemini_actual") or {}
            parts = voice.split(":")
            model_key = parts[1] if len(parts) >= 3 else (ga.get("model_key") or "?")
            info = job.get("info")
            language = (getattr(info, "language", "") or "").split("-")[0].lower() if info else ""
            payment = job.get("payment") or {}
            charged = float(payment.get("total_eur", 0) or 0)
            if charged <= 0:
                charged = float(job.get("payment_amount_eur", 0) or 0)
            google_cost_actual = float(ga.get("google_cost_eur", 0.0) or 0.0)
            try:
                should = gemini_tts.compute_user_price_eur(google_cost_actual, model_key)
                should_have_been = float(should.get("user_price_eur", 0.0))
            except Exception:
                should_have_been = 0.0
            delta_eur = round(should_have_been - charged, 4)
            rate_raw = job.get("rate", "+0%")
            try:
                rate_pct = int(str(rate_raw).replace("%", "").replace("+", "").strip() or 0)
            except Exception:
                rate_pct = 0
            rec = {
                "ts": job.get("started_at") or now_iso,
                "job_id": job_id,
                "model_key": model_key,
                "language": language,
                "rate_pct": rate_pct,
                "chars_total": int(ga.get("chars", 0) or 0),
                "audio_seconds_actual": round(float(ga.get("audio_seconds", 0) or 0), 2),
                "google_cost_eur_actual": round(google_cost_actual, 4),
                "user_price_eur_charged": round(charged, 4),
                "user_price_eur_should_have_been": round(should_have_been, 2),
                "delta_eur": delta_eur,
                "margin_eur_actual": round(charged - google_cost_actual, 4),
                "outcome": "running",
                "_live": True,
            }
            out.append(rec)
        except Exception:
            continue
    return out


# Outcome che rappresentano un rimborso TOTALE all'utente: il ricavo
# effettivo e' 0 a prescindere da `user_price_eur_charged` originario.
# I `cancelled_partial` sono trattati a parte tramite `cancel_retained_eur`.
_FULL_REFUND_OUTCOMES = frozenset({
    "failed_refunded",
    "failed_quota_refunded",
    "failed_budget_refunded",
    "failed_quality_refunded",
    "preflight_blocked_refunded",
    "cancelled_refunded",
})


def _compute_paypal_fee_eur(revenue_eur, payment_method):
    """Stima fee PayPal addebitata sul ricavo PER QUESTO record.

    Applica la formula PayPal standard (`% gross + fissa`) solo se il record
    è stato pagato via PayPal e ha ricavo > 0; altrimenti ritorna 0:
    - `voucher`: PayPal non e' coinvolto (lo era stato all'origine del voucher,
      ma la fee era gia` stata pagata in quella transazione → non si addebita
      due volte qui).
    - rimborso totale (revenue_eur = 0): la fee viene quasi sempre trattenuta
      da PayPal anche sui rimborsi, ma per l'audit del margine sul singolo job
      consideriamo 0 (non c'e' ricavo da decurtare).
    - free (nessun pagamento): 0.
    - record legacy senza `payment_method`: 0 (conservativo: non assumiamo
      paypal in assenza di dato).
    """
    try:
        rev = float(revenue_eur or 0)
    except (TypeError, ValueError):
        return 0.0
    if rev <= 0:
        return 0.0
    if (payment_method or "").lower() != "paypal":
        return 0.0
    if gemini_tts is None:
        return 0.0
    try:
        pct = float(gemini_tts.PAYPAL_PERCENT_FEE)
        fixed = float(gemini_tts.PAYPAL_FIXED_FEE_EUR)
    except (AttributeError, TypeError, ValueError):
        return 0.0
    return round(rev * pct / 100.0 + fixed, 4)


def _apply_cancel_effective(rec):
    """Augmenta il record con `_eff_*` che riflettono i rimborsi reali.

    - `cancelled_partial`: ricavo = quota trattenuta (`cancel_retained_eur`),
      a copertura del consumato; eventuale eccedenza restituita.
    - `*_refunded` (rimborso totale: qualita`/quota/budget/preflight/cancel
      pre-attivita`/failure generico): ricavo = 0, quindi margine = -costo.
    - altri (completed, running, ...): ricavo = `user_price_eur_charged`.

    Aggiunge inoltre `_paypal_fee_eur` e `_net_margin_eur` (margine al netto
    delle fee PayPal: zero per voucher/free).
    """
    if not isinstance(rec, dict):
        return rec
    charged = float(rec.get("user_price_eur_charged", 0) or 0)
    cost = float(rec.get("google_cost_eur_actual", 0) or 0)
    should = float(rec.get("user_price_eur_should_have_been", 0) or 0)
    cancel_retained = rec.get("cancel_retained_eur")
    outcome = rec.get("outcome") or ""
    if outcome in _FULL_REFUND_OUTCOMES:
        revenue = 0.0
    elif cancel_retained is not None:
        revenue = float(cancel_retained or 0)
    else:
        revenue = charged
    rec["_eff_revenue_eur"] = round(revenue, 4)
    rec["_eff_margin_eur"] = round(revenue - cost, 4)
    rec["_eff_delta_eur"] = round(should - revenue, 4)
    rec["_eff_delta_pct"] = round((rec["_eff_delta_eur"] / cost * 100), 2) if cost > 0 else 0.0
    fee = _compute_paypal_fee_eur(revenue, rec.get("payment_method", ""))
    rec["_paypal_fee_eur"] = round(fee, 4)
    rec["_net_margin_eur"] = round(revenue - cost - fee, 4)
    return rec


@app.route("/admin/api/gemini_cost_audit", methods=["GET"])
def admin_api_gemini_cost_audit():
    """List Gemini TTS audit records with filters + aggregates. Admin-only.

    Restituisce, in aggiunta ai record persistiti su JSONL, righe sintetiche
    `outcome="running"` per i job Gemini attualmente in corso (snapshot live).
    Ogni record viene arricchito con campi `_eff_*` che applicano i rimborsi
    da cancellazione anticipata su ricavo/margine/delta.
    """
    if not ADMIN_TOKEN:
        return jsonify({"error": "Admin UI disabled"}), 404
    if not _admin_auth_ok(_admin_auth_from_request()):
        time.sleep(0.5)
        return jsonify({"error": "Unauthorized"}), 401

    import gemini_cost_audit
    model = request.args.get("model")
    language = request.args.get("language")
    outcome = request.args.get("outcome")
    date_from = request.args.get("date_from")
    date_to = request.args.get("date_to")
    try:
        limit = max(1, min(int(request.args.get("limit", 200)), 1000))
        offset = max(0, int(request.args.get("offset", 0)))
    except (TypeError, ValueError):
        return jsonify({"error": "invalid limit/offset"}), 400

    def _norm(v):
        return v if v and v != "all" else None

    persisted = list(gemini_cost_audit.iter_records(
        model=_norm(model), language=_norm(language), outcome=_norm(outcome),
        date_from=date_from, date_to=date_to,
    ))
    persisted_ids = {r.get("job_id") for r in persisted}

    # Inietta i job in corso: rispetta i filtri model/lang; outcome="running"
    # e' visibile solo quando outcome=all o esattamente "running".
    live = []
    out_filter = _norm(outcome)
    if out_filter in (None, "running"):
        for r in _synth_running_gemini_audit_records():
            if r.get("job_id") in persisted_ids:
                continue
            if _norm(model) and r.get("model_key") != _norm(model):
                continue
            if _norm(language) and r.get("language") != _norm(language):
                continue
            live.append(r)

    recs = live + persisted
    for r in recs:
        _apply_cancel_effective(r)
    total = len(recs)
    page = recs[offset:offset + limit]

    # Aggregati ricomputati sui record arricchiti (solo "completed" + i live
    # in corso, esclusi rimborsi totali). Ricavo/margine usano _eff_* per
    # tener conto dei rimborsi da cancel.
    agg_n = 0
    agg_rev = 0.0
    agg_cost = 0.0
    agg_delta = 0.0
    agg_fee = 0.0
    for r in recs:
        oc = r.get("outcome") or ""
        if oc not in ("completed", "running", "cancelled_partial"):
            continue
        agg_n += 1
        agg_rev += float(r.get("_eff_revenue_eur", 0) or 0)
        agg_cost += float(r.get("google_cost_eur_actual", 0) or 0)
        agg_delta += float(r.get("_eff_delta_eur", 0) or 0)
        agg_fee += float(r.get("_paypal_fee_eur", 0) or 0)
    agg = {
        "count": agg_n,
        "revenue_eur": round(agg_rev, 4),
        "google_cost_eur": round(agg_cost, 4),
        "margin_eur": round(agg_rev - agg_cost, 4),
        "paypal_fees_eur": round(agg_fee, 4),
        "net_margin_eur": round(agg_rev - agg_cost - agg_fee, 4),
        "delta_pct_avg": round((agg_delta / agg_cost * 100), 2) if agg_cost > 0 else 0.0,
        "filters": {"model": _norm(model), "language": _norm(language),
                    "date_from": date_from, "date_to": date_to},
    }
    return jsonify({"records": page, "count": total, "aggregates": agg})


@app.route("/admin/api/gemini_kill_switch", methods=["GET", "POST"])
def admin_api_gemini_kill_switch():
    """Kill-switch admin per disattivare runtime le voci PREMIUM.

    GET  -> stato corrente {disabled, reason, updated_at, capability_ok}.
    POST -> attiva/disattiva. Body JSON: {"disabled": bool, "reason": str?}.

    Quando disabled=True, gemini_tts.is_available() ritorna False e il
    pannello Voci PREMIUM scompare dalla UI utente (incluse stime e flusso
    di pagamento Premium, che gia` gateano su is_available()).
    """
    if not ADMIN_TOKEN:
        return jsonify({"error": "Admin UI disabled"}), 404
    if not _admin_auth_ok(_admin_auth_from_request()):
        time.sleep(0.5)
        return jsonify({"error": "Unauthorized"}), 401

    if gemini_tts is None:
        return jsonify({"error": "Gemini TTS module not loaded"}), 503

    if request.method == "POST":
        data = request.get_json(silent=True) or {}
        disabled = bool(data.get("disabled", False))
        reason = str(data.get("reason", "") or "")
        state = gemini_tts.set_admin_disabled(disabled, reason)
        # Invalida la cache voci cosi` la prossima /api/voices riflette subito
        # il cambio (rimozione/aggiunta dell'optgroup PREMIUM).
        _invalidate_voices_cache()
        _log_activity("", "", "ADMIN_GEMINI_KILLSWITCH", "", _get_client_ip(),
                      "disabled" if disabled else "enabled", reason[:80])
        print(f"[admin] Gemini PREMIUM kill-switch: "
              f"{'DISABLED' if disabled else 'ENABLED'} reason={reason!r}")
        return jsonify({**state, "capability_ok": _gemini_capability_ok()})

    return jsonify({
        **gemini_tts.admin_disabled_state(),
        "capability_ok": _gemini_capability_ok(),
    })


def _gemini_capability_ok():
    """True se Gemini TTS e' tecnicamente configurato (a prescindere dal
    kill-switch). Usato dal pannello admin per distinguere 'spento per scelta'
    da 'non configurato'.
    """
    if gemini_tts is None:
        return False
    try:
        return bool(gemini_tts.is_capability_available())
    except Exception:
        return False


@app.route("/admin/api/gemini_cost_audit/languages", methods=["GET"])
def admin_api_gemini_cost_audit_languages():
    """Restituisce le lingue distinte realmente presenti nei record audit Gemini.

    Usato dalla pagina /admin/audit-tts per popolare dinamicamente il filtro
    lingua con i codici effettivamente generati, evitando di confondere
    "lingue UI" con "lingue di generazione TTS".
    """
    if not ADMIN_TOKEN:
        return jsonify({"error": "Admin UI disabled"}), 404
    if not _admin_auth_ok(_admin_auth_from_request()):
        time.sleep(0.5)
        return jsonify({"error": "Unauthorized"}), 401

    import gemini_cost_audit
    seen = set()
    for rec in gemini_cost_audit.iter_records():
        lang = (rec.get("language") or "").strip().lower()
        if lang:
            seen.add(lang)
    return jsonify({"languages": sorted(seen)})


@app.route("/admin/api/gemini_cost_audit/recalc-params", methods=["GET"])
def admin_api_gemini_recalc_params():
    """Aggrega audit records completed per (model, lang) e suggerisce tuning."""
    if not ADMIN_TOKEN:
        return jsonify({"error": "Admin UI disabled"}), 404
    if not _admin_auth_ok(_admin_auth_from_request()):
        time.sleep(0.5)
        return jsonify({"error": "Unauthorized"}), 401

    import gemini_cost_audit
    # Doppio raggruppamento:
    #   - groups_global: (model, lang)              — vista aggregata storica
    #   - groups_rate:   (model, lang, rate_step)   — calibrazione per velocità
    # La velocità incide direttamente sul prezzo proposto (estimate_audio_seconds
    # scala con rate_pct), quindi i delta vanno monitorati anche per rate.
    groups_global = {}
    groups_rate = {}
    for rec in gemini_cost_audit.iter_records(outcome="completed"):
        model = rec.get("model_key") or "?"
        lang = rec.get("language") or "?"
        # I record precedenti l'introduzione di rate_step potrebbero non averlo:
        # li trattiamo come rate_step=0 (velocità "normale").
        try:
            rstep = int(rec.get("rate_step") if rec.get("rate_step") is not None else 0)
        except Exception:
            rstep = 0
        groups_global.setdefault((model, lang), []).append(rec)
        groups_rate.setdefault((model, lang, rstep), []).append(rec)

    def _label_for(avg_delta_pct):
        if avg_delta_pct > 5:
            return "margine alto, valuta riduzione tariffa utente"
        if avg_delta_pct < -5:
            return "margine in perdita, valuta aumento tariffa utente o sec_per_kchars"
        return "parametri OK"

    def _rate_label(step):
        # Mappa step -> etichetta UI (cfr. SPEED_KEYS in app.js)
        return {-3: "vs", -2: "s", -1: "ss", 0: "n", 1: "sf", 2: "f", 3: "vf"}.get(int(step), str(step))

    # DELTA% = delta_eur / costo Google (formula del ricarico applicato).
    # Calcolato sui totali del gruppo: piu` robusto della media semplice di
    # percentuali (outlier su record con costo Google molto piccolo) e
    # coerente sui record storici salvati con la vecchia formula.
    def _avg_delta_pct(recs):
        d_sum = sum(float(r.get("delta_eur") or 0) for r in recs)
        c_sum = sum(float(r.get("google_cost_eur_actual") or 0) for r in recs)
        return (d_sum / c_sum * 100.0) if c_sum > 0 else 0.0

    suggestions = []
    suggestions.append("=== Aggregato globale (model / lang) ===")
    if not groups_global:
        suggestions.append("  (nessun record disponibile)")
    for (model, lang), recs in sorted(groups_global.items()):
        if len(recs) < 3:
            suggestions.append(f"  [{model} / {lang}] (n={len(recs)}) campioni insufficienti (servono >=3)")
            continue
        avg = _avg_delta_pct(recs)
        suggestions.append(f"  [{model} / {lang}] (n={len(recs)}) avg delta {avg:+.2f}% — {_label_for(avg)}")

    suggestions.append("")
    suggestions.append("=== Per velocità (model / lang / rate_step) ===")
    if not groups_rate:
        suggestions.append("  (nessun record disponibile)")
    for (model, lang, rstep), recs in sorted(groups_rate.items()):
        n = len(recs)
        avg = _avg_delta_pct(recs)
        rate_pcts = [int(r.get("rate_pct") or 0) for r in recs]
        rp_min, rp_max = (min(rate_pcts), max(rate_pcts)) if rate_pcts else (0, 0)
        rp_str = f"{rp_min:+d}%" if rp_min == rp_max else f"{rp_min:+d}%..{rp_max:+d}%"
        head = f"  [{model} / {lang} / step={rstep:+d} ({_rate_label(rstep)}) {rp_str}] (n={n})"
        if n < 3:
            suggestions.append(f"{head} campioni insufficienti")
        else:
            suggestions.append(f"{head} avg delta {avg:+.2f}% — {_label_for(avg)}")
    return jsonify({
        "suggestions": suggestions,
        "groups_total": len(groups_global),
        "groups_evaluated": sum(1 for g in groups_global.values() if len(g) >= 3),
        "groups_by_rate_total": len(groups_rate),
        "groups_by_rate_evaluated": sum(1 for g in groups_rate.values() if len(g) >= 3),
    })


@app.route("/api/voices")
def api_voices():
    try:
        voices = get_voices()
        # Includi info budget Google TTS come chiave speciale _google_tts
        if google_tts is not None and google_tts.is_available():
            used, remaining, limit = google_tts.get_usage()
            voices["_google_tts"] = {
                "available": remaining > 0,
                "chars_remaining": remaining,
                "chars_limit": limit,
            }
        # Stato voci PREMIUM: distingue "non configurato" (capability_ok=False)
        # da "spento per scelta admin" (admin_disabled=True). Serve alla UI per
        # mostrare il tab Premium con popup di manutenzione invece di nasconderlo,
        # cosi` l'utente non vede combo vuote che fanno sembrare l'app rotta.
        if gemini_tts is not None:
            try:
                cap_ok = bool(gemini_tts.is_capability_available())
                admin_state = gemini_tts.admin_disabled_state()
                voices["_premium_status"] = {
                    "capability_ok": cap_ok,
                    "admin_disabled": bool(admin_state.get("disabled", False)),
                }
            except Exception:
                pass
        return jsonify(voices)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/community/stats/today")
def api_community_stats_today():
    """Conteggio audiolibri completati oggi (cache 60s)."""
    return jsonify({"count": _stats_today_count()})


@app.route("/api/community/stats/month")
def api_community_stats_month():
    """Aggregato mensile per lingua TTS (cache 5min).
    Schema: {monthly: int, top: [{lang, count}, ...], other: int}."""
    return jsonify(_stats_month_by_lang())


# ─── COMMUNITY: NEWS ────────────────────────────────────────────────
_NEWS_TAGS = {"feature", "fix", "info"}
_NEWS_LANGS = {"it", "en", "fr", "es", "de", "zh", "hi"}


def _sanitize_text(s, maxlen):
    """Strip HTML tags + collapse whitespace + truncate."""
    if not isinstance(s, str):
        return ""
    s = re.sub(r"<[^>]+>", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s[:maxlen]


@app.route("/api/community/news", methods=["GET"])
def api_community_news():
    """Lista news pubbliche (top 10 non archiviate, sort desc per created_at)."""
    items = community_store.news().all(include_archived=False)
    items = sorted(items, key=lambda x: x.get("created_at", 0), reverse=True)[:10]
    out = []
    for it in items:
        out.append({
            "id": it.get("id"),
            "tag": it.get("tag", "info"),
            "title": it.get("title", ""),
            "body": it.get("body", ""),
            "lang": it.get("lang", "en"),
            "title_i18n": it.get("title_i18n") or {},
            "body_i18n": it.get("body_i18n") or {},
            "banner": bool(it.get("banner", False)),
            "created_at": it.get("created_at", 0),
        })
    return jsonify({"items": out})


def _translate_news_async(item_id: str, title: str, body: str) -> None:
    """Background: translate a news item title+body, persist."""
    payload = {}
    if title and title.strip():
        payload["title"] = title
    if body and body.strip():
        payload["body"] = body
    if not payload:
        return
    if not community_translator.is_available():
        return
    result = community_translator.translate(payload)
    if not result:
        return
    patch: dict = {}
    if "title" in payload:
        patch["title_i18n"] = {
            lg: (result.get(lg) or {}).get("title", "")
            for lg in community_translator.LANGS
        }
    if "body" in payload:
        patch["body_i18n"] = {
            lg: (result.get(lg) or {}).get("body", "")
            for lg in community_translator.LANGS
        }
    if patch:
        try:
            community_store.news().update(item_id, patch)
        except Exception as e:
            print(f"[news] translation persist failed for {item_id}: {e!s}")


@app.route("/admin/api/news", methods=["POST"])
def admin_api_news_create():
    if not _admin_auth_ok(_admin_auth_from_request()):
        return ("forbidden", 403)
    body = request.get_json(silent=True) or {}
    tag = (body.get("tag") or "info").strip().lower()
    if tag not in _NEWS_TAGS:
        return jsonify({"error": "invalid tag"}), 400
    title = _sanitize_text(body.get("title"), 200)
    if not title:
        return jsonify({"error": "title required"}), 400
    text = _sanitize_text(body.get("body"), 2000)
    lang = (body.get("lang") or "en").strip().lower()
    if lang not in _NEWS_LANGS:
        return jsonify({"error": "invalid lang"}), 400
    banner = bool(body.get("banner", False))
    item = community_store.news().add({
        "tag": tag, "title": title, "body": text,
        "lang": lang, "banner": banner,
    })
    # translate title+body into all UI langs (best-effort, async)
    threading.Thread(
        target=_translate_news_async,
        args=(item["id"], title, text),
        daemon=True,
    ).start()
    return jsonify(item), 200


@app.route("/admin/api/news/<item_id>", methods=["POST"])
def admin_api_news_update(item_id):
    if not _admin_auth_ok(_admin_auth_from_request()):
        return ("forbidden", 403)
    body = request.get_json(silent=True) or {}
    action = body.get("action")
    store = community_store.news()
    if action == "archive":
        ok = store.archive(item_id)
    elif action == "unarchive":
        ok = store.unarchive(item_id)
    elif action == "delete":
        ok = store.delete(item_id)
    elif action == "toggle_banner":
        cur = store.get(item_id)
        if not cur:
            return ("not found", 404)
        ok = store.update(item_id, {"banner": not cur.get("banner", False)}) is not None
    else:
        return jsonify({"error": "invalid action"}), 400
    return ("ok", 200) if ok else ("not found", 404)


@app.route("/admin/api/news/list", methods=["GET"])
def admin_api_news_list():
    if not _admin_auth_ok(_admin_auth_from_request()):
        return ("forbidden", 403)
    items = community_store.news().all(include_archived=True)
    items = sorted(items, key=lambda x: x.get("created_at", 0), reverse=True)
    return jsonify({"items": items})


# ─── COMMUNITY: FEEDBACK ────────────────────────────────────────────
import hashlib

_feedback_rate_lock = threading.Lock()
_feedback_rate: dict[str, list[float]] = {}  # ip_hash -> list[ts]
_FB_LIMIT_HOUR = 1
_FB_LIMIT_DAY = 5
_feedback_email_lock = threading.Lock()
_feedback_email_last = 0.0
_FB_EMAIL_THROTTLE = 1800.0  # 30 min

_IP_SALT = os.environ.get("ABM_IP_SALT") or "abm-default-salt-v1"


# ─── Rate limit generico IP-based (sliding window) ─────────────────
# Usato per endpoint costosi (upload, preview audio) per limitare DoS.
_ip_rl_lock = threading.Lock()
_ip_rl_buckets: dict[str, dict[str, list[float]]] = {}


def _client_ip() -> str:
    """Estrae IP client (rispetta X-Forwarded-For del reverse proxy)."""
    xff = request.headers.get("X-Forwarded-For", "")
    if xff:
        return xff.split(",")[0].strip()
    return request.remote_addr or ""


def _ip_rl_check(bucket: str, ip: str, limit_per_min: int, limit_per_hour: int):
    """Sliding-window rate limit IP-based. Ritorna (allowed, retry_after_sec).
    bucket: identificativo logico (es. 'analyze', 'preview').
    """
    if not ip:
        return True, 0
    now = time.time()
    with _ip_rl_lock:
        per_bucket = _ip_rl_buckets.setdefault(bucket, {})
        hits = per_bucket.get(ip, [])
        hits = [t for t in hits if now - t < 3600]
        last_min = [t for t in hits if now - t < 60]
        if len(last_min) >= limit_per_min:
            retry = 60 - int(now - last_min[0])
            per_bucket[ip] = hits
            return False, max(1, retry)
        if len(hits) >= limit_per_hour:
            retry = 3600 - int(now - hits[0])
            per_bucket[ip] = hits
            return False, max(1, retry)
        hits.append(now)
        per_bucket[ip] = hits
    return True, 0


# Default limits (override via env per ops emergency)
_ANALYZE_RL_PER_MIN = int(os.environ.get("ABM_ANALYZE_RL_PER_MIN", "5"))
_ANALYZE_RL_PER_HOUR = int(os.environ.get("ABM_ANALYZE_RL_PER_HOUR", "30"))
_PREVIEW_RL_PER_MIN = int(os.environ.get("ABM_PREVIEW_RL_PER_MIN", "20"))
_PREVIEW_RL_PER_HOUR = int(os.environ.get("ABM_PREVIEW_RL_PER_HOUR", "200"))


def _hash_ip(ip: str) -> str:
    h = hashlib.sha256((_IP_SALT + ":" + (ip or "")).encode("utf-8")).hexdigest()
    return h[:16]


def _feedback_check_rate(ip_hash: str) -> bool:
    """True se il client può inviare ora; False se sopra il limite."""
    now = time.time()
    # consulta il persistent store: i feedback cancellati (hard-delete)
    # non devono bloccare il re-inserimento
    items = community_store.feedback().all(include_archived=False)
    recent = [it for it in items if it.get("ip_hash") == ip_hash and (now - it.get("created_at", 0)) < 86400]
    last_hour = sum(1 for it in recent if (now - it.get("created_at", 0)) < 3600)
    last_day = len(recent)
    if last_hour >= _FB_LIMIT_HOUR or last_day >= _FB_LIMIT_DAY:
        return False
    with _feedback_rate_lock:
        hist = _feedback_rate.get(ip_hash, [])
        hist = [t for t in hist if now - t < 86400]
        hist.append(now)
        _feedback_rate[ip_hash] = hist
    return True


def _notify_admin_new_feedback(item: dict, comment_it: str | None = None, unvalidated: bool = False) -> None:
    """Throttled (30 min) email all'admin per nuovo feedback.

    Se ``comment_it`` è fornito, viene usato come corpo del commento
    (tipicamente la traduzione italiana prodotta dall'LLM); altrimenti
    si usa il commento originale presente in ``item``.
    """
    global _feedback_email_last
    if not ADMIN_EMAIL:
        return
    now = time.time()
    with _feedback_email_lock:
        if now - _feedback_email_last < _FB_EMAIL_THROTTLE:
            return
        _feedback_email_last = now
    try:
        rating = int(item.get("rating", 0))
        unvalidated_flag = unvalidated or bool(item.get("moderation_unvalidated", False))
        stars = "★" * rating + "☆" * (5 - rating)
        name = html_mod.escape(item.get("name") or "Anonimo")
        original = item.get("comment") or ""
        chosen_text = comment_it if (comment_it and comment_it.strip()) else original
        comment = html_mod.escape(chosen_text)
        # mostra anche l'originale se diverso dalla traduzione usata
        original_block = ""
        if comment_it and original and comment_it.strip() != original.strip():
            original_block = (
                f"<p style='font-size:.9em;color:#666;margin-top:10px'>"
                f"<b>Originale:</b></p>"
                f"<p style='border-left:3px solid #ccc;padding-left:8px;color:#666'>"
                f"{html_mod.escape(original)}</p>"
            )
        unvalidated_banner = ""
        if unvalidated_flag:
            unvalidated_banner = (
                "<p style='background:#fff3cd;border:1px solid #ffc107;padding:8px 12px;"
                "border-radius:4px;color:#856404;margin-bottom:12px'>"
                "<b>Attenzione:</b> questo commento non è stato validato dal sistema LLM."
                "</p>"
            )
        body = (
            f"{unvalidated_banner}"
            f"<p><b>Nuovo feedback ricevuto:</b></p>"
            f"<p>Voto: <span style='font-size:1.2em'>{stars}</span> ({rating}/5)</p>"
            f"<p>Nome: {name}</p>"
            f"<p>Commento (IT):</p><p style='border-left:3px solid #d9a441;padding-left:8px'>"
            f"{comment or '<i>(nessun commento)</i>'}</p>"
            f"{original_block}"
            f"<p style='font-size:.85em;color:#888'>ID: {item.get('id','')}</p>"
        )
        subject = f"[ABM] Nuovo feedback: {rating}★"
        if unvalidated_flag:
            subject = "[NON VALIDATO — LLM offline] " + subject
        email_service._send_email(ADMIN_EMAIL, subject, body)
    except Exception as e:
        print(f"[feedback] admin email failed: {e!s}")


@app.route("/api/community/feedback", methods=["GET"])
def api_community_feedback_list():
    """Lista feedback pubblici + statistiche aggregate."""
    items = community_store.feedback().all(include_archived=False)
    items = sorted(items, key=lambda x: x.get("created_at", 0), reverse=True)
    total = len(items)
    if total > 0:
        avg = round(sum(int(it.get("rating", 0)) for it in items) / total, 2)
    else:
        avg = 0.0
    histogram = [0, 0, 0, 0, 0]  # rating 1..5 -> idx 0..4
    for it in items:
        r = int(it.get("rating", 0))
        if 1 <= r <= 5:
            histogram[r - 1] += 1
    public_items = []
    for it in items[:50]:
        public_items.append({
            "id": it.get("id"),
            "rating": it.get("rating"),
            "name": it.get("name") or "",
            "comment": it.get("comment") or "",
            "comment_lang": it.get("comment_lang") or "",
            "comment_i18n": it.get("comment_i18n") or {},
            "created_at": it.get("created_at", 0),
            "admin_reply_at": it.get("admin_reply_at", 0),
            "admin_reply_lang": it.get("admin_reply_lang") or "",
            "admin_reply_text": it.get("admin_reply_text") or "",
            "admin_reply_i18n": it.get("admin_reply_i18n") or {},
        })
    return jsonify({"items": public_items, "avg": avg, "total": total, "histogram": histogram})


def _process_new_feedback(item: dict) -> None:
    """Background: translate the comment (best-effort), persist the i18n
    fields, then send the admin email using the Italian translation.

    Always emails the admin (subject to the existing throttle), even if
    translation is unavailable or empty.
    """
    item_id = item.get("id") or ""
    comment = (item.get("comment") or "").strip()
    print(f"[feedback] post-process id={item_id} comment_len={len(comment)} "
          f"llm_available={community_translator.is_available()}")
    comment_it: str | None = None
    if comment and community_translator.is_available():
        try:
            result = community_translator.translate({"comment": comment})
        except Exception as e:
            print(f"[feedback] translation call raised: {e!s}")
            result = None
        print(f"[feedback] translation result id={item_id}: "
              f"{'ok' if result else 'FAILED'}")
        if result:
            patch = {
                "comment_lang": result.get("source_lang") or "",
                "comment_i18n": {
                    lg: (result.get(lg) or {}).get("comment", "")
                    for lg in community_translator.LANGS
                },
            }
            try:
                community_store.feedback().update(item_id, patch)
            except Exception as e:
                print(f"[feedback] translation persist failed for {item_id}: {e!s}")
            comment_it = (patch["comment_i18n"].get("it") or "").strip() or None
    # admin email — always, throttled internally
    try:
        _notify_admin_new_feedback(item, comment_it=comment_it, unvalidated=item.get("moderation_unvalidated", False))
    except Exception as e:
        print(f"[feedback] admin notify failed: {e!s}")


@app.route("/api/community/feedback", methods=["POST"])
def api_community_feedback_create():
    body = request.get_json(silent=True) or {}
    # honeypot
    if body.get("website"):
        return ("", 204)
    # validate rating
    try:
        rating = int(body.get("rating", 0))
    except (TypeError, ValueError):
        rating = 0
    if rating < 1 or rating > 5:
        return jsonify({"error": "missing_rating"}), 400
    raw_name = (body.get("name") or "").strip()
    if len(raw_name) > 100:
        return jsonify({"error": "name_too_long"}), 400
    raw_comment = (body.get("comment") or "").strip()
    if len(raw_comment) > 500:
        return jsonify({"error": "comment_too_long"}), 400
    name = _sanitize_text(raw_name, 100)
    comment = _sanitize_text(raw_comment, 500)
    # rate limit
    ip = (request.headers.get("X-Forwarded-For", "").split(",")[0].strip()
          or request.remote_addr or "")
    ip_hash = _hash_ip(ip)
    if not _feedback_check_rate(ip_hash):
        return jsonify({"error": "rate_limit"}), 429
    # moderation gate
    mod_result = community_moderator.validate(name, comment)
    if not mod_result.get("approved", True):
        return jsonify({"error": "inappropriate_content"}), 400
    delete_token = secrets.token_urlsafe(24)
    item = community_store.feedback().add({
        "rating": rating,
        "name": name,
        "comment": comment,
        "ip_hash": ip_hash,
        "delete_token": delete_token,
        "moderation_unvalidated": mod_result.get("unvalidated", False),
    })
    # background: translate comment (if any) then email admin with IT version
    threading.Thread(target=_process_new_feedback, args=(item,), daemon=True).start()
    # Return the delete_token so the client can store it locally (only the
    # original poster, who has it, can later delete the comment).
    return jsonify({"id": item["id"], "delete_token": delete_token}), 200


@app.route("/api/community/feedback/<item_id>", methods=["DELETE"])
def api_community_feedback_delete(item_id: str):
    """Self-delete: requires the delete_token returned at creation time."""
    body = request.get_json(silent=True) or {}
    token = (body.get("token") or request.headers.get("X-Delete-Token") or "").strip()
    if not token:
        return jsonify({"error": "missing token"}), 400
    item = community_store.feedback().get(item_id)
    if not item:
        return jsonify({"error": "not found"}), 404
    expected = item.get("delete_token") or ""
    if not expected or not hmac.compare_digest(expected, token):
        return jsonify({"error": "forbidden"}), 403
    community_store.feedback().delete(item_id)
    return jsonify({"ok": True}), 200


@app.route("/admin/api/feedback/list", methods=["GET"])
def admin_api_feedback_list():
    if not _admin_auth_ok(_admin_auth_from_request()):
        return ("forbidden", 403)
    items = community_store.feedback().all(include_archived=True)
    items = sorted(items, key=lambda x: x.get("created_at", 0), reverse=True)
    return jsonify({"items": items})


@app.route("/admin/api/feedback/translate-missing", methods=["POST"])
def admin_api_feedback_translate_missing():
    """Backfill: translate any feedback items that have a comment but no
    populated comment_i18n. Synchronous; returns a summary so the admin
    can confirm what happened.
    """
    if not _admin_auth_ok(_admin_auth_from_request()):
        return ("forbidden", 403)
    if not community_translator.is_available():
        return jsonify({"error": "llm unavailable"}), 503
    items = community_store.feedback().all(include_archived=True)
    updated = 0
    failed = 0
    skipped = 0
    for it in items:
        comment = (it.get("comment") or "").strip()
        if not comment:
            skipped += 1
            continue
        i18n = it.get("comment_i18n") or {}
        # consider populated if at least one non-empty translation exists
        has_any = any((i18n.get(lg) or "").strip() for lg in community_translator.LANGS)
        if has_any:
            skipped += 1
            continue
        try:
            result = community_translator.translate({"comment": comment})
        except Exception as e:
            print(f"[feedback-backfill] translate raised for {it.get('id')}: {e!s}")
            result = None
        if not result:
            failed += 1
            continue
        patch = {
            "comment_lang": result.get("source_lang") or "",
            "comment_i18n": {
                lg: (result.get(lg) or {}).get("comment", "")
                for lg in community_translator.LANGS
            },
        }
        try:
            community_store.feedback().update(it.get("id"), patch)
            updated += 1
        except Exception as e:
            print(f"[feedback-backfill] persist failed for {it.get('id')}: {e!s}")
            failed += 1
    return jsonify({"updated": updated, "failed": failed, "skipped": skipped,
                    "total": len(items)})


@app.route("/admin/api/feedback/<item_id>", methods=["POST"])
def admin_api_feedback_update(item_id):
    if not _admin_auth_ok(_admin_auth_from_request()):
        return ("forbidden", 403)
    body = request.get_json(silent=True) or {}
    action = body.get("action")
    store = community_store.feedback()
    if action == "archive":
        ok = store.archive(item_id)
    elif action == "unarchive":
        ok = store.unarchive(item_id)
    elif action == "delete":
        ok = store.delete(item_id)
    else:
        return jsonify({"error": "invalid action"}), 400
    return ("ok", 200) if ok else ("not found", 404)


@app.route("/admin/api/feedback/<item_id>/reply", methods=["POST"])
def admin_api_feedback_reply(item_id):
    """Create or update an admin reply to a feedback item.
    Translates the reply into all 7 UI languages via LLM.
    First call creates the reply, subsequent calls edit it (preserving
    the original `admin_reply_at` timestamp and tracking edits with
    `admin_reply_edited_at`).
    """
    if not _admin_auth_ok(_admin_auth_from_request()):
        return ("forbidden", 403)
    body = request.get_json(silent=True) or {}
    reply_text = (body.get("reply") or "").strip()
    if not reply_text:
        return jsonify({"error": "reply text required"}), 400
    if len(reply_text) > 2000:
        return jsonify({"error": "reply text exceeds 2000 characters"}), 400
    store = community_store.feedback()
    existing = store.get(item_id)
    if not existing:
        return jsonify({"error": "feedback item not found"}), 404
    if not community_translator.is_available():
        return jsonify({"error": "llm unavailable"}), 503
    try:
        result = community_translator.translate({"reply": reply_text})
    except Exception as e:
        print(f"[feedback-reply] translate raised for {item_id}: {e!s}")
        result = None
    if not result:
        return jsonify({"error": "translation failed, please retry"}), 500
    now = int(time.time())
    prior_at = int(existing.get("admin_reply_at", 0) or 0)
    patch = {
        "admin_reply_text": reply_text,
        "admin_reply_lang": "it",
        "admin_reply_i18n": {
            lg: (result.get(lg) or {}).get("reply", "")
            for lg in community_translator.LANGS
        },
        "admin_reply_at": prior_at if prior_at > 0 else now,
    }
    if prior_at > 0:
        patch["admin_reply_edited_at"] = now
    try:
        community_store.feedback().update(item_id, patch)
    except Exception as e:
        print(f"[feedback-reply] persist failed for {item_id}: {e!s}")
        return jsonify({"error": "persist failed"}), 500
    return jsonify({"ok": True, "at": patch["admin_reply_at"], "edited": prior_at > 0})


@app.route("/admin/api/feedback/<item_id>/reply", methods=["DELETE"])
def admin_api_feedback_reply_delete(item_id):
    """Remove an admin reply from a feedback item."""
    if not _admin_auth_ok(_admin_auth_from_request()):
        return ("forbidden", 403)
    store = community_store.feedback()
    existing = store.get(item_id)
    if not existing:
        return jsonify({"error": "feedback item not found"}), 404
    patch = {
        "admin_reply_text": "",
        "admin_reply_lang": "",
        "admin_reply_i18n": {},
        "admin_reply_at": 0,
        "admin_reply_edited_at": 0,
    }
    try:
        community_store.feedback().update(item_id, patch)
    except Exception as e:
        print(f"[feedback-reply] delete failed for {item_id}: {e!s}")
        return jsonify({"error": "persist failed"}), 500
    return jsonify({"ok": True})


@app.route("/admin/community", methods=["GET"])
def admin_community_page():
    if not ADMIN_TOKEN:
        return ("Admin community UI disabled.", 404,
                {"Content-Type": "text/plain; charset=utf-8"})
    token = _admin_auth_from_request()
    if not _admin_auth_ok(token):
        return (_render_admin_gate("Community Admin", "/admin/community"),
                200, {"Content-Type": "text/html; charset=utf-8"})
    html = r"""<!DOCTYPE html>
<html lang="it"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<meta name="robots" content="noindex,nofollow">
<title>Admin · Community</title>
<style>
:root{--bg:#0f172a;--panel:#1e293b;--ink:#e2e8f0;--muted:#94a3b8;--accent:#8b5cf6;--ok:#10b981;--err:#ef4444;--warn:#f59e0b;}
*{box-sizing:border-box}
body{margin:0;font-family:system-ui,sans-serif;background:var(--bg);color:var(--ink);padding:20px;max-width:1200px;margin:0 auto}
h1{margin:0 0 20px;font-size:1.5rem}
.tabs{display:flex;gap:6px;margin-bottom:18px;border-bottom:1px solid #334155}
.tab{padding:10px 18px;background:transparent;color:var(--muted);border:none;cursor:pointer;font-size:.95rem;border-bottom:2px solid transparent}
.tab.active{color:var(--accent);border-bottom-color:var(--accent)}
.panel{background:var(--panel);border-radius:10px;padding:20px;margin-bottom:20px}
.panel h2{margin:0 0 14px;font-size:1.05rem;color:var(--accent)}
label{display:block;font-size:.85rem;color:var(--muted);margin:8px 0 4px}
input,select,textarea{width:100%;padding:9px 12px;background:#0f172a;border:1px solid #334155;color:var(--ink);border-radius:6px;font:inherit;font-size:.95rem}
input:focus,select:focus,textarea:focus{outline:none;border-color:var(--accent)}
button{padding:9px 16px;background:var(--accent);color:#fff;border:none;border-radius:6px;cursor:pointer;font-weight:600;font-size:.9rem}
button:hover{filter:brightness(1.1)}
button.sm{padding:5px 10px;font-size:.75rem;font-weight:500}
button.secondary{background:#334155}
button.danger{background:var(--err)}
.row3{display:grid;grid-template-columns:1fr 1fr 1fr;gap:12px}
@media(max-width:700px){.row3{grid-template-columns:1fr}}
.msg{padding:8px 12px;border-radius:6px;margin:10px 0;font-size:.85rem}
.msg.ok{background:rgba(16,185,129,.15);color:var(--ok);border:1px solid var(--ok)}
.msg.err{background:rgba(239,68,68,.15);color:var(--err);border:1px solid var(--err)}
table{width:100%;border-collapse:collapse;font-size:.82rem}
th{text-align:left;padding:8px;border-bottom:2px solid #334155;color:var(--muted);font-weight:500}
td{padding:8px;border-bottom:1px solid #1e293b;vertical-align:top}
tr.archived{opacity:.45}
.tag-feature{background:#8b5cf6;color:#fff;padding:2px 7px;border-radius:999px;font-size:.7rem;font-weight:700}
.tag-fix{background:#10b981;color:#fff;padding:2px 7px;border-radius:999px;font-size:.7rem;font-weight:700}
.tag-info{background:#64748b;color:#fff;padding:2px 7px;border-radius:999px;font-size:.7rem;font-weight:700}
.stars{color:#e6a92a;letter-spacing:1px}
.kpis{display:flex;gap:18px;flex-wrap:wrap;margin-bottom:18px}
.kpi{flex:1;min-width:140px;background:var(--panel);border-radius:8px;padding:12px}
.kpi-label{font-size:.75rem;color:var(--muted)}
.kpi-val{font-size:1.4rem;font-weight:700;color:var(--ink)}
.section{display:none}
.section.active{display:block}
.toolbar{display:flex;gap:10px;align-items:center;margin:10px 0}
.toolbar label{margin:0;display:flex;align-items:center;gap:6px;font-size:.85rem}
.toolbar input[type=checkbox]{width:auto}
.banner-pill{background:var(--warn);color:#000;padding:1px 6px;border-radius:999px;font-size:.7rem;font-weight:700}
.reply-modal-overlay{position:fixed;inset:0;background:rgba(0,0,0,.6);display:flex;align-items:center;justify-content:center;z-index:999}
.reply-modal-overlay[hidden]{display:none}
.reply-modal{background:var(--panel);border-radius:12px;padding:24px;max-width:500px;width:90%;max-height:90vh;overflow-y:auto}
.reply-modal h3{margin:0 0 16px;font-size:1.1rem;color:var(--accent)}
.reply-modal textarea{width:100%;min-height:120px;resize:vertical;font:inherit}
.reply-modal .chars{margin-top:6px;font-size:.75rem;color:var(--muted);text-align:right}
.reply-modal .err{color:var(--err);margin:10px 0;font-size:.85rem}
.reply-modal-btns{display:flex;gap:10px;margin-top:14px;justify-content:flex-end}
</style>
</head>
<body>
<h1>Admin · Community</h1>
<div class="tabs">
  <button class="tab active" data-tab="feedback">Feedback</button>
  <button class="tab" data-tab="news">News</button>
</div>

<section class="section active" id="sectFeedback">
  <div class="kpis" id="fbKpis"></div>
  <div class="toolbar">
    <label><input type="checkbox" id="fbShowArch"> Mostra archiviati</label>
    <button class="secondary sm" onclick="loadFb()">Aggiorna</button>
  </div>
  <table>
    <thead><tr><th>Data</th><th>★</th><th>Nome</th><th>Commento</th><th>IP</th><th></th></tr></thead>
    <tbody id="fbBody"></tbody>
  </table>
</section>

<section class="section" id="sectNews">
  <div class="panel">
    <h2>Nuova news</h2>
    <div class="row3">
      <div>
        <label>Tag</label>
        <select id="nTag"><option value="feature">feature</option><option value="fix">fix</option><option value="info">info</option></select>
      </div>
      <div>
        <label>Lingua</label>
        <select id="nLang"><option>it</option><option>en</option><option>fr</option><option>es</option><option>de</option><option>zh</option><option>hi</option></select>
      </div>
      <div>
        <label>&nbsp;</label>
        <label style="display:flex;align-items:center;gap:6px;color:var(--ink)"><input type="checkbox" id="nBanner"> Mostra come banner</label>
      </div>
    </div>
    <label>Titolo</label>
    <input id="nTitle" maxlength="200">
    <label>Testo</label>
    <textarea id="nBody" rows="4" maxlength="2000"></textarea>
    <div style="margin-top:10px"><button id="nPublish" onclick="createNews()">Pubblica</button></div>
    <div id="nMsg"></div>
  </div>
  <div class="toolbar">
    <label><input type="checkbox" id="nShowArch"> Mostra archiviate</label>
    <button class="secondary sm" onclick="loadNews()">Aggiorna</button>
  </div>
  <table>
    <thead><tr><th>Data</th><th>Lang</th><th>Tag</th><th>Titolo</th><th>Banner</th><th></th></tr></thead>
    <tbody id="nBody2"></tbody>
  </table>
</section>

<script>
const TOKEN=localStorage.getItem('abm_admin_token')||new URLSearchParams(location.search).get('token')||'';
if(!TOKEN){alert('Admin token mancante');}else{localStorage.setItem('abm_admin_token',TOKEN);}
const HDR={'X-Admin-Token':TOKEN,'Content-Type':'application/json'};
function esc(s){return (s||'').replace(/[<>&"]/g,c=>({'<':'&lt;','>':'&gt;','&':'&amp;','"':'&quot;'}[c]));}
function fmtDate(ts){return new Date(ts*1000).toLocaleString();}

document.querySelectorAll('.tab').forEach(t=>t.addEventListener('click',()=>{
  document.querySelectorAll('.tab').forEach(x=>x.classList.remove('active'));
  document.querySelectorAll('.section').forEach(x=>x.classList.remove('active'));
  t.classList.add('active');
  document.getElementById('sect'+t.dataset.tab[0].toUpperCase()+t.dataset.tab.slice(1)).classList.add('active');
  if(t.dataset.tab==='feedback') loadFb();
  else loadNews();
}));

async function loadFb(){
  const r=await fetch('/admin/api/feedback/list',{headers:HDR});
  if(!r.ok){document.getElementById('fbBody').innerHTML='<tr><td colspan=6>Errore</td></tr>';return;}
  const d=await r.json();
  const showArch=document.getElementById('fbShowArch').checked;
  let items=d.items||[];
  window._fbItems = items;
  if(!showArch) items=items.filter(it=>!it.archived);
  // KPI
  const all=d.items||[];
  const tot=all.length;
  const avg=tot?(all.reduce((a,it)=>a+(it.rating||0),0)/tot).toFixed(2):'—';
  const hist=[0,0,0,0,0]; all.forEach(it=>{const r=it.rating||0;if(r>=1&&r<=5) hist[r-1]++;});
  document.getElementById('fbKpis').innerHTML=`
    <div class="kpi"><div class="kpi-label">Totale</div><div class="kpi-val">${tot}</div></div>
    <div class="kpi"><div class="kpi-label">Media</div><div class="kpi-val">${avg}</div></div>
    <div class="kpi"><div class="kpi-label">5★</div><div class="kpi-val">${hist[4]}</div></div>
    <div class="kpi"><div class="kpi-label">4★</div><div class="kpi-val">${hist[3]}</div></div>
    <div class="kpi"><div class="kpi-label">≤3★</div><div class="kpi-val">${hist[0]+hist[1]+hist[2]}</div></div>`;
  const tb=document.getElementById('fbBody');
  tb.innerHTML='';
  for(const it of items){
    const tr=document.createElement('tr');
    if(it.archived) tr.className='archived';
    const stars='★'.repeat(it.rating||0)+'☆'.repeat(5-(it.rating||0));
    tr.innerHTML=`<td>${fmtDate(it.created_at)}</td>
      <td><span class="stars">${stars}</span></td>
      <td>${esc(it.name||'')}</td>
      <td>
        <div class="fb-it">${esc(((it.comment_i18n||{}).it)||it.comment||'')}</div>
        <div class="fb-orig" style="display:none;font-size:.85rem;color:var(--muted);margin-top:4px;">
          <span style="font-size:.7rem;text-transform:uppercase;border:1px solid var(--muted);padding:1px 4px;border-radius:4px;">${esc(it.comment_lang||'orig')}</span>
          ${esc(it.comment||'')}
        </div>
        <button class="sm secondary fb-toggle" style="margin-top:6px;">Originale</button>
      </td>
      <td><code style="font-size:.75rem">${esc(it.ip_hash||'')}</code></td>
      <td>
        ${it.admin_reply_at > 0
          ? `<button class="sm secondary" data-id="${it.id}" onclick="openReplyModal(this)" title="${fmtDate(it.admin_reply_edited_at||it.admin_reply_at)}">Modifica risposta</button>`
          : `<button class="sm" style="background:var(--accent)" data-id="${it.id}" onclick="openReplyModal(this)">Rispondi</button>`
        }
        <button class="sm secondary" data-id="${it.id}" data-act="${it.archived?'unarchive':'archive'}">${it.archived?'Riattiva':'Archivia'}</button>
        <button class="sm danger" data-id="${it.id}" data-act="delete">Elimina</button>
      </td>`;
    tb.appendChild(tr);
  }
  tb.querySelectorAll('button[data-act]').forEach(b=>b.addEventListener('click',async()=>{
    if(b.dataset.act==='delete'&&!confirm('Eliminare definitivamente?')) return;
    const r=await fetch('/admin/api/feedback/'+b.dataset.id,{method:'POST',headers:HDR,body:JSON.stringify({action:b.dataset.act})});
    if(r.ok) loadFb(); else alert('Errore');
  }));
  tb.querySelectorAll('.fb-toggle').forEach(b=>b.addEventListener('click',()=>{
    const wrap=b.closest('td').querySelector('.fb-orig');
    const hidden=wrap.style.display==='none';
    wrap.style.display=hidden?'block':'none';
    b.textContent=hidden?'Nascondi originale':'Originale';
  }));
}

async function loadNews(){
  const r=await fetch('/admin/api/news/list',{headers:HDR});
  if(!r.ok){document.getElementById('nBody2').innerHTML='<tr><td colspan=6>Errore</td></tr>';return;}
  const d=await r.json();
  const showArch=document.getElementById('nShowArch').checked;
  let items=d.items||[];
  if(!showArch) items=items.filter(it=>!it.archived);
  const tb=document.getElementById('nBody2');
  tb.innerHTML='';
  for(const it of items){
    const tr=document.createElement('tr');
    if(it.archived) tr.className='archived';
    tr.innerHTML=`<td>${fmtDate(it.created_at)}</td>
      <td>${esc(it.lang||'')}</td>
      <td><span class="tag-${esc(it.tag)}">${esc(it.tag)}</span></td>
      <td>${esc(it.title||'')}</td>
      <td>${it.banner?'<span class="banner-pill">BANNER</span>':''}</td>
      <td>
        <button class="sm secondary" data-id="${it.id}" data-act="toggle_banner">Toggle banner</button>
        <button class="sm secondary" data-id="${it.id}" data-act="${it.archived?'unarchive':'archive'}">${it.archived?'Riattiva':'Archivia'}</button>
        <button class="sm danger" data-id="${it.id}" data-act="delete">Elimina</button>
      </td>`;
    tb.appendChild(tr);
  }
  tb.querySelectorAll('button[data-act]').forEach(b=>b.addEventListener('click',async()=>{
    if(b.dataset.act==='delete'&&!confirm('Eliminare definitivamente?')) return;
    const r=await fetch('/admin/api/news/'+b.dataset.id,{method:'POST',headers:HDR,body:JSON.stringify({action:b.dataset.act})});
    if(r.ok) loadNews(); else alert('Errore');
  }));
}

async function createNews(){
  const body={
    tag:document.getElementById('nTag').value,
    lang:document.getElementById('nLang').value,
    banner:document.getElementById('nBanner').checked,
    title:document.getElementById('nTitle').value,
    body:document.getElementById('nBody').value,
  };
  const msg=document.getElementById('nMsg');
  if(!body.title.trim()){msg.innerHTML='<div class="msg err">Titolo richiesto</div>';return;}
  const btn=document.getElementById('nPublish');
  const origLabel=btn?btn.textContent:'';
  if(btn){btn.disabled=true;btn.textContent='⏳ Pubblicazione…';}
  msg.innerHTML='<div class="msg">Pubblicazione in corso…</div>';
  try{
    const r=await fetch('/admin/api/news',{method:'POST',headers:HDR,body:JSON.stringify(body)});
    if(r.ok){
      msg.innerHTML='<div class="msg ok">Pubblicata</div>';
      document.getElementById('nTitle').value='';
      document.getElementById('nBody').value='';
      document.getElementById('nBanner').checked=false;
      loadNews();
    } else {
      const e=await r.json().catch(()=>({error:'errore'}));
      msg.innerHTML='<div class="msg err">'+esc(e.error||'errore')+'</div>';
    }
  } finally {
    if(btn){btn.disabled=false;btn.textContent=origLabel;}
  }
}

document.getElementById('fbShowArch').addEventListener('change',loadFb);
document.getElementById('nShowArch').addEventListener('change',loadNews);

let _replyItemId = null;
let _replyIsEdit = false;
function openReplyModal(btn){
  const id = btn.dataset.id;
  _replyItemId = id;
  const item = (window._fbItems || []).find(it => it.id === id);
  if(!item){return;}
  _replyIsEdit = (item.admin_reply_at||0) > 0;
  document.getElementById('replyOriginal').textContent =
    ((item.comment_i18n||{}).it)||item.comment||'(senza commento)';
  const txt = document.getElementById('replyText');
  txt.value = _replyIsEdit ? (item.admin_reply_text || '') : '';
  document.getElementById('replyCharCount').textContent = String(txt.value.length);
  document.getElementById('replyErr').hidden = true;
  document.getElementById('replyTitle').textContent = _replyIsEdit ? 'Modifica risposta' : 'Rispondi al commento';
  document.getElementById('replySubmit').textContent = _replyIsEdit ? 'Salva modifiche' : 'Invia risposta';
  document.getElementById('replyDelete').hidden = !_replyIsEdit;
  document.getElementById('replyModal').hidden = false;
  txt.focus();
  txt.addEventListener('input', function(){document.getElementById('replyCharCount').textContent=this.value.length;}, {once: true});
  document.getElementById('replyCancel').addEventListener('click', closeReplyModal);
}
function closeReplyModal(){
  const modal = document.getElementById('replyModal');
  if(modal){ modal.hidden = true; }
  _replyItemId = null;
  _replyIsEdit = false;
}
async function submitReply(){
  if(!_replyItemId) return;
  const text = document.getElementById('replyText').value.trim();
  if(!text){document.getElementById('replyErr').textContent='Testo richiesto';document.getElementById('replyErr').hidden=false;return;}
  if(text.length > 2000){document.getElementById('replyErr').textContent='Max 2000 caratteri';document.getElementById('replyErr').hidden=false;return;}
  const btn = document.getElementById('replySubmit');
  const origLabel = btn ? btn.textContent : '';
  btn.disabled = true;
  btn.textContent = '⏳ Traduzione in corso…';
  const errEl = document.getElementById('replyErr');
  errEl.hidden = true;
  try {
    const r = await fetch('/admin/api/feedback/'+_replyItemId+'/reply',{
      method:'POST', headers:HDR,
      body:JSON.stringify({reply:text})
    });
    const d = await r.json().catch(()=>({}));
    if(!r.ok){
      errEl.textContent = d.error || ('Errore '+r.status);
      errEl.hidden = false;
    } else {
      closeReplyModal();
      loadFb();
    }
  } catch(e){
    errEl.textContent = 'Errore di rete';
    errEl.hidden = false;
  } finally {
    btn.disabled = false;
    btn.textContent = origLabel;
  }
}
async function deleteReply(){
  if(!_replyItemId) return;
  if(!confirm('Eliminare la risposta? L\'azione non è reversibile.')) return;
  const btn = document.getElementById('replyDelete');
  const origLabel = btn ? btn.textContent : '';
  btn.disabled = true;
  btn.textContent = '⏳ Eliminazione…';
  const errEl = document.getElementById('replyErr');
  errEl.hidden = true;
  try {
    const r = await fetch('/admin/api/feedback/'+_replyItemId+'/reply',{
      method:'DELETE', headers:HDR
    });
    const d = await r.json().catch(()=>({}));
    if(!r.ok){
      errEl.textContent = d.error || ('Errore '+r.status);
      errEl.hidden = false;
    } else {
      closeReplyModal();
      loadFb();
    }
  } catch(e){
    errEl.textContent = 'Errore di rete';
    errEl.hidden = false;
  } finally {
    btn.disabled = false;
    btn.textContent = origLabel;
  }
}

loadFb();
</script>
<div class="reply-modal-overlay" id="replyModal" hidden>
  <div class="reply-modal">
    <h3 id="replyTitle">Rispondi al commento</h3>
    <div id="replyOriginal" style="font-size:.85rem;color:var(--muted);margin-bottom:12px;padding:8px;background:#0f172a;border-radius:6px;max-height:100px;overflow-y:auto"></div>
    <textarea id="replyText" maxlength="2000" placeholder="Scrivi la risposta in italiano..." rows="5"></textarea>
    <div class="chars"><span id="replyCharCount">0</span>/2000</div>
    <div class="err" id="replyErr" hidden></div>
    <div class="reply-modal-btns" style="display:flex;justify-content:space-between;gap:8px;align-items:center;">
      <button class="danger" id="replyDelete" onclick="deleteReply()" hidden>Elimina risposta</button>
      <div style="display:flex;gap:8px;margin-left:auto;">
        <button class="secondary" id="replyCancel">Annulla</button>
        <button id="replySubmit" onclick="submitReply()">Invia risposta</button>
      </div>
    </div>
  </div>
</div>
</body></html>"""
    return html, 200, {"Content-Type": "text/html; charset=utf-8"}


@app.route("/api/admin/google_tts_status")
def api_admin_google_tts_status():
    """Endpoint admin: stato dettagliato Google TTS (consumo locale + cloud).
    Forza una riconciliazione on-demand se ?reconcile=1.
    Richiede admin token: il path /admin/ implica scope ristretto e l'endpoint
    espone metriche operative (caratteri consumati / limiti) che possono
    rivelare la capacita' residua del servizio a competitor o attaccanti.
    """
    if not _admin_auth_ok(_admin_auth_from_request()):
        return jsonify({"error": "Unauthorized"}), 403
    if google_tts is None:
        return jsonify({"error": "google_tts module not loaded"}), 503
    if not google_tts.is_available():
        return jsonify({"available": False, "reason": "credentials missing or SDK not installed"}), 200

    used, remaining, limit = google_tts.get_usage()
    response = {
        "available": True,
        "local": {
            "chars_used": used,
            "chars_remaining": remaining,
            "chars_limit": limit,
            "percent_used": round(100.0 * used / limit, 2) if limit else 0,
        },
        "monitoring": google_tts.get_reconcile_status(),
    }

    # Riconciliazione on-demand
    if request.args.get("reconcile") == "1":
        result = google_tts.reconcile_with_cloud_monitoring()
        response["reconcile_result"] = result

    # Diagnostica metriche Cloud Monitoring (per capire quale filtro usare)
    if request.args.get("diagnose") == "1":
        if hasattr(google_tts, "diagnose_monitoring"):
            response["diagnose"] = google_tts.diagnose_monitoring()
        else:
            response["diagnose"] = {"error": "diagnose_monitoring not available"}

    return jsonify(response)


def _file_hash(path):
    """Return MD5 hex digest of a file, streaming in chunks."""
    import hashlib
    h = hashlib.md5()
    with open(path, "rb") as f:
        while True:
            chunk = f.read(8192)
            if not chunk:
                break
            h.update(chunk)
    return h.hexdigest()


@app.route("/api/analyze", methods=["POST"])
def api_analyze():
    # Rate-limit IP-based: previene spam upload / DoS.
    _allowed, _retry = _ip_rl_check(
        "analyze", _client_ip(), _ANALYZE_RL_PER_MIN, _ANALYZE_RL_PER_HOUR
    )
    if not _allowed:
        return jsonify({"error": "rate_limit", "retry_after": _retry}), 429
    if "epub" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    file = request.files["epub"]
    fname_lower = file.filename.lower()
    is_txt = fname_lower.endswith(".txt")
    is_epub = fname_lower.endswith(".epub")
    is_pdf = fname_lower.endswith(".pdf")
    is_abm = fname_lower.endswith(".abm")
    if not is_epub and not is_txt and not is_pdf and not is_abm:
        return jsonify({"error": "File must be .epub, .pdf, .txt or .abm"}), 400
    if is_pdf and parse_pdf is None:
        return jsonify({"error": "PDF support not available. Install pymupdf: pip install pymupdf"}), 400

    # Sanitize filename for disk storage (Security: prevent Path Traversal)
    from werkzeug.utils import secure_filename
    safe_name = secure_filename(file.filename)
    if not safe_name:
        # Fallback if secure_filename results in empty string (e.g. only non-ascii chars)
        safe_name = str(uuid.uuid4())[:8] + "_" + fname_lower

    job_id = _new_job_id()
    work_dir = UPLOAD_DIR / job_id
    work_dir.mkdir(exist_ok=True)
    file_path = work_dir / safe_name
    file.save(str(file_path))

    # Magic-bytes validation: blocca file con estensione mentita (es. .exe rinominato .epub).
    # I parser a valle gia` rifiuterebbero, ma fail-fast evita storage waste + log inutili.
    try:
        with open(str(file_path), "rb") as _f:
            _head = _f.read(8)
    except Exception:
        _head = b""
    if (is_epub or is_abm) and not _head.startswith(b"PK"):
        try: file_path.unlink()
        except Exception: pass
        try: work_dir.rmdir()
        except Exception: pass
        return jsonify({"error": "Invalid file: not a valid ZIP-based archive"}), 400
    if is_pdf and not _head.startswith(b"%PDF-"):
        try: file_path.unlink()
        except Exception: pass
        try: work_dir.rmdir()
        except Exception: pass
        return jsonify({"error": "Invalid file: not a valid PDF"}), 400
    if is_txt:
        # TXT non ha magic; rifiuta solo se contiene byte chiaramente binari nel head
        # (NUL byte). Accetta UTF-8 BOM (\xef\xbb\xbf), UTF-16 BOM, ecc.
        if b"\x00" in _head and not (_head.startswith(b"\xff\xfe") or _head.startswith(b"\xfe\xff")):
            try: file_path.unlink()
            except Exception: pass
            try: work_dir.rmdir()
            except Exception: pass
            return jsonify({"error": "Invalid file: not a valid text file"}), 400

    client_id = _get_client_id()
    file_hash = _file_hash(str(file_path))

    # Duplicate upload detection: if this client already has an active job for the same file,
    # block if running; reuse if only analyzed/optimized.
    existing_jid = None
    existing_job = None
    with _jobs_lock:
        for jid, job in jobs.items():
            if job.get("client_id") == client_id and job.get("file_hash") == file_hash:
                existing_jid = jid
                existing_job = job
                break

    if existing_job:
        status = existing_job.get("status", "")
        if status in ("optimizing", "generating"):
            return jsonify({
                "existing_job_id": existing_jid,
                "status": status,
                "is_running": True,
                "progress_current": existing_job.get("progress_current", 0),
                "progress_total": existing_job.get("progress_total", 0),
                "opt_progress_current": existing_job.get("opt_progress_current", 0),
                "opt_progress_total": existing_job.get("opt_progress_total", 0),
            })
        if status in ("analyzed", "optimized"):
            # Reuse existing analyzed/optimized job
            info = existing_job["info"]
            _lang_re = (getattr(info, "language", None) or "it")[:2].lower()
            chapters = []
            _total_secs_re = 0.0
            for ch in info.chapters:
                _secs = _estimate_chapter_seconds(ch, _lang_re)
                _total_secs_re += _secs
                chapters.append({
                    "index": ch.index, "title": ch.title,
                    "words": ch.word_count, "chars": ch.char_count,
                    "estimated_minutes": round(_secs / 60.0, 1),
                })
            return jsonify({
                "job_id": existing_jid, "title": info.title, "author": info.author,
                "language": info.language,
                "file_type": "abm" if is_abm else ("txt" if is_txt else ("pdf" if is_pdf else "epub")),
                "has_cover": bool(existing_job.get("cover_thumb")),
                "total_chapters": len(info.chapters), "total_words": info.total_words,
                "total_chars": info.total_chars,
                "estimated_minutes": round(_total_secs_re / 60.0, 1),
                "chapters": chapters,
                "preview_text": existing_job.get("preview_text", ""),
                "llm_available": _llm_available(),
                "ai_optimized": existing_job.get("ai_optimized", False),
                "optimized_chapters": existing_job.get("optimized_chapters", []),
            })

    abm_cover_info = None  # cover data extracted from .abm file
    try:
        if is_abm:
            info, abm_cover_info = parse_abm(str(file_path))
        elif is_txt:
            info = parse_txt(str(file_path))
        elif is_pdf:
            info = parse_pdf(str(file_path))
        else:
            info = parse_epub(str(file_path))
    except Exception as e:
        label = "ABM" if is_abm else ("TXT" if is_txt else ("PDF" if is_pdf else "EPUB"))
        return jsonify({"error": f"{label} parse error: {e}"}), 400

    if not info.chapters:
        return jsonify({"error": "No content found."}), 400

    # NOTE: the ABM_MAX_TEXT_CHARS cap is no longer enforced here. We show the
    # book regardless of its total size so the user can browse chapters and
    # narrow the selection. The actual cap is applied at /api/generate and
    # /api/optimize on the *selected* chapters, where it matters for output
    # size and LLM cost.

    with _jobs_lock:
        jobs[job_id] = {"status": "analyzed", "epub_path": str(file_path), "info": info,
                         "last_poll": time.time(), "original_filename": file.filename,
                         "client_id": _get_client_id(), "client_ip": _get_client_ip(),
                         "browser_lang": _get_browser_lang(),
                         "optimized_chapters": [], "file_hash": file_hash}

    # Extract cover thumbnail for preview (EPUB or ABM; PDF/TXT have no embedded cover)
    has_cover = False
    if is_abm and abm_cover_info:
        # Cover from .abm archive
        cover_data = abm_cover_info["data"]
        cover_filename = abm_cover_info["filename"]
        is_png = cover_filename.lower().endswith(".png")
        ext = ".png" if is_png else ".jpg"
        mime = "image/png" if is_png else "image/jpeg"
        cover_out = str(work_dir / ("cover_thumb" + ext))
        with open(cover_out, "wb") as cf:
            cf.write(cover_data)
        has_cover = True
        jobs[job_id]["cover_thumb"] = cover_out
        jobs[job_id]["cover_mime"] = mime
    elif is_epub:
        cover_path, cover_mime = _extract_cover_for_preview(str(file_path), str(work_dir))
        if cover_path and os.path.exists(cover_path):
            has_cover = True
            jobs[job_id]["cover_thumb"] = cover_path
            jobs[job_id]["cover_mime"] = cover_mime

    _log_activity(job_id, file.filename, "ANALYZE",
                  jobs[job_id]["client_id"], jobs[job_id]["client_ip"],
                  browser_lang=jobs[job_id].get("browser_lang", ""))

    _lang_new = (getattr(info, "language", None) or "it")[:2].lower()
    chapters = []
    _total_secs_new = 0.0
    for ch in info.chapters:
        _secs = _estimate_chapter_seconds(ch, _lang_new)
        _total_secs_new += _secs
        chapters.append({
            "index": ch.index, "title": ch.title,
            "words": ch.word_count, "chars": ch.char_count,
            "estimated_minutes": round(_secs / 60.0, 1),
        })
    # Override total estimated minutes for response consistency with per-chapter values.
    _total_minutes_new = round(_total_secs_new / 60.0, 1)

    #  -  -  Preview text  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  - 
    # EPUB: salta il front matter e usa un capitolo interno con contenuto narrativo reale.
    # TXT:  usa il primo contenuto disponibile.
    # Lunghezza target: 200-300 caratteri, troncata a fine frase.
    def _pick_preview_text(chapters_list, is_txt_file):
        from epub_to_tts import is_content_chapter as _icc
        if not chapters_list:
            return ""
        if is_txt_file:
            for ch in chapters_list:
                raw = (ch.text or "").strip()
                if len(raw) >= 150:
                    return raw
            return ""
        # EPUB: filtra front matter con la stessa euristica usata in epub_to_tts
        valid = [ch for ch in chapters_list
                 if _icc(ch.text or "", ch.title or "") and (ch.word_count or 0) >= 80]
        if not valid:
            for ch in chapters_list:
                raw = (ch.text or "").strip()
                if len(raw) >= 150:
                    return raw
            return ""
        # Secondo capitolo valido (più probabile contenuto narrativo, non introduzione)
        target = valid[1] if len(valid) > 1 else valid[0]
        return (target.text or "").strip()

    def _trim_preview(text, min_chars=400, max_chars=600):
        """Tronca tra min e max caratteri a fine frase, oppure all'ultimo spazio."""
        import re as _re
        text = _re.sub(r'\s+', ' ', text).strip()
        if len(text) <= max_chars:
            return text
        window = text[min_chars:max_chars]
        m = _re.search(r'[.!?]["""»\)\s]', window)
        cut = (min_chars + m.start() + 1) if m else text.rfind(' ', min_chars, max_chars)
        if cut <= 0:
            cut = max_chars
        return text[:cut].rstrip()

    # Solo TXT puro usa la logica "primo capitolo con >=150 char": PDF e ABM
    # hanno capitoli strutturati e beneficiano del filtro front-matter + scelta
    # del secondo capitolo valido (narrativa, non introduzione).
    raw_preview = _pick_preview_text(info.chapters, is_txt)
    preview_text = _trim_preview(raw_preview) if raw_preview else ""
    # Store for /api/preview_audio
    jobs[job_id]["preview_text"] = preview_text
    # ----------------------------------------------------------------------

    # Detect if .abm was already AI-optimized
    abm_ai_optimized = False
    abm_manifest = None
    if is_abm:
        try:
            import zipfile
            with zipfile.ZipFile(str(file_path), "r") as zf:
                m = json.loads(zf.read("manifest.json").decode("utf-8"))
                abm_ai_optimized = m.get("ai_optimized", False)
                abm_manifest = m
        except Exception:
            pass
    if abm_ai_optimized:
        jobs[job_id]["ai_optimized"] = True
        if abm_manifest:
            jobs[job_id]["optimized_chapters"] = [c["index"] for c in abm_manifest.get("chapters", [])]

    return jsonify({
        "job_id": job_id, "title": info.title, "author": info.author,
        "language": info.language,
        "file_type": "abm" if is_abm else ("txt" if is_txt else ("pdf" if is_pdf else "epub")),
        "has_cover": has_cover,
        "total_chapters": len(info.chapters), "total_words": info.total_words,
        "total_chars": info.total_chars,
        "estimated_minutes": _total_minutes_new,
        "chapters": chapters,
        "preview_text": preview_text,
        "llm_available": _llm_available(),
        "ai_optimized": abm_ai_optimized,
        "optimized_chapters": jobs[job_id].get("optimized_chapters", []),
        "max_text_chars": MAX_TEXT_CHARS,
        "max_gemini_text_chars": MAX_GEMINI_TEXT_CHARS,
    })


@app.route("/api/preview_audio/<job_id>")
def api_preview_audio(job_id):
    """Serve l'MP3 di anteprima come endpoint GET.
    Il browser può usare l'URL direttamente come audio.src  -  nessun problema di autoplay policy.
    Il timeout è gestito da concurrent.futures (funziona sempre, a differenza di asyncio.wait_for).
    """
    if not job_id:
        return jsonify({"error": "Job non trovato"}), 404
    _job, _err, _sc = _check_job_owner(job_id)
    if _err is not None:
        return _err, _sc

    # Rate-limit IP-based: anteprime sono costose (genera TTS sample),
    # impedisce abuso e burst di generazione voci diverse.
    _allowed, _retry = _ip_rl_check(
        "preview", _client_ip(), _PREVIEW_RL_PER_MIN, _PREVIEW_RL_PER_HOUR
    )
    if not _allowed:
        return jsonify({"error": "rate_limit", "retry_after": _retry}), 429

    voice = request.args.get("voice", "it-IT-IsabellaNeural")
    rate  = request.args.get("rate",  "+0%")
    style = (request.args.get("style") or "").strip()[:200]

    # Se il client passa selected_chapters, l'anteprima deve essere un estratto
    # dei capitoli selezionati (coerente con il pannello "Voci PREMIUM"). Altrimenti
    # usa il preview_text fallback memorizzato all'upload.
    try:
        sel_raw = request.args.getlist("selected_chapters")
        sel_idxs = [int(x) for x in sel_raw if str(x).strip()]
    except (TypeError, ValueError):
        sel_idxs = []
    preview_text = ""
    if sel_idxs:
        info_pv = jobs[job_id].get("info")
        all_chs_pv = list(getattr(info_pv, "chapters", []) or []) if info_pv else []
        by_idx_pv = {ch.index: ch for ch in all_chs_pv}
        sel_chs = [by_idx_pv[i] for i in sel_idxs if i in by_idx_pv]
        if sel_chs:
            try:
                from epub_to_tts import is_content_chapter as _icc_pv
                valid = [c for c in sel_chs
                         if _icc_pv(c.text or "", c.title or "")
                         and (c.word_count or 0) >= 80]
            except Exception:
                valid = [c for c in sel_chs if (c.word_count or 0) >= 80]
            if not valid:
                valid = [c for c in sel_chs if ((c.text or "").strip())]
            if valid:
                target = valid[1] if len(valid) > 1 else valid[0]
                raw = (target.text or "").strip()
                import re as _re_pv
                raw = _re_pv.sub(r"\s+", " ", raw).strip()
                # Tronca tra 400 e 600 char a fine frase (riallinea a _trim_preview).
                if len(raw) > 600:
                    _win = raw[400:600]
                    _m = _re_pv.search(r'[.!?]["”“»\)\s]', _win)
                    _cut = (400 + _m.start() + 1) if _m else raw.rfind(" ", 400, 600)
                    if _cut <= 0:
                        _cut = 600
                    preview_text = raw[:_cut].rstrip()
                else:
                    preview_text = raw
    if not preview_text:
        preview_text = jobs[job_id].get("preview_text", "")
    if not preview_text:
        return jsonify({"error": "Nessun testo di anteprima disponibile"}), 400

    # Per Gemini riduciamo il testo a ~20-30 sec di audio (250-400 char) per
    # contenere il costo per-token (input + output sono fatturati).
    if voice.startswith("gemini:"):
        import re as _re
        _t = _re.sub(r'\s+', ' ', preview_text).strip()
        if len(_t) > 400:
            _window = _t[250:400]
            _m = _re.search(r'[.!?]["”“»\)\s]', _window)
            _cut = (250 + _m.start() + 1) if _m else _t.rfind(' ', 250, 400)
            if _cut <= 0:
                _cut = 400
            preview_text = _t[:_cut].rstrip()
        else:
            preview_text = _t

    work_dir = UPLOAD_DIR / job_id
    work_dir.mkdir(exist_ok=True)
    # Cache per (voice, rate, style, selezione): il nome del file è derivato da un
    # hash della chiave, così tornare a una combinazione già generata serve il
    # file cached invece di rigenerare (e per Gemini non consuma il preview cap).
    # La selezione capitoli entra nella chiave perché il testo varia con essa.
    sel_key = ",".join(str(i) for i in sorted(sel_idxs)) if sel_idxs else ""
    cache_key = f"{voice}|{rate}|{style}|{sel_key}"
    key_hash = hashlib.sha1(cache_key.encode("utf-8")).hexdigest()[:16]
    preview_path = work_dir / f"preview_{key_hash}.mp3"

    if preview_path.exists() and preview_path.stat().st_size > 0:
        return send_file(str(preview_path), mimetype="audio/mpeg",
                         as_attachment=False, download_name="preview.mp3",
                         conditional=True)

    # Genera l'MP3 in un thread separato con timeout reale di 30 secondi.
    # concurrent.futures.Future.result(timeout=) interrompe l'attesa indipendentemente
    # da asyncio  -  risolve il caso in cui edge-tts si blocca sulla connessione TCP.
    use_google_preview = google_tts is not None and google_tts.is_google_voice(voice)
    use_gemini_preview = gemini_tts is not None and voice.startswith("gemini:")
    client_id = "anon"

    # Preview cap per Gemini (rolling 24h per cookie)
    if use_gemini_preview:
        if not gemini_tts.is_available():
            return jsonify({"error": "gemini_tts_not_configured"}), 503
        client_id = _get_client_id() or "anon"
        used, remaining, reset_ts = gemini_tts.check_preview_cap(client_id)
        if remaining <= 0:
            reset_in = max(0, int(reset_ts - time.time()))
            return jsonify({
                "error": "preview_cap_exceeded",
                "used": used,
                "cap": gemini_tts.PREVIEW_CAP_PER_DAY,
                "reset_in_seconds": reset_in,
            }), 429
        # Preflight RPD: se il modello non ha quota per anche solo 1 chunk,
        # falliamo immediatamente con 503 invece di lasciare la call al
        # synthesize() che andrebbe in errore dopo aver consumato tempo.
        # Senza questo check, su flash25 con RPD esaurito l'utente vedeva
        # solo lo spinner per ~30s prima di un 504 generico.
        try:
            parts = voice.split(":")
            _model_key_pf = parts[1] if len(parts) >= 3 else "flash25"
            _pf = gemini_tts.preflight_can_run(_model_key_pf, 1)
            if not _pf.get("ok"):
                return jsonify({
                    "error": ("Il modello voci PREMIUM ha esaurito la quota "
                              "giornaliera. Riprova piu` tardi o seleziona "
                              "un modello differente."),
                    "code": "quota_exhausted",
                    "model_key": _model_key_pf,
                    "retry_after_sec": int(_pf.get("retry_after_sec") or 0),
                    "available": _pf.get("available"),
                }), 503
        except Exception as _pf_err:
            # Preflight non-fatal: log e prosegui. Meglio rischiare un 504
            # downstream che bloccare un preview legittimo.
            print(f"[preview] preflight check error (non-fatal): {_pf_err}")

    def _generate():
        if use_gemini_preview:
            # Native output is PCM — convert to MP3 inline for browser playback.
            pcm_tmp = str(preview_path) + ".pcm"
            try:
                # max_attempts=1: il path preview ha timeout client 30s. Se
                # Gemini restituisce EMPTY-RESPONSE con finish_reason=OTHER
                # (modello fermato per ragioni non specificate, tipico su
                # combo voce/rate/lingua poco stabili come flash25), i 3
                # retry default + backoff saturano il timeout → 504 lato
                # browser. Falliamo veloce: il caller (preview_audio)
                # converte l'errore in 502 con messaggio utile e l'utente
                # puo` ritentare manualmente o cambiare voce/modello.
                result = gemini_tts.synthesize(preview_text, voice, output_path=pcm_tmp,
                                                style_instruction=style or None,
                                                rate=rate,
                                                max_attempts=1)
                pcm_to_mp3([pcm_tmp], str(preview_path))
                # Costo Google REALE della preview (token reali x rate per MTok).
                _preview_cost_eur = 0.0
                try:
                    _bd = gemini_tts.google_cost_breakdown(
                        result.get("input_tokens", 0),
                        result.get("output_tokens", 0),
                        result.get("model_key", "flash25"),
                    )
                    _preview_cost_eur = float(_bd.get("total_eur", 0.0) or 0.0)
                except Exception as e:
                    print(f"[preview] google_cost_breakdown failed (non-fatal): {e}")
                try:
                    gemini_tts.record_usage(
                        result.get("model_key", "flash25"),
                        len(preview_text),
                        result.get("input_tokens", 0),
                        result.get("output_tokens", 0),
                        _preview_cost_eur,
                        0.0,
                    )
                except Exception as e:
                    print(f"[preview] gemini_tts.record_usage failed (non-fatal): {e}")
                # Sample rate empirico: lingua TTS scelta (NON metadata libro).
                # Priorita`: query `lang` -> job.opt_lang/gen_lang -> info.language.
                # Senza questo, una preview con voce italiana su EPUB arabo
                # registrava sample "ar" inquinando l'empirical rate per "it".
                try:
                    _job_pv = jobs[job_id]
                    _job_info = _job_pv.get("info")
                    _q_lang = (request.args.get("lang") or "").strip().split("-")[0].lower()
                    _preview_lang = (
                        _q_lang
                        or (_job_pv.get("opt_lang") or "").strip().split("-")[0].lower()
                        or (_job_pv.get("gen_lang") or "").strip().split("-")[0].lower()
                        or (getattr(_job_info, "language", None) or "it").split("-")[0].lower()
                    )[:2]
                    _norm_chars = len(gemini_tts._normalize_text(preview_text))
                    gemini_tts.record_rate_sample(
                        _norm_chars,
                        result.get("audio_seconds_real", 0.0),
                        _preview_lang,
                        result.get("model_key", "flash25"),
                        rate_pct=rate,
                    )
                except Exception as e:
                    print(f"[preview] gemini_tts.record_rate_sample failed (non-fatal): {e}")
                gemini_tts.increment_preview(client_id)
            finally:
                if os.path.exists(pcm_tmp):
                    try:
                        os.remove(pcm_tmp)
                    except OSError:
                        pass
        elif use_google_preview:
            google_tts.synthesize(preview_text, voice, rate, str(preview_path))
            # Deduce i caratteri dell'anteprima dal budget
            google_tts.deduct_chars(len(preview_text))
        else:
            import edge_tts
            loop = asyncio.new_event_loop()
            try:
                async def _run():
                    communicate = edge_tts.Communicate(
                        text=preview_text, voice=voice, rate=rate
                    )
                    await communicate.save(str(preview_path))
                loop.run_until_complete(_run())
            finally:
                loop.close()

    # Wrapper timeout model-aware: flash31 (gemini-3.1-flash-tts-preview) e`
    # strutturalmente piu` lento di flash25 (RPM cap 3/300 vs 10/750 + audio
    # gen piu` lenta lato Google). Senza maggiorazione, il wrapper a 30s
    # strozza prima del timeout HTTP Google (60s per flash31) e produce
    # 504 spuri anche su preview legittime. flash25 resta a 30s.
    _wrapper_timeout = 30
    if use_gemini_preview:
        try:
            _mk = voice.split(":")[1] if voice.startswith("gemini:") else ""
            if _mk == "flash31":
                _wrapper_timeout = int(os.environ.get(
                    "ABM_GEMINI_PREVIEW_TIMEOUT_SEC_FLASH31", "65"))
        except Exception:
            pass
    try:
        with concurrent.futures.ThreadPoolExecutor(max_workers=1) as ex:
            ex.submit(_generate).result(timeout=_wrapper_timeout)
    except concurrent.futures.TimeoutError:
        return jsonify({"error": f"Timeout: il servizio TTS non ha risposto in {_wrapper_timeout} secondi."}), 504
    except Exception as e:
        # EMPTY-RESPONSE: Gemini ha risposto senza audio (finish_reason=OTHER
        # o simili). Tipicamente combo voce/rate/lingua poco stabile su un
        # modello specifico (es. flash25). Restituiamo 502 con messaggio
        # actionable invece di 500 generico.
        if gemini_tts is not None and isinstance(e, getattr(gemini_tts, "GeminiEmptyResponse", ())):
            _fr = getattr(e, "finish_reason", None) or "unknown"
            return jsonify({
                "error": (f"Il modello TTS non ha prodotto audio (motivo: {_fr}). "
                          f"Riprova, oppure cambia voce o modello (la stessa voce su "
                          f"un modello diverso spesso funziona)."),
                "code": "empty_response",
                "finish_reason": _fr,
            }), 502
        # Quota giornaliera RPD raggiunta server-wide per il modello. Diverso
        # dal preview_cap per-client (429): qui e` un limite Gemini globale.
        # Restituiamo 503 con code dedicato per messaggio actionable.
        if gemini_tts is not None and isinstance(e, getattr(gemini_tts, "GeminiQuotaExhausted", ())):
            _ra = getattr(e, "retry_after_sec", None)
            return jsonify({
                "error": ("Il modello voci PREMIUM ha raggiunto il limite "
                          "giornaliero. Riprova piu` tardi o seleziona un "
                          "modello differente."),
                "code": "quota_exhausted",
                "retry_after_sec": _ra,
            }), 503
        # Budget EUR daily/per-job superato (raro in preview, ma possibile se
        # il budget e` molto stretto). Stesso treatment di quota_exhausted ma
        # con code distinto per logging lato server.
        if gemini_tts is not None and isinstance(e, getattr(gemini_tts, "GeminiBudgetExceeded", ())):
            return jsonify({
                "error": ("Le voci PREMIUM hanno raggiunto il budget "
                          "giornaliero. Riprova piu` tardi o seleziona "
                          "una voce Standard."),
                "code": "budget_exceeded",
            }), 503
        # HTTP timeout: il client genai non ha ricevuto risposta entro il
        # timeout configurato (ABM_GEMINI_HTTP_TIMEOUT_MS, default 25s).
        # Distinguiamo dal ThreadPoolExecutor timeout (504) perche` qui
        # abbiamo info sulla causa (API lenta/irraggiungibile).
        _ctx = getattr(e, "__context__", None)
        _is_http_timeout = False
        try:
            import httpx
            _is_http_timeout = isinstance(e, httpx.TimeoutException) or isinstance(_ctx, httpx.TimeoutException)
        except ImportError:
            pass
        if _is_http_timeout:
            return jsonify({
                "error": ("Il servizio voci PREMIUM non risponde. Riprova "
                          "tra qualche secondo o seleziona un modello/voce "
                          "differente."),
                "code": "http_timeout",
            }), 504
        return jsonify({"error": f"Errore generazione anteprima: {e}"}), 500

    if not preview_path.exists():
        return jsonify({"error": "File MP3 non generato."}), 500

    return send_file(str(preview_path), mimetype="audio/mpeg",
                     as_attachment=False, download_name="preview.mp3",
                     conditional=True)

@app.route("/api/cover/<job_id>")
def api_cover(job_id):
    """Serve the extracted cover thumbnail for preview."""
    job, err, sc = _check_job_owner(job_id)
    if err is not None:
        return "", sc if sc == 404 else 403
    cover_path = job.get("cover_thumb")
    if not cover_path or not os.path.exists(cover_path):
        return "", 404
    mime = job.get("cover_mime", "image/jpeg")
    return send_file(cover_path, mimetype=mime)


@app.route("/api/export_abm/<job_id>")
def api_export_abm(job_id):
    """Export cleaned text as .abm project file (ZIP with manifest + chapters + cover)."""
    _job, _err, _sc = _check_job_owner(job_id)
    if _err is not None:
        return _err, _sc
    with _jobs_lock:
        if job_id not in jobs:
            return jsonify({"error": "Job not found"}), 404
        job = jobs[job_id]
        info = job.get("info")
        if not info or not info.chapters:
            return jsonify({"error": "No book data available"}), 400
        # Snapshot data needed outside lock
        _job_data = {
            "optimized_chapters": job.get("optimized_chapters"),
            "selected_chapters": job.get("selected_chapters"),
            "cover_thumb": job.get("cover_thumb"),
            "original_filename": job.get("original_filename", ""),
        }

    import zipfile
    import io
    from datetime import datetime, timezone

    buf = io.BytesIO()
    safe_title = _safe_filename(info.title) or "project"

    # Align with _generate_optimized_abm: prefer cumulative optimized_chapters,
    # fall back to current selected_chapters, else include all.
    optimized = _job_data["optimized_chapters"]
    selected = _job_data["selected_chapters"]
    if optimized:
        chapter_set = set(optimized)
    elif selected:
        chapter_set = set(selected)
    else:
        chapter_set = None

    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        # Build chapter files and manifest entries
        chapters_manifest = []
        for ch in info.chapters:
            if chapter_set and ch.index not in chapter_set:
                continue
            ch_safe = _safe_filename(ch.title)[:50] or f"ch_{ch.index}"
            ch_filename = f"{ch.index:03d}_{ch_safe}.txt"
            zf.writestr(f"chapters/{ch_filename}", ch.text)
            chapters_manifest.append({
                "index": ch.index,
                "filename": ch_filename,
                "title": ch.title,
                "word_count": ch.word_count,
            })

        # Cover
        has_cover = False
        cover_file = ""
        cover_path = _job_data["cover_thumb"]
        if cover_path and os.path.exists(cover_path):
            cover_ext = ".png" if cover_path.endswith(".png") else ".jpg"
            cover_file = f"cover{cover_ext}"
            with open(cover_path, "rb") as cf:
                zf.writestr(cover_file, cf.read())
            has_cover = True

        # Manifest
        manifest = {
            "format": "audiobook-maker-project",
            "format_version": "1.0",
            "title": info.title,
            "author": getattr(info, "author", ""),
            "language": getattr(info, "language", ""),
            "has_cover": has_cover,
            "cover_file": cover_file,
            "exported_at": datetime.now(timezone.utc).isoformat(),
            "original_filename": _job_data["original_filename"],
            "ai_optimized": bool(job.get("ai_optimized")),
            "chapters": chapters_manifest,
        }
        zf.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False, indent=2))

    buf.seek(0)
    suffix = "_optimized" if job.get("ai_optimized") else ""
    download_name = f"{safe_title}{suffix}.abm"

    _log_activity(job_id, job.get("original_filename", ""), "EXPORT_ABM",
                  job.get("client_id", ""), job.get("client_ip", ""),
                  browser_lang=job.get("browser_lang", ""))

    return _apply_no_cache(send_file(buf, mimetype="application/zip", as_attachment=True,
                                      download_name=download_name))


@app.route("/api/generate", methods=["POST"])
def api_generate():
    data = request.json
    job_id = data.get("job_id")
    voice = data.get("voice", "it-IT-IsabellaNeural")
    rate = data.get("rate", "+0%")
    single_file = data.get("single_file", True)
    output_format = data.get("output_format", "m4b")
    podcast_base_url = (data.get("podcast_base_url") or "").strip()
    selected_chapters = data.get("selected_chapters")  # list of chapter indices, or None

    # Refuse Gemini voices when the module is missing or the API key is not configured.
    if voice and voice.startswith("gemini:"):
        if gemini_tts is None or not gemini_tts.is_available():
            return jsonify({"error": "gemini_tts_not_configured"}), 400

    job, err, sc = _check_job_owner(job_id)
    if err is not None:
        if sc == 404:
            return jsonify({"error": "Session expired. Re-upload file."}), 400
        return err, sc

    # Store format and podcast URL for email/download handlers
    job["output_format"] = output_format
    if output_format == "zip_rss":
        job["notify_download_type"] = "podcast"
        job["notify_base_url"] = podcast_base_url

    # Maintenance suspend check BEFORE payment preflight so we never consume
    # a payment token when the system can't accept the job anyway.
    if _suspend_new_jobs:
        return jsonify({"error": "System under maintenance. Please try again in a few minutes."}), 503

    # ----- F3: Gemini payment preflight -----
    payment_token = (data.get("payment_token") or "").strip()
    style_instruction = (data.get("gemini_style_instruction") or "")[:200]
    if voice and voice.startswith("gemini:"):
        # Recompute server-side total (mirror api_combined_estimate)
        info_pre = job.get("info")
        all_chs_pre = list(getattr(info_pre, "chapters", []) or [])
        sel = selected_chapters or []
        if sel:
            _by_index_pre = {ch.index: ch for ch in all_chs_pre}
            chs_pre = [_by_index_pre[i] for i in sel if i in _by_index_pre]
        else:
            chs_pre = all_chs_pre
        # Cap caratteri PRIMA di qualsiasi prenotazione budget o consumo del
        # pagamento. Per le voci PREMIUM il cap (MAX_GEMINI_TEXT_CHARS) e` piu`
        # restrittivo: verificarlo qui garantisce che un libro troppo grande non
        # porti MAI a riservare budget o consumare il token PayPal/voucher per
        # poi essere rifiutato dal cap a valle (riga ~6448) senza rimborso.
        _max_chars_pre = _effective_max_text_chars(voice, job)
        _sel_chars_pre = sum(getattr(ch, "char_count", 0) for ch in chs_pre)
        if _sel_chars_pre > _max_chars_pre:
            return jsonify({
                "error": f"Selection too large: {_sel_chars_pre:,} characters "
                         f"(limit {_max_chars_pre:,}). Please reduce the chapter selection.",
                "error_code": "selection_too_large",
                "chars_selected": _sel_chars_pre,
                "chars_limit": _max_chars_pre,
            }), 413
        # Lingua: priorita` (1) override UI da body request > (2) metadata libro
        # > (3) "it". Stessa logica usata da /api/combined_estimate e
        # /api/paypal_create_order_gemini: indispensabile per evitare amount
        # mismatch fra preflight pagamento e stima frontend (file TXT senza
        # metadata e EPUB/PDF con dc:language errato producono altrimenti
        # stime divergenti -> total_eur_pre sotto soglia -> payment_token
        # ignorato -> job["payment"] mai impostato -> audit charged=0).
        _ui_lang_pre = (data.get("lang") or "").strip().split("-")[0].lower()
        lang_pre = (_ui_lang_pre
                    or (getattr(info_pre, "language", "") or "").split("-")[0].lower()
                    or "it")
        # Persisti la lingua TTS effettivamente scelta dall'utente. Serve
        # all'audit Gemini (`_audit_language`) per non registrare la lingua
        # metadata del libro quando la voce TTS opera su una lingua diversa
        # (es. libro arabo con voce italiana -> audit deve mostrare "it").
        # Coesiste con `opt_lang` (settato da /api/optimize) come fallback.
        if _ui_lang_pre:
            job["gen_lang"] = _ui_lang_pre
        try:
            # Il rate scelto influisce sulla stima (estimate_audio_seconds scala
            # con rate_pct): il ricalcolo server-side deve usarlo per allinearsi
            # alla stima vista dall'utente e validare correttamente il pagamento.
            est_pre = gemini_tts.estimate_book_cost(chs_pre, voice, language=lang_pre, rate_pct=rate)
            gemini_eur_pre = round(est_pre["user_price_eur"], 2)
            # Persisti la stima sul job: serve all'audit Gemini per popolare
            # i campi *_est (input_tokens_est, output_tokens_est, audio_seconds_est,
            # google_cost_eur_est) altrimenti sempre 0 nel JSONL.
            job["gemini_estimate"] = est_pre
        except Exception as e:
            return jsonify({"error": f"estimate failed: {e}"}), 500
        llm_eur_pre = 0.0
        if data.get("ai_opt_enabled"):
            chars_pre = sum(len(getattr(c, "text", "") or "") for c in chs_pre)
            llm_eur_pre = round((chars_pre / 1_000_000.0) * LLM_RATE_EUR_PER_MCHAR, 2)
        total_eur_pre = round(gemini_eur_pre + llm_eur_pre, 2)

        # ----- Pre-flight budget guard (Google cost server-side) -----
        # Indipendente dal pagamento utente: questo è il cap interno sui costi
        # Google effettivi (per-job + daily). Se sforato, blocchiamo il job
        # PRIMA di chiamare l'API, così non spendiamo nulla.
        try:
            _google_cost_pre = float(est_pre.get("google_cost_eur", 0.0) or 0.0)
            preflight = gemini_tts.preflight_budget_check(_google_cost_pre)
            if preflight.get("warning"):
                print(f"[{job_id}] Budget warning (preflight): {preflight['warning']}")
            # Atomic reservation: blocca race fra job concorrenti che vedrebbero
            # lo stesso `spent` (audit JSONL viene scritto solo a fine job).
            gemini_tts.reserve_budget(job_id, _google_cost_pre)
        except gemini_tts.GeminiBudgetExceeded as _bex:
            return jsonify({
                "error": "budget_exceeded",
                "scope": getattr(_bex, "scope", "unknown"),
                "message": str(_bex),
                "estimated_eur": getattr(_bex, "estimated_eur", _google_cost_pre),
                "cap_eur": getattr(_bex, "cap_eur", None),
                "used_eur": getattr(_bex, "used_eur", None),
            }), 429
        except Exception as _bgenerr:
            # Errore non-budget: log e prosegui (graceful degradation).
            print(f"[{job_id}] preflight_budget_check raised non-budget error: "
                  f"{_bgenerr}")

        # ----- Pre-flight RPD check SINCRONO (prima di consumare il payment) -----
        # Eseguito qui per evitare la race condition fra il thread async di
        # run_generation (che setta status=error + marker) e lo stream SSE in
        # /api/progress (che a volte chiude il loop al primo tick prima che il
        # marker sia visibile). Bloccando qui rispondiamo a /api/generate
        # direttamente con error_code=gemini_overload e il frontend mostra il
        # popup senza fare affidamento sul polling SSE.
        try:
            _info_chs_for_plan = chs_pre
            _info_lang_for_plan = lang_pre
            _max_chars_pf = _pick_chunk_max_chars(voice, _info_lang_for_plan)
            _max_bytes_pf = _pick_chunk_max_bytes(voice)
            class _PlanInfo:
                pass
            _plan_info = _PlanInfo()
            _plan_info.chapters = _info_chs_for_plan
            _plan_for_pf = _plan_chunks(_plan_info, max_chars=_max_chars_pf, max_bytes=_max_bytes_pf)
            _total_chunks_pf = len(_plan_for_pf)
            _parts_v_pf = (voice or "").split(":")
            _model_key_pf = _parts_v_pf[1] if len(_parts_v_pf) >= 3 else "flash25"
            _pf_sync = gemini_tts.preflight_can_run(_model_key_pf, _total_chunks_pf)
            # Log RPD status (richiesta utente) — sempre, anche se OK.
            _cap_v_pf = _pf_sync.get("cap", 0)
            if _cap_v_pf and _cap_v_pf > 0:
                print(f"[{job_id}] RPD status [{_model_key_pf}]: "
                      f"used={_pf_sync.get('used', 0)}/{_cap_v_pf}, "
                      f"reserve={_pf_sync.get('reserve', 0)}, "
                      f"available={_pf_sync.get('available', 0)}, "
                      f"needed={_pf_sync.get('needed', 0)} "
                      f"-> {'OK' if _pf_sync.get('ok') else 'BLOCK (shortfall=' + str(_pf_sync.get('shortfall', 0)) + ')'}")
            else:
                print(f"[{job_id}] RPD status [{_model_key_pf}]: no local cap "
                      f"(ABM_GEMINI_RPD_{_model_key_pf.upper()}=0), needed={_pf_sync.get('needed', 0)}")
        except Exception as _pf_sync_err:
            print(f"[{job_id}] Sync preflight check error (non-fatal, proceeding): {_pf_sync_err}")
            _pf_sync = {"ok": True}
        if not _pf_sync.get("ok"):
            _reason_sync = (f"preflight_block_sync: model={_model_key_pf} "
                            f"needed={_pf_sync.get('needed')} "
                            f"available={_pf_sync.get('available')} "
                            f"shortfall={_pf_sync.get('shortfall')} "
                            f"cap={_pf_sync.get('cap')} used={_pf_sync.get('used')} "
                            f"reserve={_pf_sync.get('reserve')}")
            print(f"[{job_id}] PREFLIGHT BLOCK (sync) -> {_reason_sync}")
            # Nessun payment consumato a questo punto: niente refund da emettere.
            # Admin alert (informativo) per parita' con il blocco async.
            try:
                _admin_alert_text = (
                    f"[ABM-ADMIN] Gemini TTS — BLOCCO PREVENTIVO RPD (sync)\n\n"
                    f"job_id={job_id} model={_model_key_pf}\n"
                    f"needed={_pf_sync.get('needed')} "
                    f"available={_pf_sync.get('available')} "
                    f"shortfall={_pf_sync.get('shortfall')}\n"
                    f"cap={_pf_sync.get('cap')} used={_pf_sync.get('used')} "
                    f"reserve={_pf_sync.get('reserve')}\n"
                    f"Nessun payment consumato (blocco prima del consume).\n"
                )
                _admin_email = os.environ.get("ABM_ADMIN_EMAIL", "")
                if _admin_email:
                    email_service.send_email(_admin_email,
                                             "[ABM-ADMIN] Preflight RPD block (sync)",
                                             _admin_alert_text)
            except Exception as _ae:
                print(f"[{job_id}] Admin alert (sync) failed: {_ae}")
            try: gemini_tts.release_reservation(job_id)
            except Exception: pass
            return jsonify({
                "error": "Generation not started: PREMIUM voices are temporarily overloaded.",
                "error_code": "gemini_overload",
                "retry_after_sec": int(_pf_sync.get("retry_after_sec") or 0),
                "model_key": _model_key_pf,
                "refund_method": "",
            }), 429

        threshold_pre = float(os.environ.get("ABM_GEMINI_FREE_THRESHOLD_EUR", "0.50"))
        if total_eur_pre > threshold_pre:
            if not payment_token:
                try: gemini_tts.release_reservation(job_id)
                except Exception: pass
                return jsonify({
                    "error": "payment_required",
                    "total_eur": total_eur_pre,
                    "threshold_eur": threshold_pre,
                }), 402
            try:
                _pay_method = payment.consume_payment_token(
                    payment_token, total_eur_pre, job_id, purpose="gemini"
                )
            except ValueError as _pay_err:
                try: gemini_tts.release_reservation(job_id)
                except Exception: pass
                return jsonify({"error": f"payment_invalid: {_pay_err}"}), 400
            # Stash payment info on job for refund + audit
            job["payment"] = {
                "token": payment_token,
                "total_eur": total_eur_pre,
                "method": _pay_method,
                "ts": time.time(),
                "gemini_est": est_pre,
                "llm_eur": llm_eur_pre,
            }
        # Stash style for run_generation
        if style_instruction:
            job["gemini_style_instruction"] = style_instruction

    #  -  -  Atomic concurrency check + status claim  -  -
    client_id = job.get("client_id", "")
    client_ip = job.get("client_ip", "")
    with _jobs_lock:
        if job["status"] not in ("analyzed", "optimized"):
            _refund_payment_on_orphan(job_id, job, "status_conflict")
            try: gemini_tts.release_reservation(job_id)
            except Exception: pass
            return jsonify({"error": "Generation already running or completed."}), 400
        if client_id and MAX_CONCURRENT_PER_CLIENT > 0:
            if _active_generating_for_client_unlocked(client_id) >= MAX_CONCURRENT_PER_CLIENT:
                _refund_payment_on_orphan(job_id, job, "concurrent_limit")
                try: gemini_tts.release_reservation(job_id)
                except Exception: pass
                return jsonify({
                    "error": f"Concurrent generation limit reached ({MAX_CONCURRENT_PER_CLIENT}).",
                    "error_code": "concurrent_limit",
                    "max": MAX_CONCURRENT_PER_CLIENT,
                    "active": _active_generating_for_client_unlocked(client_id),
                }), 429
        # Atomically claim the slot
        job["status"] = "generating"
        # Save voice in job for logging
        job["voice"] = voice

    # Store format and podcast URL for email/download handlers
    job["output_format"] = output_format
    if output_format == "zip_rss":
        job["notify_download_type"] = "podcast"
        job["notify_base_url"] = podcast_base_url

    info = job["info"]

    # Filter chapters if a subset was selected
    job["selected_chapters"] = selected_chapters  # store for ABM export
    if selected_chapters:
        selected_set = set(selected_chapters)
        filtered = [ch for ch in info.chapters if ch.index in selected_set]
        if not filtered:
            return jsonify({"error": "No chapters selected."}), 400
        # Create a lightweight copy of info with filtered chapters
        info = copy(info)
        info.chapters = filtered
        info.total_words = sum(ch.word_count for ch in filtered)
        info.estimated_duration_minutes = info.total_words / 150

    # Hard cap on TTS-bound text size for THIS run: applied solo alla selezione.
    # 1 char ~= 50-100 byte di MP3, quindi il limite mantiene l'output sotto
    # ~75-150 MB. Per voci PREMIUM (gemini:) usiamo MAX_GEMINI_TEXT_CHARS
    # (default 800k, piu' restrittivo) data la maggior pressione su cost/RPM.
    max_text_chars = _effective_max_text_chars(voice, job)
    selected_chars = sum(ch.char_count for ch in info.chapters)
    if selected_chars > max_text_chars:
        with _jobs_lock:
            if job["status"] == "generating":
                job["status"] = "optimized" if job.get("ai_optimized") else "analyzed"
        # Rete di sicurezza: i cap a monte (create_order_gemini + blocco
        # pre-consume) dovrebbero impedire di arrivare qui con un pagamento gia`
        # consumato. Resta pero` il caso del testo espanso dall'ottimizzazione
        # LLM oltre il cap DOPO il consume: in quel caso rimborsa il pagamento e
        # rilascia la prenotazione budget Gemini, cosi` non si trattiene denaro
        # per un job non generabile.
        _refund_payment_on_orphan(job_id, job, "selection_too_large")
        try: gemini_tts.release_reservation(job_id)
        except Exception: pass
        return jsonify({
            "error": f"Selection too large: {selected_chars:,} characters "
                     f"(limit {max_text_chars:,}). Please reduce the chapter selection.",
            "error_code": "selection_too_large",
            "chars_selected": selected_chars,
            "chars_limit": max_text_chars,
        }), 413

    #  -  -  Pre-allocazione atomica budget Google Cloud TTS  -  -
    # Verifica E deduce immediatamente i caratteri richiesti, così conversioni
    # parallele non possono passare lo stesso check. Il refund della parte
    # non consumata avviene in run_generation in caso di errore/cancellazione.
    if google_tts is not None and google_tts.is_google_voice(voice):
        total_chars_needed = sum(ch.char_count for ch in info.chapters)
        ok, remaining_after = google_tts.reserve_chars(total_chars_needed)
        if not ok:
            with _jobs_lock:
                if job["status"] == "generating":
                    job["status"] = "analyzed" if job.get("ai_optimized") else "analyzed"
            return jsonify({
                "error": f"Google TTS monthly limit: {remaining_after:,} chars remaining, "
                         f"but this book needs {total_chars_needed:,} chars.",
                "error_code": "google_tts_budget",
                "chars_needed": total_chars_needed,
                "chars_remaining": remaining_after,
            }), 429
        # Memorizza i caratteri prenotati nel job per il refund
        job["google_tts_reserved"] = total_chars_needed
        print(f"[{job_id}] Google TTS: reserved {total_chars_needed:,} chars "
              f"(remaining: {remaining_after:,})")
        # Invalida la cache voci: se il budget si avvicina allo zero, le voci
        # potrebbero scomparire al prossimo /api/voices
        _invalidate_voices_cache()

    # Increment generation epoch to invalidate any stale threads
    job["gen_epoch"] = job.get("gen_epoch", 0) + 1
    thread = threading.Thread(
        target=run_generation, args=(job_id, info, voice, rate, single_file),
        kwargs={'output_format': output_format, 'podcast_base_url': podcast_base_url,
                'gemini_style_instruction': job.get("gemini_style_instruction")},
        daemon=True
    )
    thread.start()
    _log_activity(job_id, job.get("original_filename", ""), "GENERATE",
                  client_id, client_ip, voice,
                  browser_lang=job.get("browser_lang", ""))
    _admin_notify_generation(job_id, info, voice, job.get("original_filename", ""))
    return jsonify({"status": "started"})


@app.route("/api/job_status/<job_id>")
def api_job_status(job_id):
    job, err, sc = _check_job_owner(job_id)
    if err is not None:
        return ({"error": "Not found"} if sc == 404 else {"error": "Forbidden"}), sc

    st = job.get("status", "")
    cur, tot = 0, 0
    pct = 0
    if st == "optimizing":
        total_chars = job.get("opt_total_chars", 1)
        done_chars = job.get("opt_processed_chars", 0)
        cur_ch_chars = job.get("opt_current_chapter_chars", 0)
        streamed = min(job.get("opt_streamed_chars", 0), cur_ch_chars)
        worked = done_chars + streamed
        pct = min(99, int(worked / total_chars * 100))
        cur, tot = worked, total_chars
    elif st == "generating":
        cur = job.get("progress_current", 0)
        tot = job.get("progress_total", 0)
        pct = int(cur / tot * 100) if tot > 0 else 0
    
    return {
        "status": st,
        "current": cur,
        "total": tot,
        "pct": pct,
        "message": job.get("progress_message", "") or job.get("opt_progress_message", "")
    }


@app.route("/api/progress/<job_id>")
def api_progress(job_id):
    # Ownership check fuori dallo stream (la `request` non è disponibile dentro il generator).
    _job_pre, _err_pre, _sc_pre = _check_job_owner(job_id)
    if _err_pre is not None:
        return _err_pre, _sc_pre

    def stream():
        while True:
            if job_id not in jobs:
                yield f"data: {json.dumps({'status': 'error', 'error': 'Job not found'})}\n\n"
                break
            job = jobs[job_id]
            # Heartbeat: segna che un client sta ascoltando
            job["last_poll"] = time.time()
            payload = {
                "status": job.get("status", "unknown"),
                "progress_current": job.get("progress_current", 0),
                "progress_total": job.get("progress_total", 0),
                "progress_message": job.get("progress_message", ""),
                "current_chapter": job.get("current_chapter", ""),
                "current_chapter_num": job.get("current_chapter_num", 0),
                "total_chapters": job.get("total_chapters", 0),
                "elapsed_seconds": job.get("elapsed_seconds", 0),
                "bytes_generated": job.get("bytes_generated", 0),
                "processed_chars": job.get("processed_chars", 0),
                "total_chars": job.get("total_chars", 0),
            }
            # Espone l'importo pagato (quota refundabile) e il metodo: serve
            # al frontend per reidratare _payState dopo un reload e mostrare
            # "Importo versato" corretto nel modal di cancel. Usa total_eur
            # (l'importo effettivamente consumato dal payment_token) perche'
            # e' la stessa base che _CancelledError handler in
            # generation_engine.py passa a cancel_policy.compute_cancel_retention.
            _paym_sse = job.get("payment") or {}
            if _paym_sse:
                try:
                    payload["paid_eur"] = round(float(_paym_sse.get("total_eur", 0.0) or 0.0), 2)
                except (TypeError, ValueError):
                    payload["paid_eur"] = 0.0
                payload["paid_method"] = _paym_sse.get("method", "")
            if job.get("status") == "error":
                # Errore generico verso il client: il dettaglio resta nei log server-side.
                payload["error"] = "generation_failed"
                # Eccezione: per il pre-flight block delle voci PREMIUM,
                # esponiamo un error_kind strutturato cosi' il frontend puo'
                # mostrare un popup specifico e riportare l'utente alla scelta
                # voce (invece di redirect generico alla home).
                _pf_block = job.get("gemini_preflight_block")
                if _pf_block:
                    payload["error_kind"] = "gemini_overload"
                    payload["retry_after_sec"] = int(_pf_block.get("retry_after_sec") or 0)
                    _paym = job.get("payment") or {}
                    payload["refund_method"] = _paym.get("method", "")
                    # Codice voucher di rimborso solo se PayPal (per voucher
                    # il riaccredito e' silenzioso sul codice originale).
                    if _paym.get("method") == "paypal":
                        _vcode = job.get("refund_voucher_code")
                        if _vcode:
                            payload["refund_voucher_code"] = _vcode
                yield f"data: {json.dumps(payload)}\n\n"
                break
            if job.get("status") == "cancelled" or job.get("cancelled"):
                payload["status"] = "cancelled"
                # Espone metadati cancel volontario voci PREMIUM: refund summary
                # (paid/retained/refund/progress) + link al MP3 parziale se
                # generato. Vedi T7 generation_engine.py:_CancelledError branch.
                _cm = job.get("cancel_meta")
                if isinstance(_cm, dict):
                    payload["cancel_meta"] = {
                        "paid_eur": _cm.get("paid_eur", 0),
                        "retained_eur": _cm.get("retained_eur", 0),
                        "refund_eur": _cm.get("refund_eur", 0),
                        "progress_pct": _cm.get("progress_pct", 0),
                        "partial_audio_delivered": bool(
                            _cm.get("partial_audio_delivered", False)),
                    }
                _pdl = job.get("partial_download_url")
                if _pdl:
                    payload["partial_download_url"] = _pdl
                yield f"data: {json.dumps(payload)}\n\n"
                break
            if job.get("status") == "done":
                payload["output_name"] = job.get("output_name", "output")
                payload["has_podcast"] = job.get("podcast_ready", False)
                # Reconnection fallback: se output_m4b o optimized_abm_path non sono
                # impostati, cerca SOLO dentro la cartella della generazione corrente
                # (job["output_dir"] = output_{gen_epoch}/). Mai fare scan globale su
                # tutti gli output_*/ perché erediteresti file di run precedenti e li
                # mostreresti come scaricabili per la run corrente.
                _cur_output = job.get("output_dir")
                if _cur_output and os.path.isdir(_cur_output):
                    _cur_path = Path(_cur_output)
                    if not job.get("output_m4b"):
                        _m4bs = list(_cur_path.glob("*.m4b"))
                        if _m4bs:
                            job["output_m4b"] = str(_m4bs[0])
                    if not job.get("optimized_abm_path"):
                        _abms = list(_cur_path.glob("*.abm"))
                        if _abms:
                            job["optimized_abm_path"] = str(_abms[0])

                payload["output_m4b"] = bool(job.get("output_m4b"))
                payload["has_abm"] = bool(job.get("ai_optimized")) or (bool(job.get("optimized_abm_path")) and os.path.exists(job.get("optimized_abm_path", "")))
                payload["failed_chunks"] = job.get("failed_chunks", 0)
                yield f"data: {json.dumps(payload)}\n\n"
                break
            yield f"data: {json.dumps(payload)}\n\n"
            time.sleep(1)

    return Response(
        stream_with_context(stream()),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.route("/api/cancel/<job_id>", methods=["POST"])
def api_cancel(job_id):
    """Cancella un job in corso.

    Per voci Gemini, il cancel volontario e' bloccato oltre la soglia
    ABM_GEMINI_CANCEL_LOCK_PCT (default 70). Vedi
    docs/superpowers/specs/2026-05-25-cancel-gemini-floor-design.md.
    """
    _job, _err, _sc = _check_job_owner(job_id)
    if _err is not None:
        if _sc == 404:
            return jsonify({"status": "not_found"}), 404
        return _err, _sc
    with _jobs_lock:
        job = jobs[job_id]
        voice = job.get("voice", "") or job.get("opt_voice", "") or ""
        is_gemini = voice.startswith("gemini:")
        if is_gemini:
            try:
                lock_pct = int(os.environ.get("ABM_GEMINI_CANCEL_LOCK_PCT", "70"))
            except (TypeError, ValueError):
                lock_pct = 70
            if 0 < lock_pct < 100:
                from generation_engine import _progress_pct
                pct = _progress_pct(job)
                if pct > lock_pct:
                    return jsonify({
                        "error": "cancel_locked_progress",
                        "progress_pct": pct,
                        "lock_pct": lock_pct,
                    }), 409
        force = request.args.get("force") == "1"
        if job.get("email_registered") and not force:
            print(f"[{job_id}] Cancel ignored  -  email registered for background processing")
            return jsonify({"status": "ignored_email_registered"})
        job["cancelled"] = True
        # NB: non incrementiamo gen_epoch qui. Il bump epoch e' riservato
        # al riavvio della generazione (/api/generate linea ~5548) per
        # invalidare worker thread orfane. Bumparlo sul cancel volontario
        # farebbe finire run_generation nel ramo STALE del _CancelledError
        # handler, saltando refund + partial-audio email.
        job["status"] = "analyzed"
    return jsonify({"status": "cancelling"})


@app.route("/api/cancel_preview/<job_id>", methods=["GET"])
def api_cancel_preview(job_id):
    """Snapshot sincrono dei parametri di cancel per la modale di conferma.

    Il client legge da qui paid_eur/progress_pct invece di affidarsi a
    _payState (in-memory, perso al reload) o all'evento SSE (latenza: dopo
    F5 il primo evento puo' arrivare dopo il click su cancel, lasciando
    il modal con "Importo versato: 0.00 EUR"). Server e' single source of
    truth: legge payment.total_eur, identico a quanto consumato dal
    refund handler in generation_engine._handle_cancelled_error.
    """
    _job, _err, _sc = _check_job_owner(job_id)
    if _err is not None:
        if _sc == 404:
            return jsonify({"status": "not_found"}), 404
        return _err, _sc
    with _jobs_lock:
        if job_id not in jobs:
            return jsonify({"status": "not_found"}), 404
        job = jobs[job_id]
        paym = job.get("payment") or {}
        try:
            paid_eur = round(float(paym.get("total_eur", 0.0) or 0.0), 2)
        except (TypeError, ValueError):
            paid_eur = 0.0
        paid_method = paym.get("method", "") or ""
        try:
            from generation_engine import _progress_pct
            pct = _progress_pct(job)
        except Exception:
            pct = 0
        try:
            lock_pct = int(os.environ.get("ABM_GEMINI_CANCEL_LOCK_PCT", "70"))
        except (TypeError, ValueError):
            lock_pct = 70
    return jsonify({
        "paid_eur": paid_eur,
        "paid_method": paid_method,
        "progress_pct": pct,
        "lock_pct": lock_pct,
        "locked": (0 < lock_pct < 100 and pct > lock_pct),
    })


@app.route("/api/heartbeat/<job_id>", methods=["POST"])
def api_heartbeat(job_id):
    """Keep-alive: il client segnala che è ancora sulla pagina."""
    _job, _err, _sc = _check_job_owner(job_id)
    if _err is not None:
        return "", _sc
    with _jobs_lock:
        if job_id in jobs:
            jobs[job_id]["last_poll"] = time.time()
            return "", 204
    return "", 404


@app.route("/api/reset_to_chapters/<job_id>", methods=["POST"])
def api_reset_to_chapters(job_id):
    """Reset a completed job back to 'analyzed' so the user can select different chapters."""
    _job, _err, _sc = _check_job_owner(job_id)
    if _err is not None:
        return _err, _sc
    with _jobs_lock:
        if job_id not in jobs:
            return jsonify({"error": "Job not found"}), 404
        job = jobs[job_id]
        if job.get("status") != "done":
            return jsonify({"error": "Job is not in completed state"}), 400
        if not job.get("info") or not job["info"].chapters:
            return jsonify({"error": "Book data no longer available. Please re-upload the file."}), 400

    # Per-epoch layout: /api/generate writes into work_dir/output_{epoch}/,
    # including the .abm snapshot. Reset is fully non-destructive for outputs
    # — the next generation bumps gen_epoch and writes into a fresh dir,
    # leaving previous epochs (audio + .abm) intact for any active email
    # tokens. The cleanup loop purges orphan dirs after retention.
    #
    # The only .abm file we still delete here is a stray copy at work_dir
    # root left over from an optimization phase that never ran a generation
    # (the work_dir-root path is no longer in use once output_dir exists).
    work_dir = UPLOAD_DIR / job_id
    for abm in work_dir.glob("*.abm"):
        try:
            abm.unlink()
            print(f"[reset] Removed work_dir-root ABM: {abm}")
        except OSError:
            pass

    # Reset job state (inside lock)
    with _jobs_lock:
        job["status"] = "analyzed"
        job["last_poll"] = time.time()
        for key in ("output_files", "output_name", "output_zip", "output_file",
                    "output_m4b",
                    "podcast_ready", "podcast_safe_name", "podcast_mp3s",
                    "progress_current", "progress_total", "progress_message",
                    "processed_chars", "total_chars", "bytes_generated",
                    "start_time", "elapsed_seconds", "current_chapter",
                    "current_chapter_num", "total_chapters",
                    "downloaded_at", "email_sent_at", "email_registered",
                    "failed_chunks", "cancelled",
                    "opt_cancelled", "opt_progress_current", "opt_progress_total",
                    "opt_progress_message", "opt_current_chapter", "opt_current_chapter_num",
                    "opt_processed_chars", "opt_total_chars", "opt_streamed_chars",
                    "opt_current_chapter_chars", "opt_elapsed_seconds", "opt_completed_at",
                    "optimized_abm_path", "optimized_abm_name",
                    "selected_chapters",
                    "opt_auto_generate", "opt_single_file", "opt_output_format",
                    "opt_podcast_base_url", "opt_voice", "opt_rate", "opt_lang",
                    "email_token"):
            job.pop(key, None)
    # Keep: info, epub_path, cover_thumb, cover_mime, original_filename, preview_text,
    #        client_id, client_ip, voice (so preview still works)

    _log_activity(job_id, job.get("original_filename", ""), "RESET_CHAPTERS",
                  job.get("client_id", ""), job.get("client_ip", ""),
                  job.get("voice", ""), job.get("browser_lang", ""))

    return jsonify({"status": "ok"})


@app.route("/api/register_email", methods=["POST"])
def api_register_email():
    """Register email for job completion notification."""
    import re
    data = request.json or {}
    job_id = data.get("job_id", "")
    email = (data.get("email") or "").strip().lower()
    download_type = data.get("download_type", "audio")  # "audio" or "podcast"
    base_url = (data.get("base_url") or "").strip()

    job, err, sc = _check_job_owner(job_id)
    if err is not None:
        return err, sc

    if not email or not re.match(r'^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$', email):
        return jsonify({"error": "Invalid email address"}), 400

    if not _smtp_available():
        return jsonify({"error": "Email service not configured on this server"}), 503

    job["notify_email"] = email
    job["notify_download_type"] = download_type
    job["notify_base_url"] = base_url
    job["notify_lang"] = data.get("lang", "en")
    # Keep job alive indefinitely while generating (disable heartbeat-based cleanup)
    job["email_registered"] = True
    # Marker 'pending' su disco: protegge la dir dal cleanup orfani di altri worker
    # per tutta la lavorazione, finché _send_completion_email lo sovrascriverà
    # con il timestamp.
    _write_email_pending_marker(UPLOAD_DIR / job_id)

    print(f"[{job_id}] Email notification registered: {email} (type: {download_type})")
    _log_activity(job_id, job.get("original_filename", ""), "EMAIL_REGISTERED",
                  job.get("client_id", ""), job.get("client_ip", ""),
                  job.get("voice", ""), job.get("browser_lang", ""))

    # Persist client_id → email per fallback su job futuri (difesa da UI fallita)
    client_id = job.get("client_id", "")
    if client_id:
        _client_emails[client_id] = email
        _save_client_emails()

    return jsonify({"status": "registered", "email": email})


@app.route("/api/email_available")
def api_email_available():
    """Check if email notification is available (SMTP configured)."""
    return jsonify({"available": _smtp_available()})


# ----------------------------------------------------------------------
# LLM TEXT OPTIMIZATION API
# ----------------------------------------------------------------------

@app.route("/api/llm_available")
def api_llm_available():
    """Check if LLM text optimization is available (DeepSeek configured)."""
    return jsonify({
        "available": _llm_available(),
        "paypal_available": _paypal_available(),
        "paypal_client_id": PAYPAL_CLIENT_ID if _paypal_available() else "",
        "paypal_mode": PAYPAL_MODE,
        "rate_eur_per_mchar": LLM_RATE_EUR_PER_MCHAR,
        "free_threshold_eur": LLM_FREE_THRESHOLD_EUR,
        "voucher_bonus_percent": VOUCHER_BONUS_PERCENT,
        "voucher_expiry_days": VOUCHER_EXPIRY_DAYS,
    })


def _parse_selected_chapters(raw_data):
    """Utility ultra-robusta per estrarre indici interi da qualsiasi input."""
    if raw_data is None: return []
    indices = []
    if isinstance(raw_data, str) and "," in raw_data:
        items = raw_data.split(",")
    elif isinstance(raw_data, list):
        items = raw_data
    else:
        items = [raw_data]
    for item in items:
        if item is None: continue
        if isinstance(item, (int, float)):
            indices.append(int(item))
        elif isinstance(item, str):
            parts = item.replace("[", "").replace("]", "").split(",")
            for p in parts:
                p = p.strip()
                if p.isdigit(): indices.append(int(p))
    return sorted(list(set(indices)))

@app.route("/api/optimize_estimate/<job_id>")
def api_optimize_estimate(job_id):
    job, err, sc = _check_job_owner(job_id)
    if err is not None: return err, sc
    info = job.get("info")
    if not info or not info.chapters: return jsonify({"error": "No book data"}), 400
    raw_sel = request.args.getlist("selected_chapters") + request.args.getlist("selected_chapters[]")
    selected_indices = _parse_selected_chapters(raw_sel)
    already = set(job.get("optimized_chapters", []))
    if raw_sel:
        total_chars = sum(ch.char_count for ch in info.chapters if ch.index in selected_indices and ch.index not in already)
    else:
        total_chars = sum(ch.char_count for ch in info.chapters if ch.index not in already)
    cost = _estimate_llm_cost_eur(total_chars)

    # Pre-validate the output-size cap against the full selected set so the
    # user is informed before being asked to pay. La voce influisce sul cap
    # (Gemini -> MAX_GEMINI_TEXT_CHARS, altrimenti MAX_TEXT_CHARS): il
    # frontend passa ?voice=... quando la conosce; in assenza si usa lo standard.
    voice_q = (request.args.get("voice") or "").strip()
    max_text_chars = _max_text_chars_for_voice(voice_q)
    if raw_sel:
        selected_set_cap = set(selected_indices)
    else:
        selected_set_cap = {ch.index for ch in info.chapters}
    selected_chars_total = sum(
        ch.char_count for ch in info.chapters if ch.index in selected_set_cap
    )
    if selected_chars_total > max_text_chars:
        return jsonify({
            "error": f"Selection too large: {selected_chars_total:,} characters "
                     f"(limit {max_text_chars:,}). Please reduce the chapter selection.",
            "error_code": "selection_too_large",
            "chars_selected": selected_chars_total,
            "chars_limit": max_text_chars,
        }), 413

    return jsonify({
        "chars": total_chars, "cost_eur": cost,
        "requires_payment": cost > LLM_FREE_THRESHOLD_EUR,
        "free_threshold_eur": LLM_FREE_THRESHOLD_EUR,
        "rate_eur_per_mchar": LLM_RATE_EUR_PER_MCHAR,
        "optimized_chapters": list(already),
    })


@app.route("/api/paypal_create_order", methods=["POST"])
def api_paypal_create_order():
    if not _paypal_available(): return jsonify({"error": "PayPal not configured"}), 503
    data = request.json or {}; job_id = data.get("job_id", "")
    if job_id not in jobs: return jsonify({"error": "Job not found"}), 404
    job = jobs[job_id]; info = job.get("info")
    if not info: return jsonify({"error": "No book data"}), 400
    selected_chapters = _parse_selected_chapters(data.get("selected_chapters"))
    if selected_chapters:
        total_chars = sum(ch.char_count for ch in info.chapters if ch.index in selected_chapters)
    else:
        total_chars = sum(ch.char_count for ch in info.chapters)
    cost = _estimate_llm_cost_eur(total_chars)
    if cost <= LLM_FREE_THRESHOLD_EUR:
        return jsonify({"error": "Payment not required for this job"}), 400

    book_title = getattr(info, "title", "") or "Audiobook"
    description = f"AI text optimization  -  {book_title[:60]}"
    try:
        order = _paypal_create_order(cost, description, custom_id=job_id)
    except Exception as e:
        print(f"[paypal] create_order failed: {e}")
        return jsonify({"error": f"PayPal error: {e}"}), 500

    # Diagnostic: log links/payee info to help troubleshoot sandbox issues
    try:
        pu = (order.get("purchase_units") or [{}])[0]
        payee = pu.get("payee", {})
        print(f"[paypal] order created: id={order.get('id')} status={order.get('status')} "
              f"amount={pu.get('amount',{}).get('value')} {pu.get('amount',{}).get('currency_code')} "
              f"payee_email={payee.get('email_address')} payee_id={payee.get('merchant_id')}")
    except Exception:
        pass

    return jsonify({
        "order_id": order.get("id"),
        "amount_eur": cost,
        "status": order.get("status"),
    })


@app.route("/api/paypal_debug_order/<order_id>", methods=["GET"])
def api_paypal_debug_order(order_id):
    """Diagnostic: fetch order from PayPal to inspect payee/status/etc.
    Admin-only: returns payer PII (email, name, address) — must never be public."""
    if not _admin_auth_ok(_admin_auth_from_request()):
        return jsonify({"error": "Unauthorized"}), 403
    import requests
    if not _paypal_available():
        return jsonify({"error": "PayPal not configured"}), 503
    try:
        token = _paypal_get_access_token()
        r = requests.get(
            f"{PAYPAL_API_BASE}/v2/checkout/orders/{order_id}",
            headers={"Authorization": f"Bearer {token}"},
            timeout=15,
        )
        return jsonify({"status_code": r.status_code, "body": r.json()}), r.status_code
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/paypal_capture_order", methods=["POST"])
def api_paypal_capture_order():
    """Capture an approved PayPal order. Returns payment_token on success."""
    if not _paypal_available():
        return jsonify({"error": "PayPal not configured"}), 503
    data = request.json or {}
    order_id = (data.get("order_id") or "").strip()
    job_id = (data.get("job_id") or "").strip()
    if not order_id:
        return jsonify({"error": "Missing order_id"}), 400

    # Atomic flow: idempotency + capture + amount reconciliation + store
    # serializzato da payment._capture_lock per prevenire double-capture race.
    try:
        result = payment.capture_and_store_order(order_id, job_id=job_id)
    except payment.CaptureAmountMismatchError as e:
        print(f"[paypal] AMOUNT MISMATCH order={order_id} job={job_id}: {e}")
        _log_activity(job_id, jobs.get(job_id, {}).get("original_filename", ""),
                      "PAYMENT_AMOUNT_MISMATCH", "", "", "", str(e))
        return jsonify({"error": "Payment amount mismatch (refused)"}), 400
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        print(f"[paypal] capture_order failed: {e}")
        return jsonify({"error": f"PayPal capture error: {e}"}), 500

    amount_eur = result["amount_eur"]
    email = result["email"]

    if not result.get("already_captured"):
        _log_activity(job_id, jobs.get(job_id, {}).get("original_filename", ""),
                      "PAYMENT_CAPTURED", "", "", "", "")
        # Send receipt email (non-blocking best-effort)
        if email and _smtp_available():
            try:
                _send_payment_receipt_email(order_id, email, amount_eur, jobs.get(job_id, {}))
            except Exception as e:
                print(f"[paypal] receipt email failed: {e}")

    return jsonify({
        "payment_token": order_id,
        "amount_eur": amount_eur,
        "email": email,
        "already_captured": result.get("already_captured", False),
    })


@app.route("/api/voucher_validate", methods=["POST"])
def api_voucher_validate():
    """Validate a voucher code + email. Returns payment_token if valid.

    Protetto da rate limit per IP (5/min, 30/ora) e lockout per email dopo N fallimenti.
    Ogni tentativo viene loggato come VOUCHER_ATTEMPT.
    """
    data = request.json or {}
    code = (data.get("code") or "").strip().upper()
    email = (data.get("email") or "").strip().lower()
    ip = request.headers.get("X-Forwarded-For", request.remote_addr or "").split(",")[0].strip()
    purpose = (data.get("purpose") or "any")
    try:
        amount_required = float(data.get("amount_eur") or 0)
    except (ValueError, TypeError):
        amount_required = 0.0

    #  -  Rate limit check  -
    allowed, retry_after, reason = _voucher_rl_check(ip, email)
    if not allowed:
        _log_activity("", "", f"VOUCHER_ATTEMPT_BLOCKED:{reason}", "", ip, "", "")
        resp = jsonify({
            "valid": False,
            "reason": reason or "rate_limit",
            "error": "Too many attempts. Please try later.",
            "retry_after": retry_after,
        })
        resp.status_code = 429
        resp.headers["Retry-After"] = str(retry_after)
        return resp

    #  -  Validation logic  -
    outcome = "OK"
    status = 200
    body = None
    if not code or not email:
        outcome, status, body = "MISSING_FIELDS", 400, {"error": "Code and email required", "valid": False, "reason": "missing_fields"}
    elif code not in payment._vouchers:
        outcome, status, body = "NOT_FOUND", 404, {"error": "Voucher not found", "valid": False, "reason": "not_found"}
    else:
        v = payment._vouchers[code]
        remaining = _voucher_remaining(v)
        if v.get("expires_at", 0) < time.time():
            outcome, status, body = "EXPIRED", 400, {"error": "Voucher expired", "valid": False, "reason": "expired"}
        elif v.get("email", "").lower() != email:
            outcome, status, body = "EMAIL_MISMATCH", 400, {"error": "Email does not match voucher", "valid": False, "reason": "email_mismatch"}
        elif remaining < 0.01:
            outcome, status, body = "USED", 400, {"error": "Voucher fully used", "valid": False, "reason": "used"}
        elif amount_required > 0 and remaining < amount_required:
            outcome, status, body = "INSUFFICIENT", 400, {
                "valid": False,
                "reason": "insufficient",
                "error": "Voucher balance insufficient",
                "remaining_eur": remaining,
                "required_eur": amount_required,
            }
        else:
            # Saldo residuo: l'UI lo usa come "amount_eur" spendibile.
            body = {
                "valid": True,
                "payment_token": code,
                "amount_eur": remaining,
                "remaining_eur": remaining,
                "original_amount_eur": round(float(v.get("amount_eur", 0) or 0), 2),
                "expires_at": v.get("expires_at"),
                "purpose_requested": purpose,
            }

    success = (outcome == "OK")
    _voucher_rl_record_result(email, success)
    # Log in forma strutturata (usiamo i campi esistenti: voice=code masked, browser_lang=outcome)
    code_masked = (code[:4] + "...") if code else ""
    _log_activity("", "", "VOUCHER_ATTEMPT", "", ip, code_masked, outcome)
    return jsonify(body), status


@app.route("/api/gemini_estimate", methods=["POST"])
def api_gemini_estimate():
    """Stima costo Voci PREMIUM per il job corrente, capitoli selezionati."""
    import gemini_tts as _gemini_tts_mod
    data = request.get_json(silent=True) or {}
    job_id = data.get("job_id", "")
    voice_id = data.get("voice_id", "")
    selected = data.get("selected_chapters") or []
    rate = data.get("rate", "+0%")
    ui_lang = (data.get("lang") or "").strip().split("-")[0].lower()

    if not voice_id.startswith("gemini:"):
        return jsonify({"error": "voice_id must be a Gemini voice"}), 400
    with _jobs_lock:
        job = jobs.get(job_id)
    if not job:
        return jsonify({"error": "job not found"}), 404
    info = job.get("info")
    if info is None or not getattr(info, "chapters", None):
        return jsonify({"error": "job has no chapters"}), 400

    all_chs = list(info.chapters)
    if selected:
        _by_index = {ch.index: ch for ch in all_chs}
        chs = [_by_index[i] for i in selected if i in _by_index]
    else:
        chs = all_chs
    if not chs:
        return jsonify({"error": "no chapters selected"}), 400

    # Lingua: priorita` (1) override UI da "Impostazioni audio" > (2) metadata
    # libro > (3) "it". L'UI vince perche' governa anche cluster rate-log e
    # ratio chars/token: necessario per TXT (mai metadata) e per metadata errati.
    lang = ui_lang or (getattr(info, "language", "") or "").split("-")[0].lower() or "it"
    try:
        est = _gemini_tts_mod.estimate_book_cost(chs, voice_id, language=lang, rate_pct=rate)
    except Exception as e:
        return jsonify({"error": f"estimate failed: {e}"}), 500

    return jsonify({
        "chars_total": est["chars_total"],
        "audio_seconds_est": est["audio_seconds_est"],
        "estimated_audio_minutes": round(est["estimated_audio_minutes"], 1),
        "user_price_eur": est["user_price_eur"],
        "is_free": est["is_free"],
        "model_key": est["model_key"],
        "model_label": est["model_label"],
        "language": est["language"],
        "rate_step": est.get("rate_step", 0),
        "breakdown": {
            "input_tokens_est": est["input_tokens_est"],
            "output_tokens_est": est["output_tokens_est"],
            "google_cost_eur": est["google_cost_eur"],
            "margin_percent": est["margin_percent"],
        },
    })


@app.route("/api/combined_estimate", methods=["POST"])
def api_combined_estimate():
    """Stima combinata Voci PREMIUM + ottimizzazione testo AI."""
    import gemini_tts as _gemini_tts_mod
    data = request.get_json(silent=True) or {}
    job_id = data.get("job_id", "")
    voice_id = data.get("voice_id", "")
    selected = data.get("selected_chapters") or []
    ai_opt = bool(data.get("ai_opt_enabled", False))
    rate = data.get("rate", "+0%")
    ui_lang = (data.get("lang") or "").strip().split("-")[0].lower()

    with _jobs_lock:
        job = jobs.get(job_id)
    if not job:
        return jsonify({"error": "job not found"}), 404
    info = job.get("info")
    if info is None or not getattr(info, "chapters", None):
        return jsonify({"error": "no chapters"}), 400

    all_chs = list(info.chapters)
    if selected:
        _by_index = {ch.index: ch for ch in all_chs}
        chs = [_by_index[i] for i in selected if i in _by_index]
    else:
        chs = all_chs
    if not chs:
        return jsonify({"error": "no chapters"}), 400

    # Lingua: priorita` (1) override UI da "Impostazioni audio" > (2) metadata
    # libro > (3) "it". L'UI vince perche' governa anche cluster rate-log e
    # ratio chars/token: necessario per TXT (mai metadata) e per metadata errati.
    lang = ui_lang or (getattr(info, "language", "") or "").split("-")[0].lower() or "it"

    gemini_eur = 0.0
    gemini_breakdown = {}
    rate_step = 0
    if voice_id.startswith("gemini:"):
        try:
            est = _gemini_tts_mod.estimate_book_cost(chs, voice_id, language=lang, rate_pct=rate)
        except Exception as e:
            return jsonify({"error": f"estimate failed: {e}"}), 500
        gemini_eur = round(est["user_price_eur"], 2)
        rate_step = est.get("rate_step", 0)
        gemini_breakdown = {
            "chars": est["chars_total"],
            "audio_minutes": round(est["estimated_audio_minutes"], 1),
            "google_cost_eur": est["google_cost_eur"],
            "model_label": est["model_label"],
            "rate_step": rate_step,
        }

    llm_eur = 0.0
    llm_breakdown = {}
    if ai_opt:
        chars = sum(len(getattr(c, "text", "") or "") for c in chs)
        llm_rate = LLM_RATE_EUR_PER_MCHAR
        llm_eur = round((chars / 1_000_000.0) * llm_rate, 2)
        llm_breakdown = {"chars": chars, "rate_eur_per_mchar": llm_rate}

    total = round(gemini_eur + llm_eur, 2)
    threshold = float(os.environ.get("ABM_GEMINI_FREE_THRESHOLD_EUR", "0.50"))

    # Pre-flight RPD check ANTICIPATO (prima ancora di proporre il pagamento).
    # Cosi' l'utente che ha selezionato una voce PREMIUM saturata vede subito
    # l'avviso "non disponibile" senza passare per il flusso pagamento/PayPal.
    overload_info = None
    if voice_id.startswith("gemini:"):
        try:
            _max_chars_cb = _pick_chunk_max_chars(voice_id, lang)
            _max_bytes_cb = _pick_chunk_max_bytes(voice_id)
            class _PlanInfoCB:
                pass
            _pi = _PlanInfoCB()
            _pi.chapters = chs
            _plan_cb = _plan_chunks(_pi, max_chars=_max_chars_cb, max_bytes=_max_bytes_cb)
            _total_chunks_cb = len(_plan_cb)
            _parts_cb = voice_id.split(":")
            _model_key_cb = _parts_cb[1] if len(_parts_cb) >= 3 else "flash25"
            _pf_cb = _gemini_tts_mod.preflight_can_run(_model_key_cb, _total_chunks_cb)
            if not _pf_cb.get("ok"):
                overload_info = {
                    "model_key": _model_key_cb,
                    "retry_after_sec": int(_pf_cb.get("retry_after_sec") or 0),
                    "needed": _pf_cb.get("needed"),
                    "available": _pf_cb.get("available"),
                }
                print(f"[{job_id}] combined_estimate: gemini_overloaded "
                      f"[{_model_key_cb}] needed={_pf_cb.get('needed')} "
                      f"available={_pf_cb.get('available')}")
        except Exception as _ce_pf_err:
            print(f"[{job_id}] combined_estimate preflight error (non-fatal): {_ce_pf_err}")

    return jsonify({
        "gemini_eur": gemini_eur,
        "llm_eur": llm_eur,
        "total_eur": total,
        "is_free": total <= threshold,
        "threshold_eur": threshold,
        "rate_step": rate_step,
        "gemini_breakdown": gemini_breakdown,
        "llm_breakdown": llm_breakdown,
        "gemini_overloaded": overload_info is not None,
        "gemini_overload_info": overload_info,
        "paypal_available": _paypal_available(),
        "paypal_client_id": PAYPAL_CLIENT_ID if _paypal_available() else "",
        "paypal_mode": PAYPAL_MODE,
    })


@app.route("/api/paypal_create_order_gemini", methods=["POST"])
def api_paypal_create_order_gemini():
    """Create a PayPal order for Voci PREMIUM (+ optional AI text optimization).

    Server-side amount check: recomputes the combined estimate
    (gemini_eur + llm_eur) and rejects the request if the client-supplied
    amount differs by more than 0.01 EUR. On match, calls
    payment._paypal_create_order and returns {order_id, amount, status}.
    """
    import payment as _payment_mod
    import gemini_tts as _gemini_tts_mod

    data = request.get_json(silent=True) or {}
    job_id = (data.get("job_id") or "").strip()
    voice_id = (data.get("voice_id") or "").strip()
    selected = data.get("selected_chapters") or []
    ai_opt = bool(data.get("ai_opt_enabled", False))
    rate = data.get("rate", "+0%")
    ui_lang = (data.get("lang") or "").strip().split("-")[0].lower()
    try:
        requested_amount = float(data.get("amount_eur") or 0)
    except (TypeError, ValueError):
        return jsonify({"error": "invalid amount_eur"}), 400

    with _jobs_lock:
        job = jobs.get(job_id)
    if not job:
        return jsonify({"error": "job not found"}), 404
    info = job.get("info")
    if info is None or not getattr(info, "chapters", None):
        return jsonify({"error": "no chapters"}), 400

    all_chs = list(info.chapters)
    if selected:
        _by_index = {ch.index: ch for ch in all_chs}
        chs = [_by_index[i] for i in selected if i in _by_index]
    else:
        chs = all_chs
    if not chs:
        return jsonify({"error": "no chapters"}), 400

    # Cap caratteri PRIMA di creare l'ordine PayPal. Un libro che supera il cap
    # della voce PREMIUM (MAX_GEMINI_TEXT_CHARS, default 800k) non potra` mai
    # essere generato da /api/generate (cap a riga ~6448), quindi NON deve
    # nemmeno arrivare a creare/catturare un ordine: altrimenti l'utente paga
    # e il job viene poi rifiutato senza che il denaro sia stato consumato.
    _max_chars_voice = _effective_max_text_chars(voice_id, job)
    _sel_chars_voice = sum(getattr(ch, "char_count", 0) for ch in chs)
    if _sel_chars_voice > _max_chars_voice:
        return jsonify({
            "error": f"Selection too large: {_sel_chars_voice:,} characters "
                     f"(limit {_max_chars_voice:,}). Please reduce the chapter selection.",
            "error_code": "selection_too_large",
            "chars_selected": _sel_chars_voice,
            "chars_limit": _max_chars_voice,
        }), 413

    # Lingua: stessa priorita` di /api/combined_estimate (UI > metadata > "it").
    # Deve essere identica per evitare amount mismatch sul server-side check.
    lang = ui_lang or (getattr(info, "language", "") or "").split("-")[0].lower() or "it"

    gemini_eur = 0.0
    if voice_id.startswith("gemini:"):
        try:
            # rate_pct: la stima dipende dalla velocità scelta, quindi va
            # passata anche qui per coerenza con /api/combined_estimate.
            est = _gemini_tts_mod.estimate_book_cost(chs, voice_id, language=lang, rate_pct=rate)
        except Exception as e:
            return jsonify({"error": f"estimate failed: {e}"}), 500
        gemini_eur = round(est["user_price_eur"], 2)

    llm_eur = 0.0
    if ai_opt:
        chars = sum(len(getattr(c, "text", "") or "") for c in chs)
        rate = LLM_RATE_EUR_PER_MCHAR
        llm_eur = round((chars / 1_000_000.0) * rate, 2)

    server_total = round(gemini_eur + llm_eur, 2)

    if abs(server_total - requested_amount) > 0.01:
        return jsonify({
            "error": f"amount mismatch (server={server_total}, client={requested_amount})",
            "server_amount_eur": server_total,
            "client_amount_eur": requested_amount,
        }), 400

    book_title = getattr(info, "title", "") or "Audiobook"
    description = f"Audiobook Maker - Voci PREMIUM - {book_title[:60]}"
    try:
        order = _payment_mod._paypal_create_order(
            amount_eur=server_total,
            description=description,
            custom_id=f"gemini:{job_id}",
        )
    except Exception as e:
        print(f"[paypal] gemini create_order failed: {e}")
        return jsonify({"error": f"paypal create failed: {e}"}), 500

    return jsonify({
        "order_id": order.get("id"),
        "amount": server_total,
        "status": order.get("status"),
    })


@app.route("/api/optimize", methods=["POST"])
def api_optimize():
    if not _llm_available(): return jsonify({"error": "LLM optimization not available"}), 503
    data = request.json or {}; job_id = data.get("job_id"); batch = data.get("batch", False); auto_generate = data.get("auto_generate", False); email = (data.get("email") or "").strip().lower()
    lang = data.get("lang")
    job, err, sc = _check_job_owner(job_id)
    if err is not None:
        if sc == 404:
            return jsonify({"error": "Session expired"}), 400
        return err, sc
    info = job.get("info")

    # Check sospensione nuovi processi (admin toggle)
    if _suspend_new_jobs:
        return jsonify({"error": "System under maintenance. Please try again in a few minutes."}), 503

    # Lingua TTS selezionata in UI: e' la fonte autoritativa per scegliere
    # il prompt LLM (prompt_tts_<lang>.md), perche' l'ottimizzazione deve
    # produrre testo adatto alla voce TTS scelta, non alla lingua dell'input.
    # Es: input EN ottimizzato per voce IT -> serve prompt_tts_it.md.
    if lang:
        job["opt_lang"] = (lang.split("-")[0] or "").lower() or None

    client_id = job.get("client_id", "")
    # Atomic concurrency check + status claim for optimization
    with _jobs_lock:
        if job["status"] not in ("analyzed",):
            return jsonify({"error": "Optimization already running or completed."}), 400
        if client_id and MAX_CONCURRENT_LLM_PER_CLIENT > 0:
            if _active_optimizing_for_client_unlocked(client_id) >= MAX_CONCURRENT_LLM_PER_CLIENT:
                return jsonify({
                    "error": f"Concurrent optimization limit reached ({MAX_CONCURRENT_LLM_PER_CLIENT}).",
                    "error_code": "concurrent_optimize_limit",
                    "max": MAX_CONCURRENT_LLM_PER_CLIENT,
                    "active": _active_optimizing_for_client_unlocked(client_id),
                }), 429
        job["status"] = "optimizing"
    raw_selected = data.get("selected_chapters")
    selected_chapters = _parse_selected_chapters(raw_selected)
    already = set(job.get("optimized_chapters", []))
    if selected_chapters:
        chapters_to_optimize = [idx for idx in selected_chapters if idx not in already]
    else:
        chapters_to_optimize = [ch.index for ch in info.chapters if ch.index not in already] if info else []
    total_chars = sum(ch.char_count for ch in info.chapters if ch.index in chapters_to_optimize)
    print(f"[{job_id}] OPTIMIZE raw selected_chapters: {raw_selected!r} -> parsed: {selected_chapters!r} -> to_optimize: {chapters_to_optimize!r}")
    if not chapters_to_optimize:
        return jsonify({"status": "already_optimized", "optimized_chapters": list(already)})

    # Hard cap on text size for the final audio output, applied to the full
    # selected set (already-optimized + to-optimize). Blocks early so the
    # user doesn't pay for LLM optimization on a selection that cannot be
    # rendered to audio. Per auto_generate con voce PREMIUM si applica il
    # cap MAX_GEMINI_TEXT_CHARS (piu' restrittivo).
    max_text_chars = _max_text_chars_for_voice(data.get("voice", ""))
    if info is not None:
        if selected_chapters:
            selected_set_for_cap = set(selected_chapters)
        else:
            selected_set_for_cap = {ch.index for ch in info.chapters}
        selected_chars_total = sum(
            ch.char_count for ch in info.chapters if ch.index in selected_set_for_cap
        )
        if selected_chars_total > max_text_chars:
            # Release the "optimizing" status claimed above so the user can
            # retry with a smaller selection.
            with _jobs_lock:
                if job.get("status") == "optimizing":
                    job["status"] = "analyzed"
            return jsonify({
                "error": f"Selection too large: {selected_chars_total:,} characters "
                         f"(limit {max_text_chars:,}). Please reduce the chapter selection.",
                "error_code": "selection_too_large",
                "chars_selected": selected_chars_total,
                "chars_limit": max_text_chars,
            }), 413
    estimated_cost = _estimate_llm_cost_eur(total_chars)
    # Flusso combinato (auto_generate + voce Gemini): il payment_token
    # copre sia LLM che Gemini. Il branch standalone LLM sotto NON deve
    # consumarlo per la sola quota LLM — la gestione e` delegata al
    # blocco "Combined payment" piu` avanti.
    _is_combined_gemini = (auto_generate
                           and data.get("voice", "").startswith("gemini:")
                           and gemini_tts is not None)
    if estimated_cost > LLM_FREE_THRESHOLD_EUR and not _is_combined_gemini:
        payment_token = (data.get("payment_token") or "").strip()
        if not payment_token:
            return jsonify({
                "error": "Payment required for this optimization.",
                "error_code": "payment_required",
                "estimated_cost_eur": estimated_cost,
                "chars": total_chars,
            }), 402
        # Validate payment_token (PayPal order_id or voucher code).
        # Sec: check-and-set atomico sotto lock per impedire doppio uso del medesimo token
        # da parte di richieste concorrenti. La persistenza su disco avviene fuori dal lock
        # perché _save_payments() riacquisisce _payments_lock (non rientrante).
        valid = False
        if payment_token in payment._payments:
            _claimed_pay = None
            with payment._payments_lock:
                pay = payment._payments.get(payment_token)
                if pay and not pay.get("used") and pay.get("amount_eur", 0) >= estimated_cost:
                    pay["used"] = True
                    pay["used_at"] = time.time()
                    pay["used_job_id"] = job_id
                    _claimed_pay = pay
            if _claimed_pay is not None:
                _save_payments()
                job["payment_token"] = payment_token
                job["payment_type"] = "paypal"
                job["payment_email"] = _claimed_pay.get("email", "")
                job["payment_amount_eur"] = _claimed_pay.get("amount_eur", 0)
                valid = True
        elif payment_token in payment._vouchers:
            v = payment._vouchers[payment_token]
            remaining = _voucher_remaining(v)
            if v.get("expires_at", 0) > time.time() and remaining >= estimated_cost - 0.01:
                try:
                    new_remaining = _voucher_consume(payment_token, estimated_cost, job_id=job_id)
                except ValueError as _ve:
                    return jsonify({
                        "error": f"Voucher not spendable: {_ve}",
                        "error_code": "invalid_payment",
                    }), 402
                job["payment_token"] = payment_token
                job["payment_type"] = "voucher"
                job["payment_email"] = v.get("email", "")
                job["payment_amount_eur"] = round(float(estimated_cost), 2)
                job["voucher_remaining_after"] = new_remaining
                valid = True
        if not valid:
            return jsonify({
                "error": "Invalid or already-used payment token.",
                "error_code": "invalid_payment",
            }), 402

    # ----- Combined payment (LLM + Gemini in auto_generate flow) -----
    # Il flusso combinato usa UN unico token (PayPal order o voucher) che
    # copre entrambe le quote LLM + Gemini. Il branch standalone LLM sopra
    # e` stato saltato (_is_combined_gemini=True), quindi gestiamo il
    # consumo qui per l'intero ammontare. Vedi md_files/ttsgemini.md.
    if _is_combined_gemini:
        _combined_token = (data.get("payment_token_combined")
                           or data.get("payment_token") or "").strip()
        if _combined_token:
            # Ricalcolo quota Gemini server-side per validare l'importo.
            _voice_for_est = data.get("voice", "")
            _rate_for_est = data.get("rate", "+0%")
            _ui_lang_for_est = (lang or "").split("-")[0].lower() if lang else ""
            _lang_for_est = (_ui_lang_for_est
                             or (getattr(info, "language", "") or "").split("-")[0].lower()
                             or "it")
            # Capitoli selezionati per la generazione (stessa logica frontend
            # combined_estimate: subset se selected_chapters, altrimenti tutti).
            _all_chs = list(getattr(info, "chapters", []) or [])
            _sel_list = _parse_selected_chapters(data.get("selected_chapters"))
            if _sel_list:
                _by_idx = {ch.index: ch for ch in _all_chs}
                _chs_for_est = [_by_idx[i] for i in _sel_list if i in _by_idx]
            else:
                _chs_for_est = _all_chs
            try:
                _est_gemini = gemini_tts.estimate_book_cost(
                    _chs_for_est, _voice_for_est,
                    language=_lang_for_est, rate_pct=_rate_for_est,
                )
                _gemini_eur_quota = round(_est_gemini.get("user_price_eur", 0.0), 2)
            except Exception as _e_est:
                print(f"[{job_id}] combined-payment estimate failed: {_e_est}")
                _est_gemini = None
                _gemini_eur_quota = 0.0
            _expected_total = round(_gemini_eur_quota + estimated_cost, 2)
            # Soglia per richiedere il pagamento: ABM_GEMINI_FREE_THRESHOLD_EUR
            # (allineata con /api/generate). Sotto soglia il job e' free.
            _threshold_combined = float(
                os.environ.get("ABM_GEMINI_FREE_THRESHOLD_EUR", "0.50")
            )
            if _expected_total > _threshold_combined:
                if not _combined_token:
                    with _jobs_lock:
                        if job.get("status") == "optimizing":
                            job["status"] = "analyzed"
                    return jsonify({
                        "error": "Payment required for generation.",
                        "error_code": "payment_required",
                        "total_eur": _expected_total,
                        "gemini_eur": _gemini_eur_quota,
                        "llm_eur": estimated_cost,
                        "threshold_eur": _threshold_combined,
                    }), 402
                # Validazione + consume del token combinato.
                _consumed = False
                if _combined_token in payment._payments:
                    with payment._payments_lock:
                        _pay = payment._payments.get(_combined_token)
                        if (_pay and not _pay.get("used")
                                and float(_pay.get("amount_eur", 0)) + 0.05
                                >= _expected_total):
                            _pay["used"] = True
                            _pay["used_at"] = time.time()
                            _pay["used_job_id"] = job_id
                            _consumed_method = "paypal"
                            _consumed_email = _pay.get("email", "") or ""
                            _consumed = True
                    if _consumed:
                        _save_payments()
                elif _combined_token in payment._vouchers:
                    try:
                        payment._voucher_consume(_combined_token, _expected_total,
                                                 job_id=job_id)
                        _v = payment._vouchers.get(_combined_token, {})
                        _consumed_method = "voucher"
                        _consumed_email = _v.get("email", "") or ""
                        _consumed = True
                    except ValueError as _vc_err:
                        print(f"[{job_id}] combined voucher consume failed: {_vc_err}")
                if _consumed:
                    # Stash payment per:
                    # - audit Gemini (_write_gemini_audit legge job["payment"])
                    # - refund su cancel/error (_refund_gemini_payment)
                    # total_eur = quota Gemini. La quota LLM e` in payment["llm_eur"];
                    # l'audit Gemini legge solo la quota voce, non il combinato.
                    job["payment"] = {
                        "token": _combined_token,
                        "total_eur": _gemini_eur_quota,
                        "method": _consumed_method,
                        "ts": time.time(),
                        "gemini_est": _est_gemini,
                        "llm_eur": float(estimated_cost),
                        "source": "combined_optimize_autogen",
                    }
                    job["payment_token"] = _combined_token
                    job["payment_type"] = _consumed_method
                    job["payment_email"] = _consumed_email
                    job["payment_amount_eur"] = _expected_total
                    # Snapshot stima pre-LLM su job["gemini_estimate"]: serve
                    # all'audit per allineare i campi *_est al prezzo lockato
                    # in payment["total_eur"]. Senza questo snapshot,
                    # _finalize_optimization_complete ricalcolerebbe la stima
                    # su testo post-LLM (potenzialmente piu` lungo/corto),
                    # distorcendo delta_pct/margin nell'audit JSONL.
                    job["gemini_estimate"] = _est_gemini
                    print(f"[{job_id}] combined payment consumed at /api/optimize: "
                          f"gemini={_gemini_eur_quota:.2f}€ + llm={estimated_cost:.2f}€ "
                          f"= {_expected_total:.2f}€ ({_consumed_method})")
                else:
                    print(f"[{job_id}] WARNING: combined payment token "
                          f"{_combined_token[:12]}... not consumable "
                          f"(expected_total={_expected_total:.2f}€)")

    # Batch mode requires email
    if batch:
        import re as _re
        if not email or not _re.match(r'^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$', email):
            return jsonify({"error": "Valid email required for batch mode"}), 400
        if not _smtp_available():
            return jsonify({"error": "Email service not configured on this server"}), 503
        job["notify_email"] = email
        job["notify_lang"] = data.get("lang", "en")
        job["email_registered"] = True
        _write_email_pending_marker(UPLOAD_DIR / job_id)

    # Store auto-generate params for batch mode
    if auto_generate:
        job["opt_auto_generate"] = True
        job["opt_voice"] = data.get("voice", "it-IT-IsabellaNeural")
        job["opt_rate"] = data.get("rate", "+0%")
        job["opt_single_file"] = data.get("single_file", True)
        job["opt_output_format"] = data.get("output_format", "m4b")
        job["opt_podcast_base_url"] = (data.get("podcast_base_url") or "").strip()
        if job["opt_output_format"] == "zip_rss":
            job["notify_download_type"] = "podcast"
            job["notify_base_url"] = job["opt_podcast_base_url"]
        else:
            job["notify_download_type"] = "audio"
            job["notify_base_url"] = ""
    else:
        job["opt_auto_generate"] = False

    thread = threading.Thread(
        target=run_optimization, args=(job_id, chapters_to_optimize), daemon=True
    )
    thread.start()

    _log_activity(job_id, job.get("original_filename", ""), "OPTIMIZE",
                  client_id, job.get("client_ip", ""), "",
                  browser_lang=job.get("browser_lang", ""))

    return jsonify({"status": "started", "batch": batch, "auto_generate": auto_generate})


@app.route("/api/optimize_progress/<job_id>")
def api_optimize_progress(job_id):
    """SSE endpoint for LLM optimization progress."""
    _job_pre, _err_pre, _sc_pre = _check_job_owner(job_id)
    if _err_pre is not None:
        return _err_pre, _sc_pre
    def stream():
        while True:
            if job_id not in jobs:
                yield f"data: {json.dumps({'status': 'error', 'error': 'Job not found'})}\n\n"
                break
            job = jobs[job_id]
            job["last_poll"] = time.time()
            status = job.get("status", "unknown")
            payload = {
                "status": status,
                "opt_progress_current": job.get("opt_progress_current", 0),
                "opt_progress_total": job.get("opt_progress_total", 0),
                "opt_progress_message": job.get("opt_progress_message", ""),
                "opt_current_chapter": job.get("opt_current_chapter", ""),
                "opt_current_chapter_num": job.get("opt_current_chapter_num", 0),
                "opt_processed_chars": job.get("opt_processed_chars", 0),
                "opt_streamed_chars": job.get("opt_streamed_chars", 0),
                "opt_current_chapter_chars": job.get("opt_current_chapter_chars", 0),
                "opt_total_chars": job.get("opt_total_chars", 0),
                "opt_total_chars_extended": job.get("opt_total_chars_extended", job.get("opt_total_chars", 0)),
                "opt_elapsed_seconds": round(time.time() - job["opt_start_time"]) if job.get("opt_start_time") else job.get("opt_elapsed_seconds", 0),
            }
            if status == "error":
                # Errore generico verso il client; il dettaglio resta nei log server-side.
                payload["error"] = "optimization_failed"
                yield f"data: {json.dumps(payload)}\n\n"
                break
            if status == "cancelled" or job.get("opt_cancelled"):
                payload["status"] = "cancelled"
                yield f"data: {json.dumps(payload)}\n\n"
                break
            if status == "optimized":
                payload["ai_optimized"] = True
                yield f"data: {json.dumps(payload)}\n\n"
                break
            # If auto_generate kicked in, status is now "generating" or "done"
            if status in ("generating", "done"):
                payload["ai_optimized"] = True
                payload["auto_generate_started"] = True
                yield f"data: {json.dumps(payload)}\n\n"
                break
            yield f"data: {json.dumps(payload)}\n\n"
            time.sleep(2)

    return Response(
        stream_with_context(stream()),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.route("/api/cancel_optimize/<job_id>", methods=["POST"])
def api_cancel_optimize(job_id):
    """Cancel an LLM optimization in progress."""
    _job, _err, _sc = _check_job_owner(job_id)
    if _err is not None:
        if _sc == 404:
            return jsonify({"status": "not_found"}), 404
        return _err, _sc
    with _jobs_lock:
        if job_id in jobs:
            job = jobs[job_id]
            if job.get("status") == "optimizing":
                job["opt_cancelled"] = True
                return jsonify({"status": "cancelling"})
    return jsonify({"status": "not_found"}), 404


@app.route("/api/active_jobs")
def api_active_jobs():
    """Return list of currently generating jobs (for admin monitor).
    Protected by admin token.
    """
    if not ADMIN_TOKEN:
        return jsonify({"error": "Admin monitor disabled"}), 404
    if not _admin_auth_ok(_admin_auth_from_request()):
        return jsonify({"error": "Unauthorized"}), 401

    from datetime import datetime
    with _jobs_lock:
        snapshot = list(jobs.items())
    active = []
    for jid, job in snapshot:
        if job.get("status") in ("generating", "analyzed", "optimizing", "optimized"):
            info = job.get("info")
            title = ""
            if info:
                title = getattr(info, "title", "") or ""
            if not title:
                title = job.get("original_filename", jid)
            start_ts = job.get("start_time", 0)
            active.append({
                "title": title,
                "started": datetime.fromtimestamp(start_ts).strftime("%Y-%m-%d %H:%M:%S") if start_ts else " - ",
                "status": job.get("status", ""),
                "progress": job.get("progress_current", 0),
                "total": job.get("progress_total", 0),
                "chapter": job.get("current_chapter", ""),
            })
    return jsonify({"jobs": active, "count": len(active)})


@app.route("/dl/<token>")
def token_download_page(token):
    """Serve download page for email-linked token."""
    token_info = _download_tokens.get(token)
    if not token_info:
        return _render_dl_expired_page(), 410

    lang = token_info.get("lang", "en")
    created_at = token_info["created_at"]
    elapsed = time.time() - created_at
    # Retention per-token: PREMIUM (is_gemini) usa GEMINI_FILE_RETENTION_SEC.
    _ret = _retention_for_token_info(token_info)

    # Check retention expiration
    if elapsed > _ret:
        _download_tokens.pop(token, None)
        _save_tokens()
        return _render_dl_expired_page(lang, retention_hours=round(_ret / 3600)), 410

    # Check job exists in memory OR files still on disk
    job_id = token_info["job_id"]
    job_dir = UPLOAD_DIR / job_id
    dl_type = token_info.get("download_type", "audio")
    job_in_memory = job_id in jobs and jobs[job_id].get("status") in ("done", "optimized")
    files_on_disk = job_dir.exists()

    if not job_in_memory and not files_on_disk:
        _download_tokens.pop(token, None)
        _save_tokens()
        return _render_dl_expired_page(lang, retention_hours=round(_ret / 3600)), 410

    remaining_sec = max(60, int(_ret - elapsed))
    remaining_h = remaining_sec // 3600
    remaining_m = (remaining_sec % 3600) // 60
    if remaining_h > 0:
        remaining_str = f"~{remaining_h}h {remaining_m}min" if remaining_m > 0 else f"~{remaining_h}h"
    else:
        remaining_str = f"~{remaining_m} min"

    # Book title: from job in memory or from token snapshot
    if job_in_memory and jobs[job_id].get("info"):
        book_title = jobs[job_id]["info"].title or ""
    else:
        book_title = token_info.get("book_title", "")

    # M4B availability: regola di precedenza per non rompere l'isolamento
    # per-epoch (un token email punta a UNA specifica generazione, il job vivo
    # può aver prodotto altri file con epoche più nuove):
    #   1) token snapshot -> output_m4b (path assoluto)
    #   2) ricostruzione per basename dentro la cartella per-epoch del token
    #   3) (legacy) job vivo -> output_m4b
    #   4) (legacy) glob ricorsivo nella cartella del job, SOLO se il token
    #      NON ha info di epoch (token vecchi creati prima del layout per-epoch).
    m4b_available = False
    m4b_path_snap = token_info.get("output_m4b", "")
    if m4b_path_snap and os.path.exists(m4b_path_snap):
        m4b_available = True
    elif m4b_path_snap:
        # Per-epoch reconstruction: output_{epoch}/<basename>.m4b
        candidate = job_dir / Path(m4b_path_snap).parent.name / Path(m4b_path_snap).name
        if candidate.exists():
            m4b_available = True
    else:
        # Token legacy senza snapshot M4B: live job o glob ricorsivo come
        # last-resort. Non rompe l'isolamento per-epoch perche' arriva qui
        # solo se il token NON aveva originariamente un output_m4b.
        if job_in_memory:
            m4b_path_mem = jobs[job_id].get("output_m4b", "")
            if m4b_path_mem and os.path.exists(m4b_path_mem):
                m4b_available = True
        if not m4b_available:
            m4bs = list(job_dir.glob("**/*.m4b"))
            m4b_available = len(m4bs) > 0

    abm_path_snap = token_info.get("optimized_abm_path", "")
    has_abm = bool(abm_path_snap) and os.path.exists(abm_path_snap)
    if not has_abm and abm_path_snap:
        candidate = job_dir / Path(abm_path_snap).parent.name / Path(abm_path_snap).name
        if candidate.exists():
            has_abm = True

    output_format = token_info.get("output_format", "")
    if not output_format and job_in_memory:
        output_format = jobs[job_id].get("output_format", "")

    return _render_dl_page(token, book_title, remaining_str,
                           token_info["download_type"], lang,
                           m4b_available=m4b_available, has_abm=has_abm,
                           output_format=output_format,
                           retention_hours=round(_ret / 3600))


@app.route("/dl/<token>/abm")
def token_do_download_abm(token):
    """Serve the optimized .abm file for a token (when available)."""
    token_info = _download_tokens.get(token)
    if not token_info:
        return "Link scaduto", 410
    _ret = _retention_for_token_info(token_info)
    if time.time() - token_info["created_at"] > _ret:
        _download_tokens.pop(token, None)
        _save_tokens()
        return f"Link scaduto  -  i file sono stati cancellati dopo {_ret // 3600} ore", 410
    job_id = token_info.get("job_id", "")
    abm_name = token_info.get("optimized_abm_name", "optimized.abm")
    # Always serve the .abm captured in this token's snapshot. Each generation
    # epoch writes its own output_{epoch}/foo_optimized.abm; regenerating from
    # the live job state would overwrite with the latest cumulative selection
    # and break per-epoch isolation across sibling email tokens.
    abm_path = token_info.get("optimized_abm_path", "")
    if abm_path and not os.path.exists(abm_path) and job_id:
        job_dir = UPLOAD_DIR / job_id
        # Reconstruct by basename within the snapshot's epoch dir
        candidate = job_dir / Path(abm_path).parent.name / Path(abm_path).name
        if candidate.exists():
            abm_path = str(candidate)
        else:
            # Last-resort: legacy flat layout at work_dir root
            alt = job_dir / os.path.basename(abm_path)
            if alt.exists():
                abm_path = str(alt)
    if not abm_path or not os.path.exists(abm_path):
        return "File not available", 404
    if not _is_resume_or_probe_request():
        _log_activity(token_info.get("job_id", ""), token_info.get("original_filename", ""),
                      "DOWNLOAD_OPT_ABM", "", "", "", "")
    _mark_token_downloaded(token_info)
    return _send_file_throttled(abm_path, as_attachment=True, download_name=abm_name, no_cache=True)


@app.route("/dl/<token>/m4b")
def token_do_download_m4b(token):
    """Execute the actual M4B file download via token.

    Allineato a /api/download/<job>?type=m4b:
    - glob ricorsivo ("**/*.m4b") sulla job dir
    - fallback su MP3 con header X-Fallback se M4B non esiste
    - sync di job["output_m4b"] quando il file è trovato via glob
    """
    token_info = _download_tokens.get(token)
    if not token_info:
        return "Link scaduto", 410

    job_id = token_info["job_id"]
    _ret = _retention_for_token_info(token_info)
    if time.time() - token_info["created_at"] > _ret:
        _download_tokens.pop(token, None)
        _save_tokens()
        return f"Link scaduto  -  i file sono stati cancellati dopo {_ret // 3600} ore", 410

    # Per-epoch isolation: il token email punta a UNA specifica generazione.
    # Non usiamo MAI lo stato del job vivo (job["output_m4b"]) come fonte
    # primaria, perche' una rigenerazione successiva potrebbe averlo aggiornato
    # a un'epoch piu' nuova, esponendo file diversi al link email.
    job = jobs.get(job_id)
    if job:
        job["last_poll"] = time.time()
        job["downloaded_at"] = time.time()

    m4b_path = token_info.get("output_m4b", "")
    job_dir = UPLOAD_DIR / job_id

    # 1) Path reconstruction per-epoch: cerca il basename del file dentro la
    # cartella output_{epoch}/ catturata nello snapshot del token.
    if m4b_path and not os.path.exists(m4b_path):
        candidate = job_dir / Path(m4b_path).parent.name / Path(m4b_path).name
        if candidate.exists():
            m4b_path = str(candidate)

    # 2) Legacy fallback: per token CREATI PRIMA dell'introduzione del layout
    # per-epoch (snapshot privo di output_m4b), usiamo un glob ricorsivo come
    # ultima spiaggia. Non rompe l'isolamento: ci arriviamo solo se il token
    # non aveva originariamente un path snapshotato.
    if (not m4b_path) and (not token_info.get("output_m4b")):
        m4bs = list(job_dir.glob("**/*.m4b"))
        if m4bs:
            m4b_path = str(m4bs[0])

    safe_name = _safe_filename(token_info.get("book_title", "audiolibro"))

    if m4b_path and os.path.exists(m4b_path):
        if request.method != "HEAD" and not request.headers.get("Range"):
            _log_activity(job_id, token_info.get("original_filename", ""), "DOWNLOAD_M4B_TOKEN",
                          "", "", "", "")
        _mark_token_downloaded(token_info)
        return _send_file_throttled(m4b_path, as_attachment=True, download_name=f"{safe_name}.m4b")

    # Fallback MP3 (coerente con /api/download): l'M4B non c'è (conversione fallita
    # o non ancora pronta), serviamo l'MP3 segnalandolo al client via X-Fallback.
    print(f"[dl] M4B totally missing for job {job_id}. Falling back to MP3.")
    mp3_path = ""
    if job:
        mp3_path = (job.get("output_files") or [""])[0]
    if not mp3_path or not os.path.exists(mp3_path):
        mp3s = list(job_dir.glob("**/*.mp3"))
        if mp3s:
            mp3_path = str(mp3s[0])

    if mp3_path and os.path.exists(mp3_path):
        if request.method != "HEAD" and not request.headers.get("Range"):
            _log_activity(job_id, token_info.get("original_filename", ""), "DOWNLOAD_M4B_TOKEN_FALLBACK_MP3",
                          "", "", "", "")
        _mark_token_downloaded(token_info)
        resp = _send_file_throttled(mp3_path, as_attachment=True, download_name=f"{safe_name}.mp3")
        try:
            resp.headers["X-Fallback"] = "mp3"
            prev = resp.headers.get("Access-Control-Expose-Headers", "")
            resp.headers["Access-Control-Expose-Headers"] = (prev + ", X-Fallback").lstrip(", ")
        except Exception:
            pass
        return resp

    # Diagnostica completa: nessun M4B e nessun MP3 disponibile.
    print(f"[dl/m4b] 404 token={token} job={job_id} "
          f"token_m4b={token_info.get('output_m4b','')!r} "
          f"output_dirs={[d.name for d in _iter_output_dirs(job_dir)]}")
    return "M4B file not available", 404


@app.route("/dl/<token>/download")
def token_do_download(token):
    """Execute the actual file download via token."""
    token_info = _download_tokens.get(token)
    if not token_info:
        return "Link scaduto", 410

    job_id = token_info["job_id"]
    _ret = _retention_for_token_info(token_info)
    if time.time() - token_info["created_at"] > _ret:
        _download_tokens.pop(token, None)
        _save_tokens()
        return f"Link scaduto  -  i file sono stati cancellati dopo {_ret // 3600} ore", 410

    # Try to get data from job in memory, otherwise use token snapshot
    job = jobs.get(job_id)
    if job:
        job["last_poll"] = time.time()
        job["downloaded_at"] = time.time()

    dl_type = token_info.get("download_type", "audio")

    # Diagnostic logging
    job_dir = UPLOAD_DIR / job_id
    print(f"[dl] Token download: job={job_id}, type={dl_type}, "
          f"job_in_memory={job is not None}, "
          f"job_dir_exists={job_dir.exists()}, "
          f"stored_zip={token_info.get('output_zip', '')[:80]}, "
          f"UPLOAD_DIR={UPLOAD_DIR}")

    try:
        #  -  -  OPTIMIZED ABM download  -  -
        if dl_type == "optimized_abm":
            abm_name = token_info.get("optimized_abm_name", "optimized.abm")
            # Serve the .abm captured in this token's snapshot — see
            # /dl/<token>/abm comment for the per-epoch isolation rationale.
            abm_path = token_info.get("optimized_abm_path", "")
            if abm_path and not os.path.exists(abm_path):
                cand = job_dir / Path(abm_path).parent.name / Path(abm_path).name
                if cand.exists():
                    abm_path = str(cand)
                else:
                    alt = job_dir / os.path.basename(abm_path)
                    if alt.exists():
                        abm_path = str(alt)
            if abm_path and os.path.exists(abm_path):
                if job:
                    job["downloaded_at"] = time.time()
                if not _is_resume_or_probe_request():
                    _log_activity(job_id, token_info.get("original_filename", ""),
                                  "DOWNLOAD_OPT_ABM", "", "", "", "")
                _mark_token_downloaded(token_info)
                return _apply_no_cache(send_file(abm_path, as_attachment=True, download_name=abm_name))
            return "File not found", 404

        #  -  -  PODCAST download  -  - 
        is_podcast = dl_type == "podcast" and (
            (job and job.get("podcast_ready")) or token_info.get("podcast_ready"))

        if is_podcast:
            return _serve_podcast_download(token_info, job, job_id)

        #  -  -  AUDIO download  -  - 
        return _serve_audio_download(token_info, job, job_id)

    except Exception as e:
        print(f"[dl/{token}] ERROR in download: {e}")
        import traceback
        traceback.print_exc()
        return f"Errore durante il download. Riprova tra qualche istante.", 500


def _serve_audio_download(token_info, job, job_id):
    """Serve audio download bound to this token's epoch.

    The token snapshot holds absolute paths into `output_{epoch}/`. We always
    try those first; only fall back to live job state or directory scans if
    the snapshot path is genuinely gone (e.g. data-dir migration). This keeps
    sibling email tokens from leaking each other's files.
    """
    output_name = token_info.get("output_name", "audiobook.zip")
    orig = token_info.get("original_filename", "")
    job_dir = UPLOAD_DIR / job_id

    def _do_log():
        if _is_resume_or_probe_request():
            return
        _log_activity(job_id, orig, "DOWNLOAD_EMAIL",
                      job.get("client_id", "") if job else "",
                      job.get("client_ip", "") if job else "",
                      job.get("voice", "") if job else "",
                      job.get("browser_lang", "") if job else "")
        # Disattiva la protezione no-download per voci PREMIUM (cleanup loop).
        _mark_token_downloaded(token_info)

    output_zip = token_info.get("output_zip", "")
    output_file = token_info.get("output_file", "")

    # 1. Exact paths from token snapshot
    if output_zip and os.path.exists(output_zip):
        _do_log()
        return _send_file_throttled(output_zip, as_attachment=True, download_name=output_name)
    if output_file and os.path.exists(output_file):
        _do_log()
        return _send_file_throttled(output_file, as_attachment=True, download_name=output_name)

    # 2. Path reconstruction within the snapshot's epoch dir (data-dir moved)
    for p in (output_zip, output_file):
        if not p:
            continue
        cand = job_dir / Path(p).parent.name / Path(p).name
        if cand.exists():
            print(f"[dl] Path reconstructed: {p} -> {cand}")
            _do_log()
            return _send_file_throttled(str(cand), as_attachment=True, download_name=output_name)

    # 3. Live job state (only when snapshot path missing — older runs may have
    #    been cleaned up; we still try to serve *something* for this job).
    if job:
        orig = job.get("original_filename", orig)
        if job.get("output_zip") and os.path.exists(job["output_zip"]):
            print(f"[dl] Snapshot missing; falling back to live job output_zip")
            _do_log()
            return _send_file_throttled(job["output_zip"], as_attachment=True,
                             download_name=job.get("output_name", output_name))
        if job.get("output_files") and os.path.exists(job["output_files"][0]):
            print(f"[dl] Snapshot missing; falling back to live job output_files[0]")
            _do_log()
            return _send_file_throttled(job["output_files"][0], as_attachment=True,
                             download_name=job.get("output_name", output_name))

    # 4. Fallback: scan job directory for downloadable files
    if job_dir.exists():
        print(f"[dl] Scanning {job_dir} for downloadable files...")
        zips = sorted(job_dir.glob("*.zip")) + sorted(_find_files_in_outputs(job_dir, "*.zip"))
        # Exclude podcast zips
        zips = [z for z in zips if "_podcast" not in z.name]
        if zips:
            found = str(zips[0])
            print(f"[dl] Fallback: found ZIP {found}")
            _do_log()
            return _send_file_throttled(found, as_attachment=True,
                             download_name=output_name or os.path.basename(found))
        # Look for MP3s across all output dirs, then root
        mp3s = sorted(_find_files_in_outputs(job_dir, "*.mp3"))
        if not mp3s:
            mp3s = sorted(job_dir.glob("*.mp3"))
        if len(mp3s) == 1:
            found = str(mp3s[0])
            print(f"[dl] Fallback: found single MP3 {found}")
            _do_log()
            return _send_file_throttled(found, as_attachment=True,
                             download_name=output_name or os.path.basename(found))
        elif len(mp3s) > 1:
            # Multiple MP3s: create a ZIP on the fly
            src_dir = str(mp3s[0].parent)
            zip_file = shutil.make_archive(str(job_dir / "download"), "zip", src_dir)
            print(f"[dl] Fallback: created ZIP from {len(mp3s)} MP3s -> {zip_file}")
            _do_log()
            return _send_file_throttled(zip_file, as_attachment=True,
                             download_name=output_name or "audiobook.zip")

    print(f"[dl] No files found for job {job_id} (job_dir exists: {job_dir.exists()})")
    print(f"[dl]   stored output_zip: {output_zip}")
    print(f"[dl]   stored output_file: {output_file}")
    print(f"[dl]   UPLOAD_DIR: {UPLOAD_DIR}")
    return "File non più disponibili", 410

    print(f"[dl] No files found for job {job_id}")
    return "File non più disponibili", 410


def _generate_podcast_index_html(podcast_dir, title, author, cover_file, rss_fname, mp3_files, language="en"):
    """Generate an index.html landing page for the podcast folder (required by Netlify)."""
    lang = language[:2] if language else "en"
    _labels = {
        "it": {"heading": "Podcast", "by": "di", "subscribe": "Iscriviti al Podcast",
               "copy": "Copia URL feed", "copied": "Copiato!",
               "episodes": "Episodi", "listen": "Ascolta",
               "instructions": "Copia l'URL del feed RSS e incollalo nella tua app podcast preferita (Pocket Casts, Apple Podcasts, AntennaPod, Overcast...).",
               "footer": "Generato con Audiobook Maker"},
        "en": {"heading": "Podcast", "by": "by", "subscribe": "Subscribe to Podcast",
               "copy": "Copy feed URL", "copied": "Copied!",
               "episodes": "Episodes", "listen": "Listen",
               "instructions": "Copy the RSS feed URL and paste it in your favorite podcast app (Pocket Casts, Apple Podcasts, AntennaPod, Overcast...).",
               "footer": "Generated with Audiobook Maker"},
        "fr": {"heading": "Podcast", "by": "de", "subscribe": "S'abonner au Podcast",
               "copy": "Copier l'URL du flux", "copied": "Copié !",
               "episodes": "Épisodes", "listen": "Écouter",
               "instructions": "Copiez l'URL du flux RSS et collez-la dans votre app podcast (Pocket Casts, Apple Podcasts, AntennaPod, Overcast...).",
               "footer": "Généré avec Audiobook Maker"},
        "es": {"heading": "Podcast", "by": "de", "subscribe": "Suscríbete al Podcast",
               "copy": "Copiar URL del feed", "copied": "¡Copiado!",
               "episodes": "Episodios", "listen": "Escuchar",
               "instructions": "Copia la URL del feed RSS y pégala en tu app de podcast favorita (Pocket Casts, Apple Podcasts, AntennaPod, Overcast...).",
               "footer": "Generado con Audiobook Maker"},
        "de": {"heading": "Podcast", "by": "von", "subscribe": "Podcast abonnieren",
               "copy": "Feed-URL kopieren", "copied": "Kopiert!",
               "episodes": "Episoden", "listen": "Anhören",
               "instructions": "Kopieren Sie die RSS-Feed-URL und fügen Sie sie in Ihre Podcast-App ein (Pocket Casts, Apple Podcasts, AntennaPod, Overcast...).",
               "footer": "Erstellt mit Audiobook Maker"},
        "zh": {"heading": "播客", "by": "作者", "subscribe": "订阅播客",
               "copy": "复制订阅URL", "copied": "已复制！",
               "episodes": "章节", "listen": "收听",
               "instructions": "复制RSS订阅URL并将其粘贴到您的播客应用程序中（Pocket Casts，Apple Podcasts，AntennaPod，Overcast...）。",
               "footer": "由Audiobook Maker生成"},
        "hi": {"heading": "पॉडकास्ट", "by": "लेखक", "subscribe": "पॉडकास्ट सब्सक्राइब करें",
               "copy": "फ़ीड URL कॉपी करें", "copied": "कॉपी हो गया!",
               "episodes": "एपिसोड", "listen": "सुनें",
               "instructions": "RSS फ़ीड URL कॉपी करें और इसे अपने पसंदीदा पॉडकास्ट ऐप (Pocket Casts, Apple Podcasts, AntennaPod, Overcast...) में पेस्ट करें।",
               "footer": "Audiobook Maker से जनरेट किया गया"},
    }
    lb = _labels.get(lang, _labels["en"])

    # Build episode list
    sorted_mp3 = sorted([os.path.basename(f) for f in mp3_files if os.path.exists(f)])
    episodes_html = ""
    for i, mp3 in enumerate(sorted_mp3, 1):
        display_name = mp3.rsplit(".", 1)[0].replace("_", " ").replace("-", " ")
        episodes_html += f'<tr><td style="padding:10px 12px;border-bottom:1px solid #eee;color:#666;width:40px;text-align:center">{i}</td><td style="padding:10px 12px;border-bottom:1px solid #eee">{display_name}</td><td style="padding:10px 12px;border-bottom:1px solid #eee;text-align:right"><a href="{mp3}" style="color:#2c7bb6;text-decoration:none">&#9654; {lb["listen"]}</a></td></tr>'

    cover_tag = ""
    if cover_file:
        cover_tag = f'<img src="{cover_file}" alt="Cover" style="width:200px;height:200px;object-fit:cover;border-radius:12px;box-shadow:0 4px 20px rgba(0,0,0,.15)">'

    safe_title = (title or "Audiobook").replace('"', '&quot;').replace('<', '&lt;')
    safe_author = (author or "").replace('"', '&quot;').replace('<', '&lt;')

    html = f'''<!DOCTYPE html>
<html lang="{lang}">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{safe_title} - {lb["heading"]}</title>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,sans-serif;background:#f5f7fa;color:#333;line-height:1.6}}
.hero{{background:linear-gradient(135deg,#1a3c5e 0%,#2c7bb6 100%);color:#fff;padding:50px 20px 40px;text-align:center}}
.hero h1{{font-size:1.8rem;margin:16px 0 4px}}
.hero .author{{opacity:.8;font-size:1rem}}
.container{{max-width:680px;margin:0 auto;padding:20px}}
.feed-box{{background:#fff;border-radius:12px;padding:24px;margin:-30px auto 24px;box-shadow:0 2px 12px rgba(0,0,0,.08);position:relative;z-index:1}}
.feed-box h2{{font-size:1.1rem;margin-bottom:8px;color:#1a3c5e}}
.feed-url{{display:flex;gap:8px;margin:12px 0}}
.feed-url input{{flex:1;padding:10px 14px;border:1px solid #ddd;border-radius:8px;font-family:monospace;font-size:.85rem;background:#f8f8f8;color:#333}}
.feed-url button{{padding:10px 20px;background:#2c7bb6;color:#fff;border:none;border-radius:8px;cursor:pointer;font-weight:600;font-size:.85rem;white-space:nowrap;transition:background .2s}}
.feed-url button:hover{{background:#1a5a8a}}
.instructions{{font-size:.88rem;color:#666;margin-top:8px}}
.episodes{{background:#fff;border-radius:12px;padding:24px;box-shadow:0 2px 12px rgba(0,0,0,.08);margin-bottom:24px}}
.episodes h2{{font-size:1.1rem;margin-bottom:16px;color:#1a3c5e}}
.episodes table{{width:100%;border-collapse:collapse}}
.footer{{text-align:center;color:#aaa;font-size:.8rem;padding:20px}}
</style>
</head>
<body>
<div class="hero">
{cover_tag}
<h1>{safe_title}</h1>
{f'<div class="author">{lb["by"]} {safe_author}</div>' if safe_author else ''}
</div>
<div class="container">
<div class="feed-box">
<h2>&#x1F399;&#xFE0F; {lb["subscribe"]}</h2>
<div class="feed-url">
<input type="text" id="feedUrl" value="{rss_fname}" readonly onclick="this.select()">
<button onclick="copyFeed()">{lb["copy"]}</button>
</div>
<div class="instructions">{lb["instructions"]}</div>
</div>
<div class="episodes">
<h2>{lb["episodes"]} ({len(sorted_mp3)})</h2>
<table>{episodes_html}</table>
</div>
<div class="footer">{lb["footer"]}</div>
</div>
<script>
function copyFeed(){{
  const inp=document.getElementById('feedUrl');
  navigator.clipboard.writeText(inp.value).then(()=>{{
    const btn=document.querySelector('.feed-url button');
    btn.textContent='{lb["copied"]}';
    setTimeout(()=>btn.textContent='{lb["copy"]}',2000);
  }});
}}
// Update feed URL with full path on load
window.addEventListener('load',()=>{{
  const inp=document.getElementById('feedUrl');
  const base=window.location.href.replace(/\\/[^\\/]*$/,'/');
  inp.value=base+'{rss_fname}';
}});
</script>
</body>
</html>'''

    index_path = os.path.join(str(podcast_dir), "index.html")
    with open(index_path, "w", encoding="utf-8") as f:
        f.write(html)
    return index_path


def _serve_podcast_download(token_info, job, job_id):
    """Serve podcast download from job in memory or token snapshot on disk."""

    # If output ZIP already has RSS embedded (zip_rss format), serve it directly
    output_zip = token_info.get("output_zip", "")
    if output_zip and os.path.exists(output_zip):
        # Check if RSS is embedded (job in memory) or trust the zip_rss output
        if (job and job.get("podcast_rss_included")) or not job:
            print(f"[dl] Podcast: serving existing ZIP with embedded RSS: {output_zip}")
            if job:
                job["last_poll"] = time.time()
                job["downloaded_at"] = time.time()
            orig = token_info.get("original_filename", job.get("original_filename", "") if job else "")
            if not _is_resume_or_probe_request():
                _log_activity(job_id, orig, "DOWNLOAD_EMAIL_PODCAST",
                              job.get("client_id", "") if job else "",
                              job.get("client_ip", "") if job else "",
                              job.get("voice", "") if job else "", job.get("browser_lang", "") if job else "")
            _mark_token_downloaded(token_info)
            return _send_file_throttled(output_zip, as_attachment=True,
                             download_name=os.path.basename(output_zip))

    base_url = token_info.get("base_url", "")

    # Always source paths from the token snapshot (per-epoch isolation). The
    # live job state reflects the LATEST generation only and would cause
    # sibling email tokens to serve each other's files.
    mp3_files = token_info.get("podcast_mp3s", [])
    safe_name = token_info.get("podcast_safe_name", "audiolibro")
    epub_path = token_info.get("epub_path", "")
    p_info_title = token_info.get("podcast_info_title", "")
    p_info_author = token_info.get("podcast_info_author", "")
    p_info_lang = token_info.get("podcast_info_language", "")

    # Reconstruct epub_path if stored path doesn't exist (data dir may have changed)
    if epub_path and not os.path.exists(epub_path):
        reconstructed = str(UPLOAD_DIR / job_id / os.path.basename(epub_path))
        if os.path.exists(reconstructed):
            print(f"[dl] epub_path reconstructed: {epub_path} -> {reconstructed}")
            epub_path = reconstructed

    # Verify MP3 files exist; fallback: reconstruct paths under any output_{epoch}/ dir
    mp3_files = [f for f in mp3_files if os.path.exists(f)]
    if not mp3_files:
        job_dir = UPLOAD_DIR / job_id
        raw_mp3s = token_info.get("podcast_mp3s", [])
        for d in _iter_output_dirs(job_dir):
            candidates = []
            for old_path in raw_mp3s:
                cand = d / os.path.basename(old_path)
                if cand.exists():
                    candidates.append(str(cand))
            if candidates:
                mp3_files = candidates
                print(f"[dl] Podcast path reconstruction: {len(mp3_files)} MP3s found in {d}")
                break
    if not mp3_files:
        # Final fallback: scan all output dirs for any MP3s (newest first)
        job_dir = UPLOAD_DIR / job_id
        for d in _iter_output_dirs(job_dir):
            found = sorted([str(f) for f in d.glob("*.mp3")])
            if found:
                mp3_files = found
                print(f"[dl] Podcast scan fallback: found {len(mp3_files)} MP3s in {d}")
                break
    if not mp3_files:
        return "File non più disponibili", 410

    # Create a minimal info object for RSS generation
    # Use real info object when job is in memory (has chapters for RSS titles),
    # otherwise create minimal stub for token-based downloads
    if job and job.get("podcast_info"):
        info = job["podcast_info"]
    else:
        class _MiniInfo:
            pass
        info = _MiniInfo()
        info.title = p_info_title
        info.author = p_info_author
        info.language = p_info_lang
        info.chapters = []  # No chapter objects available; RSS will use "Episode N" fallback

    # Place cached & temp podcast artifacts inside this epoch's output dir so
    # sibling email tokens don't share a single zip at work_dir root.
    epoch_dir = Path(mp3_files[0]).parent if mp3_files else UPLOAD_DIR / job_id
    work_dir = epoch_dir

    # If a podcast zip was already built for this epoch, serve it directly
    cached_zip = epoch_dir / f"{safe_name}_podcast.zip"
    if cached_zip.exists() and cached_zip.stat().st_size > 0:
        print(f"[dl] Serving cached podcast zip: {cached_zip}")
        _mark_token_downloaded(token_info)
        return _send_file_throttled(str(cached_zip), as_attachment=True,
                         download_name=f"{safe_name}_podcast.zip")

    # Build podcast package in a unique temp dir to avoid race conditions
    import uuid as _uuid
    podcast_dir = epoch_dir / f"podcast_{_uuid.uuid4().hex[:8]}"
    podcast_dir.mkdir(parents=True, exist_ok=True)
    try:
        for mp3 in mp3_files:
            if os.path.exists(mp3):
                shutil.copy2(mp3, str(podcast_dir / os.path.basename(mp3)))
        cover_file = ""
        cover_path = str(podcast_dir / "cover.jpg")
        if epub_path and os.path.exists(epub_path):
            if _extract_cover_from_epub(epub_path, cover_path, target_size=1400):
                cover_file = "cover.jpg"
            else:
                raw_path, raw_mime = _extract_cover_for_preview(epub_path, str(podcast_dir))
                if raw_path and os.path.exists(raw_path):
                    ext = ".png" if raw_mime == "image/png" else ".jpg"
                    final_cover = str(podcast_dir / ("cover" + ext))
                    if raw_path != final_cover:
                        shutil.move(raw_path, final_cover)
                    cover_file = "cover" + ext
                else:
                    _generate_fallback_cover(cover_path, title=info.title or "", author=info.author or "")
                    if os.path.exists(cover_path) and os.path.getsize(cover_path) > 0:
                        cover_file = "cover.jpg"
        rss_fname = f"{safe_name}_podcast.xml"
        rss_path = str(podcast_dir / rss_fname)
        _generate_podcast_rss(info, mp3_files, rss_path,
                              base_url=base_url, cover_filename=cover_file,
                              rss_filename=rss_fname)
        _generate_podcast_index_html(podcast_dir, info.title, info.author,
                                     cover_file, rss_fname, mp3_files,
                                     language=getattr(info, 'language', '') or token_info.get('language', 'en'))
        podcast_zip = shutil.make_archive(
            str(work_dir / f"{safe_name}_podcast"), "zip", str(podcast_dir))
    finally:
        shutil.rmtree(str(podcast_dir), ignore_errors=True)
    orig = token_info.get("original_filename", job.get("original_filename", "") if job else "")
    if not _is_resume_or_probe_request():
        _log_activity(job_id, orig, "DOWNLOAD_EMAIL_PODCAST",
                      job.get("client_id", "") if job else "",
                      job.get("client_ip", "") if job else "",
                      job.get("voice", "") if job else "", job.get("browser_lang", "") if job else "")
    _mark_token_downloaded(token_info)
    return _send_file_throttled(podcast_zip, as_attachment=True,
                     download_name=f"{safe_name}_podcast.zip")


def _render_dl_expired_page(lang="en", retention_hours=0):
    expired_t = _DL_PAGES_I18N.get("expired", {})
    t = expired_t.get(lang, expired_t.get("en", {}))
    # Se il chiamante non passa la retention reale del token (es. token gia`
    # rimosso e non recuperabile), usa il default standard come fallback.
    # Vale come "almeno X ore sono passate"; per token Gemini il caller passa 48.
    if not retention_hours:
        retention_hours = int(EMAIL_FILE_RETENTION_SEC / 3600)
    p1_text = t['p1'].replace("{h}", str(int(retention_hours)))
    return f"""<!DOCTYPE html><html lang="{lang}"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<link rel="icon" type="image/svg+xml" href="{FAVICON_B64}">
<title>Audiobook Maker  -  {t['title']}</title>
<style>
body{{font-family:system-ui,-apple-system,sans-serif;display:flex;justify-content:center;
align-items:center;min-height:100vh;margin:0;background:#f8f9fa;color:#333}}
.box{{text-align:center;padding:48px;max-width:500px;background:white;border-radius:16px;
box-shadow:0 4px 24px rgba(0,0,0,.08)}}
h1{{font-size:3rem;margin:0 0 16px}}
h2{{color:#e74c3c;margin:0 0 16px}}
p{{color:#666;line-height:1.6}}
a{{color:#3b82f6;text-decoration:none;font-weight:600}}
a:hover{{text-decoration:underline}}
</style></head><body>
<div class="box">
<h1>&#x23F0;</h1>
<h2>{t['h2']}</h2>
<p>{p1_text}</p>
<p>{t['p2']}</p>
<p><a href="/">&#x1F3A7; Audiobook Maker</a></p>
</div></body></html>"""


def _render_dl_deleted_page(lang="en"):
    deleted_t = _DL_PAGES_I18N.get("deleted", {})
    t = deleted_t.get(lang, deleted_t.get("en", {}))
    return f"""<!DOCTYPE html><html lang="{lang}"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<link rel="icon" type="image/svg+xml" href="{FAVICON_B64}">
<title>Audiobook Maker  -  {t['title']}</title>
<style>
body{{font-family:system-ui,-apple-system,sans-serif;display:flex;justify-content:center;
align-items:center;min-height:100vh;margin:0;background:#f8f9fa;color:#333}}
.box{{text-align:center;padding:48px;max-width:500px;background:white;border-radius:16px;
box-shadow:0 4px 24px rgba(0,0,0,.08)}}
h1{{font-size:3rem;margin:0 0 16px}}
h2{{color:#e74c3c;margin:0 0 16px}}
p{{color:#666;line-height:1.6}}
a{{color:#3b82f6;text-decoration:none;font-weight:600}}
a:hover{{text-decoration:underline}}
</style></head><body>
<div class="box">
<h1>&#x1F5D1;&#xFE0F;</h1>
<h2>{t['h2']}</h2>
<p>{t['p1']}</p>
<p>{t['p2']}</p>
<p><a href="/">&#x1F3A7; Audiobook Maker</a></p>
</div></body></html>"""


def _render_dl_cooldown_page(lang="en", seconds=60, back_url=None):
    cooldown_t = _DL_PAGES_I18N.get("cooldown", {})
    t = cooldown_t.get(lang, cooldown_t.get("en", {}))
    p2 = t['p2'].format(s=seconds)
    auto_refresh_ms = max(seconds + 2, 5) * 1000
    # Al termine del countdown torniamo alla pagina con i bottoni di download (non al
    # file URL che ha generato il cooldown — altrimenti il timer riparte all'infinito).
    if back_url is None:
        try:
            path = request.path or ""
            m = re.match(r'^(/dl/[^/]+)(?:/.*)?$', path)
            if m:
                back_url = m.group(1)
            else:
                ref = request.referrer or ""
                same_origin = ref.startswith(request.host_url) if ref else False
                back_url = ref if same_origin else "/"
        except Exception:
            back_url = "/"
    import json as _json
    back_url_js = _json.dumps(back_url)
    return f"""<!DOCTYPE html><html lang="{lang}"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<link rel="icon" type="image/svg+xml" href="{FAVICON_B64}">
<title>Audiobook Maker  -  {t['title']}</title>
<style>
body{{font-family:system-ui,-apple-system,sans-serif;display:flex;justify-content:center;
align-items:center;min-height:100vh;margin:0;background:#f8f9fa;color:#333}}
.box{{text-align:center;padding:48px;max-width:500px;background:white;border-radius:16px;
box-shadow:0 4px 24px rgba(0,0,0,.08)}}
h1{{font-size:3rem;margin:0 0 16px}}
h2{{color:#f39c12;margin:0 0 16px}}
p{{color:#666;line-height:1.6}}
a{{color:#3b82f6;text-decoration:none;font-weight:600}}
a:hover{{text-decoration:underline}}
.countdown{{font-size:2.5rem;font-weight:700;color:#f39c12;margin:16px 0}}
</style></head><body>
<div class="box">
<h1>&#x23F3;</h1>
<h2>{t['h2']}</h2>
<p>{t['p1']}</p>
<p>{p2}</p>
<div class="countdown" id="cd">{seconds}</div>
<p style="font-size:0.85rem;color:#999">{t['auto']}</p>
<p style="margin-top:16px"><a href={back_url_js}>&#x1F3A7; Audiobook Maker</a></p>
</div>
<script>
(function(){{
  var backUrl = {back_url_js};
  var s={seconds},el=document.getElementById('cd');
  var iv=setInterval(function(){{
    s--;if(s<=0){{clearInterval(iv);location.href=backUrl;return;}}
    el.textContent=s;
  }},1000);
  setTimeout(function(){{location.href=backUrl}},{auto_refresh_ms});
}})();
</script>
</body></html>"""


def _render_dl_page(token, book_title, remaining_str, dl_type, lang="en", m4b_available=False, has_abm=False, output_format="", retention_hours=0):
    download_t = _DL_PAGES_I18N.get("download", {})
    t = dict(download_t.get(lang, download_t.get("en", {})))

    # Sec (XSS): book_title proviene dai metadati EPUB/PDF (controllati dall'autore del file).
    # Tutte le interpolazioni nel template devono passare per html.escape, altrimenti un
    # `<dc:title>` malevolo iniettato lato uploader produce XSS sulla pagina /dl/<token>.
    import html as _html
    book_title = _html.escape(str(book_title or ""), quote=True)
    remaining_str = _html.escape(str(remaining_str or ""), quote=True)

    # Single audio button matching post-generation page style
    # SVG download icon for single-file formats, emoji for ZIP

    audio_btn_html = ""
    type_label = ""

    _format_labels = {
        "m4b": {
            "it": ('<svg class="btn-icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round" width="20" height="20"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg> <span>Scarica audiolibro (M4B)</span>', "Audiobook (M4B)"),
            "en": ('<svg class="btn-icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round" width="20" height="20"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg> <span>Download audiobook (M4B)</span>', "Audiobook (M4B)"),
            "fr": ('<svg class="btn-icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round" width="20" height="20"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg> <span>T&eacute;l&eacute;charger l&rsquo;audiobook (M4B)</span>', "Audiobook (M4B)"),
            "es": ('<svg class="btn-icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round" width="20" height="20"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg> <span>Descargar audiolibro (M4B)</span>', "Audiobook (M4B)"),
            "de": ('<svg class="btn-icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round" width="20" height="20"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg> <span>H&ouml;rbuch herunterladen (M4B)</span>', "Audiobook (M4B)"),
            "zh": ('<svg class="btn-icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round" width="20" height="20"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg> <span>下载有声读物 (M4B)</span>', "Audiobook (M4B)"),
            "hi": ('<svg class="btn-icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round" width="20" height="20"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg> <span>ऑडियोबुक डाउनलोड करें (M4B)</span>', "Audiobook (M4B)"),
        },
        "mp3": {
            "it": ('<svg class="btn-icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round" width="20" height="20"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg> <span>Scarica audiolibro (MP3)</span>', "Audiobook (MP3)"),
            "en": ('<svg class="btn-icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round" width="20" height="20"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg> <span>Download audiobook (MP3)</span>', "Audiobook (MP3)"),
            "fr": ('<svg class="btn-icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round" width="20" height="20"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg> <span>T&eacute;l&eacute;charger l&rsquo;audiobook (MP3)</span>', "Audiobook (MP3)"),
            "es": ('<svg class="btn-icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round" width="20" height="20"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg> <span>Descargar audiolibro (MP3)</span>', "Audiobook (MP3)"),
            "de": ('<svg class="btn-icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round" width="20" height="20"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg> <span>H&ouml;rbuch herunterladen (MP3)</span>', "Audiobook (MP3)"),
            "zh": ('<svg class="btn-icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round" width="20" height="20"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg> <span>下载有声读物 (MP3)</span>', "Audiobook (MP3)"),
            "hi": ('<svg class="btn-icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round" width="20" height="20"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg> <span>ऑडियोबुक डाउनलोड करें (MP3)</span>', "Audiobook (MP3)"),
        },
        "zip": {
            "it": ("&#x2B07;&#xFE0F; <span>Scarica audiolibro (ZIP)</span>", "Audiobook (ZIP)"),
            "en": ("&#x2B07;&#xFE0F; <span>Download audiobook (ZIP)</span>", "Audiobook (ZIP)"),
            "fr": ("&#x2B07;&#xFE0F; <span>T&eacute;l&eacute;charger l&rsquo;audiobook (ZIP)</span>", "Audiobook (ZIP)"),
            "es": ("&#x2B07;&#xFE0F; <span>Descargar audiolibro (ZIP)</span>", "Audiobook (ZIP)"),
            "de": ("&#x2B07;&#xFE0F; <span>H&ouml;rbuch herunterladen (ZIP)</span>", "Audiobook (ZIP)"),
            "zh": ("&#x2B07;&#xFE0F; <span>下载有声读物 (ZIP)</span>", "Audiobook (ZIP)"),
            "hi": ("&#x2B07;&#xFE0F; <span>ऑडियोबुक डाउनलोड करें (ZIP)</span>", "Audiobook (ZIP)"),
        },
    }

    if dl_type == "optimized_abm":
        type_label = "Optimized Project (.abm)"
    elif dl_type == "podcast":
        # Podcast: ZIP with chapter MP3s + RSS
        audio_btn_html = '<a href="/dl/{}/download" class="btn">{}</a>'.format(
            token, t.get("btn_no_m4b", "&#x2B07;&#xFE0F; Download podcast"))
        type_label = "Podcast"
    elif dl_type in ("audio", "chapters"):
        # Determine format: prefer output_format from job, fallback to m4b detection.
        # "chapters" è inviato dal frontend per gli output multi-file (ZIP per capitoli):
        # forza fmt="zip" per saltare la rilevazione M4B quando il formato esplicito
        # mancasse dal token (es. token persistiti prima del fix di _save_tokens).
        fmt = output_format if output_format in ("m4b", "mp3", "zip", "zip_rss") else None
        if not fmt and dl_type == "chapters":
            fmt = "zip"
        elif not fmt and m4b_available:
            fmt = "m4b"
        elif not fmt:
            fmt = "zip"

        # Coerenza con la realtà del filesystem: se l'utente aveva chiesto M4B ma
        # il file non esiste (es. conversione PCM->AAC fallita su Gemini, oppure
        # output_format snapshottato come 'm4b' senza che il muxing sia avvenuto),
        # degradiamo l'etichetta/route a MP3 anziché esporre un link M4B che il
        # backend dovrà servire via fallback silenzioso.
        if fmt == "m4b" and not m4b_available:
            fmt = "mp3"

        if fmt == "m4b":
            label_data = _format_labels["m4b"]
            btn_url = f"/dl/{token}/m4b"
        elif fmt == "mp3":
            label_data = _format_labels["mp3"]
            btn_url = f"/dl/{token}/download"
        else:
            label_data = _format_labels["zip"]
            btn_url = f"/dl/{token}/download"

        btn_label, type_label = label_data.get(lang, label_data.get("en", label_data.get("it")))
        audio_btn_html = f'<p><a href="{btn_url}" class="btn">{btn_label}</a></p>'

    # ABM button (only if AI optimization was active)
    abm_btn_html = ""
    if has_abm:
        _abm_labels = {
            "it": "&#x1F4DD;&#xFE0F; Scarica testo ottimizzato (.abm)",
            "en": "&#x1F4DD;&#xFE0F; Download optimized text (.abm)",
            "fr": "&#x1F4DD;&#xFE0F; T&eacute;l&eacute;charger le texte optimis&eacute; (.abm)",
            "es": "&#x1F4DD;&#xFE0F; Descargar texto optimizado (.abm)",
            "de": "&#x1F4DD;&#xFE0F; Optimierten Text herunterladen (.abm)",
            "zh": "&#x1F4DD;&#xFE0F; 下载优化文本 (.abm)",
            "hi": "&#x1F4DD;&#xFE0F; ऑप्टिमाइज़्ड टेक्स्ट डाउनलोड करें (.abm)",
        }
        abm_label = _abm_labels.get(lang, _abm_labels["en"])
        abm_btn_html = f'<p><a href="/dl/{token}/abm" class="btn btn-abm">{abm_label}</a></p>'

    # Retention totale: deve riflettere _ret reale del token (es. Gemini=48h,
    # standard=18h), non un valore hardcoded. Senza questo, la riga "Dopo X ore"
    # mostrava sempre "24" anche quando il countdown sopra reportava ~48h.
    warn_text = t["warn"].replace("{r}", remaining_str).replace("{h}", str(int(retention_hours)))

    share_url = BASE_URL or "https://audiobook-maker.com"
    share_text_js = t.get("share_text", "").replace("\\", "\\\\").replace("'", "\\'").replace('"', '\\"')

    return f"""<!DOCTYPE html><html lang="{lang}"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<link rel="icon" type="image/svg+xml" href="{FAVICON_B64}">
<title>Audiobook Maker - {t['title']}</title>
<style>
body{{font-family:system-ui,-apple-system,sans-serif;display:flex;justify-content:center;
align-items:center;min-height:100vh;margin:0;background:#f8f9fa;color:#333}}
.box{{text-align:center;padding:48px;max-width:500px;background:white;border-radius:16px;
box-shadow:0 4px 24px rgba(0,0,0,.08)}}
h1{{font-size:3rem;margin:0 0 16px}}
h2{{color:#2c3e50;margin:0 0 8px}}
.title{{color:#666;font-style:italic;margin:0 0 24px}}
.btn{{display:inline-flex;align-items:center;justify-content:center;gap:8px;padding:16px 32px;background:#3b82f6;color:white;
text-decoration:none;border-radius:8px;font-weight:600;font-size:18px;
transition:background .2s;border:none;cursor:pointer}}
.btn:hover{{background:#2563eb}}
.btn.is-loading{{opacity:.85;cursor:wait;pointer-events:none}}
.btn .spin{{display:inline-block;width:16px;height:16px;border:2px solid currentColor;border-top-color:transparent;border-radius:50%;animation:spin .7s linear infinite;vertical-align:-2px;opacity:.85}}
@keyframes spin{{to{{transform:rotate(360deg)}}}}
.dl-toast{{position:fixed;top:18px;left:50%;transform:translateX(-50%) translateY(-8px);background:#1f2937;color:#fff;padding:10px 18px;border-radius:8px;font-size:.9rem;font-weight:500;box-shadow:0 6px 22px rgba(0,0,0,.18);opacity:0;transition:opacity .25s,transform .25s;z-index:9999;pointer-events:none}}
.dl-toast.show{{opacity:1;transform:translateX(-50%) translateY(0)}}
.dl-toast .spin{{margin-right:8px;border-color:rgba(255,255,255,.4);border-top-color:transparent}}
.btn-abm{{background:rgba(196,122,42,.10);color:#b8804a;border:1px solid #d4b68c;padding:15px 28px;font-size:1rem;font-weight:600;max-width:300px;margin-top:12px;white-space:normal;word-break:keep-all}}
.btn-abm:hover{{background:rgba(196,122,42,.14);color:#c47a2a;border-color:#c47a2a}}
.btn-icon{{vertical-align:middle;margin-right:4px}}
.warn{{color:#e74c3c;font-weight:600;margin-top:24px;font-size:.9rem}}
.type{{display:inline-block;padding:4px 12px;background:#e8f4f8;border-radius:12px;
font-size:.85rem;color:#2980b9;margin-bottom:16px}}
.share-row{{margin-top:28px;padding-top:20px;border-top:1px solid #eee;text-align:center}}
.share-label{{font-size:.85rem;color:#999;margin-bottom:12px}}
.share-icons{{display:flex;justify-content:center;gap:12px;flex-wrap:wrap}}
.share-icons a,.share-icons button{{width:40px;height:40px;border-radius:50%;display:inline-flex;
align-items:center;justify-content:center;border:1px solid #ddd;background:#f8f9fa;color:#666;
cursor:pointer;transition:all .2s;text-decoration:none;padding:0}}
.share-icons a:hover,.share-icons button:hover{{border-color:currentColor;
transform:translateY(-2px);box-shadow:0 3px 10px rgba(0,0,0,.08)}}
.share-icons svg{{width:20px;height:20px;fill:currentColor}}
#shX{{color:#14171a}}#shFb{{color:#1877F2}}#shWa{{color:#25D366}}#shTg{{color:#26A5E4}}#shLi{{color:#0A66C2}}#shRd{{color:#FF4500}}
.share-copy-wrap{{position:relative;display:inline-flex}}
.share-copied{{position:absolute;bottom:calc(100% + 6px);left:50%;transform:translateX(-50%);
background:#333;color:#fff;font-size:.72rem;padding:3px 8px;border-radius:4px;
white-space:nowrap;pointer-events:none;opacity:0;transition:opacity .2s}}
.share-copied.show{{opacity:1}}
.donate-panel{{margin-top:20px;padding:18px 20px;background:linear-gradient(135deg,#fffaf4,#fff3e0);
border:1px solid #e8c99a;border-radius:12px;text-align:center}}
.donate-title{{font-size:.97rem;font-weight:700;color:#2c2a26;margin-bottom:6px}}
.donate-body{{font-size:.82rem;color:#6b6760;line-height:1.5;margin-bottom:14px}}
.donate-btns{{display:grid;grid-template-columns:1fr 1fr;gap:10px}}
.donate-btn{{display:flex;align-items:center;justify-content:center;gap:6px;padding:10px;border-radius:8px;
font-size:.85rem;font-weight:600;text-decoration:none;transition:all .2s;border:1.5px solid transparent}}
.donate-coffee{{background:#ffdd00;color:#1a1400;border-color:#e5c800}}
.donate-coffee:hover{{background:#ffd000;transform:translateY(-2px);box-shadow:0 4px 12px rgba(255,208,0,.4)}}
.donate-paypal{{background:#003087;color:#fff;border-color:#002070}}
.donate-paypal:hover{{background:#002070;transform:translateY(-2px);box-shadow:0 4px 12px rgba(0,48,135,.35)}}
@media (max-width:520px){{
  body{{align-items:flex-start;padding:12px 0}}
  .box{{padding:24px 16px;border-radius:12px;max-width:100%;margin:0 8px}}
  h1{{font-size:2.4rem}}
  h2{{font-size:1.25rem;line-height:1.3}}
  .title{{font-size:.85rem;word-break:break-word;overflow-wrap:anywhere}}
  .btn{{padding:14px 18px;font-size:16px;width:100%;box-sizing:border-box;white-space:normal;line-height:1.25}}
  .btn-abm{{padding:13px 16px;font-size:.92rem;max-width:100%;width:100%;box-sizing:border-box}}
  .donate-panel{{padding:14px 14px;margin-top:18px}}
  .donate-title{{font-size:.95rem;line-height:1.3}}
  .donate-body{{font-size:.78rem}}
  .donate-btns{{grid-template-columns:1fr;gap:8px}}
  .donate-btn{{padding:12px 14px;font-size:.9rem;width:100%;box-sizing:border-box}}
}}
</style></head><body>
<div class="box">
<h1>&#x1F3A7;</h1>
<h2>{t['h2']}</h2>
<p class="title">{book_title}</p>
<p class="type">{type_label}</p>
{audio_btn_html}
{abm_btn_html}
<div class="warn">{warn_text}</div>
<div class="donate-panel">
  <div class="donate-title" id="donTitle"></div>
  <div class="donate-body"><span id="donBody"></span> <b id="donBodyBold"></b></div>
  <div class="donate-btns">
    <a href="https://buymeacoffee.com/audiobookmaker" target="_blank" rel="noopener" class="donate-btn donate-coffee">☕ <span id="donCoffee"></span></a>
    <a href="https://www.paypal.com/paypalme/gfrangiamone" target="_blank" rel="noopener" class="donate-btn donate-paypal">💙 <span id="donPaypal"></span></a>
  </div>
</div>
<div class="share-row">
<div class="share-label">{t['share']}</div>
<div class="share-icons">
  <a id="shX" href="#" target="_blank" rel="noopener" title="X / Twitter"><svg viewBox="0 0 24 24"><path d="M18.244 2.25h3.308l-7.227 8.26 8.502 11.24H16.17l-5.214-6.817L4.99 21.75H1.68l7.73-8.835L1.254 2.25H8.08l4.713 6.231zm-1.161 17.52h1.833L7.084 4.126H5.117z"/></svg></a>
  <a id="shFb" href="#" target="_blank" rel="noopener" title="Facebook"><svg viewBox="0 0 24 24"><path d="M24 12.073c0-6.627-5.373-12-12-12s-12 5.373-12 12c0 5.99 4.388 10.954 10.125 11.854v-8.385H7.078v-3.47h3.047V9.43c0-3.007 1.792-4.669 4.533-4.669 1.312 0 2.686.235 2.686.235v2.953H15.83c-1.491 0-1.956.925-1.956 1.874v2.25h3.328l-.532 3.47h-2.796v8.385C19.612 23.027 24 18.062 24 12.073z"/></svg></a>
  <a id="shWa" href="#" target="_blank" rel="noopener" title="WhatsApp"><svg viewBox="0 0 24 24"><path d="M17.472 14.382c-.297-.149-1.758-.867-2.03-.967-.273-.099-.471-.148-.67.15-.197.297-.767.966-.94 1.164-.173.199-.347.223-.644.075-.297-.15-1.255-.463-2.39-1.475-.883-.788-1.48-1.761-1.653-2.059-.173-.297-.018-.458.13-.606.134-.133.298-.347.446-.52.149-.174.198-.298.298-.497.099-.198.05-.371-.025-.52-.075-.149-.669-1.612-.916-2.207-.242-.579-.487-.5-.669-.51-.173-.008-.371-.01-.57-.01-.198 0-.52.074-.792.372-.272.297-1.04 1.016-1.04 2.479 0 1.462 1.065 2.875 1.213 3.074.149.198 2.096 3.2 5.077 4.487.709.306 1.262.489 1.694.625.712.227 1.36.195 1.871.118.571-.085 1.758-.719 2.006-1.413.248-.694.248-1.289.173-1.413-.074-.124-.272-.198-.57-.347m-5.421 7.403h-.004a9.87 9.87 0 01-5.031-1.378l-.361-.214-3.741.982.998-3.648-.235-.374a9.86 9.86 0 01-1.51-5.26c.001-5.45 4.436-9.884 9.888-9.884 2.64 0 5.122 1.03 6.988 2.898a9.825 9.825 0 012.893 6.994c-.003 5.45-4.437 9.884-9.885 9.884m8.413-18.297A11.815 11.815 0 0012.05 0C5.495 0 .16 5.335.157 11.892c0 2.096.547 4.142 1.588 5.945L.057 24l6.305-1.654a11.882 11.882 0 005.683 1.448h.005c6.554 0 11.89-5.335 11.893-11.893a11.821 11.821 0 00-3.48-8.413z"/></svg></a>
  <a id="shTg" href="#" target="_blank" rel="noopener" title="Telegram"><svg viewBox="0 0 24 24"><path d="M11.944 0A12 12 0 000 12a12 12 0 0012 12 12 12 0 0012-12A12 12 0 0012 0zm4.962 7.224c.1-.002.321.023.465.14a.506.506 0 01.171.325c.016.093.036.306.02.472-.18 1.898-.962 6.502-1.36 8.627-.168.9-.499 1.201-.82 1.23-.696.065-1.225-.46-1.9-.902-1.056-.693-1.653-1.124-2.678-1.8-1.185-.78-.417-1.21.258-1.91.177-.184 3.247-2.977 3.307-3.23.007-.032.014-.15-.056-.212s-.174-.041-.249-.024c-.106.024-1.793 1.14-5.061 3.345-.479.33-.913.49-1.302.48-.428-.008-1.252-.241-1.865-.44-.752-.245-1.349-.374-1.297-.789.027-.216.325-.437.893-.663 3.498-1.524 5.83-2.529 6.998-3.014 3.332-1.386 4.025-1.627 4.476-1.635z"/></svg></a>
  <a id="shLi" href="#" target="_blank" rel="noopener" title="LinkedIn"><svg viewBox="0 0 24 24"><path d="M20.447 20.452h-3.554v-5.569c0-1.328-.027-3.037-1.852-3.037-1.853 0-2.136 1.445-2.136 2.939v5.667H9.351V9h3.414v1.561h.046c.477-.9 1.637-1.85 3.37-1.85 3.601 0 4.267 2.37 4.267 5.455v6.286zM5.337 7.433a2.062 2.062 0 01-2.063-2.065 2.064 2.064 0 112.063 2.065zm1.782 13.019H3.555V9h3.564v11.452zM22.225 0H1.771C.792 0 0 .774 0 1.729v20.542C0 23.227.792 24 1.771 24h20.451C23.2 24 24 23.227 24 22.271V1.729C24 .774 23.2 0 22.222 0h.003z"/></svg></a>
  <a id="shRd" href="#" target="_blank" rel="noopener" title="Reddit"><svg viewBox="0 0 24 24"><path d="M12 0C5.373 0 0 5.373 0 12c0 5.084 3.163 9.426 7.627 11.174-.105-.949-.2-2.405.042-3.441.218-.937 1.407-5.965 1.407-5.965s-.359-.719-.359-1.782c0-1.668.967-2.914 2.171-2.914 1.023 0 1.518.769 1.518 1.69 0 1.029-.655 2.568-.994 3.995-.283 1.194.599 2.169 1.777 2.169 2.133 0 3.772-2.249 3.772-5.495 0-2.873-2.064-4.882-5.012-4.882-3.414 0-5.418 2.561-5.418 5.207 0 1.031.397 2.138.893 2.738a.36.36 0 01.083.345l-.333 1.36c-.053.22-.174.267-.402.161-1.499-.698-2.436-2.889-2.436-4.649 0-3.785 2.75-7.262 7.929-7.262 4.163 0 7.398 2.967 7.398 6.931 0 4.136-2.607 7.464-6.227 7.464-1.216 0-2.359-.631-2.75-1.378l-.748 2.853c-.271 1.043-1.002 2.35-1.492 3.146C9.57 23.812 10.763 24 12 24c6.627 0 12-5.373 12-12 0-6.628-5.373-12-12-12z"/></svg></a>
  <div class="share-copy-wrap">
    <button id="shCopy" title="Copy link"><svg viewBox="0 0 24 24"><path d="M16 1H4c-1.1 0-2 .9-2 2v14h2V3h12V1zm3 4H8c-1.1 0-2 .9-2 2v14c0 1.1.9 2 2 2h11c1.1 0 2-.9 2-2V7c0-1.1-.9-2-2-2zm0 16H8V7h11v14z"/></svg></button>
    <span class="share-copied" id="shCopiedTip">{t['copied']}</span>
  </div>
</div>
</div>
</div>
<script>
(function(){{
  /* ── Donate i18n (browser language) ── */
  var DL={{
    it:{{title:'\u2764\ufe0f Ti \u00e8 stato utile questo strumento?',body:'Audiobook Maker \u00e8 un\u2019app gratuita, non richiede iscrizione ed \u00e8 senza pubblicit\u00e0. Una donazione aiuta a coprire i costi di esercizio e lo sviluppo di nuove funzionalit\u00e0.',bodyBold:'Per donazioni da \u20ac5 o pi\u00f9, riceverai un coupon di valore equivalente da utilizzare per i servizi PREMIUM.',coffee:'Offrimi un caff\u00e8',paypal:'Donazione PayPal'}},
    fr:{{title:'\u2764\ufe0f Cet outil vous a \u00e9t\u00e9 utile\u00a0?',body:'Audiobook Maker est un logiciel open source, gratuit, sans inscription et sans publicit\u00e9. Un don aide \u00e0 couvrir les frais d\u2019exploitation et le d\u00e9veloppement de nouvelles fonctionnalit\u00e9s.',bodyBold:'Pour les dons de 5\u20ac ou plus, vous recevrez un coupon de valeur \u00e9quivalente \u00e0 utiliser pour les services PREMIUM.',coffee:'Offrez-moi un caf\u00e9',paypal:'Don PayPal'}},
    es:{{title:'\u2764\ufe0f \u00bfTe ha resultado \u00fatil esta herramienta?',body:'Audiobook Maker es software open source, gratuito, sin registro y sin publicidad. Una donaci\u00f3n ayuda a cubrir los costes operativos y el desarrollo de nuevas funciones.',bodyBold:'Para donaciones de 5\u20ac o m\u00e1s, recibir\u00e1s un cup\u00f3n de valor equivalente para usar en los servicios PREMIUM.',coffee:'Inv\u00edtame a un caf\u00e9',paypal:'Donaci\u00f3n PayPal'}},
    de:{{title:'\u2764\ufe0f War dieses Tool n\u00fctzlich f\u00fcr dich?',body:'Audiobook Maker ist Open Source, kostenlos, ohne Registrierung und werbefrei. Eine Spende hilft, die Betriebskosten und die Entwicklung neuer Funktionen zu decken.',bodyBold:'F\u00fcr Spenden ab 5\u20ac erhalten Sie einen Gutschein im gleichen Wert f\u00fcr PREMIUM-Dienste.',coffee:'Kauf mir einen Kaffee',paypal:'PayPal-Spende'}},
    zh:{{title:'\u2764\ufe0f \u8fd9\u4e2a\u5de5\u5177\u5bf9\u60a8\u6709\u5e2e\u52a9\u5417\uff1f',body:'Audiobook Maker \u662f\u5f00\u6e90\u8f6f\u4ef6\uff0c\u514d\u8d39\u3001\u65e0\u9700\u6ce8\u518c\u4e14\u65e0\u5e7f\u544a\u3002\u6350\u8d60\u6709\u52a9\u4e8e\u652f\u4ed8\u8fd0\u8425\u6210\u672c\u548c\u65b0\u529f\u80fd\u7684\u5f00\u53d1\u3002',bodyBold:'\u6350\u8d605\u6b27\u5143\u6216\u4ee5\u4e0a\uff0c\u60a8\u5c06\u83b7\u5f97\u7b49\u503c\u4f18\u60e0\u5238\u7528\u4e8e\u9ad8\u7ea7\u670d\u52a1\u3002',coffee:'\u8bf7\u6211\u559D\u5496\u5561',paypal:'PayPal \u6350\u6b3e'}},
    en:{{title:'\u2764\ufe0f Did you find this tool useful?',body:'Audiobook Maker is open source, free, no registration required and ad-free. A donation helps cover operating costs and the development of new features.',bodyBold:'For donations of \u20ac5 or more, you\u2019ll receive a coupon of equal value to use for PREMIUM services.',coffee:'Buy me a coffee',paypal:'PayPal donation'}},
    hi:{{title:'\u2764\ufe0f \u0915\u094d\u092f\u093e \u092f\u0939 \u091f\u0942\u0932 \u0906\u092a\u0915\u0947 \u0932\u093f\u090f \u0909\u092a\u092f\u094b\u0917\u0940 \u0930\u0939\u093e?',body:'Audiobook Maker \u090f\u0915 \u0913\u092a\u0928 \u0938\u094b\u0930\u094d\u0938 \u0910\u092a \u0939\u0948, \u092e\u0941\u092b\u094d\u0924, \u092c\u093f\u0928\u093e \u092a\u0902\u091c\u0940\u0915\u0930\u0923 \u0914\u0930 \u092c\u093f\u0928\u093e \u0935\u093f\u091c\u094d\u091e\u093e\u092a\u0928. \u0910\u0915 \u0926\u093e\u0928 \u0938\u0902\u091a\u093e\u0932\u0928 \u0932\u093e\u0917\u0924 \u0914\u0930 \u0928\u0908 \u0938\u0941\u0935\u093f\u0927\u093e\u0913\u0902 \u0915\u0947 \u0935\u093f\u0915\u093e\u0938 \u092e\u0947\u0902 \u092e\u0926\u0926 \u0915\u0930\u0924\u093e \u0939\u0948.',bodyBold:'\u20ac5 \u092f\u093e \u0905\u0927\u093f\u0915 \u0915\u0947 \u0926\u093e\u0928 \u092a\u0930, \u0906\u092a\u0915\u094b PREMIUM \u0938\u0947\u0935\u093e\u0913\u0902 \u0915\u0947 \u0932\u093f\u090f \u0938\u092e\u093e\u0928 \u092e\u0942\u0932\u094d\u092f \u0915\u093e \u0915\u0942\u092a\u0928 \u092e\u093f\u0932\u0947\u0917\u093e.',coffee:'\u092e\u0941\u091d\u0947 \u0915\u0949\u092b\u093c\u0940 \u092a\u093f\u0932\u093e\u090f\u0902',paypal:'PayPal \u0926\u093e\u0928'}}
  }};
  var bl=(navigator.language||navigator.userLanguage||'en').toLowerCase().split('-')[0];
  var d=DL[bl]||DL['en'];
  document.getElementById('donTitle').textContent=d.title;
  document.getElementById('donBody').textContent=d.body;
  document.getElementById('donBodyBold').textContent=d.bodyBold;
  document.getElementById('donCoffee').textContent=d.coffee;
  document.getElementById('donPaypal').textContent=d.paypal;
  /* ── Share links ── */
  var S='{share_url}';
  var T='{share_text_js}';
  var u=encodeURIComponent(S);
  var tx=encodeURIComponent(T);
  var f=encodeURIComponent(T+' '+S);
  document.getElementById('shX').href='https://x.com/intent/tweet?text='+tx+'&url='+u;
  document.getElementById('shFb').href='https://www.facebook.com/sharer/sharer.php?u='+u;
  document.getElementById('shWa').href='https://wa.me/?text='+f;
  document.getElementById('shTg').href='https://t.me/share/url?url='+u+'&text='+tx;
  document.getElementById('shLi').href='https://www.linkedin.com/sharing/share-offsite/?url='+u;
  document.getElementById('shRd').href='https://www.reddit.com/submit?url='+u+'&title='+tx;
  document.getElementById('shCopy').onclick=function(){{
    navigator.clipboard.writeText(S).then(function(){{
      var tip=document.getElementById('shCopiedTip');
      tip.classList.add('show');
      setTimeout(function(){{tip.classList.remove('show')}},2000);
    }});
  }};
  /* ── Download feedback: spinner sul bottone + toast (file grandi possono richiedere alcuni secondi) ── */
  var DLS={{
    it:{{preparing:'Preparazione in corso…',hint:'Il download partirà a breve. Per file grandi può richiedere qualche secondo.'}},
    en:{{preparing:'Preparing download…',hint:'Your download will start shortly. Large files may take a few seconds.'}},
    fr:{{preparing:'Préparation du téléchargement…',hint:'Le téléchargement va commencer. Les fichiers volumineux peuvent prendre quelques secondes.'}},
    es:{{preparing:'Preparando la descarga…',hint:'La descarga comenzará en breve. Los archivos grandes pueden tardar unos segundos.'}},
    de:{{preparing:'Download wird vorbereitet…',hint:'Der Download startet gleich. Bei großen Dateien kann es einige Sekunden dauern.'}},
    zh:{{preparing:'正在准备下载…',hint:'下载即将开始，大文件可能需要几秒钟。'}},
    hi:{{preparing:'डाउनलोड तैयार हो रहा है…',hint:'डाउनलोड जल्द शुरू होगा। बड़ी फ़ाइलों में कुछ सेकंड लग सकते हैं।'}}
  }};
  var dls=DLS[bl]||DLS['en'];
  var toast=document.createElement('div');
  toast.className='dl-toast';
  document.body.appendChild(toast);
  var toastTimer=null;
  function showToast(msg){{
    toast.innerHTML='<span class="spin"></span>'+msg;
    requestAnimationFrame(function(){{toast.classList.add('show')}});
    clearTimeout(toastTimer);
    toastTimer=setTimeout(function(){{toast.classList.remove('show')}},6500);
  }}
  document.querySelectorAll('a.btn').forEach(function(a){{
    a.addEventListener('click',function(){{
      if(a.classList.contains('is-loading'))return;
      var origHtml=a.innerHTML;
      a.classList.add('is-loading');
      a.innerHTML='<span class="spin"></span><span>'+dls.preparing+'</span>';
      showToast(dls.hint);
      setTimeout(function(){{
        a.classList.remove('is-loading');
        a.innerHTML=origHtml;
      }},9000);
    }});
  }});
}})();
</script></body></html>"""


@app.route("/api/download/<job_id>")
def api_download(job_id):
    job, err, sc = _check_job_owner(job_id)
    if err is not None:
        return ("Job not found" if sc == 404 else "Forbidden"), sc
    if job.get("status") != "done":
        return "Not ready", 400

    # Necessario per i fallback fisici (ricerca file su disco). In assenza di
    # questa definizione il ramo M4B->MP3 sollevava NameError -> HTTP 500.
    job_dir = UPLOAD_DIR / job_id

    download_type = request.args.get("type", "").lower()
    
    # Refresh heartbeat  -  evita che il cleanup rimuova il job durante il download
    job["last_poll"] = time.time()
    job["downloaded_at"] = time.time()
    
    log_type = "DOWNLOAD"
    if download_type == "m4b":
        log_type = "DOWNLOAD_M4B"
    elif download_type == "zip":
        log_type = "DOWNLOAD_ZIP"
    elif download_type == "abm":
        log_type = "DOWNLOAD_ABM"

    def _do_log():
        if _is_resume_or_probe_request():
            return
        _log_activity(job_id, job.get("original_filename", ""), log_type,
                      job.get("client_id", ""), job.get("client_ip", ""),
                      job.get("voice", ""), job.get("browser_lang", ""))

    if download_type == "abm":
        # Always regenerate from cumulative in-memory state to avoid stale chapters
        if job.get("ai_optimized"):
            try:
                opt_ch = job.get("optimized_chapters", [])
                print(f"[{job_id}] ABM download: regenerating (optimized_chapters={opt_ch})")
                abm_path, abm_name = generation_engine._generate_optimized_abm(job_id)
                job["optimized_abm_path"] = abm_path
                job["optimized_abm_name"] = abm_name
                _do_log()
                return _send_file_throttled(abm_path, as_attachment=True, download_name=abm_name, no_cache=True, bypass_throttle=True)
            except Exception as e:
                print(f"[{job_id}] On-demand ABM generation failed: {e}")
        # Fallback: serve existing file if any (pre-regeneration or non-AI-optimized)
        abm_path = job.get("optimized_abm_path")
        if abm_path and os.path.exists(abm_path):
            _do_log()
            safe_name = _safe_filename(job["info"].title) or "progetto"
            return _send_file_throttled(abm_path, as_attachment=True, download_name=f"{safe_name}.abm", no_cache=True, bypass_throttle=True)
        return "Optimized ABM project file not found", 404

    if download_type == "m4b":
        m4b_path = job.get("output_m4b")
        print(f"[debug] Download M4B requested. Path in job: {m4b_path}")

        if m4b_path and os.path.exists(m4b_path):
            _do_log()
            safe_name = _safe_filename(job["info"].title) or "audiolibro"
            print(f"[debug] M4B file found! Serving: {m4b_path}")
            return _send_file_throttled(m4b_path, as_attachment=True, download_name=f"{safe_name}.m4b", no_cache=True, bypass_throttle=True)
        else:
            # Physical search fallback: SOLO dentro output_dir della run corrente,
            # mai uno scan globale su tutti gli output_*/ (servirebbe un m4b di
            # una run precedente).
            print(f"[debug] M4B not found at registered path. Searching in current output_dir...")
            cur_out = job.get("output_dir")
            m4b_files = []
            if cur_out and os.path.isdir(cur_out):
                m4b_files = list(Path(cur_out).glob("*.m4b"))
            if m4b_files:
                actual_m4b = str(m4b_files[0])
                job["output_m4b"] = actual_m4b
                print(f"[debug] M4B found via physical search: {actual_m4b}")
                _do_log()
                safe_name = _safe_filename(job["info"].title) or "audiolibro"
                return _send_file_throttled(actual_m4b, as_attachment=True, download_name=f"{safe_name}.m4b", no_cache=True, bypass_throttle=True)

            print(f"[debug] M4B totally missing. Falling back to MP3.")
            # Fallback to single MP3 if M4B is missing
            # output_files puo' essere [] (assembly fallito): evita IndexError.
            _out_files = job.get("output_files") or []
            mp3_path = _out_files[0] if _out_files else ""
            if not mp3_path or not os.path.exists(mp3_path):
                mp3s = list(job_dir.glob("**/*.mp3"))
                if mp3s:
                    mp3_path = str(mp3s[0])
            if mp3_path and os.path.exists(mp3_path):
                resp = _send_file_throttled(mp3_path, as_attachment=True,
                                            download_name=f"{_safe_filename(job['info'].title)}.mp3",
                                            no_cache=True, bypass_throttle=True)
                try:
                    resp.headers["X-Fallback"] = "mp3"
                    prev = resp.headers.get("Access-Control-Expose-Headers", "")
                    resp.headers["Access-Control-Expose-Headers"] = (prev + ", X-Fallback").lstrip(", ")
                except Exception:
                    pass
                return resp
            return "File not found", 404

    if download_type == "zip":
        if "output_zip" in job and os.path.exists(job["output_zip"]):
            _do_log()
            zip_name = job.get("output_name", "audiobook.zip")
            if not zip_name.endswith(".zip"):
                 zip_name = _safe_filename(job["info"].title) + ".zip"
            return _send_file_throttled(job["output_zip"], as_attachment=True, download_name=zip_name, no_cache=True, bypass_throttle=True)
        return "ZIP file not found", 404

    # Default logic (compatibility or auto-detect)
    # Prefer M4B if it seems to be the intended primary output
    if job.get("output_name", "").endswith(".m4b") and job.get("output_m4b") and os.path.exists(job["output_m4b"]):
        _do_log()
        return _send_file_throttled(job["output_m4b"], as_attachment=True, download_name=job["output_name"], no_cache=True, bypass_throttle=True)

    if "output_zip" in job and os.path.exists(job["output_zip"]):
        _do_log()
        return _send_file_throttled(job["output_zip"], as_attachment=True, download_name=job["output_name"], no_cache=True, bypass_throttle=True)

    if job.get("output_files") and os.path.exists(job["output_files"][0]):
        _do_log()
        return _send_file_throttled(job["output_files"][0], as_attachment=True, download_name=job["output_name"], no_cache=True, bypass_throttle=True)

    return "File not found", 404

@app.route("/api/download_podcast/<job_id>")
def api_download_podcast(job_id):
    job, err, sc = _check_job_owner(job_id)
    if err is not None:
        return ("Job not found" if sc == 404 else "Forbidden"), sc
    if job.get("status") != "done":
        return "Not ready", 400
    if not job.get("podcast_ready"):
        return "Podcast not available for this job", 400

    # If RSS already embedded in output ZIP (zip_rss format), serve it directly
    if job.get("podcast_rss_included") and job.get("output_zip") and os.path.exists(job["output_zip"]):
        print(f"[{job_id}] Podcast download: serving existing ZIP with embedded RSS")
        job["last_poll"] = time.time()
        job["downloaded_at"] = time.time()
        if not _is_resume_or_probe_request():
            _log_activity(job_id, job.get("original_filename", ""), "DOWNLOAD_PODCAST",
                          job.get("client_id", ""), job.get("client_ip", ""),
                          job.get("voice", ""), job.get("browser_lang", ""))
        return _send_file_throttled(job["output_zip"], as_attachment=True,
                         download_name=job.get("output_name", "podcast.zip"), no_cache=True, bypass_throttle=True)

    # Sec (SSRF / content injection): il base_url degli <enclosure> del feed RSS è critico.
    # Se accettato dal client, un attaccante può fare pubblicare/usare il feed con enclosure
    # puntate a un host arbitrario (phishing podcast, malware audio). Forziamo il valore
    # server-side da ABM_BASE_URL e ignoriamo qualunque parametro utente.
    base_url = (BASE_URL or "").rstrip("/")
    if not base_url:
        return "Server misconfigured: ABM_BASE_URL not set", 503

    job["last_poll"] = time.time()
    job["downloaded_at"] = time.time()

    info = job["podcast_info"]
    mp3_files = job["podcast_mp3s"]
    safe_name = job["podcast_safe_name"]
    work_dir = Path(job["epub_path"]).parent

    # Build podcast ZIP on-the-fly with the user-provided base URL
    podcast_dir = work_dir / "podcast"
    podcast_dir.mkdir(exist_ok=True)
    try:
        for mp3 in mp3_files:
            if os.path.exists(mp3):
                shutil.copy2(mp3, str(podcast_dir / os.path.basename(mp3)))

        # Cover art: extract from EPUB (try Pillow for 1400px square, fallback to raw)
        cover_file = ""
        cover_path = str(podcast_dir / "cover.jpg")
        epub_path = job["epub_path"]

        # Strategy 1: Pillow resize to 1400px square (iTunes compliant)
        if _extract_cover_from_epub(epub_path, cover_path, target_size=1400):
            cover_file = "cover.jpg"
            print(f"[{job_id}] Podcast cover: Pillow 1400px ({os.path.getsize(cover_path)} bytes)")
        else:
            # Strategy 2: raw extraction via _extract_cover_for_preview (works without Pillow)
            print(f"[{job_id}] Podcast cover: _extract_cover_from_epub failed, trying raw extraction")
            raw_path, raw_mime = _extract_cover_for_preview(epub_path, str(podcast_dir))
            if raw_path and os.path.exists(raw_path):
                cover_file = os.path.basename(raw_path)
                # Rename to cover.jpg/cover.png for consistency
                ext = ".png" if raw_mime == "image/png" else ".jpg"
                final_cover = str(podcast_dir / ("cover" + ext))
                if raw_path != final_cover:
                    shutil.move(raw_path, final_cover)
                cover_file = "cover" + ext
                print(f"[{job_id}] Podcast cover: raw extraction OK ({os.path.getsize(final_cover)} bytes)")
            else:
                # Strategy 3: generate fallback cover
                print(f"[{job_id}] Podcast cover: raw extraction failed, generating fallback")
                _generate_fallback_cover(cover_path,
                                         title=info.title or "",
                                         author=info.author or "")
                if os.path.exists(cover_path) and os.path.getsize(cover_path) > 0:
                    cover_file = "cover.jpg"
                    print(f"[{job_id}] Podcast cover: fallback generated ({os.path.getsize(cover_path)} bytes)")
                else:
                    print(f"[{job_id}] Podcast cover: all strategies failed, no cover in podcast")

        rss_fname = f"{safe_name}_podcast.xml"
        rss_path = str(podcast_dir / rss_fname)
        _generate_podcast_rss(info, mp3_files, rss_path,
                              base_url=base_url, cover_filename=cover_file,
                              rss_filename=rss_fname)

        _generate_podcast_index_html(podcast_dir, info.title, info.author,
                                     cover_file, rss_fname, mp3_files,
                                     language=getattr(info, 'language', 'en'))

        # Verify ZIP contents before creating archive
        zip_contents = list(podcast_dir.iterdir())
        print(f"[{job_id}] Podcast ZIP contents: {[f.name for f in zip_contents]}")

        podcast_zip = shutil.make_archive(
            str(work_dir / f"{safe_name}_podcast"), "zip", str(podcast_dir)
        )
    finally:
        shutil.rmtree(str(podcast_dir), ignore_errors=True)

    if not _is_resume_or_probe_request():
        _log_activity(job_id, job.get("original_filename", ""), "DOWNLOAD_PODCAST",
                      job.get("client_id", ""), job.get("client_ip", ""),
                      job.get("voice", ""), job.get("browser_lang", ""))
    return _send_file_throttled(podcast_zip, as_attachment=True,
                     download_name=f"{safe_name}_podcast.zip", no_cache=True, bypass_throttle=True)


# ----------------------------------------------------------------------
# HTML TEMPLATE (i18n, upload lock, ETA)
# ----------------------------------------------------------------------

# ----------------------------------------------------------------------
# HTML TEMPLATE (assembled from modular components)
# ----------------------------------------------------------------------

# ----------------------------------------------------------------------
# SEO DATA  -  usato sia per il pre-rendering server-side che per sitemap.xml
# Mantienilo allineato con seo_data.js (che gestisce il cambio lingua client-side)
# ----------------------------------------------------------------------

_SEO_DATA = {
    "it": {
        "title":   "EPUB/PDF in Audiolibro Gratis MP3/M4B | Audiobook Maker",
        "tagline": "Convertitore Gratuito da EPUB e PDF in Audiolibro",
        "subtitle":"Converti i tuoi EPUB e PDF in audiolibri con voci neurali di alta qualità",
        "desc":    "Converti i tuoi ebook EPUB e PDF in audiolibri MP3 e M4B (con capitoli incorporati) gratis con voci AI naturali. Convertitore online gratuito text-to-speech: carica il tuo libro, scegli la voce e scarica l'audiolibro professionale. Nessuna installazione, funziona dal browser.",
        "kw":      "convertitore epub audiolibro, pdf in audiolibro, creare m4b con capitoli, audiolibro gratis online, text to speech italiano, sintesi vocale libro, ebook in audio, audiolibri per dislessia, audiolibri per ipovedenti, alternativa elevenlabs gratis, generatore podcast rss, audiobook maker",
        "ld_name": "Audiobook Maker",
        "ld_desc": "Convertitore online gratuito per trasformare ebook EPUB e PDF in audiolibri MP3 e M4B con capitoli e voci neurali TTS AI. Supporta 6 lingue, selezione capitoli e generazione feed podcast RSS.",
    },
    "en": {
        "title":   "Free EPUB/PDF to MP3 & M4B Audiobook | Audiobook Maker",
        "tagline": "Free EPUB & PDF to Audiobook Converter",
        "subtitle":"Convert your EPUBs and PDFs into audiobooks with high-quality neural voices",
        "desc":    "Convert your EPUB and PDF ebooks to MP3 or M4B audiobooks (with embedded chapters) for free with natural AI voices. Free online text-to-speech converter: upload your book, choose a voice, and download your professional audiobook. No installation needed, works in your browser.",
        "kw":      "epub to audiobook converter, pdf to audiobook, m4b with chapters, free audiobook maker, text to speech audiobook, ai audiobook generator, ebook to mp3, audiobook for dyslexia, accessible audiobook, listen to PDF, elevenlabs alternative free, podcast rss generator",
        "ld_name": "Audiobook Maker",
        "ld_desc": "Free online tool to convert EPUB and PDF ebooks into MP3 and M4B audiobooks (with chapters) using neural AI TTS voices. Supports 6 languages, chapter selection, and podcast RSS feed generation.",
    },
    "fr": {
        "title":   "EPUB/PDF en Livre Audio Gratuit MP3/M4B | Audiobook Maker",
        "tagline": "Convertisseur Gratuit EPUB & PDF en Livre Audio",
        "subtitle":"Convertissez vos EPUB et PDF en livres audio avec des voix neurales",
        "crumb":   "Convertisseur en ligne",
        "desc":    "Convertissez vos ebooks EPUB et PDF en livres audio MP3 et M4B (avec chapitres) gratuitement avec des voix IA naturelles. Convertisseur en ligne gratuit text-to-speech : téléchargez votre livre, choisissez une voix et téléchargez votre livre audio professionnel. Aucune installation, fonctionne dans le navigateur.",
        "kw":      "convertisseur epub livre audio, pdf en livre audio, créer m4b avec chapitres, livre audio gratuit en ligne, text to speech français, synthèse vocale livre, ebook en audio, livre audio dyslexie, livre audio malvoyants, alternative elevenlabs gratuit, générateur podcast rss, audiobook maker",
        "ld_name": "Audiobook Maker",
        "ld_desc": "Outil en ligne gratuit pour convertir des ebooks EPUB e PDF en livres audio MP3 avec des voix neuronales TTS IA. Prend en charge 6 langues et la génération de flux RSS podcast.",
    },
    "es": {
        "title":   "EPUB/PDF a Audiolibro Gratis MP3/M4B | Audiobook Maker",
        "tagline": "Convertidor Gratuito de EPUB y PDF a Audiolibro",
        "subtitle":"Convierte tus EPUB y PDF en audiolibros con voces neurales de alta calidad",
        "desc":    "Convierte tus ebooks EPUB y PDF en audiolibros MP3 y M4B (con capítulos incorporados) gratis con voces IA naturales. Convertidor online gratuito text-to-speech: sube tu libro, elige una voz y descarga tu audiolibro profesional. Sin instalación, funciona desde el navegador.",
        "kw":      "convertidor epub audiolibro, pdf a audiolibro, crear m4b con capítulos, audiolibro gratis online, text to speech español, síntesis de voz libro, ebook en audio, audiolibro para dislexia, audiolibro para ciegos, alternativa elevenlabs gratis, generador podcast rss, audiobook maker",
        "ld_name": "Audiobook Maker",
        "ld_desc": "Herramienta online gratuita para convertir ebooks EPUB y PDF en audiolibros MP3 con voces neuronales TTS IA. Soporta 6 idiomas y generación de feed podcast RSS.",
        "crumb":   "Convertidor online",
    },
    "de": {
        "title":   "EPUB/PDF zu Hörbuch Gratis MP3/M4B | Audiobook Maker",
        "tagline": "Kostenloser EPUB- & PDF-zu-Hörbuch-Konverter",
        "subtitle":"Konvertieren Sie EPUBs und PDFs in Hörbücher mit neuronalen Stimmen",
        "desc":    "Konvertieren Sie Ihre EPUB- und PDF-E-Books kostenlos in MP3- und M4B-Hörbücher (mit eingebetteten Kapiteln) mit natürlichen KI-Stimmen. Kostenloser Online Text-to-Speech Konverter: Laden Sie Ihr Buch hoch, wählen Sie eine Stimme und laden Sie Ihr professionelles Hörbuch herunter. Keine Installation nötig, funktioniert im Browser.",
        "kw":      "epub zu hörbuch konverter, pdf zu hörbuch, m4b mit kapiteln erstellen, hörbuch erstellen kostenlos, text to speech deutsch, sprachsynthese buch, ebook in audio umwandeln, hörbuch für legasthenie, barrierefreies hörbuch, elevenlabs alternative kostenlos, podcast rss generator, audiobook maker",
        "ld_name": "Audiobook Maker",
        "ld_desc": "Kostenloses Online-Tool zum Konvertieren von EPUB- und PDF-E-Books in MP3-Hörbücher mit neuronalen KI-TTS-Stimmen. Unterstützt 6 Sprachen und Podcast-RSS-Feed-Generierung.",
        "crumb":   "Online-Konverter",
    },
    "zh": {
        "title":   "免费在线EPUB/PDF转MP3/M4B有声书转换器 | Audiobook Maker",
        "tagline": "免费EPUB和PDF转有声书转换器",
        "subtitle":"使用高品质神经网络AI语音将EPUB和PDF电子书免费转换为有声读物",
        "desc":    "在您的浏览器中免费、安全、快速地将EPUB和PDF电子书转换为高质量MP3或M4B（含章节）有声读物。由AI神经网络语音驱动。无需安装，支持章节选择和专业M4B格式输出。免费在线文字转语音转换器，支持50多种语言。",
        "kw":      "epub转有声书, pdf转有声书, m4b章节制作, 免费有声书制作, 免费在线文字转语音, 文字转语音有声书, 电子书转mp3, 电子书转音频, ai语音朗读, 阅读障碍有声书, 无障碍有声书, 有声书转换器, elevenlabs替代品, 播客rss生成器, audiobook maker",
        "ld_name": "Audiobook Maker",
        "ld_desc": "免费在线工具，利用神经网络AI文字转语音技术将EPUB和PDF电子书转换为MP3有声书。支持6种语言、章节选择和播客RSS订阅生成。",
        "crumb":   "在线转换器",
    },
    "hi": {
        "title":   "मुफ़्त EPUB/PDF से MP3/M4B ऑडियोबुक कनवर्टर | Audiobook Maker",
        "tagline": "मुफ़्त EPUB और PDF से ऑडियोबुक कनवर्टर",
        "subtitle":"उच्च गुणवत्ता वाली न्यूरल आवाज़ों के साथ अपनी EPUB और PDF को ऑडियोबुक में बदलें",
        "desc":    "अपनी EPUB और PDF ईबुक को प्राकृतिक AI आवाज़ों के साथ मुफ़्त में MP3 या M4B ऑडियोबुक (अध्यायों के साथ) में बदलें. मुफ़्त ऑनलाइन टेक्स्ट-टू-स्पीच कनवर्टर: अपनी पुस्तक अपलोड करें, एक आवाज़ चुनें, और अपनी पेशेवर ऑडियोबुक डाउनलोड करें. कोई इंस्टॉलेशन नहीं, ब्राउज़र में काम करता है.",
        "kw":      "epub से ऑडियोबुक कनवर्टर, pdf से ऑडियोबुक, अध्यायों के साथ m4b, मुफ़्त ऑडियोबुक मेकर, टेक्स्ट टू स्पीच ऑडियोबुक, ai ऑडियोबुक जनरेटर, ebook से mp3, डिस्लेक्सिया के लिए ऑडियोबुक, सुलभ ऑडियोबुक, elevenlabs मुफ़्त विकल्प, पॉडकास्ट rss जनरेटर, audiobook maker",
        "ld_name": "Audiobook Maker",
        "ld_desc": "EPUB और PDF ईबुक को न्यूरल AI TTS आवाज़ों के साथ MP3 और M4B ऑडियोबुक (अध्यायों के साथ) में बदलने के लिए मुफ़्त ऑनलाइन टूल. 7 भाषाओं, अध्याय चयन और पॉडकास्ट RSS फ़ीड जनरेशन का समर्थन.",
        "crumb":   "ऑनलाइन कनवर्टर",
    },
}

_SUPPORTED_LANGS = list(_SEO_DATA.keys())  # ['it', 'en', 'fr', 'es', 'de', 'zh', 'hi']

# Pre-rendering: una copia HTML per lingua, pronta a startup.
# Nessun costo a request-time; ogni risposta è un semplice return di stringa.
HTML_TEMPLATES: dict[str, str] = {
    lang: build_html_template(
        lang=lang,
        seo=seo,
        base_url=BASE_URL,
        version=__version__,
        updated_date=get_formatted_date(),
    )
    for lang, seo in _SEO_DATA.items()
}
# Fallback generico per URL sconosciuti
HTML_TEMPLATE = HTML_TEMPLATES["en"]

# Template dedicati per la root (/): canonical punta a BASE_URL/ (se stesso),
# non a /{lang}/. Risolve l'errore SEO "hreflang URL non usa il proprio canonical".
# Google crawla / e vede canonical=/, che corrisponde all'x-default negli hreflang.
HTML_ROOT_TEMPLATES: dict[str, str] = {
    lang: build_html_template(
        lang=lang,
        seo=seo,
        base_url=BASE_URL,
        version=__version__,
        canonical_url=f"{BASE_URL}/" if BASE_URL else "",
        updated_date=get_formatted_date(),
    )
    for lang, seo in _SEO_DATA.items()
}


def _detect_lang_from_request() -> str:
    """Rileva la lingua preferita dall'header Accept-Language del browser.

    Scorre i tag di qualità q= e restituisce la prima lingua supportata.
    Fallback: 'en'.
    """
    accept = request.headers.get("Accept-Language", "en")
    # Formato: "it-IT,it;q=0.9,en-US;q=0.8,en;q=0.7"
    tags = re.findall(r'([a-zA-Z]{2,3})(?:-[a-zA-Z0-9]+)*(?:;q=([0-9.]+))?', accept)
    # Ordina per q (default 1.0)
    ranked = sorted(tags, key=lambda t: float(t[1]) if t[1] else 1.0, reverse=True)
    for lang_tag, _ in ranked:
        lang = lang_tag.lower()
        if lang in _SUPPORTED_LANGS:
            return lang
    return "en"


@app.after_request
def _set_client_cookie(response):
    """Ensure every response carries the abm_cid cookie for client tracking.

    Sec: il cookie è il bearer di ownership su tutti gli endpoint job (vedi _check_job_owner).
    - secure=True quando la request è HTTPS (anche dietro reverse proxy) per impedire
      interception in chiaro su navigazioni HTTP plain verso lo stesso host.
    - samesite='Lax' mantenuto: il cookie deve essere inviato sui link cliccati dalle
      email di notifica (/dl/<token>) che possono provenire da altri domini.
    """
    if _CLIENT_COOKIE_NAME not in request.cookies:
        cid = str(uuid.uuid4())[:12]
        is_https = (request.scheme == "https") or (request.headers.get("X-Forwarded-Proto", "") == "https")
        response.set_cookie(
            _CLIENT_COOKIE_NAME, cid,
            max_age=_CLIENT_COOKIE_MAX_AGE,
            httponly=True,
            secure=is_https,
            samesite="Lax",
        )
    return response




# ----------------------------------------------------------------------
# AUTO-CLEANUP (deletes EPUB/PDF/TXT + MP3 files)
# ----------------------------------------------------------------------

# Regole di cancellazione:
# 1. Browser chiuso senza email registrata  →  cancella (heartbeat perso per 60s)
# 2. Utente scarica direttamente dall'UI web  →  cancella subito dopo download
# 3. Email di notifica inviata  →  mantieni 24h dall'invio, poi cancella
# 4. Job in errore o cancellato  →  cancella subito
# 5. Cartelle orfane su disco (non in jobs né in tokens)  →  cancella

CLEANUP_GRACE_AFTER_DOWNLOAD_SEC = 5 * 60  # 5 min grazia dopo download diretto
CLEANUP_HEARTBEAT_TIMEOUT_SEC = 60          # heartbeat perso per 60s = browser chiuso
CLEANUP_INTERVAL_SEC = 60                   # check every 60 seconds
CLEANUP_ORPHAN_DIR_AGE_SEC = 2 * 60 * 60   # cartelle orfane > 2h vengono rimosse

# Marker scritto nella job dir per proteggere la cartella dal cleanup orfani
# di OGNI worker Gunicorn. Il filesystem è l'unica fonte autoritativa condivisa.
# Due stati possibili nel contenuto del file:
#   - "pending"  → email registrata, lavorazione in corso. Protezione illimitata
#                  fino al cap EMAIL_PENDING_MAX_AGE_SEC (anti-orphan se worker
#                  crasha durante un job lungo).
#   - "<float>"  → email inviata al timestamp indicato. Protezione per
#                  EMAIL_FILE_RETENTION_SEC + 300s da quel timestamp.
EMAIL_MARKER_FILENAME = ".email_sent"
_EMAIL_MARKER_PENDING = "pending"
EMAIL_PENDING_MAX_AGE_SEC = 48 * 3600  # cap di sicurezza se la lavorazione si interrompe senza email

# Marker scritto in work_dir dei job Gemini falliti con refund per consentire
# analisi forense post-mortem. Contiene JSON {retain_until, created_at, kind,
# outcome, reason, job_id, days}. Sopravvive a restart del service e blocca
# TUTTI i branch di cleanup (status=error, orphan dir, token-orphan, orphan
# output) finché now < retain_until. Retention configurabile via
# ABM_GEMINI_FORENSIC_RETENTION_DAYS (default 7; 0 = disabilita).
FORENSIC_MARKER_FILENAME = ".forensic_retain.json"
try:
    FORENSIC_RETENTION_DAYS = int(os.environ.get("ABM_GEMINI_FORENSIC_RETENTION_DAYS", "7"))
except (TypeError, ValueError):
    FORENSIC_RETENTION_DAYS = 7
FORENSIC_RETENTION_DAYS = max(0, FORENSIC_RETENTION_DAYS)


def _marker_protection_window(is_gemini):
    """Finestra di protezione (sec) da incidere nel marker email, per tipo voce.

    - Gemini (True): finestra estesa no-download (prudenza massima sui file
      PREMIUM, mai cancellati prima della finestra completa).
    - Standard (False): retention base email (18h): la dir standard NON va
      sovra-protetta con la finestra Gemini.
    - Ignoto (None): ritorna None → il lettore usa il fallback conservativo
      max() (favorisce sempre la conservazione del file generato)."""
    if is_gemini is True:
        return GEMINI_FILE_RETENTION_SEC * GEMINI_NO_DOWNLOAD_RETENTION_MULTIPLIER
    if is_gemini is False:
        return EMAIL_FILE_RETENTION_SEC
    return None


def _write_email_marker(work_dir, when=None, is_gemini=None):
    """Marca una job dir come 'email inviata' (timestamp epoch in secondi).
    Sovrascrive un eventuale marker 'pending'. Idempotente.

    Formato del contenuto:
      - "<ts>"                  → legacy, lettura con fallback conservativo max().
      - "<ts>|<retention_sec>"  → self-describing: il cleanup applica la finestra
                                  CORRETTA per tipo voce senza sovra-proteggere le
                                  dir standard con la finestra Gemini (96h).
    Quando `is_gemini` è None (tipo voce ignoto) si scrive la forma legacy: il
    lettore protegge in modo conservativo. La forma self-describing si usa solo
    quando il tipo voce è noto con certezza."""
    try:
        ts = float(when) if when is not None else time.time()
        marker = Path(work_dir) / EMAIL_MARKER_FILENAME
        marker.parent.mkdir(parents=True, exist_ok=True)
        window = _marker_protection_window(is_gemini)
        if window is not None:
            marker.write_text(f"{ts:.3f}|{int(window)}", encoding="utf-8")
        else:
            marker.write_text(f"{ts:.3f}", encoding="utf-8")
    except OSError as e:
        print(f"[email-marker] write failed in {work_dir}: {e}")


def _write_email_pending_marker(work_dir):
    """Marca una job dir come 'email registrata, lavorazione in corso'.
    Non sovrascrive un marker timestamp già presente (email già inviata).
    Non riscrive se è già 'pending' (preserva mtime → cap age coerente)."""
    try:
        marker = Path(work_dir) / EMAIL_MARKER_FILENAME
        marker.parent.mkdir(parents=True, exist_ok=True)
        if marker.exists():
            try:
                existing = marker.read_text(encoding="utf-8").strip()
                if existing == _EMAIL_MARKER_PENDING:
                    return
                # Se è un timestamp valido (legacy "<ts>" o self-describing
                # "<ts>|<win>"), l'email è già stata inviata: NON degradare a
                # 'pending' (ridurrebbe la finestra di protezione del file).
                float(existing.partition("|")[0])
                return
            except (OSError, ValueError):
                pass  # contenuto illeggibile/corrotto: sovrascriviamo
        marker.write_text(_EMAIL_MARKER_PENDING, encoding="utf-8")
    except OSError as e:
        print(f"[email-marker] pending write failed in {work_dir}: {e}")


def _email_marker_protects(work_dir, now):
    """True se il marker email protegge la dir.
    Pending: protetto se mtime entro EMAIL_PENDING_MAX_AGE_SEC.
    Timestamp: protetto se entro max(EMAIL_FILE_RETENTION_SEC,
    GEMINI_FILE_RETENTION_SEC * GEMINI_NO_DOWNLOAD_RETENTION_MULTIPLIER) + 300s.
    Usiamo il max perche' il marker su disco non conosce voice/downloaded_at del job;
    moltiplichiamo per il fattore no-download per non cancellare un Gemini job dir
    mai scaricato prima della finestra estesa (default 96h)."""
    marker = Path(work_dir) / EMAIL_MARKER_FILENAME
    try:
        if not marker.exists():
            return False
    except OSError:
        return False
    try:
        content = marker.read_text(encoding="utf-8").strip()
    except OSError:
        return False
    if content == _EMAIL_MARKER_PENDING:
        try:
            mtime = marker.stat().st_mtime
        except OSError:
            return False
        return (now - mtime) < EMAIL_PENDING_MAX_AGE_SEC
    # Forma self-describing "<ts>|<retention_sec>": usa la finestra esplicita.
    # Legacy "<ts>" o contenuto corrotto: fallback conservativo al max()
    # (favorisce SEMPRE la conservazione del file generato).
    ts_str, sep, win_str = content.partition("|")
    try:
        ts = float(ts_str)
    except ValueError:
        try:
            ts = marker.stat().st_mtime
        except OSError:
            return False
    window = None
    if sep:
        try:
            window = int(win_str)
        except ValueError:
            window = None
    if window is None or window <= 0:
        window = max(
            EMAIL_FILE_RETENTION_SEC,
            GEMINI_FILE_RETENTION_SEC * GEMINI_NO_DOWNLOAD_RETENTION_MULTIPLIER,
        )
    return (now - ts) < window + 300


def _forensic_marker_protects(work_dir, now):
    """True se il marker forense protegge la work_dir dal cleanup.
    Scritto da generation_engine al refund per job Gemini falliti; sopravvive
    a restart e blocca tutti i branch di cleanup finché now < retain_until.
    """
    marker = Path(work_dir) / FORENSIC_MARKER_FILENAME
    try:
        if not marker.exists():
            return False
    except OSError:
        return False
    try:
        with marker.open("r", encoding="utf-8") as f:
            data = json.load(f)
        retain_until = float(data.get("retain_until", 0) or 0)
    except (OSError, ValueError, TypeError):
        return False
    return now < retain_until


def _cleanup_job(job_id, reason=""):
    """Remove all files for a job and delete the job entry.
    NOTA: nessun gate marker qui — questo path viene invocato solo dal branch
    per-status che opera su `jobs` locali con info complete (cancelled/error/
    done+retention-scaduta). La protezione cross-worker è nei branch orfani.

    Gate forense: se la work_dir contiene `.forensic_retain.json` valido
    (refund Gemini in attesa di analisi admin), rimuoviamo l'entry in memoria
    ma preserviamo la dir su disco finché il marker è valido.
    """
    with _jobs_lock:
        jobs.pop(job_id, None)
    work_dir = UPLOAD_DIR / job_id
    if work_dir.exists():
        if _forensic_marker_protects(work_dir, time.time()):
            print(f"[cleanup] {job_id} entry removed but dir preserved "
                  f"(forensic retention) — {reason}")
            return
        shutil.rmtree(str(work_dir), ignore_errors=True)
    print(f"[cleanup] {job_id} removed ({reason})")


def _cleanup_loop():
    """Background thread: periodically clean up finished/abandoned jobs."""
    while True:
        time.sleep(CLEANUP_INTERVAL_SEC)
        now = time.time()

        # Multi-worker: absorb tokens created by other workers before deciding
        # what to delete. Without this, this worker's view of _download_tokens
        # misses peers' tokens and the orphan-dir branch wipes their job dirs.
        _merge_tokens_from_disk()

        with _jobs_lock:
            to_remove = []
            for jid, job in list(jobs.items()):
                status = job.get("status", "")
                # Un job e' "emailato" se l'utente ha registrato l'email OPPURE
                # se una notifica e' comunque partita (path fallback post-COMPLETE
                # che imposta email_sent_at senza email_registered). Senza questo
                # OR, un job consegnato via email-fallback ricadeva nella regola
                # "downloaded + 5 min" e veniva cancellato pochi minuti dopo il
                # primo accesso, ignorando la retention 18h/48h.
                has_email = bool(job.get("email_registered")) or bool(job.get("email_sent_at"))

                if status == "cancelled":
                    # Non rimuovere job con download token ancora attivi
                    # (es. email inviata prima del cancel con link validi).
                    if not _has_active_download_tokens(jid, now):
                        to_remove.append((jid, "cancelled"))
                    continue

                if status == "error":
                    start = job.get("start_time", now)
                    if (now - start) > 120:
                        to_remove.append((jid, "error"))
                    continue

                if status == "analyzed":
                    last_poll = job.get("last_poll", job.get("start_time", now))
                    if (now - last_poll) > CLEANUP_HEARTBEAT_TIMEOUT_SEC * 30:
                        # Non rimuovere job con download token ancora attivi
                        # (es. email inviata in generazione precedente).
                        if not _has_active_download_tokens(jid, now):
                            to_remove.append((jid, "stale analyzed"))
                    continue

                if status == "optimizing":
                    if has_email:
                        continue
                    last_poll = job.get("last_poll", job.get("opt_start_time", now))
                    if (now - last_poll) > CLEANUP_HEARTBEAT_TIMEOUT_SEC:
                        job["opt_cancelled"] = True
                        to_remove.append((jid, f"heartbeat lost during optimization ({int(now - last_poll)}s)"))
                    continue

                if status == "optimized":
                    opt_done = job.get("opt_completed_at") or job.get("email_sent_at") or now
                    # Effective retention: per voci PREMIUM senza alcun download
                    # del .abm, raddoppia il timer.
                    _ret = _effective_retention_for_job(job)
                    if (now - opt_done) > _ret:
                        h = _ret // 3600
                        reason = ("optimization email retention expired" if has_email
                                  else f"optimized project retention expired ({h}h)")
                        to_remove.append((jid, reason))
                    continue

                if status == "generating":
                    if has_email:
                        continue
                    last_poll = job.get("last_poll", job.get("start_time", now))
                    if (now - last_poll) > CLEANUP_HEARTBEAT_TIMEOUT_SEC:
                        job["cancelled"] = True
                        to_remove.append((jid, f"heartbeat lost during generation ({int(now - last_poll)}s)"))
                    continue

                if status == "done":
                    dl_at = job.get("downloaded_at")
                    email_sent_at = job.get("email_sent_at")
                    last_poll = job.get("last_poll", 0)

                    if has_email and email_sent_at:
                        # Effective retention: voci PREMIUM senza download → 2x.
                        if (now - email_sent_at) > _effective_retention_for_job(job):
                            to_remove.append((jid, f"email retention expired ({int(now - email_sent_at)}s)"))
                        continue

                    if has_email and not email_sent_at:
                        continue

                    if dl_at:
                        if (now - dl_at) > CLEANUP_GRACE_AFTER_DOWNLOAD_SEC:
                            to_remove.append((jid, f"downloaded {int(now - dl_at)}s ago"))
                        continue

                    if last_poll and (now - last_poll) > CLEANUP_HEARTBEAT_TIMEOUT_SEC:
                        to_remove.append((jid, f"abandoned (heartbeat lost {int(now - last_poll)}s)"))
                        continue

        # File I/O outside lock
        for jid, reason in to_remove:
            try:
                _cleanup_job(jid, reason)
            except Exception as e:
                print(f"[cleanup] error removing {jid}: {e}")

        #  -  -  Cleanup expired download tokens  -  -
        # Retention per-token: se il token e' marcato is_gemini (voce PREMIUM)
        # vale GEMINI_FILE_RETENTION_SEC, altrimenti EMAIL_FILE_RETENTION_SEC.
        # _effective_* raddoppia per voci PREMIUM mai scaricate (protezione costo).
        with _tokens_lock:
            expired_tokens = [(t, info) for t, info in _download_tokens.items()
                              if (now - info["created_at"]) > _effective_retention_for_token_info(info) + 300]
        for t, t_info in expired_tokens:
            with _tokens_lock:
                _download_tokens.pop(t, None)
            jid = t_info.get("job_id", "")
            with _jobs_lock:
                job_in_memory = jid in jobs
            if jid:
                job_dir = UPLOAD_DIR / jid
                # Legacy: tokens created before the per-epoch refactor may have an
                # `output_archive_dir` field pointing to a manually-built archive.
                archive_rel = t_info.get("output_archive_dir", "")
                if archive_rel and job_dir.exists():
                    archive_path = job_dir / archive_rel
                    if archive_path.exists() and archive_path.is_dir():
                        shutil.rmtree(str(archive_path), ignore_errors=True)
                        print(f"[cleanup] Legacy archive removed (token expired): {archive_path}")
                if not job_in_memory and job_dir.exists():
                    if _email_marker_protects(job_dir, now):
                        continue
                    if _forensic_marker_protects(job_dir, now):
                        continue
                    shutil.rmtree(str(job_dir), ignore_errors=True)
                    print(f"[cleanup] Token-orphan dir removed: {jid}")
        if expired_tokens:
            # _save_tokens() acquires _tokens_lock internally; wrapping it here
            # would deadlock on the non-reentrant lock and freeze every later
            # caller (including the post-COMPLETE email notification).
            _save_tokens()

        #  -  -  Cleanup orphan per-epoch output dirs  -  -
        # An output_{epoch}/ directory is removable when:
        # - It is NOT the current output_dir of the job (if job is alive)
        # - AND it is not referenced by any active token's output_zip/output_file/output_m4b
        # - AND its mtime is older than the retention window
        with _jobs_lock:
            current_output_dirs = {jobs[j].get("output_dir", ""): j for j in jobs}
        with _tokens_lock:
            referenced_paths = set()
            for info in _download_tokens.values():
                for key in ("output_zip", "output_file", "output_m4b"):
                    p = info.get(key) or ""
                    if p:
                        try:
                            referenced_paths.add(str(Path(p).parent.resolve()))
                        except OSError:
                            pass
        try:
            for jdir in UPLOAD_DIR.iterdir():
                if not jdir.is_dir() or jdir.name.startswith("_"):
                    continue
                for od in jdir.iterdir():
                    if not od.is_dir():
                        continue
                    if not (od.name == "output" or od.name.startswith("output_")):
                        continue
                    try:
                        od_resolved = str(od.resolve())
                    except OSError:
                        continue
                    if od_resolved in current_output_dirs:
                        continue
                    if od_resolved in referenced_paths:
                        continue
                    try:
                        age = now - od.stat().st_mtime
                    except OSError:
                        continue
                    # Senza contesto-job, usiamo la retention piu' lunga (Gemini)
                    # moltiplicata per il fattore no-download: la dir orfana puo'
                    # appartenere a un job PREMIUM mai scaricato.
                    if age > max(
                        EMAIL_FILE_RETENTION_SEC,
                        GEMINI_FILE_RETENTION_SEC * GEMINI_NO_DOWNLOAD_RETENTION_MULTIPLIER,
                    ):
                        if _email_marker_protects(od.parent, now):
                            continue
                        if _forensic_marker_protects(od.parent, now):
                            continue
                        shutil.rmtree(str(od), ignore_errors=True)
                        print(f"[cleanup] Orphan output dir removed: {od} (age: {int(age)}s)")
        except OSError:
            pass

        #  -  -  Cleanup cartelle orfane su disco  -  -
        with _jobs_lock:
            _known_job_ids = set(jobs.keys())
        with _tokens_lock:
            _known_token_jobs = set(info.get("job_id", "") for info in _download_tokens.values())
        _all_known = _known_job_ids | _known_token_jobs
        try:
            for entry in UPLOAD_DIR.iterdir():
                if not entry.is_dir():
                    continue
                if entry.name.startswith("_"):
                    continue
                if entry.name in _all_known:
                    continue
                try:
                    dir_age = now - entry.stat().st_mtime
                except OSError:
                    continue
                if dir_age > CLEANUP_ORPHAN_DIR_AGE_SEC:
                    if _email_marker_protects(entry, now):
                        continue
                    if _forensic_marker_protects(entry, now):
                        continue
                    shutil.rmtree(str(entry), ignore_errors=True)
                    print(f"[cleanup] Orphan dir removed: {entry.name} (age: {int(dir_age)}s)")
        except OSError:
            pass

        # Flush pending admin digest (rate-limited: max 1/hour)
        _try_send_admin_digest()


# ----------------------------------------------------------------------
# ENTRY POINT
# ----------------------------------------------------------------------

# Startup: load persisted download tokens, init DeepSeek, start background threads
# (works both under __main__ and Gunicorn)
_load_tokens()
_load_payments()
_load_vouchers()
payment._migrate_paid_opt_to_paid_jobs()
payment._load_paid_jobs_done()
payment._recover_orphaned_voucher_charges(jobs)

# Load persisted client_id → email mapping for cross-job notification fallback
_load_client_emails()

# Configura il motore di generazione (spostato in generation_engine.py)
generation_engine.configure(
    jobs=jobs,
    upload_dir=UPLOAD_DIR,
    download_tokens=_download_tokens,
    save_tokens_fn=_save_tokens,
    log_activity_fn=_log_activity,
    google_tts_module=google_tts,
    invalidate_voices_cache_fn=_invalidate_voices_cache,
    jobs_lock=_jobs_lock,
    retention_sec=EMAIL_FILE_RETENTION_SEC,
    gemini_retention_sec=GEMINI_FILE_RETENTION_SEC,
    write_email_marker_fn=_write_email_marker,
    lookup_client_email_fn=_lookup_client_email
)

if _paypal_available():
    print(f"[startup] PayPal payment enabled (mode: {PAYPAL_MODE}, "
          f"rate: {LLM_RATE_EUR_PER_MCHAR} EUR/Mchar, threshold: {LLM_FREE_THRESHOLD_EUR} EUR)")
else:
    print(f"[startup] PayPal payment disabled (ABM_PAYPAL_CLIENT_ID/SECRET not set)")
_cleanup_started = False

def _google_tts_reconcile_loop():
    """Thread di background: riconcilia il contatore Google TTS con Cloud Monitoring
    ogni GOOGLE_TTS_RECONCILE_INTERVAL_SEC secondi (default 30 minuti).
    Le metriche di Cloud Monitoring hanno latenza ~5 min, quindi un intervallo
    inferiore non porta beneficio."""
    if google_tts is None:
        return
    # Attesa iniziale per non sovraccaricare lo startup
    time.sleep(60)
    while True:
        try:
            if google_tts.is_available():
                result = google_tts.reconcile_with_cloud_monitoring()
                if result is None:
                    # Monitoring non disponibile: smettiamo di provarci
                    print("[google-tts] Reconcile loop: monitoring unavailable, stopping")
                    return
        except Exception as e:
            print(f"[google-tts] Reconcile loop error: {e}")
        time.sleep(GOOGLE_TTS_RECONCILE_INTERVAL_SEC)


GOOGLE_TTS_RECONCILE_INTERVAL_SEC = int(os.environ.get("ABM_GOOGLE_TTS_RECONCILE_INTERVAL", "1800"))


def _ensure_background_threads():
    global _cleanup_started
    if _cleanup_started:
        return
    _cleanup_started = True
    threading.Thread(target=get_voices, daemon=True).start()
    threading.Thread(target=_cleanup_loop, daemon=True).start()
    if google_tts is not None:
        threading.Thread(target=_google_tts_reconcile_loop, daemon=True).start()
    
    # Verifica dipendenze audio (ffmpeg/ffprobe) per formato M4B
    ffmpeg_ok, ffprobe_ok = _check_audio_dependencies()
    if not ffmpeg_ok or not ffprobe_ok:
        missing = []
        if not ffmpeg_ok: missing.append("ffmpeg")
        if not ffprobe_ok: missing.append("ffprobe")
        print(f"WARNING: Missing critical audio dependencies: {', '.join(missing)}. "
              "M4B generation and audio duration detection will be disabled.", file=sys.stderr)
    else:
        print("[startup] Audio dependencies (ffmpeg/ffprobe) found.")

    print(f"[startup] Background threads started (data dir: {UPLOAD_DIR})")
    print(f"[startup] Max concurrent per client: {MAX_CONCURRENT_PER_CLIENT}")
    print(f"[startup] Max concurrent LLM per client: {MAX_CONCURRENT_LLM_PER_CLIENT}")
    if _llm_available():
        print(f"[startup] LLM text optimization enabled (Model: {LLM_MODEL})")
    if ADMIN_EMAIL:
        print(f"[startup] Admin digest enabled  ->  {ADMIN_EMAIL} (interval: {ADMIN_DIGEST_INTERVAL_SEC}s)")
    else:
        print("[startup] Admin digest disabled (ABM_ADMIN_EMAIL not set)")

_init_log_dedup()
_ensure_background_threads()

if __name__ == "__main__":
    PORT = int(os.environ.get("ABM_PORT", "5601"))
    DEBUG = os.environ.get("ABM_DEBUG", "0").strip().lower() in ("1", "true", "yes", "on")
    print(f"\n{'='*50}")
    print(f"  Audiobook Maker v{__version__}")
    print(f"  http://localhost:{PORT}")
    print(f"{'='*50}")
    print(f"  Script folder: {SCRIPT_DIR}")
    print(f"  Data folder:   {UPLOAD_DIR}")
    print(f"  Activity log:  {SCRIPT_DIR / 'activity_YYYY-MM.log'}")
    _max_text_chars_startup = os.environ.get("ABM_MAX_TEXT_CHARS", "1500000")
    print(f"  ABM_MAX_TEXT_CHARS: {_max_text_chars_startup} "
          f"({'env' if 'ABM_MAX_TEXT_CHARS' in os.environ else 'default'})")
    _max_gemini_text_chars_startup = os.environ.get("ABM_MAX_GEMINI_TEXT_CHARS", "800000")
    print(f"  ABM_MAX_GEMINI_TEXT_CHARS: {_max_gemini_text_chars_startup} "
          f"({'env' if 'ABM_MAX_GEMINI_TEXT_CHARS' in os.environ else 'default'})")
    _job_retention_startup = os.environ.get("ABM_JOB_RETENTION_SEC", "64800")
    print(f"  ABM_JOB_RETENTION_SEC: {_job_retention_startup}s "
          f"(~{int(_job_retention_startup)//3600}h) "
          f"({'env' if 'ABM_JOB_RETENTION_SEC' in os.environ else 'default'})")
    _gemini_retention_startup = os.environ.get("ABM_GEMINI_JOB_RETENTION_SEC", "172800")
    print(f"  ABM_GEMINI_JOB_RETENTION_SEC: {_gemini_retention_startup}s "
          f"(~{int(_gemini_retention_startup)//3600}h) "
          f"({'env' if 'ABM_GEMINI_JOB_RETENTION_SEC' in os.environ else 'default'})")
    print(f"  Debug mode: {DEBUG} "
          f"({'env ABM_DEBUG' if 'ABM_DEBUG' in os.environ else 'default off'})")
    print(f"{'='*50}\n")
    app.run(host="127.0.0.1", port=PORT, debug=DEBUG)
